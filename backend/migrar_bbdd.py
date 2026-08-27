"""
Script de migración: BBDD.xlsx -> MongoDB (ikastxiki)
Importa TODOS los jugadores (421), crea equipos a partir de "EQUIPO 25&26",
crea/deduplica familias, y crea registros de pago a partir de las cuotas.

Uso (en el VPS, dentro del venv del backend):
    cd /var/www/ikastxiki/backend
    source venv/bin/activate
    python3 /ruta/al/migrar_bbdd.py /ruta/al/base_de_datos_20251203__1_.xlsx

Por defecto hace un DRY RUN (no escribe nada, solo muestra el resumen).
Para escribir de verdad en MongoDB:
    python3 migrar_bbdd.py archivo.xlsx --commit
"""
import sys
import argparse
import uuid
from datetime import datetime, date, timezone

import openpyxl
from pymongo import MongoClient

MONGO_URL = "mongodb://127.0.0.1:27017"
DB_NAME = "ikastxiki"

CATEGORIES = [
    {"name": "Prebenjamín", "min_age": 6, "max_age": 7},
    {"name": "Benjamín", "min_age": 8, "max_age": 9},
    {"name": "Alevín", "min_age": 10, "max_age": 11},
    {"name": "Infantil", "min_age": 12, "max_age": 13},
    {"name": "Cadete", "min_age": 14, "max_age": 15},
    {"name": "Juvenil", "min_age": 16, "max_age": 18},
]


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def new_id():
    return str(uuid.uuid4())


def compute_category(birthdate_iso):
    if not birthdate_iso:
        return None
    try:
        bd = datetime.fromisoformat(birthdate_iso).date()
    except Exception:
        return None
    today = date.today()
    season_year = today.year if today.month >= 7 else today.year - 1
    age = season_year - bd.year
    for c in CATEGORIES:
        if c["min_age"] <= age <= c["max_age"]:
            return c["name"]
    if age < 6:
        return "Querubín"
    return "Senior"


def s(v):
    """String limpio o None."""
    if v is None:
        return None
    v = str(v).strip()
    return v if v else None


def to_iso_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel_path")
    ap.add_argument("--commit", action="store_true", help="Escribir de verdad en MongoDB (sin esto es dry-run)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.excel_path)
    ws = wb["BBDD"]
    headers = [c.value for c in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    def col(name):
        return headers.index(name)

    idx = {
        "nombre": col("NOMBRE"),
        "apellidos": col("APELLIDOS"),
        "fecha_nac": col("FECHA DE NACIMIENTO"),
        "centro": col("CENTRO DE PROCEDENCIA"),
        "padre_nombre": col("NOMBRE DEL PADRE"),
        "padre_tel": col("TELÉFONO DEL PADRE"),
        "madre_nombre": col("NOMBRE DE LA MADRE"),
        "madre_tel": col("TELÉFONO DE LA MADRE"),
        "direccion": col("DIRECCIÓN"),
        "email_aita": col("CORREO ELECTRÓNICO AITA"),
        "email_ama": col("CORREO ELECTRÓNICO AMA"),
        "equipo": col("EQUIPO 25&26"),
        "categoria_juego": col("CATEGORÍA DE JUEGO"),
        "categoria_raw": col("CATEGORÍA"),
        "dorsal": col("DORSAL 25&26"),
        "talla_camiseta": col("TALLA 25&26"),
        "talla_medias": col("TALLA MEDIAS 25&26"),
        "demarcacion": col("DEMARCACIÓN"),
        "cuota_a_pagar": col("CUOTA A PAGAR"),
        "cuota_pagada": col("CUOTA PAGADA"),
        "cuota_pendiente": col("CUOTA PENDIENTE DE PAGO"),
        "observ_interes": col("BESTE DATU INTERESGARRIAK/ OTROS DATOS DE INTERÉS"),
    }

    teams_map = {}     # nombre_equipo -> team doc
    families_map = {}  # clave dedupe -> family doc
    players = []
    payments = []

    skipped_no_name = 0

    for r in rows:
        nombre = s(r[idx["nombre"]])
        if not nombre:
            skipped_no_name += 1
            continue

        apellidos = s(r[idx["apellidos"]]) or ""
        fecha_nac = to_iso_date(r[idx["fecha_nac"]])
        equipo_nombre = s(r[idx["equipo"]])
        estado = "activo" if equipo_nombre else "baja"

        # --- Equipo ---
        equipo_id = None
        if equipo_nombre:
            if equipo_nombre not in teams_map:
                teams_map[equipo_nombre] = {
                    "id": new_id(),
                    "nombre": equipo_nombre,
                    "categoria": s(r[idx["categoria_juego"]]) or s(r[idx["categoria_raw"]]),
                    "temporada": "2025-2026",
                    "entrenador": None,
                    "segundo_entrenador": None,
                    "delegado": None,
                    "dias_entrenamiento": None,
                    "horario": None,
                    "campo": None,
                    "limite_jugadores": 20,
                    "estado": "activo",
                    "created_at": now_iso(),
                    "updated_at": now_iso(),
                }
            equipo_id = teams_map[equipo_nombre]["id"]

        # --- Familia (dedupe por teléfono del padre, si no por teléfono de la madre) ---
        padre_tel = s(r[idx["padre_tel"]])
        madre_tel = s(r[idx["madre_tel"]])
        dedupe_key = padre_tel or madre_tel or f"sin-telefono-{nombre}-{apellidos}"

        if dedupe_key not in families_map:
            families_map[dedupe_key] = {
                "id": new_id(),
                "progenitor1_nombre": s(r[idx["padre_nombre"]]),
                "progenitor1_telefono": padre_tel,
                "progenitor1_email": s(r[idx["email_aita"]]),
                "progenitor2_nombre": s(r[idx["madre_nombre"]]),
                "progenitor2_telefono": madre_tel,
                "progenitor2_email": s(r[idx["email_ama"]]),
                "domicilio": s(r[idx["direccion"]]),
                "contacto_principal": None,
                "preferencia_comunicacion": "email",
                "observaciones": None,
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }
        familia_id = families_map[dedupe_key]["id"]

        categoria = compute_category(fecha_nac) or s(r[idx["categoria_juego"]]) or s(r[idx["categoria_raw"]])

        player_id = new_id()
        player = {
            "id": player_id,
            "nombre": nombre,
            "apellidos": apellidos,
            "fecha_nacimiento": fecha_nac,
            "centro_escolar": s(r[idx["centro"]]),
            "progenitor1_nombre": s(r[idx["padre_nombre"]]),
            "progenitor1_telefono": padre_tel,
            "progenitor1_email": s(r[idx["email_aita"]]),
            "progenitor2_nombre": s(r[idx["madre_nombre"]]),
            "progenitor2_telefono": madre_tel,
            "progenitor2_email": s(r[idx["email_ama"]]),
            "domicilio": s(r[idx["direccion"]]),
            "foto": None,
            "categoria": categoria,
            "equipo_id": equipo_id,
            "dorsal": s(r[idx["dorsal"]]),
            "posicion": s(r[idx["demarcacion"]]),
            "estado": estado,
            "numero_licencia": None,
            "fecha_alta": None,
            "fecha_baja": None,
            "nueva_incorporacion": False,
            "segundo_hermano": False,
            "hermano_vinculado": None,
            "descuento": 0,
            "observaciones": s(r[idx["observ_interes"]]),
            "familia_id": familia_id,
            "alergias": None,
            "enfermedades": None,
            "medicacion": None,
            "seguro_medico": None,
            "contacto_emergencia": None,
            "telefono_emergencia": None,
            "observaciones_medicas": None,
            "talla_camiseta": s(r[idx["talla_camiseta"]]),
            "talla_pantalon": None,
            "talla_chandal": None,
            "talla_medias": s(r[idx["talla_medias"]]),
            "talla_calzado": None,
            "equipacion_entregada": False,
            "fecha_entrega_equipacion": None,
            "observaciones_material": None,
            "doc_dni_jugador": False,
            "doc_dni_tutor": False,
            "doc_foto": False,
            "doc_autorizacion": False,
            "doc_justificante_pago": False,
            "doc_ficha_federativa": False,
            "estado_documental": "pendiente",
            "fecha_revision_doc": None,
            "observaciones_doc": None,
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        players.append(player)

        # --- Pago ---
        cuota_a_pagar = to_float(r[idx["cuota_a_pagar"]])
        cuota_pagada = to_float(r[idx["cuota_pagada"]])
        cuota_pendiente = to_float(r[idx["cuota_pendiente"]])
        if cuota_a_pagar is not None:
            if cuota_pendiente == 0:
                pago_estado = "pagado"
            elif cuota_pagada and cuota_pagada > 0:
                pago_estado = "parcial"
            else:
                pago_estado = "pendiente"
            payments.append({
                "id": new_id(),
                "player_id": player_id,
                "concepto": "Cuota temporada 2025-2026",
                "importe_base": cuota_a_pagar,
                "descuento_hermano": 0,
                "importe_final": cuota_a_pagar,
                "forma_pago": None,
                "iban": None,
                "iban_validado": False,
                "estado": pago_estado,
                "fecha_pago": None,
                "recibo_generado": False,
                "observaciones": f"Pagado: {cuota_pagada}, Pendiente: {cuota_pendiente}" if cuota_pagada is not None else None,
                "created_at": now_iso(),
                "updated_at": now_iso(),
            })

    # ---- Resumen ----
    print(f"Filas sin nombre (omitidas): {skipped_no_name}")
    print(f"Jugadores a importar: {len(players)}")
    print(f"  - activos (con equipo 25&26): {sum(1 for p in players if p['estado'] == 'activo')}")
    print(f"  - baja (sin equipo 25&26):    {sum(1 for p in players if p['estado'] == 'baja')}")
    print(f"Equipos a crear: {len(teams_map)}")
    for t in teams_map.values():
        n = sum(1 for p in players if p["equipo_id"] == t["id"])
        print(f"  - {t['nombre']}: {n} jugadores")
    print(f"Familias a crear (deduplicadas): {len(families_map)}")
    print(f"Pagos a crear: {len(payments)}")

    if not args.commit:
        print("\n*** DRY RUN: no se ha escrito nada en MongoDB. Repite con --commit para aplicar. ***")
        return

    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]

    if teams_map:
        db.teams.insert_many(list(teams_map.values()))
    if families_map:
        db.families.insert_many(list(families_map.values()))
    if players:
        db.players.insert_many(players)
    if payments:
        db.payments.insert_many(payments)

    print("\n*** Migración aplicada en MongoDB. ***")


if __name__ == "__main__":
    main()
