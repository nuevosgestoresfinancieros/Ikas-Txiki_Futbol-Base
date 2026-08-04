"""Renderizadores en memoria para informes profesionales PDF y Excel."""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any, Mapping
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from brand_assets import BRAND_BLUE, BRAND_TEAL, pdf_logo


REPORT_LABELS = {
    "es": {
        "generated": "Generado", "filters": "Filtros aplicados", "all": "Todos",
        "no_results": "Sin resultados", "totals": "Resumen", "page": "Página",
        "season": "Temporada", "category": "Categoría", "team_id": "Equipo",
        "modality": "Modalidad", "status": "Estado", "search": "Búsqueda",
        "date_from": "Fecha inicial", "date_to": "Fecha final", "player_id": "Jugador",
        "period": "Periodo", "group_by": "Agrupación", "weekly": "Semanal",
        "monthly": "Mensual", "player": "Jugador", "team": "Equipo",
        "name": "Nombre", "surname": "Apellidos", "team_col": "Equipo",
        "category_col": "Categoría", "modality_col": "Modalidad", "number": "Dorsal",
        "status_col": "Estado", "sessions": "Sesiones", "present": "Presencias",
        "justified": "Justificadas", "unjustified": "Injustificadas", "injury": "Lesiones",
        "percentage": "Asistencia (%)", "players": "Jugadores", "teams": "Equipos",
        "groups": "Grupos", "rows": "Registros", "items": "Artículos", "delivered": "Entregados",
        "birth_date": "Fecha de nacimiento", "position": "Posición", "license": "Licencia",
        "joined_at": "Fecha de alta", "date": "Fecha", "time": "Hora", "location": "Lugar",
        "attendees": "Asistentes", "period_label": "Periodo", "opponent": "Rival",
        "condition": "Condición", "competition": "Competición", "score": "Resultado",
        "response": "Respuesta", "responded_at": "Fecha de respuesta", "late": "Fuera de plazo",
        "attendance_status": "Asistencia real", "consistent": "Coherencia",
        "capacity": "Capacidad", "occupancy": "Ocupación (%)", "type": "Tipo",
        "movement": "Movimiento", "document_status": "Estado documental",
        "missing_documents": "Documentos pendientes", "authorization_type": "Tipo de autorización",
        "signed_at": "Fecha de firma", "expires_at": "Caducidad", "missing_count": "Campos pendientes",
        "missing_fields": "Datos pendientes", "equipment_item": "Equipación",
        "shirt_size": "Talla camiseta", "shorts_size": "Talla pantalón",
        "tracksuit_size": "Talla chándal", "delivery_date": "Fecha de entrega",
        "contact_name": "Contacto", "phone": "Teléfono", "email": "Correo",
        "matches_called": "Convocatorias",
        "minutes": "Minutos", "goals": "Goles", "assists": "Asistencias",
        "yellow_cards": "Amarillas", "red_cards": "Rojas", "rating": "Valoración",
        "matches_played": "Partidos disputados", "wins": "Victorias", "draws": "Empates", "losses": "Derrotas",
        "concept": "Concepto", "expected": "Previsto", "paid": "Cobrado",
        "pending": "Pendiente", "payment_method": "Forma de pago", "payment_date": "Fecha de pago",
        "contact_type": "Tipo de contacto", "delivery": "Entrega", "movement_filter": "Movimiento",
    },
    "eu": {
        "generated": "Sortua", "filters": "Aplikatutako iragazkiak", "all": "Guztiak",
        "no_results": "Emaitzarik ez", "totals": "Laburpena", "page": "Orria",
        "season": "Denboraldia", "category": "Kategoria", "team_id": "Taldea",
        "modality": "Modalitatea", "status": "Egoera", "search": "Bilaketa",
        "date_from": "Hasiera-data", "date_to": "Amaiera-data", "player_id": "Jokalaria",
        "period": "Aldia", "group_by": "Taldekatzea", "weekly": "Astero",
        "monthly": "Hilero", "player": "Jokalaria", "team": "Taldea",
        "name": "Izena", "surname": "Abizenak", "team_col": "Taldea",
        "category_col": "Kategoria", "modality_col": "Modalitatea", "number": "Dortsala",
        "status_col": "Egoera", "sessions": "Saioak", "present": "Bertaratzeak",
        "justified": "Justifikatuak", "unjustified": "Justifikatu gabeak", "injury": "Lesioak",
        "percentage": "Asistentzia (%)", "players": "Jokalariak", "teams": "Taldeak",
        "groups": "Taldekatzeak", "rows": "Erregistroak", "items": "Artikuluak", "delivered": "Entregatuak",
        "birth_date": "Jaioteguna", "position": "Posizioa", "license": "Lizentzia",
        "joined_at": "Alta-data", "date": "Data", "time": "Ordua", "location": "Lekua",
        "attendees": "Bertaratuak", "period_label": "Aldia", "opponent": "Aurkaria",
        "condition": "Baldintza", "competition": "Lehiaketa", "score": "Emaitza",
        "response": "Erantzuna", "responded_at": "Erantzun-data", "late": "Epez kanpo",
        "attendance_status": "Benetako asistentzia", "consistent": "Koherentzia",
        "capacity": "Edukiera", "occupancy": "Okupazioa (%)", "type": "Mota",
        "movement": "Mugimendua", "document_status": "Dokumentuen egoera",
        "missing_documents": "Falta diren dokumentuak", "authorization_type": "Baimen mota",
        "signed_at": "Sinadura-data", "expires_at": "Iraungitzea", "missing_count": "Falta diren eremuak",
        "missing_fields": "Falta diren datuak", "equipment_item": "Ekipamendua",
        "shirt_size": "Kamiseta-taila", "shorts_size": "Praka-taila",
        "tracksuit_size": "Txandal-taila", "delivery_date": "Entrega-data",
        "contact_name": "Kontaktua", "phone": "Telefonoa", "email": "Posta elektronikoa",
        "matches_called": "Deialdiak",
        "minutes": "Minutuak", "goals": "Golak", "assists": "Asistentziak",
        "yellow_cards": "Txartel horiak", "red_cards": "Txartel gorriak", "rating": "Balorazioa",
        "matches_played": "Jokatutako partidak", "wins": "Garaipenak", "draws": "Berdinketak", "losses": "Porrotak",
        "concept": "Kontzeptua", "expected": "Aurreikusita", "paid": "Kobratuta",
        "pending": "Zain", "payment_method": "Ordainketa modua", "payment_date": "Ordainketa-data",
        "contact_type": "Kontaktu mota", "delivery": "Entrega", "movement_filter": "Mugimendua",
    },
}

COLUMN_KEYS = {
    "team": "team_col", "category": "category_col", "modality": "modality_col", "status": "status_col",
}
FORMULA_PREFIXES = ("=", "+", "-", "@")


def _labels(lang: str) -> dict:
    return REPORT_LABELS["eu" if lang == "eu" else "es"]


def safe_text(value: Any, missing: str = "—") -> str:
    return missing if value in (None, "") else str(value)


def excel_safe(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def safe_filename(report_id: str, lang: str, extension: str, generated_at: datetime | None = None) -> str:
    generated_at = generated_at or datetime.now(timezone.utc)
    safe_report = re.sub(r"[^a-z0-9_-]", "-", report_id.casefold()).strip("-") or "informe"
    return f"ikas-txiki_{safe_report}_{'eu' if lang == 'eu' else 'es'}_{generated_at:%Y%m%d}.{extension}"


def human_filters(filters: Mapping[str, Any], options: Mapping[str, Any], lang: str) -> list[tuple[str, str]]:
    labels = _labels(lang)
    teams = {item.get("id"): item.get("name") for item in options.get("teams", [])}
    players = {item.get("id"): item.get("name") for item in options.get("players", [])}
    result = []
    for key, value in filters.items():
        if value in (None, "", "all"):
            continue
        if key == "team_id":
            value = teams.get(value, "—")
        elif key == "player_id":
            value = players.get(value, "—")
        elif key in {"period", "group_by"}:
            value = labels.get(str(value), value)
        result.append((labels.get(key, key), safe_text(value)))
    return result


def _logo(branding: Mapping[str, Any]):
    try:
        return pdf_logo(branding.get("club_logo"), 14)
    except (OSError, ValueError):
        return None


def generate_pdf(report: Mapping[str, Any], rows: list[dict], totals: Mapping[str, Any], filters: Mapping[str, Any],
                 options: Mapping[str, Any], branding: Mapping[str, Any], lang: str = "es",
                 generated_at: datetime | None = None) -> bytes:
    lang = "eu" if lang == "eu" else "es"
    labels = _labels(lang)
    generated_at = generated_at or datetime.now(timezone.utc)
    buffer = io.BytesIO()
    page_size = landscape(A4)
    title = report.get("name", {}).get(lang) or report.get("name", {}).get("es") or "Informe"
    club = safe_text(branding.get("club_nombre"), "Ikas-Txiki")
    document = SimpleDocTemplate(buffer, pagesize=page_size, leftMargin=12 * mm, rightMargin=12 * mm,
                                 topMargin=12 * mm, bottomMargin=15 * mm, title=title,
                                 author="Ikas-Txiki Manager", subject="Informe deportivo")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=16,
                                 leading=19, textColor=colors.HexColor(BRAND_BLUE), alignment=TA_CENTER)
    small = ParagraphStyle("ReportSmall", parent=styles["Normal"], fontName="Helvetica", fontSize=7.5,
                           leading=10, textColor=colors.HexColor("#475569"))
    cell = ParagraphStyle("ReportCell", parent=styles["Normal"], fontName="Helvetica", fontSize=7.5,
                          leading=9, textColor=colors.HexColor("#1f2937"))
    story = []
    logo = _logo(branding)
    header = Table([[logo or "", Paragraph(f"<b>{escape(club)}</b>", title_style)]], colWidths=[20 * mm, 245 * mm])
    header.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 0),
                                     ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
    generated_by = safe_text(report.get("generated_by"), "—")
    story.extend([header, Spacer(1, 3 * mm), Paragraph(escape(title), title_style),
                  Paragraph(f"{labels['generated']}: {generated_at.astimezone(timezone.utc):%Y-%m-%d %H:%M UTC} · Usuario: {escape(generated_by)}", small),
                  Spacer(1, 3 * mm)])
    filter_rows = human_filters(filters, options, lang)
    filter_text = " · ".join(f"<b>{escape(name)}:</b> {escape(value)}" for name, value in filter_rows) or labels["all"]
    story.extend([Paragraph(f"<b>{labels['filters']}:</b> {filter_text}", small), Spacer(1, 2 * mm)])
    total_text = " · ".join(f"<b>{escape(labels.get(key, key))}:</b> {escape(safe_text(value))}" for key, value in totals.items())
    story.extend([Paragraph(f"<b>{labels['totals']}:</b> {total_text}", small), Spacer(1, 4 * mm)])
    columns = list(report.get("columns", []))
    table_data = [[Paragraph(f"<b>{escape(labels.get(COLUMN_KEYS.get(column, column), column))}</b>", cell) for column in columns]]
    if rows:
        for row in rows:
            table_data.append([Paragraph(escape(safe_text(row.get(column))), cell) for column in columns])
    else:
        table_data.append([Paragraph(escape(labels["no_results"]), cell)] + [""] * max(0, len(columns) - 1))
    available = page_size[0] - 24 * mm
    weights = {"name": 1.5, "surname": 1.6, "team": 1.5, "category": 1.15, "modality": .8,
               "number": .7, "status": 1.0, "sessions": .8, "present": .8, "justified": .9,
               "unjustified": 1.0, "injury": .8, "percentage": .9, "missing_documents": 2.2,
               "missing_fields": 2.0, "contact_name": 1.5, "email": 1.8, "concept": 1.5}
    total_weight = sum(weights.get(column, 1) for column in columns) or 1
    widths = [available * weights.get(column, 1) / total_weight for column in columns]
    table = LongTable(table_data, colWidths=widths, repeatRows=1, splitByRow=1)
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_TEAL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), .25, colors.HexColor("#CBD5E1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for index in range(2, len(table_data), 2):
        commands.append(("BACKGROUND", (0, index), (-1, index), colors.HexColor("#F8FAFC")))
    if not rows:
        commands.append(("SPAN", (0, 1), (-1, 1)))
    table.setStyle(TableStyle(commands))
    story.append(table)

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawString(12 * mm, 7 * mm, "Ikas-Txiki Manager")
        canvas.drawRightString(page_size[0] - 12 * mm, 7 * mm, f"{labels['page']} {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=footer)
    return buffer.getvalue()


def generate_xlsx(report: Mapping[str, Any], rows: list[dict], totals: Mapping[str, Any], filters: Mapping[str, Any],
                  options: Mapping[str, Any], branding: Mapping[str, Any], lang: str = "es",
                  generated_at: datetime | None = None) -> bytes:
    lang = "eu" if lang == "eu" else "es"
    labels = _labels(lang)
    generated_at = generated_at or datetime.now(timezone.utc)
    title = report.get("name", {}).get(lang) or report.get("name", {}).get("es") or "Informe"
    workbook = Workbook()
    workbook.properties.creator = "Ikas-Txiki Manager"
    workbook.properties.title = title
    workbook.properties.subject = "Informe deportivo"
    sheet = workbook.active
    sheet.title = re.sub(r"[^A-Za-z0-9 _-]", "", title)[:31] or ("Informe" if lang == "es" else "Txostena")
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5
    columns = list(report.get("columns", []))
    last_column = get_column_letter(max(1, len(columns)))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=BRAND_TEAL.lstrip("#"))
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[1].height = 28
    sheet["A2"] = labels["generated"]
    sheet["B2"] = generated_at.astimezone(timezone.utc).replace(tzinfo=None)
    sheet["B2"].number_format = "yyyy-mm-dd hh:mm"
    sheet["C2"] = "Usuario"
    sheet["D2"] = excel_safe(safe_text(report.get("generated_by")))
    current = 4
    sheet[f"A{current}"] = labels["filters"]
    sheet[f"A{current}"].font = Font(bold=True, color="0F172A")
    current += 1
    filter_rows = human_filters(filters, options, lang)
    if not filter_rows:
        filter_rows = [("", labels["all"])]
    for name, value in filter_rows:
        if name:
            sheet.cell(current, 1, name).font = Font(bold=True)
        sheet.cell(current, 2, excel_safe(value))
        current += 1
    current += 1
    sheet.cell(current, 1, labels["totals"]).font = Font(bold=True, color="0F172A")
    current += 1
    for name, value in totals.items():
        sheet.cell(current, 1, labels.get(name, name)).font = Font(bold=True)
        target = sheet.cell(current, 2)
        if name == "percentage" and isinstance(value, (int, float)):
            target.value = value / 100
            target.number_format = "0.0%"
        else:
            target.value = excel_safe(value)
        current += 1
    current += 1
    header_row = current
    for index, column in enumerate(columns, 1):
        target = sheet.cell(header_row, index, labels.get(COLUMN_KEYS.get(column, column), column))
        target.font = Font(bold=True, color="FFFFFF")
        target.fill = PatternFill("solid", fgColor=BRAND_TEAL.lstrip("#"))
        target.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    if rows:
        for row_index, row in enumerate(rows, header_row + 1):
            for column_index, column in enumerate(columns, 1):
                value = row.get(column)
                target = sheet.cell(row_index, column_index)
                if column == "percentage" and isinstance(value, (int, float)):
                    target.value = value / 100
                    target.number_format = "0.0%"
                else:
                    target.value = excel_safe(safe_text(value))
                target.alignment = Alignment(vertical="top", wrap_text=True)
                target.border = Border(bottom=thin)
                if row_index % 2 == 0:
                    target.fill = PatternFill("solid", fgColor="F8FAFC")
    else:
        sheet.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=max(1, len(columns)))
        sheet.cell(header_row + 1, 1, labels["no_results"])
    last_row = header_row + max(1, len(rows))
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.print_title_rows = f"{header_row}:{header_row}"
    sheet.print_area = f"A1:{last_column}{last_row}"
    widths = {"name": 22, "surname": 26, "team": 22, "category": 16, "modality": 14, "number": 11,
              "status": 16, "sessions": 12, "present": 12, "justified": 14, "unjustified": 16,
              "injury": 12, "percentage": 16, "missing_documents": 34, "missing_fields": 32,
              "contact_name": 24, "phone": 18, "email": 28, "concept": 24}
    for index, column in enumerate(columns, 1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(column, 16)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
