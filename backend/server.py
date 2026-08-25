from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Depends, Cookie, Response, Request, Query
from fastapi.security import OAuth2PasswordBearer
from fastapi.responses import StreamingResponse, FileResponse
from dotenv import load_dotenv
from jose import JWTError, jwt
from passlib.context import CryptContext
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
from urllib.parse import quote
import asyncio
import io
import base64
import copy
import html as html_lib
import json
import logging
import math
import re
import shutil
import time
import zipfile
from collections import defaultdict, deque
from pathlib import Path
from pydantic import BaseModel, Field, field_validator, model_validator
from typing import List, Optional, Dict, Any, Mapping
import uuid
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, KeepTogether
from datetime import datetime, timezone, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from pymongo import ReturnDocument
from pymongo.errors import DuplicateKeyError

from authz import (
    ROLES, ROLE_PERMISSIONS, current_user_context, enforce_permission, enforce_related_scope,
    has_permission, ids,
    merge_query, public_user, route_permission,
)
from dashboard_service import pending_callups, player_callup_status, prioritized_alerts, weekly_attendance
from attendance_service import (
    ATTENDANCE_STATES, attendance_history, attendance_rows, attendance_summary,
    attendance_trend, callup_attendance_comparison, player_percentages, repeated_absence_alerts,
)
from callup_service import (
    VALID_STATUSES, apply_response, is_late, normalize_callup, normalize_status,
    response_counts,
)
from calendar_service import (
    CALENDAR_TYPES, aggregate_calendar_events, calendar_to_ical,
    next_calendar_event, subscription_capability,
)
from portal_service import document_status, portal_attendance, portal_callups, safe_payment, safe_player, upcoming
from notification_service import dispatch_email, dispatch_telegram, make_notification, notification_enabled, provider_configuration
from brand_assets import BRAND_BLUE, BRAND_TEAL, pdf_logo
from inscription_import_service import (
    SEASON as IMPORT_SEASON, SAFE_PLAYER_FIELDS, ImportValidationError,
    analyze_rows, decode_plan, encode_plan, encrypt_iban, family_key,
    file_sha256, identity_key, masked_iban, merge_nonempty, normalize_key,
    parse_excel, _fernet,
)
from import_staging_service import (
    ALLOWED_RECORD_FIELDS, audit_event, draft_summary, effective_records, expiry, field_is_valid,
    historical_readiness, prepare_records, public_draft, valid_iban as staging_valid_iban,
)
from historical_import_adapter import (
    HISTORICAL_FORMAT, historical_quality_summary, historical_simulation,
    parse_historical_excel, prepare_historical_staging,
)
from modality_service import (
    ModalityCreateRequest, ModalityDefinition, ModalityReorderRequest,
    ModalityStatusRequest, ModalityUpdateRequest, catalog_from_settings,
    normalize_modality, validate_compatibility_catalog,
)
from report_service import (
    MAX_EXPORT_ROWS, ReportValidationError, build_report, catalog_for_role,
    enforce_export_limit, paginate, validate_report_filters,
)
from report_export_service import generate_pdf as generate_report_pdf, generate_xlsx as generate_report_xlsx, safe_filename
from statistics_service import calculate_statistics, paginate_player_rows
from match_report_service import (
    MatchReportValidationError, build_objective_statistics, dry_run_historical_rows,
    normalize_goal_event, normalize_participant, normalize_substitution, period_configuration,
    validate_goal_events, validate_participants, validate_substitutions,
)
from health_service import deployment_version
from assistant_knowledge import KNOWLEDGE_VERSION, available_modules
from assistant_service import (
    ACTION_DEFINITIONS, ExternalAssistantProvider, ProposalStore, answer_help,
    public_proposal, session_fingerprint,
)
from user_admin_service import (
    ACCOUNT_STATUSES, account_status, effective_scope, link_is_complete, normalized_key, normalized_text,
    safe_audit_detail, security_state, status_is_active, user_search_text, validate_password_strength,
)
from communication_recipient_service import (
    communication_record_active, consented_contacts, recipient_summary, usable_account,
)
from exercise_service import (
    EXERCISE_CATEGORIES, EXERCISE_STATES, INTENSITIES, RATINGS, VISIBILITIES,
    ExerciseValidationError, exercise_statistics, normalize_exercise,
    normalize_planned_exercises, normalize_template, validate_exercise_update,
)
from user_security_service import (
    INVITATION_TTL_HOURS, LOCK_DURATION_MINUTES, MAX_ACCOUNT_ATTEMPTS,
    generate_temporary_password, invitation_status, issue_token, legacy_session_allowed,
    parse_time, safe_security_audit, security_public, token_digest, token_is_usable, utcnow,
)


ROOT_DIR = Path(__file__).parent
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]
PROCESS_STARTED_MONOTONIC = time.monotonic()
DEPLOYMENT_VERSION = deployment_version(ROOT_DIR.parent)
APP_ENVIRONMENT = str(os.environ.get("APP_ENVIRONMENT") or "production").strip()

# ── Auth config ──────────────────────────────────────────────
JWT_SECRET = os.environ.get("JWT_SECRET")
JWT_ALGORITHM = "HS256"
JWT_EXPIRE_HOURS = 8
ADMIN_USER = os.environ.get("ADMIN_USER")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")

if not JWT_SECRET or len(JWT_SECRET) < 32:
    raise RuntimeError("JWT_SECRET debe existir y tener al menos 32 caracteres")
if not ADMIN_USER or not ADMIN_PASSWORD or ADMIN_PASSWORD.lower() == "admin":
    raise RuntimeError("Configura ADMIN_USER y una ADMIN_PASSWORD segura en el entorno")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
LOGIN_WINDOW_SECONDS = max(60, int(os.environ.get("LOGIN_WINDOW_SECONDS", "600")))
LOGIN_MAX_ATTEMPTS = max(3, min(20, int(os.environ.get("LOGIN_MAX_ATTEMPTS", "6"))))
ACCOUNT_LOCK_MINUTES = max(1, min(1440, int(os.environ.get("ACCOUNT_LOCK_MINUTES", str(LOCK_DURATION_MINUTES)))))
login_attempts = defaultdict(deque)
assistant_attempts = defaultdict(deque)
assistant_proposals = ProposalStore(int(os.environ.get("ASSISTANT_PROPOSAL_TTL_SECONDS", "600")))
assistant_provider = ExternalAssistantProvider(
    transport=None,
    timeout_seconds=float(os.environ.get("ASSISTANT_PROVIDER_TIMEOUT_SECONDS", "5")),
)
ASSISTANT_RATE_WINDOW_SECONDS = 60
ASSISTANT_RATE_MAX_REQUESTS = 30

def create_access_token(data: dict) -> str:
    expire = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS)
    return jwt.encode({**data, "iat": datetime.now(timezone.utc), "exp": expire}, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def load_user(username: str) -> Optional[dict]:
    if username == ADMIN_USER:
        return {
            "id": "environment-admin", "username": username, "role": "admin",
            "active": True, "assigned_team_ids": [], "language": "es",
            "notification_preferences": {}, "account_status": "active",
            "system_account": True, "read_only": True,
        }
    return await db["users"].find_one({"username": username}, {"_id": 0, "password_hash": 0})


async def record_security_event(event_type: str, user: Optional[Mapping[str, Any]], action: str,
                                reason: str, *, aggregate: Optional[dict] = None) -> None:
    """Best-effort security audit containing no target or contact data."""
    try:
        await db.internal_events.insert_one({
            "id": str(uuid.uuid4()),
            "type": event_type,
            "actor_user_id": (user or {}).get("id"),
            "actor_role": (user or {}).get("role"),
            "action": action,
            "result": "denied" if event_type.endswith(".denied") else "filtered",
            "reason": reason,
            "aggregate": {str(key): int(value) for key, value in (aggregate or {}).items()},
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        logging.getLogger(__name__).warning("Security audit event could not be persisted: %s", event_type)


async def get_current_user(request: Request):
    token = request.cookies.get("ikastxiki_session")
    if not token:
        raise HTTPException(status_code=401, detail="No autenticado")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        username = payload.get("sub")
        if not username:
            raise HTTPException(status_code=401, detail="Token inválido")
        user = await load_user(username)
        if not user:
            raise HTTPException(status_code=403, detail="Usuario sin acceso")
        if not user.get("active", True) or not status_is_active(account_status(user)):
            await record_security_event("security.authentication.denied", user, "authenticated_request", "account_inactive")
            raise HTTPException(status_code=403, detail="Usuario inactivo o sin acceso")
        locked_until = parse_time(user.get("locked_until"))
        if locked_until and locked_until > utcnow():
            await record_security_event("security.authentication.denied", user, "authenticated_request", "account_locked")
            raise HTTPException(status_code=403, detail="Usuario inactivo o sin acceso")
        if not user.get("system_account") and not legacy_session_allowed(payload, user):
            await record_security_event("security.authentication.denied", user, "authenticated_request", "session_revoked")
            raise HTTPException(status_code=401, detail="Sesión revocada")
        if payload.get("purpose") == "password_change":
            raise HTTPException(status_code=403, detail="Debes cambiar la contraseña")
        request.state.current_user = user
        current_user_context.set(user)
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Sesión expirada")

app = FastAPI(title="Ikas-Txiki Manager API")


@app.get("/api/health")
async def health():
    """Comprobación pública, rápida y de solo lectura para monitorización."""
    try:
        await asyncio.wait_for(db.command({"ping": 1, "maxTimeMS": 100}), timeout=0.2)
    except Exception:
        raise HTTPException(status_code=503, detail={
            "status": "error", "version": DEPLOYMENT_VERSION,
            "environment": APP_ENVIRONMENT, "mongodb": "error",
            "uptime_seconds": max(0, int(time.monotonic() - PROCESS_STARTED_MONOTONIC)),
            "timestamp": datetime.now(timezone.utc).isoformat(),
        })
    return {
        "status": "ok",
        "version": DEPLOYMENT_VERSION,
        "environment": APP_ENVIRONMENT,
        "mongodb": "ok",
        "uptime_seconds": max(0, int(time.monotonic() - PROCESS_STARTED_MONOTONIC)),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }

# ── Auth endpoints ────────────────────────────────────────────
@app.post("/api/auth/login")
async def login(request: Request, response: Response, data: Dict[str, Any]):
    username = data.get("username", "").strip()
    password = data.get("password", "")
    client_ip = request.client.host if request.client else "unknown"
    attempt_key = f"{client_ip}:{username.lower()}"
    now = time.monotonic()
    attempts = login_attempts[attempt_key]
    while attempts and now - attempts[0] > LOGIN_WINDOW_SECONDS:
        attempts.popleft()
    if len(attempts) >= LOGIN_MAX_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Demasiados intentos. Espera unos minutos e inténtalo de nuevo")

    # Verificar contra usuario admin del .env
    valid_user = username == ADMIN_USER and password == ADMIN_PASSWORD
    db_user = None
    # También buscar en colección users de MongoDB
    if not valid_user:
        db_user = await db["users"].find_one({"username": username})
        if db_user and db_user.get("active", True) and status_is_active(account_status(db_user)) and pwd_context.verify(password, db_user.get("password_hash", "")):
            valid_user = True
    locked_until = parse_time(db_user.get("locked_until")) if db_user else None
    if locked_until and locked_until > utcnow():
        raise HTTPException(status_code=429, detail="Acceso bloqueado temporalmente")
    if not valid_user:
        attempts.append(now)
        if db_user:
            updated = await db.users.find_one_and_update(
                {"id": db_user["id"]}, {"$inc": {"failed_login_count": 1},
                "$set": {"last_failed_login_at": now_iso()}}, return_document=ReturnDocument.AFTER,
            )
            if int((updated or {}).get("failed_login_count", 0)) >= MAX_ACCOUNT_ATTEMPTS:
                await db.users.update_one({"id": db_user["id"], "locked_until": {"$in": [None, ""]}},
                                          {"$set": {"locked_until": (utcnow() + timedelta(minutes=ACCOUNT_LOCK_MINUTES)).isoformat()}})
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")
    login_attempts.pop(attempt_key, None)
    if db_user:
        await db["users"].update_one({"id": db_user["id"]}, {"$set": {
            "last_access_at": now_iso(), "failed_login_count": 0, "locked_until": None,
        }})
    if db_user and db_user.get("must_change_password"):
        change_token = create_access_token({
            "sub": username, "purpose": "password_change", "ver": int(db_user.get("session_version", 0)),
        })
        return {"ok": False, "requires_password_change": True, "password_change_token": change_token}
    token = create_access_token({"sub": username, "sid": str(uuid.uuid4()),
                                 "ver": int((db_user or {}).get("session_version", 0))})
    response.set_cookie(
        key="ikastxiki_session",
        value=token,
        httponly=True,
        secure=True,
        samesite="strict",
        max_age=JWT_EXPIRE_HOURS * 3600,
        path="/",
    )
    user = await load_user(username)
    return {"ok": True, "user": public_user(user), "username": username}


@app.post("/api/auth/logout")
async def logout(response: Response):
    response.delete_cookie(key="ikastxiki_session", path="/")
    return {"ok": True}


@app.get("/api/auth/me")
async def me(current_user: dict = Depends(get_current_user)):
    return public_user(current_user)


class PasswordChangeRequest(BaseModel):
    token: str
    password: str
    password_confirmation: str


class RecoveryRequest(BaseModel):
    identifier: str


class TokenPasswordRequest(BaseModel):
    token: str
    password: str
    password_confirmation: str


async def consume_password_token(kind: str, request: TokenPasswordRequest) -> dict:
    if request.password != request.password_confirmation:
        raise HTTPException(status_code=422, detail="Las contraseñas no coinciden")
    validate_password_strength(request.password)
    digest = token_digest(request.token, JWT_SECRET)
    field = "invitation" if kind == "invitation" else "recovery"
    user = await db.users.find_one({f"{field}.digest": digest})
    record = (user or {}).get(field)
    if not user or not token_is_usable(record):
        raise HTTPException(status_code=400, detail="Enlace inválido o caducado")
    moment = now_iso()
    result = await db.users.update_one({
        "id": user["id"], f"{field}.digest": digest, f"{field}.used_at": None,
        f"{field}.cancelled_at": None, f"{field}.expires_at": {"$gt": moment},
    }, {"$set": {
        "password_hash": pwd_context.hash(request.password), "must_change_password": False,
        "account_status": "active", "active": True, f"{field}.used_at": moment,
        "last_password_change_at": moment, "sessions_revoked_at": moment,
    }, "$inc": {"session_version": 1}})
    if not result.modified_count:
        raise HTTPException(status_code=400, detail="Enlace inválido, caducado o ya utilizado")
    await record_user_audit(f"{kind}_completed", user, [])
    return {"ok": True}


@app.post("/api/auth/change-temporary-password")
async def change_temporary_password(request: PasswordChangeRequest):
    if request.password != request.password_confirmation:
        raise HTTPException(status_code=422, detail="Las contraseñas no coinciden")
    validate_password_strength(request.password)
    try:
        payload = jwt.decode(request.token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=400, detail="Enlace inválido o caducado")
    if payload.get("purpose") != "password_change":
        raise HTTPException(status_code=400, detail="Token no válido")
    user = await db.users.find_one({"username": payload.get("sub")})
    if not user or not user.get("must_change_password") or not legacy_session_allowed(payload, user):
        raise HTTPException(status_code=400, detail="Token ya utilizado o no válido")
    moment = now_iso()
    result = await db.users.update_one({
        "id": user["id"], "must_change_password": True,
        "session_version": int(user.get("session_version", 0)),
    }, {"$set": {
        "password_hash": pwd_context.hash(request.password), "must_change_password": False,
        "last_password_change_at": moment, "sessions_revoked_at": moment,
    }, "$inc": {"session_version": 1}})
    if not result.modified_count:
        raise HTTPException(status_code=400, detail="Token ya utilizado o no válido")
    await record_user_audit("temporary_password_changed", user, [])
    return {"ok": True}


@app.post("/api/auth/activate")
async def activate_invitation(request: TokenPasswordRequest):
    return await consume_password_token("invitation", request)


@app.post("/api/auth/recovery/request")
async def request_recovery(request: RecoveryRequest):
    identifier = normalized_key(request.identifier)
    user = await db.users.find_one({"$or": [
        {"username_normalized": identifier}, {"email_normalized": identifier},
        {"username": request.identifier.strip()},
    ]})
    if user and user.get("active", True) and status_is_active(account_status(user)):
        plain, record = issue_token(JWT_SECRET)
        await db.users.update_one({"id": user["id"]}, {"$set": {"recovery": record}})
        await record_user_audit("recovery_requested", user, [])
        if os.environ.get("SECURITY_TEST_MODE") == "1":
            return {"ok": True, "message": "Si la cuenta existe, recibirá instrucciones", "test_token": plain}
    return {"ok": True, "message": "Si la cuenta existe, recibirá instrucciones"}


@app.post("/api/auth/recovery/reset")
async def reset_recovery_password(request: TokenPasswordRequest):
    return await consume_password_token("recovery", request)


@app.get("/api/public/branding")
async def public_branding(response: Response):
    """Expone solo la identidad visual necesaria antes de iniciar sesión."""
    doc = await db.settings.find_one(
        {"id": "global"},
        {"_id": 0, "club_nombre": 1, "club_logo": 1, "temporada_actual": 1},
    )
    response.headers["Cache-Control"] = "public, max-age=300"
    return doc or {
        "club_nombre": "Ikas-Txiki",
        "club_logo": None,
        "temporada_actual": None,
    }


api_router = APIRouter(prefix="/api")


async def authorize_request(request: Request, user: dict = Depends(get_current_user)):
    resource, action = route_permission(request)
    enforce_permission(user, resource, action)
    return user


# ---------- Helpers ----------
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def new_id() -> str:
    return str(uuid.uuid4())


CATEGORIES = [
    {"name": "Prebenjamín", "min_age": 6, "max_age": 7},
    {"name": "Benjamín", "min_age": 8, "max_age": 9},
    {"name": "Alevín", "min_age": 10, "max_age": 11},
    {"name": "Infantil", "min_age": 12, "max_age": 13},
    {"name": "Cadete", "min_age": 14, "max_age": 15},
    {"name": "Juvenil", "min_age": 16, "max_age": 18},
]


def compute_category(birthdate: Optional[str]) -> Optional[str]:
    if not birthdate:
        return None
    try:
        bd = datetime.fromisoformat(birthdate).date() if "T" in birthdate or "-" in birthdate else None
        if bd is None:
            return None
    except Exception:
        try:
            bd = datetime.strptime(birthdate[:10], "%Y-%m-%d").date()
        except Exception:
            return None
    today = date.today()
    # Football season age: age reached during the season year
    season_year = today.year if today.month >= 7 else today.year - 1
    age = season_year - bd.year
    for c in CATEGORIES:
        if c["min_age"] <= age <= c["max_age"]:
            return c["name"]
    if age < 6:
        return "Querubín"
    return "Senior"


def clean(doc: dict) -> dict:
    doc.pop("_id", None)
    return doc


# ---------- Generic CRUD factory ----------
PLAYER_SCOPED_COLLECTIONS = {"payments", "authorizations", "stats", "equipment"}


async def user_player_ids(user: dict) -> list[str]:
    if user.get("role") == "player":
        return ids([user.get("player_id")])
    if user.get("role") == "family":
        explicit = ids(user.get("player_ids") or [])
        linked = await db.players.find(
            {"familia_id": user.get("family_id")}, {"_id": 0, "id": 1}
        ).to_list(100)
        return sorted(set(explicit + ids(row.get("id") for row in linked)))
    team_ids = ids(user.get("assigned_team_ids") or [])
    if team_ids:
        linked = await db.players.find(
            {"equipo_id": {"$in": team_ids}}, {"_id": 0, "id": 1}
        ).to_list(5000)
        return ids(row.get("id") for row in linked)
    return []


async def scope_for_collection(coll: str, user: Optional[dict] = None) -> Optional[dict]:
    user = user or current_user_context.get()
    if not user:
        return {"id": {"$in": []}}
    if coll == "notifications":
        return {"$or": [
            {"recipient_user_id": user.get("id")},
            {"recipient_username": user.get("username")},
        ]}
    if user.get("role") == "admin":
        return None
    role = user.get("role")
    team_ids = ids(user.get("assigned_team_ids") or [])
    player_ids = await user_player_ids(user)
    if role in {"coordinator", "coach"}:
        if coll == "teams":
            return {"id": {"$in": team_ids}}
        if coll in {"players", "matches", "trainings", "callups"}:
            return {"equipo_id": {"$in": team_ids}}
        if coll in PLAYER_SCOPED_COLLECTIONS or coll == "inscriptions":
            return {"player_id": {"$in": player_ids}}
        if coll == "families":
            family_ids = await db.players.distinct("familia_id", {"id": {"$in": player_ids}})
            return {"id": {"$in": ids(family_ids)}}
        if coll == "communications":
            return {"$or": [
                {"destinatario_tipo": "equipo", "destinatario_id": {"$in": team_ids}},
                {"destinatario_tipo": "individual", "destinatario_id": {"$in": player_ids}},
            ]}
        if coll == "club_events":
            return {"$or": [
                {"equipo_id": {"$in": team_ids}}, {"equipo_id": None}, {"equipo_id": {"$exists": False}},
            ]}
    if role in {"family", "player"}:
        if coll == "players":
            return {"id": {"$in": player_ids}}
        if coll == "families":
            return {"id": user.get("family_id")}
        team_ids = ids(await db.players.distinct("equipo_id", {"id": {"$in": player_ids}}))
        if coll == "teams":
            return {"id": {"$in": team_ids}}
        if coll in {"matches", "trainings"}:
            return {"equipo_id": {"$in": team_ids}}
        if coll == "callups":
            return {"convocados.player_id": {"$in": player_ids}}
        if coll in PLAYER_SCOPED_COLLECTIONS or coll == "inscriptions":
            return {"player_id": {"$in": player_ids}}
        if coll == "communications":
            targets = player_ids + ids([user.get("family_id")])
            return {"$or": [
                {"destinatario_tipo": "individual", "destinatario_id": {"$in": targets}},
                {"destinatario_tipo": "equipo", "destinatario_id": {"$in": team_ids}},
            ]}
        if coll == "club_events":
            return {"$or": [
                {"equipo_id": {"$in": team_ids}}, {"equipo_id": None}, {"equipo_id": {"$exists": False}},
            ]}
    return {"id": {"$in": []}}


async def ensure_data_scope(coll: str, data: dict) -> None:
    user = current_user_context.get()
    if not user or user.get("role") == "admin":
        return
    role = user.get("role")
    if role not in {"coordinator", "coach"}:
        raise HTTPException(status_code=403, detail="No puedes modificar datos de este módulo")
    team_ids = set(ids(user.get("assigned_team_ids") or []))
    if coll in {"teams", "players", "matches", "trainings", "callups"}:
        target = data.get("id") if coll == "teams" else data.get("equipo_id")
        if target not in team_ids:
            raise HTTPException(status_code=403, detail="El elemento no pertenece a tus equipos asignados")
    if coll == "club_events" and data.get("equipo_id") not in team_ids:
        raise HTTPException(status_code=403, detail="El evento no pertenece a tus equipos asignados")
    player_ids = set(await user_player_ids(user))
    if coll == "communications" and data.get("destinatario_tipo") == "categoria":
        category_team_ids = set(ids(await db.teams.distinct("id", {"categoria": data.get("destinatario_id")})))
        if not category_team_ids or not category_team_ids.issubset(team_ids):
            raise HTTPException(status_code=403, detail="La categoría contiene equipos fuera de tu ámbito")
        return
    if coll == "players" and data.get("familia_id"):
        family_scope = await scope_for_collection("families", user)
        if not await db.families.find_one(merge_query({"id": data["familia_id"]}, family_scope)):
            raise HTTPException(status_code=403, detail="La familia no pertenece a tu ámbito")
    enforce_related_scope(coll, data, team_ids, player_ids)
    if coll == "callups":
        if data.get("match_id") and not await db.matches.find_one({
            "id": data["match_id"], "equipo_id": {"$in": list(team_ids)},
        }):
            raise HTTPException(status_code=403, detail="El partido no pertenece a tu ámbito")
    if coll == "trainings" and data.get("callup_id"):
        if not await db.callups.find_one({"id": data["callup_id"], "equipo_id": {"$in": list(team_ids)}}):
            raise HTTPException(status_code=403, detail="La convocatoria no pertenece a tus equipos asignados")
    if coll in PLAYER_SCOPED_COLLECTIONS and data.get("player_id"):
        if data["player_id"] not in player_ids:
            raise HTTPException(status_code=403, detail="El jugador no pertenece a tu ámbito")


async def list_docs(coll: str, query: dict = None):
    scoped = merge_query(query, await scope_for_collection(coll))
    cursor = db[coll].find(scoped, {"_id": 0}).sort("created_at", -1)
    return await cursor.to_list(5000)


async def notification_users(team_ids: Optional[list[str]] = None, player_ids: Optional[list[str]] = None,
                             family_ids: Optional[list[str]] = None, include_admins: bool = False) -> list[dict]:
    clauses = []
    if team_ids:
        clauses.append({"assigned_team_ids": {"$in": team_ids}})
    if player_ids:
        clauses.append({"player_id": {"$in": player_ids}})
    if family_ids:
        clauses.append({"family_id": {"$in": family_ids}})
    if include_admins:
        clauses.append({"role": "admin"})
    if not clauses:
        return []
    return await db.users.find({"active": {"$ne": False}, "$or": clauses}, {"_id": 0}).to_list(5000)


async def enqueue_notifications(users: list[dict], notification_type: str, title: str, message: str,
                                link: Optional[str] = None, priority: str = "normal",
                                related: Optional[dict] = None, dedupe_key: Optional[str] = None) -> int:
    created = 0
    for user in users:
        if not notification_enabled(user.get("notification_preferences") or {}, notification_type):
            continue
        document = make_notification(user, notification_type, title, message, link, priority, related,
                                     dedupe_key=dedupe_key)
        query = {"recipient_user_id": user.get("id"), "dedupe_key": dedupe_key} if dedupe_key else {"id": document["id"]}
        result = await db.notifications.update_one(query, {"$setOnInsert": document}, upsert=True)
        created += int(bool(result.upserted_id))
    return created


async def users_for_team_players(team_id: Optional[str], player_ids: Optional[list[str]] = None,
                                 include_admins: bool = False) -> list[dict]:
    ids_for_players = player_ids or []
    if team_id and not ids_for_players:
        ids_for_players = await db.players.distinct("id", {"equipo_id": team_id})
    family_ids = await db.players.distinct("familia_id", {"id": {"$in": ids_for_players}}) if ids_for_players else []
    return await notification_users([team_id] if team_id else [], ids(ids_for_players), ids(family_ids), include_admins)


async def get_doc(coll: str, _id: str):
    query = merge_query({"id": _id}, await scope_for_collection(coll))
    doc = await db[coll].find_one(query, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="No encontrado")
    return doc


async def insert_doc(coll: str, data: dict):
    await ensure_data_scope(coll, data)
    data["id"] = new_id()
    data["created_at"] = now_iso()
    data["updated_at"] = now_iso()
    await db[coll].insert_one(dict(data))
    return clean(data)


async def update_doc(coll: str, _id: str, data: dict):
    query = merge_query({"id": _id}, await scope_for_collection(coll))
    existing = await db[coll].find_one(query)
    if not existing:
        raise HTTPException(status_code=404, detail="No encontrado")
    data = {k: v for k, v in data.items() if v is not None or k in data}
    data["updated_at"] = now_iso()
    await ensure_data_scope(coll, {**existing, **data})
    await db[coll].update_one(query, {"$set": data})
    return await get_doc(coll, _id)


async def delete_doc(coll: str, _id: str):
    query = merge_query({"id": _id}, await scope_for_collection(coll))
    res = await db[coll].delete_one(query)
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="No encontrado")
    return {"ok": True}


# ================= USERS / RBAC =================
class NotificationPreferences(BaseModel):
    in_app: bool = True
    email: bool = True
    telegram: bool = True
    callups: bool = True
    schedule_changes: bool = True
    payments: bool = True
    documents: bool = True


class UserCreate(BaseModel):
    username: str
    password: Optional[str] = None
    password_confirmation: Optional[str] = None
    first_name: str
    last_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    role: str
    account_status: str = "active"
    assigned_team_ids: List[str] = Field(default_factory=list)
    assigned_category_ids: List[str] = Field(default_factory=list)
    player_id: Optional[str] = None
    family_id: Optional[str] = None
    language: str = "es"
    notification_preferences: NotificationPreferences = Field(default_factory=NotificationPreferences)
    access_method: str = "password"

    @field_validator("username")
    @classmethod
    def validate_username(cls, value: str):
        value = value.strip()
        if len(value) < 3 or not value.replace("_", "").replace("-", "").isalnum():
            raise ValueError("El usuario debe tener al menos 3 caracteres alfanuméricos")
        return value

    @field_validator("password")
    @classmethod
    def validate_password(cls, value: Optional[str]):
        return validate_password_strength(value) if value else None

    @model_validator(mode="after")
    def validate_access_credentials(self):
        if self.access_method == "password":
            if not self.password or self.password != self.password_confirmation:
                raise ValueError("Las contraseñas no coinciden")
        return self

    @field_validator("first_name", "last_name")
    @classmethod
    def validate_identity(cls, value: str):
        value = normalized_text(value)
        if len(value) < 2:
            raise ValueError("Nombre y apellidos son obligatorios")
        return value

    @field_validator("email")
    @classmethod
    def validate_email(cls, value: Optional[str]):
        value = normalized_key(value)
        if value and ("@" not in value or value.startswith("@") or value.endswith("@")):
            raise ValueError("Correo no válido")
        return value or None

    @field_validator("phone")
    @classmethod
    def validate_phone(cls, value: Optional[str]):
        value = normalized_text(value)
        if value and (len(value) < 7 or len(value) > 24 or not re.fullmatch(r"[+() 0-9.-]+", value)):
            raise ValueError("Teléfono no válido")
        return value or None

    @field_validator("access_method")
    @classmethod
    def validate_access_method(cls, value: str):
        if value not in {"password", "temporary_password", "invitation", "pending"}:
            raise ValueError("Método de acceso no válido")
        return value

    @field_validator("role")
    @classmethod
    def validate_role(cls, value: str):
        if value not in ROLES:
            raise ValueError("Rol no válido")
        return value

    @field_validator("language")
    @classmethod
    def validate_language(cls, value: str):
        if value not in {"es", "eu"}:
            raise ValueError("Idioma no válido")
        return value

    @field_validator("account_status")
    @classmethod
    def validate_status(cls, value: str):
        if value not in ACCOUNT_STATUSES:
            raise ValueError("Estado de cuenta no válido")
        return value


class UserUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[str] = None
    account_status: Optional[str] = None
    assigned_team_ids: Optional[List[str]] = None
    assigned_category_ids: Optional[List[str]] = None
    player_id: Optional[str] = None
    family_id: Optional[str] = None
    language: Optional[str] = None
    notification_preferences: Optional[NotificationPreferences] = None

    @field_validator("language")
    @classmethod
    def validate_language(cls, value: Optional[str]):
        if value is not None and value not in {"es", "eu"}:
            raise ValueError("Idioma no válido")
        return value

    @field_validator("role")
    @classmethod
    def validate_update_role(cls, value: Optional[str]):
        if value is not None and value not in ROLES:
            raise ValueError("Rol no válido")
        return value

    @field_validator("account_status")
    @classmethod
    def validate_update_status(cls, value: Optional[str]):
        if value is not None and value not in ACCOUNT_STATUSES:
            raise ValueError("Estado de cuenta no válido")
        return value

    @field_validator("first_name", "last_name")
    @classmethod
    def validate_update_identity(cls, value: Optional[str]):
        if value is not None and len(normalized_text(value)) < 2:
            raise ValueError("Nombre y apellidos no válidos")
        return normalized_text(value) if value is not None else None

    @field_validator("email")
    @classmethod
    def validate_update_email(cls, value: Optional[str]):
        value = normalized_key(value)
        if value and ("@" not in value or value.startswith("@") or value.endswith("@")):
            raise ValueError("Correo no válido")
        return value or None

    @field_validator("phone")
    @classmethod
    def validate_update_phone(cls, value: Optional[str]):
        value = normalized_text(value)
        if value and (len(value) < 7 or len(value) > 24 or not re.fullmatch(r"[+() 0-9.-]+", value)):
            raise ValueError("Teléfono no válido")
        return value or None


async def validate_user_relationships(data: dict) -> dict:
    role = data.get("role")
    allow_incomplete = data.get("account_status") in {"pending_activation", "incomplete_link", "deactivated"}
    team_ids = sorted(set(ids(data.get("assigned_team_ids") or [])))
    category_ids = sorted(set(ids(data.get("assigned_category_ids") or [])))
    if any(normalized_key(value) == "no aplica" for value in [*team_ids, *category_ids]):
        raise HTTPException(status_code=422, detail="NO APLICA no es una vinculación válida")
    if role == "admin":
        team_ids, category_ids = [], []
        data["player_id"], data["family_id"], data["linked_player_ids"] = None, None, []
    if role == "player" and not data.get("player_id") and not allow_incomplete:
        raise HTTPException(status_code=422, detail="El rol jugador requiere un jugador asociado")
    if role == "family" and not data.get("family_id") and not allow_incomplete:
        raise HTTPException(status_code=422, detail="El rol familia requiere una familia asociada")
    if role in {"coordinator", "coach"} and not team_ids and not allow_incomplete:
        raise HTTPException(status_code=422, detail="El rol requiere al menos un equipo asignado")
    if role in {"coordinator", "coach"}:
        found = set(ids(await db.teams.distinct("id", {"id": {"$in": team_ids}})))
        if found != set(team_ids):
            raise HTTPException(status_code=422, detail="Uno o varios equipos no existen")
    if role == "coordinator" and category_ids:
        known_categories = set(ids(await db.teams.distinct("categoria")))
        if not set(category_ids).issubset(known_categories):
            raise HTTPException(status_code=422, detail="Una o varias categorías no existen")
        compatible = set(ids(await db.teams.distinct("id", {
            "id": {"$in": team_ids}, "categoria": {"$in": category_ids},
        })))
        if compatible != set(team_ids):
            raise HTTPException(status_code=422, detail="Hay equipos incompatibles con las categorías asignadas")
    if role == "family" and data.get("family_id"):
        family = await db.families.find_one({"id": data.get("family_id")}, {"_id": 0, "id": 1})
        if not family:
            raise HTTPException(status_code=422, detail="La familia asociada no existe")
        data["linked_player_ids"] = ids(await db.players.distinct("id", {"familia_id": data["family_id"]}))
        data["player_id"], team_ids, category_ids = None, [], []
    if role == "player" and data.get("player_id"):
        player = await db.players.find_one({"id": data.get("player_id")}, {"_id": 0, "id": 1})
        if not player:
            raise HTTPException(status_code=422, detail="El jugador asociado no existe")
        data["family_id"], data["linked_player_ids"], team_ids, category_ids = None, [], [], []
    if role not in {"family", "player"}:
        data["linked_player_ids"] = []
    if role != "admin" and not link_is_complete({**data, "assigned_team_ids": team_ids,
                                                   "assigned_category_ids": category_ids}):
        data["account_status"], data["active"] = "incomplete_link", False
    data["assigned_team_ids"] = team_ids
    data["assigned_category_ids"] = category_ids
    return data


async def record_user_audit(action: str, target: dict, changed_fields: list[str] | None = None,
                            previous: dict | None = None) -> None:
    actor = current_user_context.get() or {}
    await db.internal_events.insert_one({
        "id": new_id(), "type": f"user.{action}", "actor_user_id": actor.get("id"),
        "actor_role": actor.get("role"), "target_user_id": target.get("id"),
        "detail": safe_audit_detail(action, changed_fields, previous, target), "created_at": now_iso(),
    })


def system_admin_public() -> dict:
    return public_user({
        "id": "environment-admin", "username": None, "first_name": "Administrador del sistema",
        "last_name": "", "role": "admin", "active": True, "account_status": "active",
        "system_account": True, "read_only": True,
        "system_label": "Configurado en el servidor", "language": "es",
    })


def secured_public_user(user: dict) -> dict:
    return {**public_user(user), **security_public(user),
            "security_state": security_state(user), "link_complete": link_is_complete(user)}


async def ensure_admin_protection(existing: dict, candidate: dict) -> None:
    actor = current_user_context.get() or {}
    removing_admin = existing.get("role") == "admin" and (
        candidate.get("role") != "admin" or not status_is_active(account_status(candidate))
    )
    same_actor = actor.get("id") == existing.get("id") or actor.get("username") == existing.get("username")
    if removing_admin and same_actor:
        raise HTTPException(status_code=409, detail="No puedes retirar tu propio acceso administrativo")
    if removing_admin:
        others = await db.users.count_documents({
            "id": {"$ne": existing.get("id")}, "role": "admin", "active": {"$ne": False},
            "account_status": {"$nin": ["suspended", "deactivated"]},
        })
        if others == 0:
            raise HTTPException(status_code=409, detail="No se puede retirar al último administrador persistente")


@api_router.get("/users/permissions")
async def get_permission_matrix():
    return {
        role: {resource: sorted(actions) for resource, actions in resources.items()}
        for role, resources in ROLE_PERMISSIONS.items()
    }


@api_router.post("/users")
async def create_user(user: UserCreate):
    data = user.model_dump()
    data.pop("password_confirmation")
    access_method = data.pop("access_method", "password")
    if access_method in {"invitation", "pending"}:
        data["account_status"], data["active"] = "pending_activation", False
    data["username"] = normalized_text(data["username"])
    data["username_normalized"] = normalized_key(data["username"])
    data["email_normalized"] = normalized_key(data.get("email")) or None
    data = await validate_user_relationships(data)
    if data["username"] == ADMIN_USER or await db.users.find_one({"username_normalized": data["username_normalized"]}):
        raise HTTPException(status_code=409, detail="El nombre de usuario ya existe")
    if data.get("email_normalized") and await db.users.find_one({"email_normalized": data["email_normalized"]}):
        raise HTTPException(status_code=409, detail="El correo ya está asociado a otra cuenta")
    password = data.pop("password") or generate_temporary_password()
    must_change = access_method == "temporary_password"
    data.update({
        "id": new_id(), "password_hash": pwd_context.hash(password),
        "active": status_is_active(data["account_status"]),
        "must_change_password": must_change, "session_version": 0,
        "failed_login_count": 0, "locked_until": None,
        "created_at": now_iso(), "updated_at": now_iso(), "last_access_at": None,
    })
    data["notification_preferences"] = dict(data["notification_preferences"])
    await db.users.insert_one(dict(data))
    await record_user_audit("created", data, list(data.keys()))
    response = secured_public_user(data)
    if access_method == "temporary_password":
        response.update({"temporary_password": password, "show_once": True})
    elif access_method == "invitation":
        plain, record = issue_token(JWT_SECRET, ttl_minutes=0, ttl_hours=INVITATION_TTL_HOURS)
        await db.users.update_one({"id": data["id"]}, {"$set": {"invitation": record}})
        response.update({"invitation_token": plain, "expires_at": record["expires_at"], "show_once": True, "delivery": "not_sent"})
    return response


@api_router.get("/users")
async def get_users(
    page: Optional[int] = Query(default=None, ge=1), page_size: int = Query(default=25, ge=1, le=100),
    search: str = "", role: str = "", status: str = "", team_id: str = "",
    last_access: str = "", sort_by: str = "username", sort_dir: str = "asc",
):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(5000)
    public = [system_admin_public(), *[secured_public_user(user) for user in users]]
    if page is None:
        return sorted(public, key=lambda item: normalized_key(item.get("username") or item.get("first_name")))
    counters = {
        "total": len(public),
        "active": sum(account_status(user) == "active" for user in public),
        "pending": sum(account_status(user) == "pending_activation" for user in public),
        "blocked": sum(security_state(user) == "locked" for user in public),
        "deactivated": sum(account_status(user) == "deactivated" for user in public),
        "incomplete": sum(not link_is_complete(user) for user in public),
    }
    if role:
        public = [user for user in public if user.get("role") == role]
    if status:
        public = [user for user in public if account_status(user) == status or security_state(user) == status]
    if team_id:
        public = [user for user in public if team_id in (user.get("assigned_team_ids") or [])]
    if search:
        needle = normalized_key(search)
        public = [user for user in public if needle in user_search_text(user)]
    if last_access == "never":
        public = [user for user in public if not user.get("last_access_at")]
    elif last_access in {"7d", "30d", "90d"}:
        cutoff = utcnow() - timedelta(days=int(last_access[:-1]))
        public = [user for user in public if parse_time(user.get("last_access_at")) and parse_time(user.get("last_access_at")) >= cutoff]
    allowed_sort = {"username", "first_name", "role", "account_status", "last_access_at"}
    sort_key = sort_by if sort_by in allowed_sort else "username"
    public.sort(key=lambda item: normalized_key(item.get(sort_key)), reverse=sort_dir == "desc")
    start = (page - 1) * page_size
    return {"items": public[start:start + page_size], "total": len(public), "page": page,
            "page_size": page_size, "pages": max(1, math.ceil(len(public) / page_size)), "counters": counters}


@api_router.get("/users/options")
async def get_user_administration_options():
    teams = await db.teams.find({}, {"_id": 0, "id": 1, "nombre": 1, "categoria": 1, "modalidad": 1, "temporada": 1, "estado": 1}).to_list(1000)
    usable_teams = [team for team in teams if normalized_key(team.get("nombre")) not in {"", "no aplica"}
                    and normalized_key(team.get("estado") or "activo") not in {"inactivo", "archivado", "cerrado"}]
    players = await db.players.find({}, {"_id": 0, "id": 1, "nombre": 1, "apellidos": 1, "familia_id": 1, "equipo_id": 1}).to_list(5000)
    families = await db.families.find({}, {"_id": 0, "id": 1, "progenitor1_nombre": 1, "contacto_principal": 1}).to_list(5000)
    return {"teams": usable_teams, "players": players, "families": families}


@api_router.get("/users/{user_id}/effective-permissions")
async def get_effective_permissions(user_id: str):
    if user_id == "environment-admin":
        user = {"id": user_id, "role": "admin", "active": True, "system_account": True}
    else:
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="No encontrado")
    role = user.get("role", "player")
    return {
        "role": role,
        "permissions": {resource: sorted(actions) for resource, actions in ROLE_PERMISSIONS.get(role, {}).items()},
        "scope": effective_scope(user),
    }


@api_router.get("/users/{user_id}/administration-profile")
async def get_user_administration_profile(user_id: str):
    if user_id == "environment-admin":
        system = system_admin_public()
        return {"user": {**system, "security_state": "verified", "link_complete": True},
                "permissions": {resource: sorted(actions) for resource, actions in ROLE_PERMISSIONS["admin"].items()},
                "scope": effective_scope(system), "activity": [], "sessions": {"individual_tracking": False},
                "communications": {"count": 0}, "read_only": True}
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0, "invitation": 0,
                                                         "recovery": 0, "previous_invitation": 0})
    if not user:
        raise HTTPException(status_code=404, detail="No encontrado")
    events = await db.internal_events.find(
        {"target_user_id": user_id}, {"_id": 0, "id": 1, "type": 1, "actor_user_id": 1,
                                      "actor_role": 1, "detail": 1, "created_at": 1},
    ).sort("created_at", -1).to_list(100)
    communication_count = await db.communications.count_documents({"$or": [
        {"recipient_user_id": user_id}, {"recipient_username": user.get("username")},
    ]})
    role = user.get("role", "player")
    return {"user": secured_public_user(user),
            "permissions": {resource: sorted(actions) for resource, actions in ROLE_PERMISSIONS.get(role, {}).items()},
            "scope": effective_scope(user), "activity": events,
            "sessions": {"individual_tracking": False, "sessions_revoked_at": user.get("sessions_revoked_at")},
            "communications": {"count": communication_count}, "read_only": False}


@api_router.get("/users/{user_id}")
async def get_user(user_id: str):
    if user_id == "environment-admin":
        return system_admin_public()
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="No encontrado")
    return secured_public_user(user)


@api_router.put("/users/{user_id}")
async def edit_user(user_id: str, changes: UserUpdate):
    if user_id == "environment-admin":
        raise HTTPException(status_code=403, detail="La cuenta del sistema es de solo lectura")
    existing = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="No encontrado")
    data = changes.model_dump(exclude_unset=True)
    if data.get("notification_preferences") is not None:
        data["notification_preferences"] = dict(data["notification_preferences"])
    if "email" in data:
        data["email_normalized"] = normalized_key(data.get("email")) or None
        if data["email_normalized"] and await db.users.find_one({
            "id": {"$ne": user_id}, "email_normalized": data["email_normalized"],
        }):
            raise HTTPException(status_code=409, detail="El correo ya está asociado a otra cuenta")
    candidate = await validate_user_relationships({**existing, **data})
    if "account_status" in data:
        candidate["active"] = status_is_active(candidate["account_status"])
    await ensure_admin_protection(existing, candidate)
    data = {key: value for key, value in candidate.items() if existing.get(key) != value and key not in {"_id", "password_hash"}}
    if "account_status" in data:
        data["active"] = status_is_active(data["account_status"])
    data["updated_at"] = now_iso()
    await db.users.update_one({"id": user_id}, {"$set": data})
    changed = set(data)
    if "role" in changed:
        audit_action = "role_changed"
    elif changed & {"assigned_team_ids", "assigned_category_ids", "player_id", "family_id", "linked_player_ids"}:
        audit_action = "scope_changed"
    elif "account_status" in changed:
        audit_action = {
            "active": "activated", "suspended": "suspended", "deactivated": "deactivated",
            "pending_activation": "activation_pending", "incomplete_link": "link_incomplete",
        }.get(candidate.get("account_status"), "status_changed")
    else:
        audit_action = "updated"
    await record_user_audit(audit_action, candidate, list(data.keys()), existing)
    return secured_public_user(candidate)


@api_router.delete("/users/{user_id}")
async def deactivate_user(user_id: str):
    if user_id == "environment-admin":
        raise HTTPException(status_code=403, detail="La cuenta del sistema es de solo lectura")
    existing = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="No encontrado")
    candidate = {**existing, "active": False, "account_status": "deactivated"}
    await ensure_admin_protection(existing, candidate)
    await db.users.update_one({"id": user_id}, {"$set": {
        "active": False, "account_status": "deactivated", "updated_at": now_iso(),
    }})
    await record_user_audit("deactivated", candidate, ["active", "account_status"], existing)
    return {"ok": True}


security_attempts = defaultdict(deque)


def enforce_sensitive_rate(actor: dict, action: str, maximum: int = 10) -> None:
    key = f"{actor.get('id')}:{action}"
    current = time.monotonic()
    bucket = security_attempts[key]
    while bucket and current - bucket[0] > 60:
        bucket.popleft()
    if len(bucket) >= maximum:
        raise HTTPException(status_code=429, detail="Demasiadas operaciones. Inténtalo más tarde")
    bucket.append(current)


async def mutable_security_user(user_id: str) -> dict:
    if user_id == "environment-admin":
        raise HTTPException(status_code=403, detail="La cuenta del sistema es de solo lectura")
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="No encontrado")
    return user


@api_router.get("/users/{user_id}/security")
async def get_user_security(user_id: str):
    if user_id == "environment-admin":
        return {**security_public({}), "system_account": True, "read_only": True}
    return security_public(await mutable_security_user(user_id))


@api_router.post("/users/{user_id}/security/temporary-password")
async def generate_user_temporary_password(user_id: str):
    actor = current_user_context.get() or {}
    enforce_sensitive_rate(actor, "temporary-password")
    user = await mutable_security_user(user_id)
    temporary = generate_temporary_password()
    moment = now_iso()
    await db.users.update_one({"id": user_id}, {"$set": {
        "password_hash": pwd_context.hash(temporary), "must_change_password": True,
        "last_password_change_at": moment, "sessions_revoked_at": moment,
    }, "$inc": {"session_version": 1}})
    await record_user_audit("temporary_password_generated", user, [])
    return {"ok": True, "temporary_password": temporary, "show_once": True}


@api_router.post("/users/{user_id}/security/invitation")
async def generate_user_invitation(user_id: str):
    actor = current_user_context.get() or {}
    enforce_sensitive_rate(actor, "invitation")
    user = await mutable_security_user(user_id)
    email = str(user.get("email") or "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(status_code=422, detail="La cuenta necesita un correo válido para enviar la invitación")
    public_url = (os.environ.get("PUBLIC_APP_URL") or "").strip().rstrip("/")
    if not public_url:
        origin = str(os.environ.get("CORS_ORIGINS") or "").strip().strip("[]")
        public_url = origin.split(",")[0].strip().strip('"').rstrip("/")
    if not public_url:
        raise HTTPException(status_code=503, detail="No está configurada la URL pública de activación")
    plain, record = issue_token(JWT_SECRET, ttl_minutes=0, ttl_hours=INVITATION_TTL_HOURS)
    previous = user.get("invitation") or {}
    if previous and not previous.get("used_at"):
        previous = {**previous, "cancelled_at": now_iso()}
    await db.users.update_one({"id": user_id}, {"$set": {
        "invitation": record, "previous_invitation": previous or None,
        "account_status": "pending_activation", "active": False,
    }})
    link = f"{public_url}/activar?token={quote(plain)}"
    delivery = dispatch_email(
        email, "Activa tu acceso a Ikas-Txiki",
        f"Hola,\n\nActiva tu acceso y crea tu contraseña personal:\n{link}\n\n"
        f"El enlace caduca en {INVITATION_TTL_HOURS} horas. Si no esperabas este correo, puedes ignorarlo.",
    )
    delivery.update({"type": "user_access_invitation", "user_id": user_id})
    await db.delivery_logs.insert_one(delivery)
    await record_user_audit("invitation_sent" if delivery["status"] == "sent" else "invitation_delivery_failed", user, [])
    if delivery["status"] != "sent":
        raise HTTPException(status_code=502, detail="No se pudo enviar la invitación; revisa la configuración de correo")
    return {"ok": True, "delivery": "sent", "expires_at": record["expires_at"]}


@api_router.delete("/users/{user_id}/security/invitation")
async def cancel_user_invitation(user_id: str):
    user = await mutable_security_user(user_id)
    if not user.get("invitation"):
        raise HTTPException(status_code=409, detail="No existe una invitación activa")
    await db.users.update_one({"id": user_id}, {"$set": {"invitation.cancelled_at": now_iso()}})
    await record_user_audit("invitation_cancelled", user, [])
    return {"ok": True}


@api_router.post("/users/{user_id}/security/revoke-sessions")
async def revoke_user_sessions(user_id: str):
    actor = current_user_context.get() or {}
    user = await mutable_security_user(user_id)
    if actor.get("id") == user_id:
        raise HTTPException(status_code=409, detail="No puedes revocar accidentalmente tu propia sesión")
    moment = now_iso()
    await db.users.update_one({"id": user_id}, {"$set": {"sessions_revoked_at": moment},
                                                   "$inc": {"session_version": 1}})
    await record_user_audit("sessions_revoked", user, [])
    return {"ok": True}


@api_router.post("/users/{user_id}/security/lock")
async def lock_user_access(user_id: str):
    actor = current_user_context.get() or {}
    user = await mutable_security_user(user_id)
    candidate = {**user, "account_status": "suspended", "active": False}
    await ensure_admin_protection(user, candidate)
    if actor.get("id") == user_id:
        raise HTTPException(status_code=409, detail="No puedes bloquear tu propia cuenta")
    until = (utcnow() + timedelta(minutes=ACCOUNT_LOCK_MINUTES)).isoformat()
    await db.users.update_one({"id": user_id}, {"$set": {"locked_until": until}})
    await record_user_audit("locked", user, [])
    return {"ok": True, "locked_until": until}


@api_router.post("/users/{user_id}/security/unlock")
async def unlock_user_access(user_id: str):
    user = await mutable_security_user(user_id)
    await db.users.update_one({"id": user_id}, {"$set": {"locked_until": None, "failed_login_count": 0}})
    await record_user_audit("unlocked", user, [])
    return {"ok": True}


# ================= PLAYERS =================
class Player(BaseModel):
    # Datos del formulario
    fecha_inscripcion: Optional[str] = None
    email_formulario: Optional[str] = None
    nombre: str
    apellidos: Optional[str] = ""
    fecha_nacimiento: Optional[str] = None
    centro_escolar: Optional[str] = None
    # Progenitores
    progenitor1_nombre: Optional[str] = None
    progenitor1_telefono: Optional[str] = None
    progenitor1_email: Optional[str] = None
    progenitor2_nombre: Optional[str] = None
    progenitor2_telefono: Optional[str] = None
    progenitor2_email: Optional[str] = None
    domicilio: Optional[str] = None
    # Deportivos / administrativos
    foto: Optional[str] = None  # base64 data url
    categoria: Optional[str] = None
    equipo_id: Optional[str] = None
    dorsal: Optional[str] = None
    posicion: Optional[str] = None
    estado: str = "pendiente_documentacion"  # activo, baja, lesionado, pendiente_documentacion, en_prueba
    numero_licencia: Optional[str] = None
    fecha_alta: Optional[str] = None
    fecha_baja: Optional[str] = None
    nueva_incorporacion: bool = False
    segundo_hermano: bool = False
    hermano_vinculado: Optional[str] = None
    descuento: Optional[float] = 0
    observaciones: Optional[str] = None
    familia_id: Optional[str] = None
    # Salud
    alergias: Optional[str] = None
    enfermedades: Optional[str] = None
    medicacion: Optional[str] = None
    seguro_medico: Optional[str] = None
    contacto_emergencia: Optional[str] = None
    telefono_emergencia: Optional[str] = None
    observaciones_medicas: Optional[str] = None
    # Equipación
    talla_camiseta: Optional[str] = None
    talla_pantalon: Optional[str] = None
    talla_chandal: Optional[str] = None
    talla_medias: Optional[str] = None
    talla_calzado: Optional[str] = None
    equipacion_entregada: bool = False
    fecha_entrega_equipacion: Optional[str] = None
    observaciones_material: Optional[str] = None
    # Datos históricos importados. Son referencias, no sustituyen asignaciones,
    # autorizaciones firmadas ni deudas oficiales.
    historial_equipacion: Dict[str, Any] = Field(default_factory=dict)
    segunda_equipacion: Dict[str, Any] = Field(default_factory=dict)
    historial_equipos: Dict[str, Any] = Field(default_factory=dict)
    historial_federacion: Dict[str, Any] = Field(default_factory=dict)
    historial_deportivo: Dict[str, Any] = Field(default_factory=dict)
    equipo_historico_referencia: Optional[str] = None
    historial_entrenamientos: List[Any] = Field(default_factory=list)
    referencias_cuotas: Dict[str, Any] = Field(default_factory=dict)
    referencias_permisos: Dict[str, Any] = Field(default_factory=dict)
    # Documentación
    doc_dni_jugador: bool = False
    doc_dni_tutor: bool = False
    doc_foto: bool = False
    doc_autorizacion: bool = False
    doc_justificante_pago: bool = False
    doc_ficha_federativa: bool = False
    estado_documental: str = "pendiente"  # completo, pendiente, incompleto
    fecha_revision_doc: Optional[str] = None
    observaciones_doc: Optional[str] = None


@api_router.post("/players")
async def create_player(player: Player):
    data = player.model_dump()
    if not data.get("fecha_inscripcion"):
        data["fecha_inscripcion"] = now_iso()
    if data.get("fecha_nacimiento"):
        data["categoria"] = compute_category(data["fecha_nacimiento"])
    return await insert_doc("players", data)


@api_router.get("/players")
async def get_players(equipo_id: Optional[str] = None, estado: Optional[str] = None,
                      categoria: Optional[str] = None, temporada: Optional[str] = None,
                      documentacion_pendiente: bool = False, q: Optional[str] = None):
    query: Dict[str, Any] = {}
    if equipo_id:
        query["equipo_id"] = equipo_id
    if estado:
        query["estado"] = estado
    if categoria:
        query["categoria"] = categoria
    if documentacion_pendiente:
        query["estado_documental"] = {"$ne": "completo"}
    docs = await list_docs("players", query)
    if temporada:
        season_team_ids = {
            team.get("id") for team in await list_docs("teams", {"temporada": temporada})
            if team.get("id")
        }
        docs = [player for player in docs if player.get("equipo_id") in season_team_ids]
    if q:
        ql = q.lower()
        docs = [d for d in docs if ql in (f"{d.get('nombre','')} {d.get('apellidos','')}").lower()]
    return docs


@api_router.get("/players/{player_id}")
async def get_player(player_id: str):
    return await get_doc("players", player_id)


@api_router.put("/players/{player_id}")
async def edit_player(player_id: str, player: Player):
    data = player.model_dump()
    if data.get("fecha_nacimiento"):
        data["categoria"] = compute_category(data["fecha_nacimiento"])
    return await update_doc("players", player_id, data)


@api_router.delete("/players/{player_id}")
async def remove_player(player_id: str):
    return await delete_doc("players", player_id)


# ================= FAMILIES =================
class Family(BaseModel):
    progenitor1_nombre: Optional[str] = None
    progenitor1_telefono: Optional[str] = None
    progenitor1_email: Optional[str] = None
    progenitor2_nombre: Optional[str] = None
    progenitor2_telefono: Optional[str] = None
    progenitor2_email: Optional[str] = None
    domicilio: Optional[str] = None
    contacto_principal: Optional[str] = None
    preferencia_comunicacion: Optional[str] = "email"  # email, telefono, telegram
    observaciones: Optional[str] = None


@api_router.post("/families")
async def create_family(family: Family):
    return await insert_doc("families", family.model_dump())


@api_router.get("/families")
async def get_families():
    fams = await list_docs("families")
    players = await list_docs("players")
    for f in fams:
        linked = [p for p in players if p.get("familia_id") == f["id"]]
        f["num_hijos"] = len(linked)
        f["hijos"] = [{"id": p["id"], "nombre": f"{p.get('nombre','')} {p.get('apellidos','')}"} for p in linked]
    return fams


@api_router.get("/families/{family_id}")
async def get_family(family_id: str):
    return await get_doc("families", family_id)


@api_router.put("/families/{family_id}")
async def edit_family(family_id: str, family: Family):
    return await update_doc("families", family_id, family.model_dump())


@api_router.delete("/families/{family_id}")
async def remove_family(family_id: str):
    return await delete_doc("families", family_id)


# ---------- Family access invitations ----------
class FamilyAccessInvitationRequest(BaseModel):
    family_ids: List[str] = Field(min_length=1, max_length=500)


def _family_access_email(family: dict) -> str | None:
    for key in ("progenitor1_email", "progenitor2_email"):
        value = normalized_key(family.get(key))
        if value and "@" in value:
            return value
    return None


async def _family_access_username(family: dict) -> str:
    raw = normalized_key(family.get("progenitor1_nombre") or family.get("contacto_principal") or "familia")
    base = re.sub(r"[^a-z0-9]+", "", raw)[:18] or "familia"
    for suffix in range(1, 10000):
        candidate = f"{base}.{suffix}"
        if not await db.users.find_one({"username_normalized": normalized_key(candidate)}, {"_id": 1}):
            return candidate
    raise HTTPException(status_code=409, detail="No se ha podido generar un usuario único para la familia")


def _require_family_access_admin() -> dict:
    actor = current_user_context.get() or {}
    if actor.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administración puede gestionar invitaciones familiares")
    return actor


@api_router.get("/account-provisioning/families")
async def get_family_accesses():
    _require_family_access_admin()
    families = await list_docs("families")
    players = await list_docs("players")
    users = await db.users.find({"role": "family"}, {"_id": 0, "password_hash": 0}).to_list(5000)
    users_by_family = {user.get("family_id"): user for user in users if user.get("family_id")}
    rows = []
    for family in families:
        user = users_by_family.get(family.get("id"))
        email = _family_access_email(family)
        children = [player for player in players if player.get("familia_id") == family.get("id")]
        state = "active" if user and account_status(user) == "active" else (
            invitation_status((user or {}).get("invitation")) if user else ("ready" if email else "missing_email")
        )
        rows.append({
            "family_id": family.get("id"), "family_name": family.get("progenitor1_nombre") or family.get("contacto_principal") or "Familia",
            "email": email, "children": [{"id": child.get("id"), "name": f"{child.get('nombre', '')} {child.get('apellidos', '')}".strip()} for child in children],
            "user_id": (user or {}).get("id"), "username": (user or {}).get("username"),
            "status": state, "invitation_expires_at": ((user or {}).get("invitation") or {}).get("expires_at"),
        })
    return sorted(rows, key=lambda row: normalized_key(row["family_name"]))


@api_router.post("/account-provisioning/families/invitations")
async def send_family_access_invitations(request: FamilyAccessInvitationRequest):
    _require_family_access_admin()
    family_ids = sorted(set(ids(request.family_ids)))
    families = await db.families.find({"id": {"$in": family_ids}}, {"_id": 0}).to_list(len(family_ids))
    found = {family.get("id") for family in families}
    if found != set(family_ids):
        raise HTTPException(status_code=404, detail="Una o varias familias no existen")
    public_url = (os.environ.get("PUBLIC_APP_URL") or os.environ.get("CORS_ORIGINS", "").split(",")[0]).rstrip("/")
    results = []
    for family in families:
        email = _family_access_email(family)
        if not email:
            results.append({"family_id": family["id"], "status": "skipped", "reason": "missing_email"})
            continue
        children = ids(await db.players.distinct("id", {"familia_id": family["id"]}))
        existing = await db.users.find_one({"role": "family", "family_id": family["id"]}, {"_id": 0})
        email_owner = await db.users.find_one({"email_normalized": email}, {"_id": 0, "id": 1, "family_id": 1})
        if email_owner and email_owner.get("family_id") != family["id"]:
            results.append({"family_id": family["id"], "status": "skipped", "reason": "email_already_associated"})
            continue
        if existing and account_status(existing) == "active":
            results.append({"family_id": family["id"], "status": "skipped", "reason": "already_active"})
            continue
        plain, invitation = issue_token(JWT_SECRET, ttl_minutes=0, ttl_hours=INVITATION_TTL_HOURS)
        if existing:
            await db.users.update_one({"id": existing["id"]}, {"$set": {
                "email": email, "email_normalized": email, "linked_player_ids": children,
                "invitation": invitation, "account_status": "pending_activation", "active": False, "updated_at": now_iso(),
            }})
            target = {**existing, "email": email}
            action = "family_invitation_resent"
        else:
            username = await _family_access_username(family)
            target = {
                "id": new_id(), "username": username, "username_normalized": normalized_key(username),
                "first_name": normalized_text(family.get("progenitor1_nombre") or family.get("contacto_principal") or "Familia"),
                "last_name": "", "email": email, "email_normalized": email, "phone": family.get("progenitor1_telefono"),
                "role": "family", "family_id": family["id"], "linked_player_ids": children,
                "assigned_team_ids": [], "assigned_category_ids": [], "player_id": None,
                "language": "es", "notification_preferences": NotificationPreferences().model_dump(),
                "password_hash": pwd_context.hash(generate_temporary_password()), "invitation": invitation,
                "account_status": "pending_activation", "active": False, "must_change_password": False,
                "session_version": 0, "failed_login_count": 0, "locked_until": None,
                "created_at": now_iso(), "updated_at": now_iso(), "last_access_at": None,
            }
            await db.users.insert_one(dict(target))
            action = "family_invitation_created"
        link = f"{public_url}/activar?token={quote(plain)}"
        delivery = dispatch_email(email, "Activa tu acceso a Ikas-Txiki", f"Hola,\n\nActiva tu acceso familiar y crea tu contraseña personal:\n{link}\n\nEl enlace caduca en 48 horas. Si no esperabas este correo, puedes ignorarlo.")
        delivery.update({"type": "family_access_invitation", "family_id": family["id"], "user_id": target["id"]})
        await db.delivery_logs.insert_one(delivery)
        await record_user_audit(action, target, ["family_id", "email", "linked_player_ids"])
        results.append({"family_id": family["id"], "status": delivery["status"], "reason": delivery.get("error")})
    return {"ok": True, "results": results, "sent": sum(item["status"] == "sent" for item in results)}


class PlayerAccessInvitationRequest(BaseModel):
    player_ids: List[str] = Field(min_length=1, max_length=500)


def _upper_category_player(player: dict) -> bool:
    return normalized_key(player.get("categoria")) in {"juvenil", "senior", "sénior"}


@api_router.get("/account-provisioning/players")
async def get_player_accesses():
    _require_family_access_admin()
    players = await list_docs("players")
    families = {item.get("id"): item for item in await list_docs("families") if item.get("id")}
    users = await db.users.find({"role": "player"}, {"_id": 0, "password_hash": 0}).to_list(5000)
    users_by_player = {user.get("player_id"): user for user in users if user.get("player_id")}
    rows = []
    for player in players:
        if not _upper_category_player(player):
            continue
        user = users_by_player.get(player.get("id"))
        email = normalized_key(player.get("email_formulario")) or None
        family = families.get(player.get("familia_id"), {})
        state = "active" if user and account_status(user) == "active" else (
            invitation_status((user or {}).get("invitation")) if user else ("ready" if email else "missing_email")
        )
        rows.append({
            "player_id": player.get("id"),
            "player_name": f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip(),
            "category": player.get("categoria"), "email": email,
            "family_id": player.get("familia_id"),
            "family_name": family.get("progenitor1_nombre") or family.get("contacto_principal"),
            "user_id": (user or {}).get("id"), "username": (user or {}).get("username"),
            "status": state, "invitation_expires_at": ((user or {}).get("invitation") or {}).get("expires_at"),
        })
    return sorted(rows, key=lambda row: normalized_key(row["player_name"]))


@api_router.post("/account-provisioning/players/invitations")
async def send_player_access_invitations(request: PlayerAccessInvitationRequest):
    _require_family_access_admin()
    player_ids = sorted(set(ids(request.player_ids)))
    players = await db.players.find({"id": {"$in": player_ids}}, {"_id": 0}).to_list(len(player_ids))
    if {player.get("id") for player in players} != set(player_ids):
        raise HTTPException(status_code=404, detail="Uno o varios jugadores no existen")
    public_url = (os.environ.get("PUBLIC_APP_URL") or os.environ.get("CORS_ORIGINS", "").split(",")[0]).rstrip("/")
    results = []
    for player in players:
        if not _upper_category_player(player):
            results.append({"player_id": player["id"], "status": "skipped", "reason": "category_not_eligible"})
            continue
        email = normalized_key(player.get("email_formulario")) or None
        if not email or "@" not in email:
            results.append({"player_id": player["id"], "status": "skipped", "reason": "missing_personal_email"})
            continue
        existing = await db.users.find_one({"role": "player", "player_id": player["id"]}, {"_id": 0})
        email_owner = await db.users.find_one({"email_normalized": email}, {"_id": 0, "id": 1, "player_id": 1})
        if email_owner and email_owner.get("player_id") != player["id"]:
            results.append({"player_id": player["id"], "status": "skipped", "reason": "email_already_associated"})
            continue
        if existing and account_status(existing) == "active":
            results.append({"player_id": player["id"], "status": "skipped", "reason": "already_active"})
            continue
        plain, invitation = issue_token(JWT_SECRET, ttl_minutes=0, ttl_hours=INVITATION_TTL_HOURS)
        if existing:
            await db.users.update_one({"id": existing["id"]}, {"$set": {
                "email": email, "email_normalized": email, "family_id": player.get("familia_id"),
                "invitation": invitation, "account_status": "pending_activation", "active": False, "updated_at": now_iso(),
            }})
            target = {**existing, "email": email, "family_id": player.get("familia_id")}
            action = "player_invitation_resent"
        else:
            username = await _family_access_username({"progenitor1_nombre": f"{player.get('nombre', '')} {player.get('apellidos', '')}"})
            target = {
                "id": new_id(), "username": username, "username_normalized": normalized_key(username),
                "first_name": normalized_text(player.get("nombre") or "Jugador"),
                "last_name": normalized_text(player.get("apellidos") or "Ikas-Txiki"),
                "email": email, "email_normalized": email, "phone": None,
                "role": "player", "player_id": player["id"], "family_id": player.get("familia_id"),
                "linked_player_ids": [], "assigned_team_ids": [], "assigned_category_ids": [],
                "language": "es", "notification_preferences": NotificationPreferences().model_dump(),
                "password_hash": pwd_context.hash(generate_temporary_password()), "invitation": invitation,
                "account_status": "pending_activation", "active": False, "must_change_password": False,
                "session_version": 0, "failed_login_count": 0, "locked_until": None,
                "created_at": now_iso(), "updated_at": now_iso(), "last_access_at": None,
            }
            await db.users.insert_one(dict(target))
            action = "player_invitation_created"
        link = f"{public_url}/activar?token={quote(plain)}"
        delivery = dispatch_email(email, "Activa tu acceso de jugador a Ikas-Txiki", f"Hola,\n\nActiva tu acceso de jugador y crea tu contraseña personal:\n{link}\n\nEl enlace caduca en 48 horas. Si no esperabas este correo, puedes ignorarlo.")
        delivery.update({"type": "player_access_invitation", "player_id": player["id"], "user_id": target["id"]})
        await db.delivery_logs.insert_one(delivery)
        await record_user_audit(action, target, ["player_id", "family_id", "email"])
        results.append({"player_id": player["id"], "status": delivery["status"], "reason": delivery.get("error")})
    return {"ok": True, "results": results, "sent": sum(item["status"] == "sent" for item in results)}


# ================= TEAMS =================
class Team(BaseModel):
    nombre: str
    categoria: Optional[str] = None
    modalidad: Optional[str] = None
    temporada: Optional[str] = None
    entrenador: Optional[str] = None
    segundo_entrenador: Optional[str] = None
    delegado: Optional[str] = None
    dias_entrenamiento: Optional[str] = None
    horario: Optional[str] = None
    campo: Optional[str] = None
    limite_jugadores: Optional[int] = 20
    estado: str = "activo"  # activo, cerrado, pendiente


@api_router.post("/teams")
async def create_team(team: Team):
    if normalized_key(team.nombre) == "no aplica":
        raise HTTPException(status_code=422, detail="NO APLICA no es un equipo gestionable")
    return await insert_doc("teams", team.model_dump())


@api_router.get("/teams")
async def get_teams():
    teams = [
        team for team in await list_docs("teams")
        if normalized_key(team.get("nombre")) != "no aplica"
    ]
    players = await list_docs("players")
    for t in teams:
        t["num_jugadores"] = len([p for p in players if p.get("equipo_id") == t["id"]])
    return teams


@api_router.get("/teams/{team_id}")
async def get_team(team_id: str):
    team = await get_doc("teams", team_id)
    players = await list_docs("players", {"equipo_id": team_id})
    team["jugadores"] = players
    return team


@api_router.put("/teams/{team_id}")
async def edit_team(team_id: str, team: Team):
    if normalized_key(team.nombre) == "no aplica":
        raise HTTPException(status_code=422, detail="NO APLICA no es un equipo gestionable")
    return await update_doc("teams", team_id, team.model_dump())


@api_router.delete("/teams/{team_id}")
async def remove_team(team_id: str):
    return await delete_doc("teams", team_id)


# ================= MATCHES =================
class Match(BaseModel):
    temporada: Optional[str] = None
    jornada: Optional[str] = None
    fecha: Optional[str] = None
    hora: Optional[str] = None
    equipo_id: Optional[str] = None
    rival: Optional[str] = None
    condicion: str = "local"  # local, visitante
    campo: Optional[str] = None
    direccion_campo: Optional[str] = None
    tipo: str = "liga"  # liga, copa, amistoso, torneo
    estado: str = "programado"  # programado, jugado, aplazado, suspendido, cancelado
    resultado_propio: Optional[int] = None
    resultado_rival: Optional[int] = None
    observaciones: Optional[str] = None


@api_router.post("/matches")
async def create_match(match: Match):
    return await insert_doc("matches", match.model_dump())


@api_router.get("/matches")
async def get_matches(equipo_id: Optional[str] = None, estado: Optional[str] = None):
    query: Dict[str, Any] = {}
    if equipo_id:
        query["equipo_id"] = equipo_id
    if estado:
        query["estado"] = estado
    matches = await list_docs("matches", query)
    teams = {t["id"]: t["nombre"] for t in await list_docs("teams")}
    for m in matches:
        m["equipo_nombre"] = teams.get(m.get("equipo_id"), "—")
    return matches


@api_router.get("/matches/{match_id}")
async def get_match(match_id: str):
    return await get_doc("matches", match_id)


@api_router.put("/matches/{match_id}")
async def edit_match(match_id: str, match: Match):
    previous = await get_doc("matches", match_id)
    updated = await update_doc("matches", match_id, match.model_dump())
    changed = any(previous.get(field) != updated.get(field) for field in ("fecha", "hora", "campo"))
    if changed:
        users = await users_for_team_players(updated.get("equipo_id"))
        await enqueue_notifications(users, "schedule.changed", "Cambio de partido / Partida aldaketa",
                                    str(updated.get("rival") or "Partido"), "/partidos", "urgent",
                                    {"match_id": match_id}, f"schedule.changed:match:{match_id}:{updated.get('updated_at')}")
    return updated


@api_router.delete("/matches/{match_id}")
async def remove_match(match_id: str):
    return await delete_doc("matches", match_id)


# ================= MATCH REPORTS / PERFORMANCE =================
MATCH_REPORT_STATUSES = {"draft", "closed", "reopened"}


class MatchReportParticipantPayload(BaseModel):
    player_id: str
    player_name: Optional[str] = Field(default=None, max_length=300)
    called_up: bool = False
    callup_response: Optional[str] = None
    callup_response_original: Optional[str] = None
    role: str = "did_not_play"
    played: bool = False
    shirt_number: Optional[str] = Field(default=None, max_length=20)
    initial_position: Optional[str] = Field(default=None, max_length=120)
    minutes: int = Field(default=0, ge=0, le=1000)
    initial_period: Optional[str] = None
    final_period: Optional[str] = None
    period_ids: List[str] = Field(default_factory=list)
    entries: int = Field(default=0, ge=0, le=100)
    exits: int = Field(default=0, ge=0, le=100)
    changes: List[Dict[str, Any]] = Field(default_factory=list)
    goals: int = Field(default=0, ge=0, le=100)
    own_goals: int = Field(default=0, ge=0, le=100)
    incidents: List[str] = Field(default_factory=list)
    non_participation_reason: Optional[str] = Field(default=None, max_length=1000)
    exceptional_reason: Optional[str] = Field(default=None, max_length=1000)
    availability_override_reason: Optional[str] = Field(default=None, max_length=1000)
    minutes_override_reason: Optional[str] = Field(default=None, max_length=1000)
    internal_notes: Optional[str] = Field(default=None, max_length=4000)
    origin: str = "manual"
    original_value: Optional[Any] = None
    warning_confirmed: bool = False

    @field_validator("player_id", "role", "origin")
    @classmethod
    def required_match_report_text(cls, value: str):
        normalized = str(value or "").strip()
        if not normalized:
            raise ValueError("El participante contiene un campo obligatorio vacío")
        return normalized

    @field_validator(
        "internal_notes", "shirt_number", "initial_position", "non_participation_reason",
        "exceptional_reason", "availability_override_reason", "minutes_override_reason",
    )
    @classmethod
    def normalize_match_report_notes(cls, value: Optional[str]):
        normalized = str(value or "").strip()
        return normalized or None


class MatchReportSubstitutionPayload(BaseModel):
    id: Optional[str] = None
    incoming_player_id: str
    outgoing_player_id: str
    period_id: str
    minute: int = Field(ge=0, le=1000)
    notes: Optional[str] = Field(default=None, max_length=1000)


class MatchReportGoalPayload(BaseModel):
    id: Optional[str] = None
    kind: str = "player"
    scorer_player_id: Optional[str] = None
    period_id: str
    minute: int = Field(ge=0, le=1000)
    notes: Optional[str] = Field(default=None, max_length=1000)


class MatchReportSavePayload(BaseModel):
    version: int = Field(ge=1)
    participants: List[MatchReportParticipantPayload] = Field(default_factory=list, max_length=500)
    internal_notes: Optional[str] = Field(default=None, max_length=6000)
    period_configuration: Optional[Dict[str, Any]] = None
    substitutions: List[MatchReportSubstitutionPayload] = Field(default_factory=list, max_length=500)
    goal_events: List[MatchReportGoalPayload] = Field(default_factory=list, max_length=500)


class MatchReportParticipantSavePayload(BaseModel):
    version: int = Field(ge=1)
    participant: MatchReportParticipantPayload


class MatchReportClosePayload(BaseModel):
    version: int = Field(ge=1)
    confirm_warnings: bool = False
    confirm_score_discrepancy: bool = False
    score_discrepancy_reason: Optional[str] = Field(default=None, max_length=1000)


class MatchReportReopenPayload(BaseModel):
    version: int = Field(ge=1)
    reason: str = Field(min_length=3, max_length=1000)

    @field_validator("reason")
    @classmethod
    def normalize_reopen_reason(cls, value: str):
        return value.strip()


class MatchReportImportDryRunPayload(BaseModel):
    rows: List[Dict[str, Any]] = Field(min_length=1, max_length=5000)


_match_report_write_lock = asyncio.Lock()


async def ensure_match_report_indexes() -> None:
    """Índices idempotentes; se invocan solo al escribir una acta.

    La colección usa participantes embebidos para que el acta completa se
    modifique de forma atómica. La combinación ``match_id + player_id`` queda
    protegida por la unicidad del acta y la validación previa de duplicados.
    """
    await db.match_reports.create_index([("match_id", 1)], unique=True, name="match_report_match_unique")
    await db.match_reports.create_index(
        [("team_id", 1), ("status", 1), ("updated_at", -1)], name="match_report_scope_status"
    )
    await db.match_reports.create_index(
        [("participants.player_id", 1), ("status", 1)], name="match_report_player_statistics"
    )


async def _match_report_context(match_id: str) -> dict:
    actor = current_user_context.get() or {}
    match = await get_doc("matches", match_id)
    team_id = match.get("equipo_id")
    if not team_id:
        raise HTTPException(status_code=422, detail="El partido no tiene equipo asociado")
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=422, detail="El equipo del partido no existe")
    if actor.get("role") != "admin" and team_id not in set(ids(actor.get("assigned_team_ids") or [])):
        raise HTTPException(status_code=403, detail="El partido queda fuera de tu ámbito")
    catalog = await _load_modality_catalog()
    normalized_modality = normalize_modality(match.get("modalidad") or team.get("modalidad"), catalog)
    if normalized_modality.status != "recognized" or not normalized_modality.active:
        raise HTTPException(status_code=422, detail="El partido no tiene una modalidad F7 o F11 activa")
    try:
        periods = period_configuration(normalized_modality.code)
    except MatchReportValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    callup_scope = await scope_for_collection("callups", actor)
    callup = await db.callups.find_one(
        merge_query({"match_id": match_id, "equipo_id": team_id}, callup_scope), {"_id": 0}
    )
    player_scope = await scope_for_collection("players", actor)
    players = await db.players.find(
        merge_query({"equipo_id": team_id, "estado": {"$ne": "baja"}}, player_scope), {"_id": 0}
    ).sort([("apellidos", 1), ("nombre", 1)]).to_list(500)
    return {
        "actor": actor, "match": match, "team": team, "callup": callup,
        "players": players, "players_by_id": {row["id"]: row for row in players},
        "modality": normalized_modality.code, "period_configuration": periods,
    }


def _match_report_header(context: dict) -> dict:
    match = context["match"]
    team = context["team"]
    return {
        "match_id": match.get("id"),
        "temporada": match.get("temporada") or team.get("temporada"),
        "categoria": match.get("categoria") or team.get("categoria"),
        "team_id": team.get("id"),
        "team_name": team.get("nombre"),
        "rival": match.get("rival"),
        "condition": match.get("condicion"),
        "fecha": match.get("fecha"),
        "hora": match.get("hora"),
        "campo": match.get("campo"),
        "modalidad": context["modality"],
        "jornada": match.get("jornada"),
        "competicion": match.get("tipo"),
        "match_status": match.get("estado"),
        "result": {
            "own": match.get("resultado_propio"),
            "rival": match.get("resultado_rival"),
        },
    }


def _match_report_callup(context: dict) -> dict:
    callup = context.get("callup") or {}
    rows = callup.get("convocados") or []
    return {
        "exists": bool(callup), "id": callup.get("id"), "count": len(rows),
        "responses": response_counts(rows) if rows else {"confirmed": 0, "declined": 0, "pending": 0},
    }


def _initial_match_report_participants(context: dict) -> list[dict]:
    actor = context["actor"]
    now = now_iso()
    callup_items = {
        row.get("player_id"): row for row in (context.get("callup") or {}).get("convocados", [])
        if row.get("player_id")
    }
    participants = []
    for player in context["players"]:
        called = callup_items.get(player["id"])
        participants.append({
            "player_id": player["id"], "called_up": bool(called),
            "player_name": f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip(),
            "callup_response": normalize_status(called.get("estado")) if called else None,
            "callup_response_original": normalize_status(called.get("estado")) if called else None,
            "role": "did_not_play" if called else "not_called", "played": False,
            "shirt_number": player.get("dorsal"), "initial_position": player.get("posicion"),
            "minutes": 0, "initial_period": None, "final_period": None, "period_ids": [],
            "entries": 0, "exits": 0, "changes": [], "goals": 0, "own_goals": 0,
            "incidents": [], "non_participation_reason": None, "exceptional_reason": None,
            "availability_override_reason": None, "minutes_override_reason": None,
            "internal_notes": None, "origin": "manual", "original_value": None,
            "warning_confirmed": False, "created_at": now, "created_by": actor.get("id"),
            "updated_at": now, "updated_by": actor.get("id"),
        })
    return participants


def _match_report_history_event(action: str, actor: dict, version: int, *, detail: Optional[dict] = None) -> dict:
    return {
        "id": new_id(), "action": action, "actor_user_id": actor.get("id"),
        "actor_role": actor.get("role"), "created_at": now_iso(), "version": version,
        "detail": detail or {},
    }


def _match_report_player_changes(previous: list[dict], current: list[dict]) -> list[dict]:
    """Devuelve un diff deportivo acotado para el historial inmutable del acta."""
    tracked_fields = (
        "called_up", "callup_response", "callup_response_original", "role", "played", "shirt_number", "initial_position",
        "minutes", "initial_period",
        "final_period", "period_ids", "entries", "exits", "changes", "goals",
        "own_goals", "incidents", "non_participation_reason", "exceptional_reason",
        "availability_override_reason", "minutes_override_reason", "internal_notes",
        "origin", "original_value",
    )
    before_by_id = {row.get("player_id"): row for row in previous}
    changes = []
    for row in current:
        player_id = row.get("player_id")
        before = before_by_id.get(player_id, {})
        fields = {
            field: {"previous": before.get(field), "new": row.get(field)}
            for field in tracked_fields if before.get(field) != row.get(field)
        }
        if fields:
            changes.append({"player_id": player_id, "fields": fields})
    return changes


async def _record_match_report_audit(report: dict, event: dict) -> None:
    await db.internal_events.insert_one({
        "id": new_id(), "type": f"match_report.{event['action']}",
        "actor_user_id": event.get("actor_user_id"), "actor_role": event.get("actor_role"),
        "match_id": report.get("match_id"), "team_id": report.get("team_id"),
        "report_id": report.get("id"), "status": report.get("status"),
        "version": event.get("version"), "detail": event.get("detail") or {},
        "created_at": event.get("created_at"), "result": "success",
    })


def _public_match_report(report: Optional[dict], context: dict) -> dict:
    header = _match_report_header(context)
    candidates = [{
        "id": row.get("id"), "name": f"{row.get('nombre', '')} {row.get('apellidos', '')}".strip(),
        "shirt_number": row.get("dorsal"),
    } for row in context["players"]]
    candidate_ids = {row["id"] for row in candidates}
    candidates.extend({
        "id": row.get("player_id"), "name": row.get("player_name") or "—", "shirt_number": row.get("shirt_number"),
    } for row in (report or {}).get("participants") or [] if row.get("player_id") not in candidate_ids)
    warning = None if context.get("callup") else "El partido no tiene convocatoria; determina manualmente la participación real"
    public = dict(report or {})
    public.pop("_id", None)
    public["header"] = header
    public["callup"] = _match_report_callup(context)
    public["candidates"] = candidates
    public["configuration"] = (report or {}).get("period_configuration") or context["period_configuration"]
    public["setup_warning"] = warning
    public["report"] = bool(report)
    return public


def _secure_match_report_participants(payloads: list[dict], existing: dict, context: dict) -> list[dict]:
    allowed_players = context["players_by_id"]
    previous_by_id = {row.get("player_id"): row for row in existing.get("participants") or []}
    allowed_player_ids = set(allowed_players) | set(previous_by_id)
    submitted_ids = {row.get("player_id") for row in payloads}
    if not submitted_ids.issubset(allowed_player_ids):
        raise HTTPException(status_code=422, detail="El acta contiene jugadores ajenos al equipo del partido")
    callup_items = {
        row.get("player_id"): row for row in (context.get("callup") or {}).get("convocados", [])
        if row.get("player_id")
    }
    now = now_iso()
    actor = context["actor"]
    secured = []
    for payload in payloads:
        try:
            row = normalize_participant(payload)
        except MatchReportValidationError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        callup_row = callup_items.get(row["player_id"])
        row["called_up"] = bool(callup_row)
        row["callup_response"] = normalize_status(callup_row.get("estado")) if callup_row else None
        if row["minutes_override_reason"] and actor.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Solo administración puede autorizar minutos excepcionales")
        previous = previous_by_id.get(row["player_id"]) or {}
        player = allowed_players.get(row["player_id"]) or {}
        row["player_name"] = previous.get("player_name") or f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip() or None
        row["callup_response_original"] = previous.get("callup_response_original", row["callup_response"])
        # La procedencia es metadato del servidor, no un campo editable. Esto
        # evita que una edición ordinaria suplante datos importados o inyecte
        # estructuras arbitrarias en ``original_value``.
        row["origin"] = previous.get("origin") or "manual"
        row["original_value"] = previous.get("original_value")
        row.update({
            "created_at": previous.get("created_at") or now,
            "created_by": previous.get("created_by") or actor.get("id"),
            "updated_at": now, "updated_by": actor.get("id"),
        })
        secured.append(row)
    return secured


def _secure_match_report_events(
    payloads: list[dict], existing: dict, context: dict, *, kind: str,
) -> list[dict]:
    actor = context["actor"]
    now = now_iso()
    existing_rows = {row.get("id"): row for row in existing.get(kind) or [] if row.get("id")}
    allowed_players = set(context["players_by_id"]) | {
        row.get("player_id") for row in existing.get("participants") or [] if row.get("player_id")
    }
    secured = []
    for payload in payloads:
        try:
            row = normalize_substitution(payload) if kind == "substitutions" else normalize_goal_event(payload)
        except MatchReportValidationError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        related = (
            {row["incoming_player_id"], row["outgoing_player_id"]}
            if kind == "substitutions" else ({row["scorer_player_id"]} if row.get("scorer_player_id") else set())
        )
        if not related.issubset(allowed_players):
            raise HTTPException(status_code=422, detail="El movimiento contiene jugadores ajenos al equipo")
        previous = existing_rows.get(row.get("id")) or {}
        row.update({
            "id": row.get("id") or new_id(),
            "created_at": previous.get("created_at") or now,
            "created_by": previous.get("created_by") or actor.get("id"),
            "updated_at": now,
            "updated_by": actor.get("id"),
        })
        secured.append(row)
    return secured


def _derive_match_report_participant_events(
    participants: list[dict], substitutions: list[dict], goal_events: list[dict],
) -> list[dict]:
    by_player = {row["player_id"]: dict(row) for row in participants}
    for row in by_player.values():
        row.update({"entries": 0, "exits": 0, "changes": [], "goals": 0})
    for movement in substitutions:
        incoming = by_player.get(movement.get("incoming_player_id"))
        outgoing = by_player.get(movement.get("outgoing_player_id"))
        change = {"period_id": movement.get("period_id"), "minute": movement.get("minute")}
        if incoming is not None:
            incoming["entries"] += 1
            incoming["changes"].append({**change, "kind": "entry"})
        if outgoing is not None:
            outgoing["exits"] += 1
            outgoing["changes"].append({**change, "kind": "exit"})
    for goal in goal_events:
        scorer = by_player.get(goal.get("scorer_player_id"))
        if goal.get("kind") == "player" and scorer is not None:
            scorer["goals"] += 1
    return list(by_player.values())


def _match_report_validation(report: dict, context: dict, *, strict: bool) -> dict:
    config = report.get("period_configuration") or context["period_configuration"]
    try:
        normalized_config = period_configuration(context["modality"], config)
        result = validate_participants(
            report.get("participants") or [], normalized_config,
            official_own_goals=None, strict=strict,
        )
        substitutions = validate_substitutions(
            result["participants"], report.get("substitutions") or [], normalized_config,
        )
        goals = validate_goal_events(
            result["participants"], report.get("goal_events") or [], normalized_config,
            official_own_goals=context["match"].get("resultado_propio"), strict=strict,
            discrepancy_confirmed=bool(report.get("score_discrepancy_confirmed")),
            discrepancy_reason=report.get("score_discrepancy_reason"),
        )
        result["errors"].extend(substitutions["errors"])
        result["errors"].extend(goals["errors"])
        result["warnings"].extend(goals["warnings"])
        result["substitutions"] = substitutions["substitutions"]
        result["goal_events"] = goals["goal_events"]
    except MatchReportValidationError as exc:
        return {"errors": [str(exc)], "warnings": [], "participants": [], "strict": strict}
    result["configuration"] = normalized_config
    result["can_close"] = not result["errors"]
    return result


async def _match_report_find(match_id: str, context: dict) -> Optional[dict]:
    query: Dict[str, Any] = {"match_id": match_id, "team_id": context["team"]["id"]}
    return await db.match_reports.find_one(query, {"_id": 0})


async def _atomic_match_report_update(match_id: str, version: int, values: dict, event: dict, context: dict) -> dict:
    query = {
        "match_id": match_id, "team_id": context["team"]["id"], "version": version,
        "status": {"$in": ["draft", "reopened"]},
    }
    updated = await db.match_reports.find_one_and_update(
        query,
        {"$set": {**values, "updated_at": now_iso(), "updated_by": context["actor"].get("id")},
         "$inc": {"version": 1}, "$push": {"history": event}},
        return_document=ReturnDocument.AFTER, projection={"_id": 0},
    )
    if not updated:
        current = await _match_report_find(match_id, context)
        if current and current.get("status") == "closed":
            raise HTTPException(status_code=409, detail="El acta está cerrada y no admite edición ordinaria")
        if current:
            raise HTTPException(status_code=409, detail="El acta cambió en otra sesión; vuelve a cargarla antes de guardar")
        raise HTTPException(status_code=404, detail="El acta no existe")
    await _record_match_report_audit(updated, event)
    return updated


@api_router.get("/match-reports/match/{match_id}")
async def get_match_report(match_id: str):
    context = await _match_report_context(match_id)
    report = await _match_report_find(match_id, context)
    return _public_match_report(report, context)


@api_router.post("/match-reports/match/{match_id}")
async def initialize_match_report(match_id: str):
    context = await _match_report_context(match_id)
    actor = context["actor"]
    await ensure_match_report_indexes()
    now = now_iso()
    event = _match_report_history_event("created", actor, 1, detail={"source": "manual"})
    document = {
        "id": new_id(), "match_id": match_id, "team_id": context["team"]["id"],
        "status": "draft", "origin": "manual", "version": 1,
        "period_configuration": context["period_configuration"],
        "participants": _initial_match_report_participants(context), "internal_notes": None,
        "substitutions": [], "goal_events": [],
        "score_discrepancy_reason": None, "score_discrepancy_confirmed": False,
        "created_at": now, "created_by": actor.get("id"), "updated_at": now, "updated_by": actor.get("id"),
        "closed_at": None, "closed_by": None, "reopened_at": None, "reopened_by": None,
        "reopen_reason": None, "history": [event],
    }
    async with _match_report_write_lock:
        try:
            await db.match_reports.insert_one(dict(document))
        except DuplicateKeyError as exc:
            raise HTTPException(status_code=409, detail="Ya existe un acta para este partido") from exc
    await _record_match_report_audit(document, event)
    return _public_match_report(document, context)


@api_router.put("/match-reports/match/{match_id}")
async def save_match_report(match_id: str, payload: MatchReportSavePayload):
    context = await _match_report_context(match_id)
    existing = await _match_report_find(match_id, context)
    if not existing:
        raise HTTPException(status_code=404, detail="El acta no existe")
    participants = _secure_match_report_participants(
        [row.model_dump() for row in payload.participants], existing, context
    )
    substitutions = _secure_match_report_events(
        [row.model_dump() for row in payload.substitutions], existing, context, kind="substitutions",
    )
    goal_events = _secure_match_report_events(
        [row.model_dump() for row in payload.goal_events], existing, context, kind="goal_events",
    )
    participants = _derive_match_report_participant_events(participants, substitutions, goal_events)
    try:
        config = period_configuration(
            context["modality"], payload.period_configuration or existing.get("period_configuration")
        )
    except MatchReportValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    existing_config = period_configuration(context["modality"], existing.get("period_configuration"))
    if config != existing_config and context["actor"].get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administración puede cambiar la configuración del partido")
    candidate = {
        **existing, "participants": participants, "period_configuration": config,
        "substitutions": substitutions, "goal_events": goal_events,
        "score_discrepancy_reason": None, "score_discrepancy_confirmed": False,
    }
    validation = _match_report_validation(candidate, context, strict=False)
    if validation["errors"]:
        raise HTTPException(status_code=422, detail={"message": "El acta contiene datos incompatibles", **validation})
    player_changes = _match_report_player_changes(existing.get("participants") or [], participants)
    event = _match_report_history_event(
        "updated", context["actor"], payload.version + 1,
        detail={
            "player_changes": player_changes,
            "internal_notes": {
                "previous": existing.get("internal_notes"),
                "new": str(payload.internal_notes or "").strip() or None,
            } if existing.get("internal_notes") != (str(payload.internal_notes or "").strip() or None) else None,
            "substitutions": {"previous": existing.get("substitutions") or [], "new": substitutions}
            if (existing.get("substitutions") or []) != substitutions else None,
            "goal_events": {"previous": existing.get("goal_events") or [], "new": goal_events}
            if (existing.get("goal_events") or []) != goal_events else None,
            "warnings": validation["warnings"],
        },
    )
    updated = await _atomic_match_report_update(
        match_id, payload.version,
        {"participants": participants, "period_configuration": config,
         "substitutions": substitutions, "goal_events": goal_events,
         "score_discrepancy_reason": None, "score_discrepancy_confirmed": False,
         "internal_notes": str(payload.internal_notes or "").strip() or None},
        event, context,
    )
    return {**_public_match_report(updated, context), "validation": validation}


@api_router.put("/match-reports/match/{match_id}/participants/{player_id}")
async def save_match_report_participant(match_id: str, player_id: str, payload: MatchReportParticipantSavePayload):
    if payload.participant.player_id != player_id:
        raise HTTPException(status_code=409, detail="El jugador de la ruta y del contenido no coincide")
    context = await _match_report_context(match_id)
    existing = await _match_report_find(match_id, context)
    if not existing:
        raise HTTPException(status_code=404, detail="El acta no existe")
    by_player = {row.get("player_id"): row for row in existing.get("participants") or []}
    by_player[player_id] = payload.participant.model_dump()
    participants = _secure_match_report_participants(list(by_player.values()), existing, context)
    participants = _derive_match_report_participant_events(
        participants, existing.get("substitutions") or [], existing.get("goal_events") or [],
    )
    validation = _match_report_validation({**existing, "participants": participants}, context, strict=False)
    if validation["errors"]:
        raise HTTPException(status_code=422, detail={"message": "El participante contiene datos incompatibles", **validation})
    event = _match_report_history_event(
        "participant_updated", context["actor"], payload.version + 1,
        detail={
            "player_id": player_id,
            "player_changes": _match_report_player_changes(existing.get("participants") or [], participants),
            "warnings": validation["warnings"],
        },
    )
    updated = await _atomic_match_report_update(
        match_id, payload.version, {"participants": participants}, event, context
    )
    return {**_public_match_report(updated, context), "validation": validation}


@api_router.post("/match-reports/match/{match_id}/validate")
async def validate_match_report(match_id: str):
    context = await _match_report_context(match_id)
    report = await _match_report_find(match_id, context)
    if not report:
        raise HTTPException(status_code=404, detail="El acta no existe")
    return _match_report_validation(report, context, strict=True)


@api_router.post("/match-reports/match/{match_id}/close")
async def close_match_report(match_id: str, payload: MatchReportClosePayload):
    context = await _match_report_context(match_id)
    report = await _match_report_find(match_id, context)
    if not report:
        raise HTTPException(status_code=404, detail="El acta no existe")
    if report.get("status") == "closed":
        return _public_match_report(report, context)
    discrepancy_reason = str(payload.score_discrepancy_reason or "").strip() or None
    discrepancy_confirmed = bool(payload.confirm_score_discrepancy)
    if discrepancy_confirmed and context["actor"].get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administración puede justificar una discrepancia de marcador")
    candidate = {
        **report,
        "score_discrepancy_reason": discrepancy_reason,
        "score_discrepancy_confirmed": discrepancy_confirmed,
    }
    validation = _match_report_validation(candidate, context, strict=True)
    if validation["errors"]:
        raise HTTPException(status_code=422, detail={"message": "El acta no puede cerrarse", **validation})
    if validation["warnings"] and not payload.confirm_warnings:
        raise HTTPException(status_code=409, detail={
            "message": "El acta contiene advertencias que requieren confirmación", **validation,
        })
    now = now_iso()
    event = _match_report_history_event(
        "closed", context["actor"], payload.version + 1,
        detail={
            "status": {"previous": report.get("status"), "new": "closed"},
            "score_discrepancy_reason": discrepancy_reason,
            "warnings_confirmed": bool(validation["warnings"]), "warnings": validation["warnings"],
        },
    )
    updated = await db.match_reports.find_one_and_update(
        {"match_id": match_id, "team_id": context["team"]["id"], "version": payload.version,
         "status": {"$in": ["draft", "reopened"]}},
        {"$set": {"status": "closed", "closed_at": now, "closed_by": context["actor"].get("id"),
                  "score_discrepancy_reason": discrepancy_reason,
                  "score_discrepancy_confirmed": discrepancy_confirmed,
                  "updated_at": now, "updated_by": context["actor"].get("id")},
         "$inc": {"version": 1}, "$push": {"history": event}},
        return_document=ReturnDocument.AFTER, projection={"_id": 0},
    )
    if not updated:
        raise HTTPException(status_code=409, detail="El acta cambió en otra sesión; vuelve a cargarla antes de cerrar")
    await _record_match_report_audit(updated, event)
    return {**_public_match_report(updated, context), "validation": validation}


@api_router.post("/match-reports/match/{match_id}/reopen")
async def reopen_match_report(match_id: str, payload: MatchReportReopenPayload):
    context = await _match_report_context(match_id)
    actor = context["actor"]
    if actor.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo un administrador puede reabrir un acta cerrada")
    now = now_iso()
    event = _match_report_history_event(
        "reopened", actor, payload.version + 1,
        detail={"status": {"previous": "closed", "new": "reopened"}, "reason": payload.reason},
    )
    updated = await db.match_reports.find_one_and_update(
        {"match_id": match_id, "team_id": context["team"]["id"], "version": payload.version,
         "status": "closed"},
        {"$set": {"status": "reopened", "reopened_at": now, "reopened_by": actor.get("id"),
                  "reopen_reason": payload.reason, "updated_at": now, "updated_by": actor.get("id")},
         "$inc": {"version": 1}, "$push": {"history": event}},
        return_document=ReturnDocument.AFTER, projection={"_id": 0},
    )
    if not updated:
        current = await _match_report_find(match_id, context)
        if not current:
            raise HTTPException(status_code=404, detail="El acta no existe")
        if current.get("status") != "closed":
            raise HTTPException(status_code=409, detail="Solo se puede reabrir un acta cerrada")
        raise HTTPException(status_code=409, detail="El acta cambió en otra sesión; vuelve a cargarla")
    await _record_match_report_audit(updated, event)
    return _public_match_report(updated, context)


@api_router.get("/match-reports/match/{match_id}/history")
async def get_match_report_history(match_id: str):
    context = await _match_report_context(match_id)
    report = await _match_report_find(match_id, context)
    if not report:
        raise HTTPException(status_code=404, detail="El acta no existe")
    return {"match_id": match_id, "version": report.get("version"), "history": report.get("history") or []}


def _match_report_pdf(report: dict, context: dict, lang: str) -> io.BytesIO:
    language = "eu" if lang == "eu" else "es"
    settings = context.get("settings") or {}
    header = _match_report_header(context)
    players = {
        row["id"]: f"{row.get('nombre', '')} {row.get('apellidos', '')}".strip()
        for row in context["players"]
    }
    for row in report.get("participants") or []:
        players.setdefault(row.get("player_id"), row.get("player_name") or "—")
    labels = {
        "es": {
            "title": "ACTA Y RENDIMIENTO / AKTA ETA ERRENDIMENDUA",
            "match": "Partido / Partida", "date": "Fecha / Data", "field": "Campo / Zelaia",
            "result": "Resultado / Emaitza", "player": "Jugador / Jokalaria",
            "role": "Situación / Egoera", "minutes": "Minutos / Minutuak", "goals": "Goles / Golak",
            "incidents": "Incidencias / Intzidentziak", "changes": "Cambios / Aldaketak",
            "closed": "Cierre / Itxiera",
        },
        "eu": {
            "title": "AKTA ETA ERRENDIMENDUA / ACTA Y RENDIMIENTO",
            "match": "Partida / Partido", "date": "Data / Fecha", "field": "Zelaia / Campo",
            "result": "Emaitza / Resultado", "player": "Jokalaria / Jugador",
            "role": "Egoera / Situación", "minutes": "Minutuak / Minutos", "goals": "Golak / Goles",
            "incidents": "Intzidentziak / Incidencias", "changes": "Aldaketak / Cambios",
            "closed": "Itxiera / Cierre",
        },
    }[language]
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle("MatchReportTitle", parent=styles["Title"], fontSize=15, leading=18, alignment=TA_CENTER)
    story = [
        Table(
            [[pdf_logo(settings.get("club_logo"), 15), Paragraph(labels["title"], title)]],
            colWidths=[19 * mm, 158 * mm],
            style=TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]),
        ),
        Spacer(1, 5 * mm),
        Paragraph(
            f"<b>{labels['match']}:</b> {html_lib.escape(str(header.get('team_name') or '—'))} · "
            f"{html_lib.escape(str(header.get('rival') or '—'))}", styles["BodyText"],
        ),
        Paragraph(f"<b>{labels['date']}:</b> {header.get('fecha') or '—'} {header.get('hora') or ''}", styles["BodyText"]),
        Paragraph(f"<b>{labels['field']}:</b> {html_lib.escape(str(header.get('campo') or '—'))}", styles["BodyText"]),
        Paragraph(
            f"<b>{labels['result']}:</b> {header.get('result', {}).get('own') if header.get('result', {}).get('own') is not None else '—'} - "
            f"{header.get('result', {}).get('rival') if header.get('result', {}).get('rival') is not None else '—'} · "
            f"{header.get('modalidad') or '—'}",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
    ]
    role_labels = {
        "starter": "Titular / Hasierakoa", "substitute": "Suplente / Ordezkoa",
        "did_not_play": "No participa / Ez du parte hartzen", "absent": "Ausencia / Absentzia",
        "late_withdrawal": "Baja / Baja", "not_called": "No convocado / Deitu gabe",
    }
    participant_rows = [[labels["player"], labels["role"], labels["minutes"], labels["goals"], labels["incidents"]]]
    for row in report.get("participants") or []:
        participant_rows.append([
            html_lib.escape(players.get(row.get("player_id"), "—")),
            role_labels.get(row.get("role"), str(row.get("role") or "—")),
            str(row.get("minutes") or 0), str(row.get("goals") or 0),
            html_lib.escape(", ".join(row.get("incidents") or []) or "—"),
        ])
    table = Table(participant_rows, colWidths=[48 * mm, 42 * mm, 23 * mm, 18 * mm, 47 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_TEAL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), .35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"), ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    story.extend([table, Spacer(1, 5 * mm), Paragraph(f"<b>{labels['changes']}:</b>", styles["Heading3"])])
    if report.get("substitutions"):
        for movement in report["substitutions"]:
            story.append(Paragraph(
                f"{movement.get('period_id') or '—'} · {movement.get('minute', 0)}' · "
                f"{html_lib.escape(players.get(movement.get('incoming_player_id'), '—'))} ↔ "
                f"{html_lib.escape(players.get(movement.get('outgoing_player_id'), '—'))}",
                styles["BodyText"],
            ))
    else:
        story.append(Paragraph("—", styles["BodyText"]))
    story.extend([Spacer(1, 4 * mm), Paragraph(f"<b>{labels['goals']}:</b>", styles["Heading3"])])
    if report.get("goal_events"):
        goal_kind = {
            "opponent_own_goal": "Autogol rival / Aurkariaren autogola",
            "unidentified": "Sin autor / Egilerik gabe",
        }
        for goal in report["goal_events"]:
            scorer = players.get(goal.get("scorer_player_id")) or goal_kind.get(goal.get("kind"), "—")
            story.append(Paragraph(
                f"{goal.get('period_id') or '—'} · {goal.get('minute', 0)}' · {html_lib.escape(str(scorer))}",
                styles["BodyText"],
            ))
    else:
        story.append(Paragraph("—", styles["BodyText"]))
    story.extend([
        Spacer(1, 4 * mm),
        Paragraph(
            f"<b>{labels['closed']}:</b> {report.get('closed_at') or '—'} · {html_lib.escape(str(report.get('closed_by') or '—'))}",
            styles["BodyText"],
        ),
    ])
    document.build(story)
    buffer.seek(0)
    return buffer


@api_router.get("/match-reports/match/{match_id}/export.pdf")
async def export_match_report_pdf(match_id: str, lang: str = "es"):
    context = await _match_report_context(match_id)
    report = await _match_report_find(match_id, context)
    if not report:
        raise HTTPException(status_code=404, detail="El acta no existe")
    if report.get("status") != "closed":
        raise HTTPException(status_code=409, detail="Solo se puede exportar un acta cerrada")
    context["settings"] = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    event = _match_report_history_event(
        "exported", context["actor"], report.get("version", 1), detail={"format": "pdf", "lang": lang},
    )
    rendered_pdf = _match_report_pdf(report, context, lang)
    await _record_match_report_audit(report, event)
    return StreamingResponse(
        rendered_pdf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="acta_{match_id}.pdf"'},
    )


@api_router.get("/match-reports/statistics/objective")
async def get_match_report_statistics(
    temporada: Optional[str] = None, categoria: Optional[str] = None, equipo_id: Optional[str] = None,
    player_id: Optional[str] = None, modalidad: Optional[str] = None, competicion: Optional[str] = None,
    desde: Optional[str] = None, hasta: Optional[str] = None,
):
    actor = current_user_context.get() or {}
    team_ids = set(ids(actor.get("assigned_team_ids") or []))
    if actor.get("role") != "admin" and equipo_id and equipo_id not in team_ids:
        raise HTTPException(status_code=403, detail="El equipo no pertenece a tu ámbito")
    report_query: Dict[str, Any] = {"status": "closed"}
    if actor.get("role") != "admin":
        report_query["team_id"] = {"$in": list(team_ids)}
    if equipo_id:
        report_query["team_id"] = equipo_id
    reports = await db.match_reports.find(report_query, {"_id": 0}).to_list(5000)
    match_ids = ids(row.get("match_id") for row in reports)
    matches = {row["id"]: row for row in await db.matches.find(
        {"id": {"$in": match_ids}}, {"_id": 0}
    ).to_list(5000)}
    team_rows = {row["id"]: row for row in await db.teams.find({}, {"_id": 0}).to_list(1000)}
    synchronized = []
    for report in reports:
        match = matches.get(report.get("match_id"), {})
        team = team_rows.get(report.get("team_id"), {})
        synchronized.append({
            **report, "temporada": match.get("temporada") or team.get("temporada"),
            "categoria": match.get("categoria") or team.get("categoria"),
            "equipo_id": report.get("team_id"), "modalidad": match.get("modalidad") or team.get("modalidad"),
            "competicion": match.get("tipo"), "fecha": match.get("fecha"),
        })
    filters = {key: value for key, value in {
        "temporada": temporada, "categoria": categoria, "equipo_id": equipo_id,
        "player_id": player_id, "modalidad": modalidad, "competicion": competicion,
        "desde": desde, "hasta": hasta,
    }.items() if value}
    result = build_objective_statistics(synchronized, filters)
    player_ids = ids(row.get("player_id") for row in result["players"])
    players = {
        row["id"]: row
        for row in await db.players.find(
            {"id": {"$in": player_ids}},
            {"_id": 0, "id": 1, "nombre": 1, "apellidos": 1},
        ).to_list(5000)
    }
    for row in result["players"]:
        player = players.get(row["player_id"], {})
        row["player_name"] = f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip() or row.get("player_name") or "—"
    for row in result.get("teams", []):
        row["team_name"] = (team_rows.get(row["team_id"]) or {}).get("nombre") or "—"
    return result


@api_router.post("/match-reports/import/dry-run")
async def dry_run_match_report_import(payload: MatchReportImportDryRunPayload):
    match_ids = {str(row.get("match_id") or "").strip() for row in payload.rows}
    player_ids = {str(row.get("player_id") or "").strip() for row in payload.rows}
    match_rows = await db.matches.find(
        {"id": {"$in": list(match_ids)}},
        {"_id": 0, "id": 1, "equipo_id": 1, "modalidad": 1},
    ).to_list(5000)
    player_rows = await db.players.find(
        {"id": {"$in": list(player_ids)}}, {"_id": 0, "id": 1, "equipo_id": 1}
    ).to_list(5000)
    team_ids = ids(row.get("equipo_id") for row in match_rows)
    teams = {row["id"]: row for row in await db.teams.find(
        {"id": {"$in": team_ids}}, {"_id": 0, "id": 1, "modalidad": 1}
    ).to_list(5000)}
    catalog = await _load_modality_catalog()
    match_contexts = {
        row["id"]: {
            "team_id": row.get("equipo_id"),
            "modality": normalize_modality(
                row.get("modalidad") or (teams.get(row.get("equipo_id")) or {}).get("modalidad"), catalog
            ).code,
        }
        for row in match_rows
    }
    player_contexts = {
        row["id"]: {"team_id": row.get("equipo_id")} for row in player_rows
    }
    known_matches = set(match_contexts)
    known_players = set(player_contexts)
    closed_matches = set(await db.match_reports.distinct(
        "match_id", {"match_id": {"$in": list(match_ids)}, "status": "closed"}
    ))
    return dry_run_historical_rows(
        payload.rows, known_match_ids=known_matches, known_player_ids=known_players,
        closed_match_ids=closed_matches, match_contexts=match_contexts, player_contexts=player_contexts,
    )


# ================= UNIFIED CALENDAR =================
class ClubCalendarEvent(BaseModel):
    titulo: str
    tipo: str = "club_event"
    fecha: str
    hora: Optional[str] = None
    fecha_fin: Optional[str] = None
    hora_fin: Optional[str] = None
    equipo_id: Optional[str] = None
    lugar: Optional[str] = None
    descripcion: Optional[str] = None
    temporada: Optional[str] = None
    categoria: Optional[str] = None

    @field_validator("tipo")
    @classmethod
    def validate_type(cls, value: str):
        if value not in {"meeting", "club_event"}:
            raise ValueError("El tipo debe ser meeting o club_event")
        return value

    @field_validator("titulo")
    @classmethod
    def validate_title(cls, value: str):
        value = value.strip()
        if not value or len(value) > 160:
            raise ValueError("El título debe tener entre 1 y 160 caracteres")
        return value

    @field_validator("fecha", "fecha_fin")
    @classmethod
    def validate_date(cls, value: Optional[str]):
        if value:
            date.fromisoformat(value[:10])
        return value


async def calendar_payload(start: Optional[str] = None, end: Optional[str] = None,
                           equipo_id: Optional[str] = None, categoria: Optional[str] = None,
                           temporada: Optional[str] = None, tipo: Optional[str] = None):
    actor = current_user_context.get() or {}
    teams = await list_docs("teams")
    team_ids = {team.get("id") for team in teams if team.get("id")}
    if equipo_id and equipo_id not in team_ids:
        raise HTTPException(status_code=403, detail="El equipo solicitado no pertenece a tu ámbito")
    categories = {team.get("categoria") for team in teams if team.get("categoria")}
    seasons = {team.get("temporada") for team in teams if team.get("temporada")}
    if categoria and actor.get("role") != "admin" and categoria not in categories:
        raise HTTPException(status_code=403, detail="La categoría solicitada no pertenece a tu ámbito")
    if temporada and actor.get("role") != "admin" and seasons and temporada not in seasons:
        raise HTTPException(status_code=403, detail="La temporada solicitada no pertenece a tu ámbito")
    if tipo and tipo not in CALENDAR_TYPES:
        raise HTTPException(status_code=422, detail="Tipo de evento no válido")
    events = aggregate_calendar_events(
        await list_docs("matches"), await list_docs("trainings"), await list_docs("club_events"), teams,
        start, end, equipo_id, categoria, temporada, tipo,
    )
    return {
        "events": events,
        "filter_options": {
            "teams": [{"id": team.get("id"), "name": team.get("nombre"), "category": team.get("categoria"), "season": team.get("temporada")} for team in teams],
            "categories": sorted(categories), "seasons": sorted(seasons), "types": list(CALENDAR_TYPES),
        },
        "subscription": subscription_capability(),
    }


@api_router.get("/calendar/events")
async def get_calendar_events(start: Optional[str] = None, end: Optional[str] = None,
                              equipo_id: Optional[str] = None, categoria: Optional[str] = None,
                              temporada: Optional[str] = None, tipo: Optional[str] = None):
    return await calendar_payload(start, end, equipo_id, categoria, temporada, tipo)


@api_router.post("/calendar/club-events")
async def create_club_calendar_event(event: ClubCalendarEvent):
    return await insert_doc("club_events", event.model_dump())


@api_router.put("/calendar/club-events/{event_id}")
async def edit_club_calendar_event(event_id: str, event: ClubCalendarEvent):
    return await update_doc("club_events", event_id, event.model_dump())


@api_router.delete("/calendar/club-events/{event_id}")
async def delete_club_calendar_event(event_id: str):
    return await delete_doc("club_events", event_id)


@api_router.get("/calendar/export.ics")
async def export_calendar_ical(start: Optional[str] = None, end: Optional[str] = None,
                               equipo_id: Optional[str] = None, categoria: Optional[str] = None,
                               temporada: Optional[str] = None, tipo: Optional[str] = None):
    payload = await calendar_payload(start, end, equipo_id, categoria, temporada, tipo)
    content = calendar_to_ical(payload["events"])
    return Response(content=content, media_type="text/calendar; charset=utf-8", headers={
        "Content-Disposition": "attachment; filename=ikas-txiki-calendario.ics",
        "Cache-Control": "private, no-store",
    })


# ================= FAMILY / PLAYER PORTAL =================
@api_router.get("/portal")
async def get_portal():
    actor = current_user_context.get() or {}
    role = actor.get("role")
    if role not in {"family", "player"}:
        raise HTTPException(status_code=403, detail="El portal está reservado a familias y jugadores")

    raw_players = await list_docs("players")
    player_ids = {player.get("id") for player in raw_players if player.get("id")}
    teams = await list_docs("teams")
    team_map = {team.get("id"): team for team in teams}
    players = [{**safe_player(player), "team_name": team_map.get(player.get("equipo_id"), {}).get("nombre")} for player in raw_players]

    matches = await list_docs("matches")
    trainings = await list_docs("trainings")
    club_events = await list_docs("club_events")
    calendar_events = aggregate_calendar_events(matches, trainings, club_events, teams)
    callups = portal_callups(await list_docs("callups"), player_ids)
    match_map = {match.get("id"): match for match in matches}
    for callup in callups:
        match = match_map.get(callup.get("match_id"), {})
        callup["match"] = {
            key: match.get(key) for key in ("id", "fecha", "hora", "rival", "campo", "estado")
        }
        callup["team_name"] = team_map.get(callup.get("equipo_id"), {}).get("nombre")
        callup["deadline_expired"] = is_late(callup.get("response_deadline"))

    authorizations = []
    for authorization in await list_docs("authorizations"):
        authorizations.append({
            key: authorization.get(key) for key in (
                "id", "player_id", "tipo", "estado", "fecha_firma", "fecha_caducidad",
                "persona_autorizada", "observaciones",
            )
        } | {"has_signed_file": bool(authorization.get("archivo_firmado"))})

    communications = [{key: item.get(key) for key in (
        "id", "asunto", "mensaje", "canal", "fecha_envio", "created_at",
    )} for item in await list_docs("communications")]
    payments = [safe_payment(item) for item in await list_docs("payments")] if role == "family" else []

    return {
        "role": role,
        "players": players,
        "teams": [{key: team.get(key) for key in ("id", "nombre", "categoria", "temporada", "entrenador")} for team in teams],
        "schedule": upcoming(calendar_events),
        "next_activity": next_calendar_event(calendar_events),
        "callups": callups,
        "attendance": portal_attendance(trainings, player_ids),
        "payments": payments,
        "authorizations": authorizations,
        "documents": document_status(raw_players),
        "communications": communications[:20],
    }


# ================= CALLUPS (Convocatorias) =================
class ConvocadoItem(BaseModel):
    player_id: str
    estado: str = "pending"
    motivo: Optional[str] = None
    responded_at: Optional[str] = None
    responded_by_user_id: Optional[str] = None
    responded_by_username: Optional[str] = None
    responded_by_role: Optional[str] = None
    late: bool = False
    history: List[Dict[str, Any]] = Field(default_factory=list)

    @field_validator("estado")
    @classmethod
    def normalize_legacy_status(cls, value: str):
        normalized = normalize_status(value)
        if normalized not in VALID_STATUSES:
            raise ValueError("Estado no válido")
        return normalized


class Callup(BaseModel):
    match_id: str
    equipo_id: Optional[str] = None
    convocados: List[ConvocadoItem] = Field(default_factory=list)
    hora_quedada: Optional[str] = None
    lugar_quedada: Optional[str] = None
    material: Optional[str] = None
    mensaje_familias: Optional[str] = None
    response_deadline: Optional[str] = None
    player_self_response_allowed: bool = False


class CallupResponse(BaseModel):
    player_id: str
    status: str
    reason: Optional[str] = None

    @field_validator("status")
    @classmethod
    def validate_status(cls, value: str):
        normalized = normalize_status(value)
        if normalized not in {"confirmed", "declined"}:
            raise ValueError("La respuesta debe ser confirmed o declined")
        return normalized


@api_router.post("/callups")
async def create_callup(callup: Callup):
    created = await insert_doc("callups", callup.model_dump())
    actor = current_user_context.get() or {}
    await db.internal_events.insert_one({
        "id": new_id(), "type": "callup.created", "callup_id": created["id"],
        "team_id": created.get("equipo_id"), "actor_user_id": actor.get("id"),
        "created_at": now_iso(), "delivered": False,
    })
    player_ids = ids(item.get("player_id") for item in created.get("convocados", []))
    users = await users_for_team_players(created.get("equipo_id"), player_ids)
    await enqueue_notifications(users, "callup.created", "Nueva convocatoria / Deialdi berria",
                                str(created.get("instrucciones") or "Ikas-Txiki"), "/convocatorias", "high",
                                {"callup_id": created["id"]}, f"callup.created:{created['id']}")
    return normalize_callup(created)


@api_router.get("/callups")
async def get_callups(estado: Optional[str] = None, q: Optional[str] = None):
    callups = await list_docs("callups")
    actor = current_user_context.get() or {}
    visible_player_ids = set(await user_player_ids(actor)) if actor.get("role") in {"family", "player"} else None
    matches = {m["id"]: m for m in await list_docs("matches")}
    teams = {t["id"]: t["nombre"] for t in await list_docs("teams")}
    players = {p["id"]: p for p in await list_docs("players")}
    wanted = normalize_status(estado) if estado else None
    for c in callups:
        normalized = normalize_callup(c)
        c.update(normalized)
        if visible_player_ids is not None:
            c["convocados"] = [item for item in c.get("convocados", []) if item.get("player_id") in visible_player_ids]
        m = matches.get(c.get("match_id"), {})
        c["match"] = m
        c["equipo_nombre"] = teams.get(c.get("equipo_id"), "—")
        for item in c.get("convocados", []):
            player = players.get(item.get("player_id"), {})
            item["nombre"] = f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip()
            item["dorsal"] = player.get("dorsal")
        c["num_convocados"] = len(c.get("convocados", []))
        c["response_counts"] = response_counts(c.get("convocados", []))
        c["deadline_expired"] = is_late(c.get("response_deadline"))
    if wanted:
        callups = [c for c in callups if any(i.get("estado") == wanted for i in c.get("convocados", []))]
    if q:
        needle = q.casefold()
        callups = [c for c in callups if any(needle in i.get("nombre", "").casefold() for i in c.get("convocados", []))]
    return callups


@api_router.get("/callups/stats/summary")
async def get_callup_stats(temporada: Optional[str] = None, equipo_id: Optional[str] = None):
    actor = current_user_context.get() or {}
    if equipo_id and actor.get("role") in {"coordinator", "coach"} and equipo_id not in set(ids(actor.get("assigned_team_ids") or [])):
        raise HTTPException(status_code=403, detail="El equipo no pertenece a tu ámbito")
    query: Dict[str, Any] = {}
    if equipo_id:
        query["equipo_id"] = equipo_id
    callups = await list_docs("callups", query)
    matches = {m["id"]: m for m in await list_docs("matches")}
    if temporada:
        callups = [c for c in callups if matches.get(c.get("match_id"), {}).get("temporada") == temporada]
    totals = response_counts(item for c in callups for item in c.get("convocados", []))
    return {"callups": len(callups), "responses": totals, "team_id": equipo_id, "season": temporada}


@api_router.get("/callups/{callup_id}")
async def get_callup(callup_id: str):
    c = normalize_callup(await get_doc("callups", callup_id))
    actor = current_user_context.get() or {}
    if actor.get("role") in {"family", "player"}:
        visible_player_ids = set(await user_player_ids(actor))
        c["convocados"] = [item for item in c.get("convocados", []) if item.get("player_id") in visible_player_ids]
    players = {p["id"]: p for p in await list_docs("players")}
    for item in c.get("convocados", []):
        p = players.get(item["player_id"], {})
        item["nombre"] = f"{p.get('nombre','')} {p.get('apellidos','')}".strip()
        item["dorsal"] = p.get("dorsal")
        item["foto"] = p.get("foto")
    match = await db.matches.find_one({"id": c.get("match_id")}, {"_id": 0}) or {}
    team = await db.teams.find_one({"id": c.get("equipo_id")}, {"_id": 0}) or {}
    c["match"] = match
    c["equipo_nombre"] = team.get("nombre", "—")
    c["response_counts"] = response_counts(c.get("convocados", []))
    c["deadline_expired"] = is_late(c.get("response_deadline"))
    return c


@api_router.put("/callups/{callup_id}")
async def edit_callup(callup_id: str, callup: Callup):
    existing = await get_doc("callups", callup_id)
    actor = current_user_context.get() or {}
    previous_items = {item.get("player_id"): item for item in existing.get("convocados", [])}
    data = callup.model_dump()
    secured_items = []
    history_events = []
    for submitted in data.get("convocados", []):
        previous = previous_items.get(submitted.get("player_id"))
        if not previous:
            secured_items.append({**submitted, "history": []})
            continue
        old_status = normalize_status(previous.get("estado"))
        new_status = normalize_status(submitted.get("estado"))
        if old_status == new_status:
            secured_items.append({**submitted, "history": previous.get("history", []),
                                  "responded_at": previous.get("responded_at"),
                                  "responded_by_user_id": previous.get("responded_by_user_id"),
                                  "responded_by_username": previous.get("responded_by_username"),
                                  "responded_by_role": previous.get("responded_by_role")})
            continue
        changed_at = now_iso()
        change = {
            "previous_status": old_status, "new_status": new_status, "changed_at": changed_at,
            "changed_by_user_id": actor.get("id"), "changed_by_username": actor.get("username"),
            "relation": actor.get("role"), "reason": submitted.get("motivo") if new_status == "declined" else None,
            "late": is_late(data.get("response_deadline")),
        }
        secured_items.append({**submitted, "motivo": change["reason"], "responded_at": changed_at,
                              "responded_by_user_id": actor.get("id"),
                              "responded_by_username": actor.get("username"),
                              "responded_by_role": actor.get("role"), "late": change["late"],
                              "history": [*(previous.get("history") or []), change]})
        history_events.append((submitted.get("player_id"), change))
    data["convocados"] = secured_items
    updated = normalize_callup(await update_doc("callups", callup_id, data))
    if history_events:
        await db.internal_events.insert_many([{
            "id": new_id(), "type": "callup.response_updated", "callup_id": callup_id,
            "player_id": player_id, "actor_user_id": actor.get("id"), "history": change,
            "created_at": now_iso(), "delivered": False,
        } for player_id, change in history_events])
    return updated


@api_router.patch("/callups/{callup_id}/respond")
async def respond_callup(callup_id: str, response: CallupResponse):
    user = current_user_context.get() or {}
    callup = await get_doc("callups", callup_id)
    allowed_players = set(await user_player_ids(user))
    if response.player_id not in allowed_players:
        raise HTTPException(status_code=403, detail="No puedes responder por este jugador")
    if user.get("role") == "player" and not callup.get("player_self_response_allowed", False):
        raise HTTPException(status_code=403, detail="La respuesta directa del jugador no está autorizada")
    index = next((i for i, item in enumerate(callup.get("convocados", []))
                  if item.get("player_id") == response.player_id), None)
    if index is None:
        raise HTTPException(status_code=404, detail="El jugador no está convocado")
    if is_late(callup.get("response_deadline")):
        await db.internal_events.update_one(
            {"type": "callup.deadline_expired", "callup_id": callup_id},
            {"$setOnInsert": {"id": new_id(), "type": "callup.deadline_expired",
                               "callup_id": callup_id, "created_at": now_iso(), "delivered": False}},
            upsert=True,
        )
        raise HTTPException(status_code=409, detail="El plazo de respuesta ha finalizado")
    old_status = normalize_status(callup["convocados"][index].get("estado"))
    updated, history = apply_response(
        callup["convocados"][index], response.status, response.reason,
        user, callup.get("response_deadline"),
    )
    await db.callups.update_one({"id": callup_id}, {"$set": {
        f"convocados.{index}": updated, "updated_at": now_iso(),
    }})
    event_type = "callup.response_updated" if old_status != "pending" else "callup.response_registered"
    await db.internal_events.insert_one({
        "id": new_id(), "type": event_type, "callup_id": callup_id,
        "player_id": response.player_id, "actor_user_id": user.get("id"),
        "history": history, "created_at": now_iso(), "delivered": False,
    })
    staff = await notification_users([callup.get("equipo_id")], include_admins=True)
    await enqueue_notifications(staff, "callup.response", "Respuesta de convocatoria / Deialdiaren erantzuna",
                                response.status, f"/convocatorias", "normal",
                                {"callup_id": callup_id, "player_id": response.player_id},
                                f"callup.response:{callup_id}:{response.player_id}:{updated.get('responded_at')}")
    return {"player_id": response.player_id, **updated}


@api_router.get("/callups/{callup_id}/pdf")
async def export_callup_pdf(callup_id: str):
    callup = await get_doc("callups", callup_id)
    match = await db.matches.find_one({"id": callup.get("match_id")}, {"_id": 0}) or {}
    team = await db.teams.find_one({"id": callup.get("equipo_id")}, {"_id": 0}) or {}
    settings = await db.settings.find_one({"id": "global"}, {"_id": 0}) or {}
    player_ids = [i.get("player_id") for i in callup.get("convocados", [])]
    rows = await db.players.find({"id": {"$in": player_ids}}, {"_id": 0}).to_list(500)
    players = {p["id"]: p for p in rows}
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=16*mm, leftMargin=16*mm,
                            topMargin=14*mm, bottomMargin=14*mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("CallupTitle", parent=styles["Title"], fontSize=16, leading=19, alignment=TA_CENTER)
    callup_header = Table(
        [[pdf_logo(settings.get("club_logo"), 16), Paragraph(str(settings.get("club_nombre") or "Ikas-Txiki"), styles["Heading2"])]],
        colWidths=[20 * mm, 154 * mm],
    )
    callup_header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    story = [
        callup_header,
        Paragraph("CONVOCATORIA / DEIALDIA", title), Spacer(1, 4*mm),
        Paragraph(f"<b>Equipo / Taldea:</b> {html_lib.escape(str(team.get('nombre') or '—'))}", styles["BodyText"]),
        Paragraph(f"<b>Partido / Partida:</b> {html_lib.escape(str(match.get('rival') or '—'))}", styles["BodyText"]),
        Paragraph(f"<b>Fecha y hora / Data eta ordua:</b> {match.get('fecha') or '—'} {match.get('hora') or ''}", styles["BodyText"]),
        Paragraph(f"<b>Lugar / Lekua:</b> {html_lib.escape(str(match.get('campo') or callup.get('lugar_quedada') or '—'))}", styles["BodyText"]),
        Spacer(1, 5*mm),
    ]
    table_data = [["Jugador / Jokalaria", "Estado / Egoera", "Motivo / Arrazoia"]]
    labels = {"pending": "Pendiente / Zain", "confirmed": "Confirmado / Baieztatuta", "declined": "Rechazado / Ukatuta"}
    for item in callup.get("convocados", []):
        player = players.get(item.get("player_id"), {})
        name = f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip() or "—"
        table_data.append([name, labels[normalize_status(item.get("estado"))], item.get("motivo") or "—"])
    table = Table(table_data, colWidths=[78*mm, 49*mm, 48*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_TEAL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), .4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"), ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    counts = response_counts(callup.get("convocados", []))
    story.extend([table, Spacer(1, 5*mm), Paragraph(
        f"<b>Resumen / Laburpena:</b> {counts['confirmed']} confirmados / baieztatuta · "
        f"{counts['declined']} rechazados / ukatuta · {counts['pending']} pendientes / zain",
        styles["BodyText"]), Paragraph(f"Generado / Sortuta: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["BodyText"]),
    ])
    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={
        "Content-Disposition": f'attachment; filename="convocatoria_{callup_id}.pdf"'
    })


@api_router.delete("/callups/{callup_id}")
async def remove_callup(callup_id: str):
    return await delete_doc("callups", callup_id)


# ================= PAYMENTS =================
class Payment(BaseModel):
    player_id: Optional[str] = None
    concepto: Optional[str] = "Cuota temporada"
    importe_base: Optional[float] = 0
    descuento_hermano: Optional[float] = 0
    importe_final: Optional[float] = 0
    forma_pago: Optional[str] = None  # domiciliacion, transferencia, efectivo, bizum
    iban: Optional[str] = None
    iban_validado: bool = False
    estado: str = "pendiente"  # pendiente, pagado, parcial, devuelto
    fecha_pago: Optional[str] = None
    recibo_generado: bool = False
    observaciones: Optional[str] = None
    titular_cuenta: Optional[str] = None
    historical_bank_reference: bool = False
    confirmed_debt: bool = False


@api_router.post("/payments")
async def create_payment(payment: Payment):
    data = payment.model_dump()
    base = data.get("importe_base") or 0
    desc = data.get("descuento_hermano") or 0
    data["importe_final"] = round(base - desc, 2)
    return await insert_doc("payments", data)


@api_router.get("/payments")
async def get_payments(estado: Optional[str] = None):
    query = {"estado": estado} if estado else {}
    payments = await list_docs("payments", query)
    players = {p["id"]: f"{p.get('nombre','')} {p.get('apellidos','')}".strip() for p in await list_docs("players")}
    for p in payments:
        p["player_nombre"] = players.get(p.get("player_id"), "—")
        p.pop("iban_encrypted", None)
        if p.get("iban_last4"):
            p["iban"] = masked_iban(p.get("iban_last4"))
    return payments


@api_router.put("/payments/{payment_id}")
async def edit_payment(payment_id: str, payment: Payment):
    data = payment.model_dump()
    base = data.get("importe_base") or 0
    desc = data.get("descuento_hermano") or 0
    data["importe_final"] = round(base - desc, 2)
    return await update_doc("payments", payment_id, data)


@api_router.delete("/payments/{payment_id}")
async def remove_payment(payment_id: str):
    return await delete_doc("payments", payment_id)


# ================= AUTHORIZATIONS =================
class Authorization(BaseModel):
    player_id: Optional[str] = None
    tipo: str = "general"  # general, imagen, medica, desplazamientos, recogida, proteccion_datos
    persona_autorizada: Optional[str] = None
    dni_autorizada: Optional[str] = None
    firmante: Optional[str] = None
    fecha_firma: Optional[str] = None
    fecha_caducidad: Optional[str] = None
    estado: str = "pendiente"  # pendiente, firmada, caducada
    archivo_firmado: Optional[str] = None  # ruta relativa al PDF firmado subido
    observaciones: Optional[str] = None


AUTHORIZATION_PDF_TYPES = {
    "general": {
        "es": "Autorización general de participación",
        "eu": "Parte hartzeko baimen orokorra",
        "text_es": "Autorizo a mi hijo/a a participar en todas las actividades deportivas, entrenamientos y partidos organizados por el club durante la temporada, así como a recibir la atención necesaria en caso de accidente leve.",
        "text_eu": "Nire seme-alabari klubak denboraldian antolatutako kirol-jarduera, entrenamendu eta partida guztietan parte hartzeko baimena ematen diot, baita istripu arin baten kasuan beharrezko arreta jasotzeko ere.",
    },
    "imagen": {
        "es": "Autorización de uso de imagen",
        "eu": "Irudia erabiltzeko baimena",
        "text_es": "Autorizo el uso de la imagen de mi hijo/a en fotografías y vídeos del club con fines informativos y de difusión de las actividades deportivas, sin contraprestación económica y respetando en todo momento la dignidad del menor.",
        "text_eu": "Nire seme-alabaren irudia klubaren argazki eta bideoetan erabiltzeko baimena ematen dut, kirol-jardueren berri emateko eta zabaltzeko, ordain ekonomikorik gabe eta adingabearen duintasuna une oro errespetatuz.",
    },
    "medica": {
        "es": "Autorización médica básica",
        "eu": "Oinarrizko mediku-baimena",
        "text_es": "Autorizo al club a tomar las medidas oportunas en caso de urgencia médica cuando no sea posible contactar con el/los tutor/es. Declaro estar informado/a del estado de salud de mi hijo/a y no tener conocimiento de impedimento médico para la práctica deportiva.",
        "text_eu": "Klubari baimena ematen diot larrialdi mediko baten aurrean beharrezko neurriak har ditzan, tutoreekin harremanetan jartzea ezinezkoa denean. Adierazten dut nire seme-alabaren osasun-egoeraren berri dudala eta ez dudala kirola egiteko eragozpen medikorik ezagutzen.",
    },
    "desplazamientos": {
        "es": "Autorización de desplazamientos",
        "eu": "Lekualdaketa baimena",
        "text_es": "Autorizo los desplazamientos de mi hijo/a en los vehículos del club o de otros progenitores/tutores a los distintos campos y localizaciones donde se disputen los partidos y actividades organizadas por el club.",
        "text_eu": "Nire seme-alabak klubeko ibilgailuetan edo beste guraso edo tutore batzuen ibilgailuetan bidaiatzeko baimena ematen dut, partidak eta klubak antolatutako jarduerak egiten diren zelai eta lekuetara joateko.",
    },
    "recogida": {
        "es": "Autorización para recogida por terceros",
        "eu": "Hirugarrenek jasotzeko baimena",
        "text_es": "Autorizo a la persona indicada a continuación a recoger a mi hijo/a tras los entrenamientos y partidos del club, en los casos en que yo no pueda hacerlo personalmente.",
        "text_eu": "Jarraian adierazitako pertsonari nire seme-alaba entrenamendu eta partiden ondoren jasotzeko baimena ematen diot, nik neuk jaso ezin dudanean.",
    },
    "proteccion_datos": {
        "es": "Protección de datos",
        "eu": "Datuen babesa",
        "text_es": "Doy mi consentimiento informado para el tratamiento de los datos personales de mi hijo/a conforme al Reglamento (UE) 2016/679 (RGPD) y la Ley Orgánica 3/2018 (LOPDGDD). Los datos serán utilizados exclusivamente para la gestión de las actividades del club y no serán cedidos a terceros sin consentimiento expreso.",
        "text_eu": "Nire seme-alabaren datu pertsonalak tratatzeko baimen informatua ematen dut, 2016/679 (EB) Erregelamenduaren (DBEO) eta 3/2018 Lege Organikoaren arabera. Datuak klubaren jarduerak kudeatzeko baino ez dira erabiliko, eta ez zaizkie hirugarrenei lagako berariazko baimenik gabe.",
    },
}


def build_authorization_pdf(auth: dict, player: dict, settings: dict, lang: str = "es") -> io.BytesIO:
    """Genera una autorización A4 vectorial sin depender del navegador."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title="Autorización Ikas-Txiki", author=settings.get("club_nombre") or "Ikas-Txiki",
    )
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("AuthNormal", parent=styles["Normal"], fontName="Helvetica", fontSize=10, leading=15, textColor=colors.HexColor("#1f2937"))
    small = ParagraphStyle("AuthSmall", parent=normal, fontSize=8, leading=11, textColor=colors.HexColor("#6b7280"))
    title = ParagraphStyle("AuthTitle", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=15, leading=19, alignment=TA_CENTER, spaceAfter=4, textColor=colors.HexColor("#111827"))
    club_style = ParagraphStyle("AuthClub", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=17, leading=20, textColor=colors.HexColor(BRAND_BLUE))
    esc = lambda value, fallback="": html_lib.escape(str(value if value not in (None, "") else fallback))

    club_name = settings.get("club_nombre") or "Ikas-Txiki"
    club_lines = [esc(settings.get("club_direccion")), esc(settings.get("club_email")), esc(settings.get("club_telefono"))]
    club_info = " · ".join(value for value in club_lines if value)
    header_cells = []
    try:
        header_cells.append(pdf_logo(settings.get("club_logo"), 18))
    except (OSError, ValueError):
        header_cells.append(Spacer(18 * mm, 18 * mm))
    header_cells.append(Paragraph(f"<b>{esc(club_name)}</b>{'<br/><font size=8>' + club_info + '</font>' if club_info else ''}", club_style))
    header = Table([header_cells], colWidths=[22 * mm, 150 * mm])
    header.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))

    type_data = AUTHORIZATION_PDF_TYPES.get(auth.get("tipo"), {"es": auth.get("tipo") or "Autorización", "eu": auth.get("tipo") or "Baimena", "text": ""})
    type_label_es = type_data.get("es") or auth.get("tipo") or "Autorización"
    type_label_eu = type_data.get("eu") or auth.get("tipo") or "Baimena"
    player_name = f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip() or "________________"
    season = settings.get("temporada_actual") or "2025-2026"
    story = [header, Spacer(1, 5 * mm), HRFlowable(width="100%", thickness=1.5, color=colors.HexColor(BRAND_BLUE)), Spacer(1, 8 * mm)]
    story.extend([
        Paragraph(esc(type_label_es).upper(), title),
        Paragraph(esc(type_label_eu).upper(), ParagraphStyle("AuthTitleEu", parent=title, fontSize=12, leading=15)),
        Paragraph(f"Temporada / Denboraldia {esc(season)}", ParagraphStyle("Season", parent=small, alignment=TA_CENTER)),
        Spacer(1, 7 * mm),
    ])

    details = [
        [Paragraph("Padre/Madre/Tutor/a · Aita/Ama/Tutorea:", small), Paragraph(f"<b>{esc(auth.get('firmante'), '________________')}</b>", normal)],
        [Paragraph("Jugador/a · Jokalaria:", small), Paragraph(f"<b>{esc(player_name)}</b>", normal)],
        [Paragraph("Fecha de nacimiento · Jaioteguna:", small), Paragraph(esc((player.get("fecha_nacimiento") or "—")[:10]), normal)],
        [Paragraph("Categoría · Kategoria:", small), Paragraph(esc(player.get("categoria"), "—"), normal)],
    ]
    details_table = Table(details, colWidths=[52 * mm, 116 * mm], rowHeights=[9 * mm] * 4)
    details_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LINEBELOW", (1, 0), (1, -1), 0.5, colors.HexColor("#9ca3af")),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.extend([details_table, Spacer(1, 7 * mm)])

    consent = Table([[Paragraph(
        f"<b>AUTORIZO:</b><br/>{esc(type_data.get('text_es'))}<br/><br/>"
        f"<b>BAIMENA EMATEN DUT:</b><br/>{esc(type_data.get('text_eu'))}", normal
    )]], colWidths=[168 * mm])
    consent.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")), ("LINEBEFORE", (0, 0), (0, 0), 3, colors.HexColor(BRAND_BLUE)), ("BOX", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")), ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12), ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10)]))
    story.extend([consent, Spacer(1, 6 * mm)])

    if auth.get("tipo") == "recogida":
        pickup = Table([
            [Paragraph("Persona autorizada · Baimendutako pertsona:", small), Paragraph(f"<b>{esc(auth.get('persona_autorizada'), '________________')}</b>", normal)],
            [Paragraph("DNI / NIF · NAN / IFZ:", small), Paragraph(esc(auth.get("dni_autorizada"), "________________"), normal)],
        ], colWidths=[52 * mm, 116 * mm])
        pickup.setStyle(TableStyle([("LINEBELOW", (1, 0), (1, -1), 0.5, colors.HexColor("#9ca3af")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        story.extend([pickup, Spacer(1, 5 * mm)])
    if auth.get("observaciones"):
        story.extend([Paragraph(f"<b>Observaciones · Oharrak:</b> <i>{esc(auth.get('observaciones'))}</i>", normal), Spacer(1, 7 * mm)])

    signature_style = ParagraphStyle("Signature", parent=small, alignment=TA_CENTER)
    signatures = Table([
        ["", "", ""],
        [Paragraph("Firma del padre/madre/tutor/a<br/>Aita/ama/tutorearen sinadura", signature_style), Paragraph(f"Fecha · Data: {esc(auth.get('fecha_firma'), '____________')}", signature_style), Paragraph("Sello del club<br/>Klubaren zigilua", signature_style)],
    ], colWidths=[52 * mm, 52 * mm, 52 * mm], rowHeights=[18 * mm, 12 * mm], hAlign="CENTER")
    signatures.setStyle(TableStyle([("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#111827")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6)]))
    story.extend([Spacer(1, 8 * mm), KeepTogether(signatures), Spacer(1, 10 * mm), HRFlowable(width="100%", thickness=0.4, color=colors.HexColor("#d1d5db")), Spacer(1, 2 * mm), Paragraph(f"{esc(club_name)} · Documento generado / Sortutako dokumentua: {date.today().strftime('%d/%m/%Y')} · RGPD/DBEO (UE/EB) 2016/679", ParagraphStyle("Footer", parent=small, alignment=TA_CENTER, fontSize=7))])
    doc.build(story)
    buffer.seek(0)
    return buffer


@api_router.post("/authorizations")
async def create_authorization(auth: Authorization):
    return await insert_doc("authorizations", auth.model_dump())


@api_router.post("/authorizations/ensure-all")
async def ensure_all_player_authorizations():
    """Create the missing authorization set for every player in the caller's scope."""
    authorization_types = tuple(AUTHORIZATION_PDF_TYPES.keys())
    players = await list_docs("players")
    player_ids = ids(player.get("id") for player in players)
    existing = await list_docs("authorizations")
    existing_keys = {
        (item.get("player_id"), item.get("tipo"))
        for item in existing
        if item.get("player_id") in player_ids
    }
    now = now_iso()
    missing = []
    for player_id in player_ids:
        for authorization_type in authorization_types:
            if (player_id, authorization_type) in existing_keys:
                continue
            document = Authorization(
                player_id=player_id, tipo=authorization_type, estado="pendiente"
            ).model_dump()
            await ensure_data_scope("authorizations", document)
            missing.append({
                **document, "id": new_id(), "created_at": now, "updated_at": now,
            })
    if missing:
        await db.authorizations.insert_many(missing)
    return {"players": len(player_ids), "created": len(missing), "types": len(authorization_types)}


@api_router.get("/authorizations")
async def get_authorizations(estado: Optional[str] = None):
    query = {"estado": estado} if estado else {}
    auths = await list_docs("authorizations", query)
    players = {p["id"]: p for p in await list_docs("players")}
    for a in auths:
        p = players.get(a.get("player_id"), {})
        a["player_nombre"] = f"{p.get('nombre','')} {p.get('apellidos','')}".strip() or "—"
    return auths


@api_router.get("/authorizations/{auth_id}/pdf")
async def download_authorization_pdf(auth_id: str, lang: str = "es"):
    auth = await get_doc("authorizations", auth_id)
    player = await get_doc("players", auth.get("player_id"))
    settings = await get_settings()
    buffer = build_authorization_pdf(auth, player, settings, lang)
    filename = f"autorizacion_{auth_id}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.put("/authorizations/{auth_id}")
async def edit_authorization(auth_id: str, auth: Authorization):
    return await update_doc("authorizations", auth_id, auth.model_dump())


@api_router.delete("/authorizations/{auth_id}")
async def remove_authorization(auth_id: str):
    return await delete_doc("authorizations", auth_id)

@api_router.post("/authorizations/{auth_id}/upload-signed")
async def upload_signed_authorization(auth_id: str, file: UploadFile = File(...)):
    """Recibe un PDF firmado, lo guarda en disco y marca la autorización como firmada."""
    auth = await get_doc("authorizations", auth_id)
    # Eliminar archivo anterior si existe
    if auth.get("archivo_firmado"):
        old_path = UPLOADS_DIR / auth["archivo_firmado"]
        if old_path.exists():
            old_path.unlink()
    # Guardar nuevo archivo
    ext = Path(file.filename).suffix if file.filename else ".pdf"
    filename = f"auth_{auth_id}{ext}"
    file_path = UPLOADS_DIR / filename
    with open(file_path, "wb") as out:
        shutil.copyfileobj(file.file, out)
    # Actualizar documento
    await db["authorizations"].update_one(
        {"id": auth_id},
        {"$set": {"archivo_firmado": filename, "estado": "firmada", "updated_at": now_iso()}}
    )
    return {"ok": True, "archivo_firmado": filename}


@api_router.get("/authorizations/{auth_id}/signed-file")
async def get_signed_authorization(auth_id: str):
    """Devuelve el PDF firmado almacenado para una autorización."""
    auth = await get_doc("authorizations", auth_id)
    if not auth.get("archivo_firmado"):
        raise HTTPException(status_code=404, detail="No hay archivo firmado para esta autorización")
    file_path = UPLOADS_DIR / auth["archivo_firmado"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado en disco")
    return FileResponse(
        path=str(file_path),
        media_type="application/pdf",
        filename=f"autorizacion_firmada_{auth_id}.pdf"
    )


@api_router.delete("/authorizations/{auth_id}/signed-file")
async def delete_signed_authorization(auth_id: str):
    """Elimina el PDF firmado de una autorización y la vuelve a estado pendiente."""
    auth = await get_doc("authorizations", auth_id)
    if auth.get("archivo_firmado"):
        file_path = UPLOADS_DIR / auth["archivo_firmado"]
        if file_path.exists():
            file_path.unlink()
    await db["authorizations"].update_one(
        {"id": auth_id},
        {"$set": {"archivo_firmado": None, "estado": "pendiente", "updated_at": now_iso()}}
    )
    return {"ok": True}


# ================= INSCRIPTIONS =================
class Inscription(BaseModel):
    tipo: str = "alta"  # alta, renovacion
    nombre: str
    apellidos: Optional[str] = ""
    fecha_nacimiento: Optional[str] = None
    email_formulario: Optional[str] = None
    centro_escolar: Optional[str] = None
    progenitor1_nombre: Optional[str] = None
    progenitor1_telefono: Optional[str] = None
    progenitor1_email: Optional[str] = None
    progenitor2_nombre: Optional[str] = None
    progenitor2_telefono: Optional[str] = None
    progenitor2_email: Optional[str] = None
    domicilio: Optional[str] = None
    nueva_incorporacion: bool = True
    estado: str = "recibida"  # recibida, revisada, aceptada, pendiente, rechazada
    categoria: Optional[str] = None
    temporada: Optional[str] = None
    equipo_id: Optional[str] = None
    modalidad: Optional[str] = None
    equipamiento_items: List[str] = Field(default_factory=list)
    player_id: Optional[str] = None  # set when converted to player
    observaciones: Optional[str] = None


def _detect_siblings(insc: dict, players: list) -> list:
    keys = [insc.get("progenitor1_telefono"), insc.get("progenitor2_telefono"),
            insc.get("progenitor1_email"), insc.get("progenitor2_email"), insc.get("domicilio")]
    keys = [k for k in keys if k]
    matches = []
    for p in players:
        pvals = [p.get("progenitor1_telefono"), p.get("progenitor2_telefono"),
                 p.get("progenitor1_email"), p.get("progenitor2_email"), p.get("domicilio")]
        pvals = [v for v in pvals if v]
        if any(k in pvals for k in keys):
            matches.append({"id": p["id"], "nombre": f"{p.get('nombre','')} {p.get('apellidos','')}".strip()})
    return matches


@api_router.post("/inscriptions")
async def create_inscription(insc: Inscription):
    data = insc.model_dump()
    if data.get("fecha_nacimiento"):
        data["categoria"] = compute_category(data["fecha_nacimiento"])
    return await insert_doc("inscriptions", data)


@api_router.get("/inscriptions")
async def get_inscriptions(estado: Optional[str] = None):
    query = {"estado": estado} if estado else {}
    inscs = await list_docs("inscriptions", query)
    players = await list_docs("players")
    for i in inscs:
        i["posibles_hermanos"] = _detect_siblings(i, players)
    return inscs


@api_router.put("/inscriptions/{insc_id}")
async def edit_inscription(insc_id: str, insc: Inscription):
    data = insc.model_dump()
    if data.get("fecha_nacimiento"):
        data["categoria"] = compute_category(data["fecha_nacimiento"])
    return await update_doc("inscriptions", insc_id, data)


@api_router.delete("/inscriptions/{insc_id}")
async def remove_inscription(insc_id: str):
    return await delete_doc("inscriptions", insc_id)


@api_router.post("/inscriptions/{insc_id}/to-player")
async def inscription_to_player(insc_id: str):
    insc = await get_doc("inscriptions", insc_id)
    if insc.get("player_id"):
        raise HTTPException(status_code=400, detail="Ya tiene ficha de jugador")
    pdata = Player(
        nombre=insc.get("nombre"), apellidos=insc.get("apellidos") or "",
        fecha_nacimiento=insc.get("fecha_nacimiento"), email_formulario=insc.get("email_formulario"),
        centro_escolar=insc.get("centro_escolar"), domicilio=insc.get("domicilio"),
        progenitor1_nombre=insc.get("progenitor1_nombre"), progenitor1_telefono=insc.get("progenitor1_telefono"),
        progenitor1_email=insc.get("progenitor1_email"), progenitor2_nombre=insc.get("progenitor2_nombre"),
        progenitor2_telefono=insc.get("progenitor2_telefono"), progenitor2_email=insc.get("progenitor2_email"),
        nueva_incorporacion=insc.get("nueva_incorporacion", True), estado="pendiente_documentacion",
        fecha_inscripcion=insc.get("created_at"),
    ).model_dump()
    if pdata.get("fecha_nacimiento"):
        pdata["categoria"] = compute_category(pdata["fecha_nacimiento"])
    player = await insert_doc("players", pdata)
    await db.inscriptions.update_one({"id": insc_id}, {"$set": {"player_id": player["id"], "estado": "aceptada"}})
    return player


# ================= SAFE INSCRIPTION EXCEL IMPORT =================
class ImportPlanRequest(BaseModel):
    plan_token: str


class ImportConfirmRequest(ImportPlanRequest):
    confirmed: bool = False
    decisions: Dict[str, str] = Field(default_factory=dict)


class StagingRecordUpdate(BaseModel):
    field: str
    value: Any = None
    confirm_suggestion: bool = False


class StagingBulkUpdate(BaseModel):
    record_ids: List[str]
    field: str
    value: str
    confirm_suggestion: bool = False


class StagingDuplicateDecision(BaseModel):
    decision: str


class StagingHistoricalReview(BaseModel):
    decision: str


class StagingOctoberSelection(BaseModel):
    record_ids: List[str]


class StagingIncidentResolution(BaseModel):
    resolution: str


class StagingConfirmRequest(BaseModel):
    confirmed: bool = False


IMPORT_COLLECTIONS = ("teams", "families", "players", "inscriptions", "payments")


async def _import_existing() -> dict:
    return {
        name: await db[name].find({}, {"_id": 0}).to_list(20000)
        for name in IMPORT_COLLECTIONS
    }


def _public_analysis(analysis: dict, token: str) -> dict:
    return {
        "season": analysis["season"], "summary": analysis["summary"],
        "unique_count": analysis["unique_count"], "blocking_errors": analysis["blocking_errors"],
        "unresolved_conflicts": analysis["unresolved_conflicts"], "duplicate_file": analysis["duplicate_file"],
        "rows": analysis["public_rows"], "issues": analysis["issues"], "plan_token": token,
    }


def _clean_doc(document: Optional[dict]) -> Optional[dict]:
    if document is None:
        return None
    return {key: value for key, value in document.items() if key != "_id"}


def _build_import_operations(analysis: dict, existing: dict, job_id: str,
                             decisions: Dict[str, str], *, allow_pending_team: bool = False,
                             keep_family_candidates_separate: bool = False,
                             skip_payments: bool = False) -> list[dict]:
    now = now_iso()
    maps = {name: {str(doc.get("id")): _clean_doc(doc) for doc in existing[name]} for name in IMPORT_COLLECTIONS}
    team_by_name = {normalize_key(doc.get("nombre")): doc for doc in maps["teams"].values()}
    family_by_key = {family_key(doc): doc for doc in maps["families"].values()}
    player_by_key = {str(doc.get("import_identity_key") or identity_key(doc)): doc for doc in maps["players"].values()}
    inscription_by_key = {
        (doc.get("import_identity_key"), doc.get("temporada")): doc
        for doc in maps["inscriptions"].values() if doc.get("import_identity_key")
    }
    payment_by_key = {
        (doc.get("player_id"), doc.get("temporada")): doc
        for doc in maps["payments"].values() if doc.get("temporada")
    }
    operations: Dict[tuple, dict] = {}

    def schedule(collection: str, before: Optional[dict], after: dict) -> None:
        key = (collection, after["id"])
        if key in operations:
            operations[key]["after"] = _clean_doc(after)
        else:
            operations[key] = {"collection": collection, "id": after["id"],
                               "before": _clean_doc(before), "after": _clean_doc(after)}
        maps[collection][after["id"]] = after

    family_fields = {
        "progenitor1_nombre", "progenitor1_telefono", "progenitor1_email",
        "progenitor2_nombre", "progenitor2_telefono", "progenitor2_email",
        "domicilio", "observaciones",
    }
    inscription_fields = set(SAFE_PLAYER_FIELDS) | {"tipo", "equipamiento_items"}
    for result in analysis["rows"]:
        if result.get("status") in {"error", "duplicate"}:
            continue
        if result.get("status") == "conflict":
            if decisions.get(result.get("conflict_id")) != "skip":
                raise HTTPException(status_code=409, detail="Todos los conflictos requieren una decisión explícita")
            continue
        row = result["record"]
        team_key = normalize_key(row.get("equipo"))
        team = maps["teams"].get(str(row.get("equipo_id"))) if row.get("equipo_id") else None
        if not team and team_key and not allow_pending_team:
            team = team_by_name.get(team_key)
        if not team and not allow_pending_team:
            team = {
                "id": new_id(), "nombre": row["equipo"], "categoria": row["categoria"],
                "modalidad": row["modalidad"], "temporada": analysis["season"],
                "limite_jugadores": 18 if row["modalidad"] == "F7" else 25,
                "estado": "activo", "created_at": now, "updated_at": now,
            }
            schedule("teams", None, team); team_by_name[team_key] = team

        fkey = family_key(row)
        if keep_family_candidates_separate and row.get("_family_candidate_pending"):
            # Possible sibling links are a review suggestion, never a reason to
            # silently merge two households during the initial historical load.
            fkey = f"{fkey}:staging:{row.get('_staging_id') or row.get('_row')}"
        family = family_by_key.get(fkey)
        family_before = family
        if family:
            family = merge_nonempty(family, row, family_fields)
            family["updated_at"] = now
        else:
            family = merge_nonempty({"id": new_id(), "created_at": now}, row, family_fields)
            family.update({"contacto_principal": "progenitor1", "preferencia_comunicacion": "email", "updated_at": now})
        if family != family_before:
            schedule("families", family_before, family)
        family_by_key[fkey] = family

        pkey = row["import_identity_key"]
        player = player_by_key.get(pkey)
        player_before = player
        player = merge_nonempty(player or {"id": new_id(), "created_at": now}, row, SAFE_PLAYER_FIELDS)
        current_equipment = list(player.get("equipamiento_items") or [])
        for item in row.get("equipamiento_items") or []:
            if item.casefold() not in {value.casefold() for value in current_equipment}:
                current_equipment.append(item)
        player.update({
            "equipo_id": team["id"] if team else None, "familia_id": family["id"], "equipamiento_items": current_equipment,
            "import_identity_key": pkey, "import_job_id": job_id, "updated_at": now,
        })
        player.setdefault("estado", "pendiente_documentacion")
        player.setdefault("fecha_inscripcion", now)
        if player != player_before:
            schedule("players", player_before, player)
        player_by_key[pkey] = player

        ikey = (pkey, analysis["season"])
        inscription = inscription_by_key.get(ikey)
        inscription_before = inscription
        inscription = merge_nonempty(inscription or {"id": new_id(), "created_at": now}, row, inscription_fields)
        inscription.update({
            "temporada": analysis["season"], "equipo_id": team["id"] if team else None, "familia_id": family["id"],
            "player_id": player["id"], "modalidad": row["modalidad"], "import_identity_key": pkey,
            "import_job_id": job_id, "estado": inscription.get("estado", "recibida"), "updated_at": now,
        })
        if inscription != inscription_before:
            schedule("inscriptions", inscription_before, inscription)
        inscription_by_key[ikey] = inscription

        protected_bank = row.get("_bank") or {}
        if not skip_payments and (row.get("iban") or protected_bank.get("iban_encrypted")):
            pay_key = (player["id"], analysis["season"])
            payment = payment_by_key.get(pay_key)
            payment_before = payment
            payment = dict(payment or {"id": new_id(), "created_at": now, "estado": "pendiente",
                                       "concepto": f"Cuota temporada {analysis['season']}"})
            payment.update({
                "player_id": player["id"], "temporada": analysis["season"], "forma_pago": "domiciliacion",
                **(protected_bank if protected_bank.get("iban_encrypted") else encrypt_iban(row["iban"], JWT_SECRET)),
                "iban": None, "import_job_id": job_id, "updated_at": now,
            })
            schedule("payments", payment_before, payment); payment_by_key[pay_key] = payment
    return list(operations.values())


async def _apply_operations(operations: list[dict], reverse: bool = False) -> None:
    sequence = list(reversed(operations)) if reverse else operations
    applied = []
    try:
        for operation in sequence:
            target = operation["before"] if reverse else operation["after"]
            if target is None:
                await db[operation["collection"]].delete_one({"id": operation["id"]})
            else:
                await db[operation["collection"]].replace_one({"id": operation["id"]}, target, upsert=True)
            applied.append(operation)
    except Exception:
        for operation in reversed(applied):
            target = operation["after"] if reverse else operation["before"]
            if target is None:
                await db[operation["collection"]].delete_one({"id": operation["id"]})
            else:
                await db[operation["collection"]].replace_one({"id": operation["id"]}, target, upsert=True)
        raise


def _operations_snapshot(existing: dict, operations: list[dict]) -> dict:
    """Devuelve una vista en memoria de ``existing`` después de las operaciones.

    La importación histórica necesita crear primero la ficha base y, sobre esa
    misma ficha, guardar únicamente sus referencias históricas protegidas. La
    vista evita una escritura intermedia y mantiene toda la importación en la
    misma transacción compensable.
    """
    snapshot = {collection: [copy.deepcopy(item) for item in rows]
                for collection, rows in existing.items()}
    indexes = {
        collection: {item.get("id"): item for item in rows}
        for collection, rows in snapshot.items()
    }
    for operation in operations:
        collection, identifier = operation["collection"], operation["id"]
        target = copy.deepcopy(operation["after"])
        bucket = indexes.setdefault(collection, {})
        if target is None:
            bucket.pop(identifier, None)
        else:
            bucket[identifier] = target
    for collection, bucket in indexes.items():
        snapshot[collection] = list(bucket.values())
    return snapshot


def _historical_enrichment_operations(
    draft: dict, existing: dict, job_id: str, *, resolve_team_suggestions: bool = True,
) -> tuple[list[dict], dict]:
    """Añade referencias históricas a coincidencias únicas ya resueltas.

    En una publicación inicial de un Excel histórico, los nombres de equipo son
    referencias y no una orden para crear o asignar equipos. ``False`` conserva
    esa protección: las fichas se publican, pero su equipo actual queda pendiente
    de una confirmación administrativa explícita.
    """
    now = now_iso()
    players_by_name: Dict[tuple[str, str], list[dict]] = {}
    for player in existing.get("players", []):
        key = (normalize_key(player.get("nombre")), normalize_key(player.get("apellidos")))
        players_by_name.setdefault(key, []).append(player)
    teams_by_name: Dict[str, list[dict]] = {}
    for team in existing.get("teams", []):
        teams_by_name.setdefault(normalize_key(team.get("nombre")), []).append(team)
    payments = {
        (item.get("player_id"), item.get("temporada")): item
        for item in existing.get("payments", []) if item.get("historical_bank_reference")
    }
    operations, matched, unmatched, ambiguous, bank_refs = [], 0, 0, 0, 0
    team_assignments = team_pending = teams_created = 0
    for record in effective_records(draft):
        key = (normalize_key(record.get("nombre")), normalize_key(record.get("apellidos")))
        matches = players_by_name.get(key, [])
        if not matches:
            unmatched += 1
            continue
        if len(matches) != 1:
            ambiguous += 1
            continue
        before = _clean_doc(matches[0])
        player = dict(before)
        historical = record.get("historical") or {}
        player.update({
            "historial_equipacion": historical.get("equipment_history") or {},
            "segunda_equipacion": historical.get("equipment_current") or {},
            "historial_equipos": historical.get("team_history") or {},
            "historial_federacion": historical.get("federation_history") or {},
            "historial_deportivo": historical.get("sport") or {},
            "historial_entrenamientos": historical.get("schedule_history") or [],
            "referencias_cuotas": historical.get("fees") or {},
            "referencias_permisos": historical.get("consents") or {},
            "historical_import_job_id": job_id, "updated_at": now,
        })
        if not player.get("posicion") and (historical.get("sport") or {}).get("position"):
            player["posicion"] = historical["sport"]["position"]
        team_name = record.get("equipo") or record.get("equipo_anterior")
        team_key = normalize_key(team_name)
        if resolve_team_suggestions:
            # "NO APLICA" means that the player has no team. Never resolve it
            # against a legacy catalogue entry with that display name.
            if team_key == "no_aplica":
                team_matches = []
                player["equipo_id"] = None
            else:
                team_matches = teams_by_name.get(team_key, []) if team_name else []
            if not team_matches and team_key and team_key != "no_aplica":
                source_season = (historical.get("sport") or {}).get("team_assignment_source") or "2025-2026"
                team = {
                    "id": new_id(), "nombre": team_name, "categoria": record.get("categoria") or None,
                    "modalidad": record.get("modalidad") or None, "temporada": source_season,
                    "limite_jugadores": 18 if record.get("modalidad") == "F7" else 25,
                    "estado": "activo", "created_at": now, "updated_at": now,
                    "historical_import_job_id": job_id,
                }
                operations.append({"collection": "teams", "id": team["id"], "before": None, "after": team})
                teams_by_name[team_key] = [team]
                team_matches = [team]
                teams_created += 1
            if len(team_matches) == 1:
                player["equipo_id"] = team_matches[0]["id"]
                player["equipo_historico_referencia"] = team_name
                team_assignments += 1
            elif team_name and team_key != "no_aplica":
                player["equipo_historico_referencia"] = team_name
                team_pending += 1
        operations.append({"collection": "players", "id": player["id"], "before": before, "after": player})
        matched += 1

        bank = record.get("bank") or {}
        if bank.get("status") == "valid" and bank.get("iban_encrypted"):
            pay_key = (player["id"], draft["season"])
            payment_before = _clean_doc(payments.get(pay_key))
            payment = dict(payment_before or {"id": new_id(), "created_at": now})
            payment.update({
                "player_id": player["id"], "temporada": draft["season"],
                "concepto": "Cuenta bancaria histórica (sin deuda)",
                "importe_base": 0, "importe_final": 0, "descuento_hermano": 0,
                "forma_pago": "domiciliacion", "estado": "referencia",
                "iban": None, "iban_encrypted": bank["iban_encrypted"],
                "iban_last4": bank.get("iban_last4"), "iban_validado": True,
                "titular_cuenta": (historical.get("bank_reference") or {}).get("holder") or "",
                "historical_bank_reference": True, "confirmed_debt": False,
                "import_job_id": job_id, "updated_at": now,
            })
            operations.append({"collection": "payments", "id": payment["id"], "before": payment_before, "after": payment})
            payments[pay_key] = payment
            bank_refs += 1
    return operations, {
        "matched_players": matched, "unmatched_rows": unmatched,
        "ambiguous_rows": ambiguous, "bank_references": bank_refs,
        "team_assignments": team_assignments, "teams_created": teams_created,
        "team_names_pending": team_pending,
        "created_players": 0, "official_debts": 0,
    }


@api_router.get("/inscription-imports/template")
async def download_inscription_template():
    path = ROOT_DIR / "templates" / "plantilla_inscripciones_2026-2027.xlsx"
    if not path.exists():
        raise HTTPException(status_code=503, detail="Plantilla no disponible")
    return FileResponse(path, filename=path.name,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@api_router.post("/inscription-imports/analyze")
async def analyze_inscription_import(file: UploadFile = File(...), season: str = Form(...)):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se admiten archivos .xlsx")
    content = await file.read()
    digest = file_sha256(content)
    try:
        rows = parse_excel(content)
        duplicate = bool(await db.import_locks.find_one({"_id": f"{season}:{digest}"}))
        analysis = analyze_rows(rows, season, await _import_existing(), digest, duplicate)
        token = encode_plan({"season": season, "file_sha256": digest, "rows": rows,
                             "generated_at": now_iso()}, JWT_SECRET)
        return _public_analysis(analysis, token)
    except ImportValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


def _decode_and_check_plan(plan_token: str) -> dict:
    try:
        return decode_plan(plan_token, JWT_SECRET)
    except ImportValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


@api_router.post("/inscription-imports/error-report")
async def inscription_import_error_report(request: ImportPlanRequest):
    plan = _decode_and_check_plan(request.plan_token)
    duplicate = bool(await db.import_locks.find_one({"_id": f"{plan['season']}:{plan['file_sha256']}"}))
    analysis = analyze_rows(plan["rows"], plan["season"], await _import_existing(), plan["file_sha256"], duplicate)
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Errores"
    sheet.append(["Fila", "Estado", "Gravedad", "Código", "Mensaje"])
    for issue in analysis["issues"]:
        sheet.append([issue.get("row"), issue.get("status"), issue.get("severity"), issue.get("code"), issue.get("message")])
    buffer = io.BytesIO(); workbook.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=informe_importacion.xlsx"})


@api_router.post("/inscription-imports/confirm")
async def confirm_inscription_import(request: ImportConfirmRequest):
    if request.confirmed is not True:
        raise HTTPException(status_code=422, detail="La importación requiere confirmación expresa")
    plan = _decode_and_check_plan(request.plan_token)
    lock_id = f"{plan['season']}:{plan['file_sha256']}"
    existing = await _import_existing()
    duplicate = bool(await db.import_locks.find_one({"_id": lock_id}))
    analysis = analyze_rows(plan["rows"], plan["season"], existing, plan["file_sha256"], duplicate)
    if analysis["blocking_errors"]:
        raise HTTPException(status_code=409, detail="La importación contiene errores graves")
    if any(request.decisions.get(item.get("conflict_id")) != "skip" for item in analysis["rows"] if item.get("status") == "conflict"):
        raise HTTPException(status_code=409, detail="Todos los conflictos requieren una decisión explícita")
    actor = current_user_context.get() or {}
    job_id = new_id()
    operations = _build_import_operations(analysis, existing, job_id, request.decisions)
    try:
        await db.import_locks.insert_one({"_id": lock_id, "job_id": job_id, "created_at": now_iso()})
    except DuplicateKeyError as exc:
        raise HTTPException(status_code=409, detail="Este archivo ya fue importado") from exc
    job = {"id": job_id, "season": plan["season"], "file_sha256": plan["file_sha256"],
           "status": "applying", "summary": analysis["summary"], "operations": operations,
           "created_by_user_id": actor.get("id"), "created_at": now_iso(), "updated_at": now_iso()}
    try:
        await db.inscription_import_jobs.insert_one(job)
        await _apply_operations(operations)
        await db.inscription_import_jobs.update_one({"id": job_id}, {"$set": {"status": "applied", "updated_at": now_iso()}})
    except Exception as exc:
        await db.import_locks.delete_one({"_id": lock_id})
        await db.inscription_import_jobs.update_one({"id": job_id}, {"$set": {"status": "failed", "updated_at": now_iso()}}, upsert=False)
        raise HTTPException(status_code=500, detail="La importación fue revertida sin cambios parciales") from exc
    return {"ok": True, "job_id": job_id, "status": "applied", "operations": len(operations), "summary": analysis["summary"]}


@api_router.get("/inscription-imports/history")
async def inscription_import_history():
    jobs = await db.inscription_import_jobs.find({}, {"_id": 0, "operations": 0, "created_by_user_id": 0}).sort("created_at", -1).to_list(100)
    for job in jobs:
        job["file_sha256"] = f"{job.get('file_sha256', '')[:12]}…"
    return jobs


@api_router.post("/inscription-imports/{job_id}/undo")
async def undo_inscription_import(job_id: str):
    job = await db.inscription_import_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job or job.get("status") != "applied":
        raise HTTPException(status_code=409, detail="La importación no existe o ya fue deshecha")
    for operation in job.get("operations") or []:
        current = _clean_doc(await db[operation["collection"]].find_one({"id": operation["id"]}))
        if current != operation.get("after"):
            raise HTTPException(status_code=409, detail="No se puede deshacer: existen cambios posteriores")
    await _apply_operations(job.get("operations") or [], reverse=True)
    actor = current_user_context.get() or {}
    await db.inscription_import_jobs.update_one({"id": job_id}, {"$set": {
        "status": "undone", "undone_at": now_iso(), "undone_by_user_id": actor.get("id"), "updated_at": now_iso(),
    }})
    await db.import_locks.delete_one({"_id": f"{job['season']}:{job['file_sha256']}"})
    return {"ok": True, "job_id": job_id, "status": "undone"}


# ================= IMPORT PREPARATION / STAGING =================
def _staging_ttl_hours() -> int:
    try:
        return max(1, min(int(os.environ.get("IMPORT_STAGING_TTL_HOURS", "168")), 24 * 90))
    except ValueError:
        return 168


async def _ensure_staging_indexes() -> None:
    await db.import_staging.create_index("expires_at", expireAfterSeconds=0, name="expires_at_ttl")
    await db.import_staging.create_index([("status", 1), ("updated_at", -1)], name="status_updated")
    await db.import_staging.create_index("source_sha256", name="source_sha256")


async def _staging_doc(draft_id: str) -> dict:
    draft = await db.import_staging.find_one({"id": draft_id})
    if not draft:
        raise HTTPException(status_code=404, detail="El borrador no existe o ha caducado")
    return draft


def _staging_actor() -> Optional[str]:
    return (current_user_context.get() or {}).get("id")


@api_router.post("/inscription-imports/staging")
async def create_import_staging(file: UploadFile = File(...), season: str = Form(...)):
    if season != IMPORT_SEASON:
        raise HTTPException(status_code=422, detail=f"La temporada permitida es {IMPORT_SEASON}")
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se admiten archivos .xlsx")
    content = await file.read()
    digest = file_sha256(content)
    await _ensure_staging_indexes()
    previous = await db.import_staging.find_one({"source_sha256": digest, "season": season, "status": "draft"})
    if previous:
        return public_draft(previous)
    parsed_historical = None
    try:
        try:
            parsed_historical = parse_historical_excel(content, new_id)
        except ImportValidationError as historical_error:
            if "hoja histórica" not in str(historical_error):
                raise
            rows = parse_excel(content)
    except ImportValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if parsed_historical:
        records, duplicates, incidents = prepare_historical_staging(parsed_historical, JWT_SECRET)
    else:
        records, duplicates, incidents = prepare_records(rows, JWT_SECRET)
    now = datetime.now(timezone.utc)
    draft = {
        "id": new_id(), "season": season, "status": "draft", "source_sha256": digest,
        "records": records, "duplicates": duplicates, "incidents": incidents,
        "audit": [audit_event(_staging_actor(), "draft_created", {"rows": len(records)})],
        "created_by_user_id": _staging_actor(), "created_at": now, "updated_at": now,
        "expires_at": expiry(_staging_ttl_hours()),
    }
    if parsed_historical:
        existing_players = await db.players.find({}, {"_id": 0, "nombre": 1, "apellidos": 1, "fecha_nacimiento": 1}).to_list(10000)
        draft.update({
            "source_format": HISTORICAL_FORMAT, "auxiliary_rows": parsed_historical["auxiliary_rows"],
            "fuzzy_matches": parsed_historical["fuzzy_matches"], "family_candidates": parsed_historical["family_candidates"],
            "quality": historical_quality_summary(parsed_historical),
            "simulation": historical_simulation(parsed_historical, existing_players),
        })
    await db.import_staging.insert_one(draft)
    return public_draft(draft)


@api_router.get("/inscription-imports/staging")
async def list_import_staging():
    await _ensure_staging_indexes()
    drafts = await db.import_staging.find({"status": "draft"}).sort("updated_at", -1).to_list(100)
    return [public_draft(item, include_records=False) for item in drafts]


@api_router.get("/inscription-imports/staging/{draft_id}")
async def get_import_staging(draft_id: str):
    return public_draft(await _staging_doc(draft_id))


@api_router.get("/inscription-imports/staging/{draft_id}/simulation")
async def get_import_staging_simulation(draft_id: str):
    draft = await _staging_doc(draft_id)
    if draft.get("source_format") != HISTORICAL_FORMAT:
        raise HTTPException(status_code=409, detail="El borrador no utiliza el adaptador histórico")
    return {"quality": draft.get("quality", {}), "simulation": draft.get("simulation", {}),
            "summary": draft_summary(draft)}


@api_router.post("/inscription-imports/staging/{draft_id}/historical-readiness")
async def apply_historical_readiness(draft_id: str):
    """Persist deterministic readiness rules for an authenticated admin draft."""
    draft = await _staging_doc(draft_id)
    if draft.get("source_format") != HISTORICAL_FORMAT:
        raise HTTPException(status_code=409, detail="El borrador no utiliza el adaptador histórico")
    if draft.get("status") != "draft":
        raise HTTPException(status_code=409, detail="Solo se puede preparar un borrador activo")
    try:
        prepared, result = historical_readiness(draft)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    now = datetime.now(timezone.utc)
    await db.import_staging.update_one({"id": draft_id}, {"$set": {
        "records": prepared["records"], "incidents": prepared["incidents"],
        "historical_readiness": prepared["historical_readiness"], "updated_at": now,
        "expires_at": expiry(_staging_ttl_hours()),
    }, "$push": {"audit": audit_event(_staging_actor(), "historical_readiness_applied", result)}})
    return public_draft(await _staging_doc(draft_id))


def _existing_category(value: Any) -> Optional[str]:
    requested = str(value or "").strip()
    return next((item["name"] for item in CATEGORIES if item["name"] == requested), None)


def _category_key(value: Any) -> str:
    return normalize_key(value)


def _team_category_matches(team_category: Any, official_category: Any) -> bool:
    """Match legacy team labels to an official category without rewriting either."""
    requested = _existing_category(official_category)
    if not requested:
        return False
    team_key = _category_key(team_category)
    official_key = _category_key(requested)
    return team_key == official_key or team_key.startswith(official_key)


def _team_modality_matches(team_modality: Any, record_modality: Any) -> bool:
    """Accept legacy teams without modality; otherwise require the same official code."""
    team_value = normalize_key(team_modality)
    if not team_value:
        return True
    return bool(normalize_key(record_modality)) and team_value == normalize_key(record_modality)


def _team_is_usable(team: Mapping[str, Any]) -> bool:
    """Legacy teams without a status are usable; explicit non-active states are not."""
    return normalize_key(team.get("estado")) in {"", "activo", "active"}


async def _active_staging_team(team_id: Any) -> Optional[dict]:
    requested = str(team_id or "").strip()
    if not requested:
        return None
    team = await db.teams.find_one(
        {"id": requested},
        {"_id": 0, "id": 1, "nombre": 1, "categoria": 1, "estado": 1},
    )
    if not team or team.get("id") != requested or not _team_is_usable(team):
        return None
    return team


async def _active_staging_modality(value: Any) -> Optional[str]:
    """Resolve a browser value to the active official catalog code."""
    catalog = await _load_modality_catalog()
    normalized = normalize_modality(value, catalog)
    if normalized.status != "recognized" or not normalized.active:
        return None
    return normalized.code


@api_router.patch("/inscription-imports/staging/{draft_id}/records/{record_id}")
async def update_staging_record(draft_id: str, record_id: str, request: StagingRecordUpdate):
    draft = await _staging_doc(draft_id)
    current_record = next((row for row in draft.get("records", []) if row.get("id") == record_id), {})
    previous_value = current_record.get(request.field)
    additional_updates = {}
    historical = draft.get("source_format") == HISTORICAL_FORMAT
    if request.field == "iban":
        normalized = normalize_key(request.value)
        raw = str(request.value or "").replace(" ", "").upper()
        if normalized and not staging_valid_iban(raw):
            raise HTTPException(status_code=422, detail="El IBAN no es válido; no se guardó")
        bank = encrypt_iban(raw, JWT_SECRET) if raw else {}
        update_path = "records.$.bank"
        value = {"status": "valid", **bank} if bank else {"status": "pending", "iban_encrypted": None, "iban_last4": None}
        issue_field = "bank"
    elif request.field in ALLOWED_RECORD_FIELDS - {"equipamiento_items"}:
        update_path, value, issue_field = f"records.$.{request.field}", str(request.value or "").strip(), request.field
        if historical and request.field in {"categoria", "equipo", "modalidad"} and not request.confirm_suggestion:
            raise HTTPException(status_code=422, detail="La asignación requiere confirmación administrativa expresa")
        if historical and request.field in {"categoria", "equipo"}:
            if draft.get("status") != "draft":
                raise HTTPException(status_code=409, detail="Solo se puede modificar un borrador activo")
            if request.field == "categoria":
                value = _existing_category(value)
                if not value:
                    raise HTTPException(status_code=422, detail="La categoría no existe o no está activa")
                if current_record.get("equipo_id"):
                    current_team = await _active_staging_team(current_record.get("equipo_id"))
                    if not current_team or not _team_category_matches(current_team.get("categoria"), value):
                        raise HTTPException(status_code=422, detail="La categoría no es compatible con el equipo asignado")
            else:
                selected_team = await _active_staging_team(value)
                if not selected_team:
                    raise HTTPException(status_code=422, detail="El equipo no existe o está inactivo")
                if not _team_category_matches(selected_team.get("categoria"), current_record.get("categoria")):
                    raise HTTPException(status_code=422, detail="El equipo no pertenece a la categoría del registro")
                value = selected_team["nombre"]
                additional_updates["records.$.equipo_id"] = selected_team["id"]
        if request.field == "modalidad":
            if draft.get("status") != "draft":
                raise HTTPException(status_code=409, detail="Solo se puede modificar un borrador activo")
            value = await _active_staging_modality(value)
            if not value:
                raise HTTPException(status_code=422, detail="La modalidad no existe o está inactiva")
        candidate = dict(current_record)
        candidate[request.field] = value
        if not field_is_valid(candidate, request.field):
            raise HTTPException(status_code=422, detail="El valor no tiene un formato válido")
    else:
        raise HTTPException(status_code=422, detail="Campo no editable")
    now = datetime.now(timezone.utc)
    event_detail = {"record_id": record_id, "field": request.field}
    if request.field in {"modalidad", "categoria", "equipo"}:
        event_detail.update({"previous_value": previous_value or None, "new_value": value or None})
        if request.field == "equipo":
            event_detail.update({
                "previous_id": current_record.get("equipo_id"),
                "new_id": additional_updates.get("records.$.equipo_id"),
            })
    event = audit_event(_staging_actor(), "record_updated", event_detail)
    result = await db.import_staging.update_one(
        {"id": draft_id, "records.id": record_id},
        {"$set": {update_path: value, **additional_updates, "updated_at": now, "expires_at": expiry(_staging_ttl_hours())}, "$push": {"audit": event}},
    )
    if not result.modified_count:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    if value not in (None, ""):
        await db.import_staging.update_many(
            {"id": draft_id}, {"$set": {"incidents.$[issue].resolution": "corrected"}},
            array_filters=[{"issue.record_id": record_id, "issue.field": issue_field}],
        )
    return public_draft(await _staging_doc(draft_id))


@api_router.post("/inscription-imports/staging/{draft_id}/bulk")
async def bulk_update_staging(draft_id: str, request: StagingBulkUpdate):
    if request.field not in {"equipo", "categoria", "modalidad"}:
        raise HTTPException(status_code=422, detail="Asignación masiva no permitida para este campo")
    value = request.value.strip()
    if not value:
        raise HTTPException(status_code=422, detail="El valor de asignación es obligatorio")
    draft = await _staging_doc(draft_id)
    historical = draft.get("source_format") == HISTORICAL_FORMAT
    selected_team = None
    if historical and request.field in {"categoria", "equipo"}:
        if draft.get("status") != "draft":
            raise HTTPException(status_code=409, detail="Solo se puede modificar un borrador activo")
        if not request.confirm_suggestion:
            raise HTTPException(status_code=422, detail="La asignación requiere confirmación administrativa expresa")
        if request.field == "categoria":
            value = _existing_category(value)
            if not value:
                raise HTTPException(status_code=422, detail="La categoría no existe o no está activa")
        else:
            selected_team = await _active_staging_team(value)
            if not selected_team:
                raise HTTPException(status_code=422, detail="El equipo no existe o está inactivo")
            value = selected_team["nombre"]
    if request.field == "modalidad":
        if draft.get("status") != "draft":
            raise HTTPException(status_code=409, detail="Solo se puede modificar un borrador activo")
        value = await _active_staging_modality(value)
        if not value:
            raise HTTPException(status_code=422, detail="La modalidad no existe o está inactiva")
        if not request.confirm_suggestion:
            raise HTTPException(status_code=422, detail="La asignación de modalidad requiere confirmación expresa")
    selected = set(request.record_ids)
    active_record_ids = {row["id"] for row in effective_records(draft)}
    if not selected or not selected.issubset(active_record_ids):
        raise HTTPException(status_code=422, detail="Selección de registros no válida")
    selected_rows = [row for row in effective_records(draft) if row.get("id") in selected]
    if historical and request.field == "categoria":
        for row in selected_rows:
            if row.get("equipo_id"):
                current_team = await _active_staging_team(row.get("equipo_id"))
                if not current_team or not _team_category_matches(current_team.get("categoria"), value):
                    raise HTTPException(status_code=422, detail="La categoría no es compatible con un equipo asignado")
    if historical and request.field == "equipo":
        if any(not _team_category_matches(selected_team.get("categoria"), row.get("categoria")) for row in selected_rows):
            raise HTTPException(status_code=422, detail="El equipo no pertenece a la categoría de todos los registros seleccionados")
        if any(not _team_modality_matches(selected_team.get("modalidad"), row.get("modalidad")) for row in selected_rows):
            raise HTTPException(status_code=422, detail="El equipo no pertenece a la modalidad de todos los registros seleccionados")
    now = datetime.now(timezone.utc)
    assignment_changes = []
    if request.field in {"modalidad", "categoria", "equipo"}:
        assignment_changes = [
            {
                "record_id": row["id"], "previous_value": row.get(request.field) or None,
                "new_value": value,
                **({"previous_id": row.get("equipo_id"), "new_id": selected_team["id"]}
                   if request.field == "equipo" and selected_team else {}),
            }
            for row in selected_rows
        ]
    record_updates = {"records.$[row]." + request.field: value}
    if request.field == "equipo" and selected_team:
        record_updates["records.$[row].equipo_id"] = selected_team["id"]
    await db.import_staging.update_one(
        {"id": draft_id},
        {"$set": {
            **record_updates,
            **({"records.$[row].suggestion_confirmed": True} if request.field == "modalidad" else {}),
            "incidents.$[issue].resolution": "corrected", "updated_at": now,
            "expires_at": expiry(_staging_ttl_hours()),
        }, "$push": {"audit": audit_event(_staging_actor(), "bulk_updated", {
            "field": request.field, "records": len(selected),
            **({"changes": assignment_changes} if assignment_changes else {}),
        })}},
        array_filters=[{"row.id": {"$in": list(selected)}}, {"issue.record_id": {"$in": list(selected)}, "issue.field": request.field}],
    )
    return public_draft(await _staging_doc(draft_id))


@api_router.post("/inscription-imports/staging/{draft_id}/duplicates/{group_id}")
async def resolve_staging_duplicate(draft_id: str, group_id: str, request: StagingDuplicateDecision):
    allowed = {"keep_first", "keep_second", "merge", "different_people"}
    if request.decision not in allowed:
        raise HTTPException(status_code=422, detail="Decisión de duplicado no válida")
    result = await db.import_staging.update_one(
        {"id": draft_id, "duplicates.id": group_id},
        {"$set": {"duplicates.$.decision": request.decision, "updated_at": datetime.now(timezone.utc),
                  "expires_at": expiry(_staging_ttl_hours())},
         "$push": {"audit": audit_event(_staging_actor(), "duplicate_resolved", {"group_id": group_id, "decision": request.decision})}},
    )
    if not result.modified_count:
        raise HTTPException(status_code=404, detail="Duplicado no encontrado")
    return public_draft(await _staging_doc(draft_id))


@api_router.post("/inscription-imports/staging/{draft_id}/reviews/{kind}/{group_id}")
async def resolve_historical_review(draft_id: str, kind: str, group_id: str, request: StagingHistoricalReview):
    config = {
        "fuzzy": ("fuzzy_matches", {"same_person", "different_people", "leave_pending"}),
        "family": ("family_candidates", {"confirm_shared", "keep_separate"}),
    }
    if kind not in config or request.decision not in config.get(kind, (None, set()))[1]:
        raise HTTPException(status_code=422, detail="Decisión de revisión histórica no válida")
    field, _ = config[kind]
    draft = await _staging_doc(draft_id)
    if draft.get("source_format") != HISTORICAL_FORMAT:
        raise HTTPException(status_code=409, detail="El borrador no utiliza el adaptador histórico")
    if not any(item.get("id") == group_id for item in draft.get(field, [])):
        raise HTTPException(status_code=404, detail="Grupo de revisión no encontrado")
    result = await db.import_staging.update_one(
        {"id": draft_id, f"{field}.id": group_id},
        {"$set": {f"{field}.$.decision": request.decision, "updated_at": datetime.now(timezone.utc),
                  "expires_at": expiry(_staging_ttl_hours())},
         "$push": {"audit": audit_event(_staging_actor(), "historical_review", {"kind": kind, "group_id": group_id, "decision": request.decision})}},
    )
    if not result.modified_count:
        raise HTTPException(status_code=409, detail="La revisión ya estaba registrada")
    return public_draft(await _staging_doc(draft_id))


@api_router.post("/inscription-imports/staging/{draft_id}/october")
async def select_staging_october(draft_id: str, request: StagingOctoberSelection):
    draft = await _staging_doc(draft_id)
    requested = set(request.record_ids)
    valid_ids = {row["id"] for row in effective_records(draft) if row.get("tipo") == "renovacion"}
    if len(requested) > 54 or not requested.issubset(valid_ids):
        raise HTTPException(status_code=422, detail="La selección de octubre admite exactamente hasta 54 registros válidos")
    records = draft.get("records", [])
    for row in records:
        row["selected_october"] = row["id"] in requested
    await db.import_staging.update_one({"id": draft_id}, {"$set": {
        "records": records, "updated_at": datetime.now(timezone.utc), "expires_at": expiry(_staging_ttl_hours()),
    }, "$push": {"audit": audit_event(_staging_actor(), "october_selection", {"count": len(requested)})}})
    return public_draft(await _staging_doc(draft_id))


@api_router.patch("/inscription-imports/staging/{draft_id}/incidents/{incident_id}")
async def resolve_staging_incident(draft_id: str, incident_id: str, request: StagingIncidentResolution):
    if request.resolution not in {"corrected", "not_applicable"}:
        raise HTTPException(status_code=422, detail="Resolución no válida")
    draft = await _staging_doc(draft_id)
    incident = next((item for item in draft.get("incidents", []) if item.get("id") == incident_id), None)
    if not incident:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    if incident.get("blocking") and request.resolution == "not_applicable":
        raise HTTPException(status_code=422, detail="Una incidencia obligatoria debe corregirse")
    record = next((row for row in draft.get("records", []) if row.get("id") == incident.get("record_id")), {})
    if request.resolution == "corrected" and not field_is_valid(record, incident.get("field")):
        raise HTTPException(status_code=422, detail="Corrige el valor antes de cerrar la incidencia")
    await db.import_staging.update_one(
        {"id": draft_id, "incidents.id": incident_id},
        {"$set": {"incidents.$.resolution": request.resolution, "updated_at": datetime.now(timezone.utc),
                  "expires_at": expiry(_staging_ttl_hours())},
         "$push": {"audit": audit_event(_staging_actor(), "incident_resolved", {"incident_id": incident_id, "resolution": request.resolution})}},
    )
    return public_draft(await _staging_doc(draft_id))


@api_router.delete("/inscription-imports/staging/{draft_id}")
async def delete_import_staging(draft_id: str):
    result = await db.import_staging.delete_one({"id": draft_id, "status": "draft"})
    if not result.deleted_count:
        raise HTTPException(status_code=404, detail="El borrador no existe")
    return {"ok": True}


@api_router.post("/inscription-imports/staging/{draft_id}/confirm")
async def confirm_import_staging(draft_id: str, request: StagingConfirmRequest):
    if request.confirmed is not True:
        raise HTTPException(status_code=422, detail="La importación requiere confirmación expresa")
    draft = await _staging_doc(draft_id)
    summary = draft_summary(draft)
    if not summary["can_import"]:
        raise HTTPException(status_code=409, detail="El borrador mantiene bloqueos pendientes")
    historical = draft.get("source_format") == HISTORICAL_FORMAT
    if historical:
        existing = await _import_existing()
        family_candidate_ids = {
            record_id for group in draft.get("family_candidates", []) if not group.get("decision")
            for record_id in (group.get("record_ids") or [])
        }
        rows = []
        for record in effective_records(draft):
            row = {key: value for key, value in record.items() if key in ALLOWED_RECORD_FIELDS}
            row.update({
                "_row": record.get("source_row"), "_staging_id": record.get("id"),
                "equipo_id": record.get("equipo_id"),
                "_family_candidate_pending": record.get("id") in family_candidate_ids,
            })
            if record.get("identity_override"):
                row["external_id"] = record["identity_override"]
            bank = record.get("bank") or {}
            if bank.get("status") == "valid":
                row["_bank"] = {
                    "iban_encrypted": bank.get("iban_encrypted"),
                    "iban_last4": bank.get("iban_last4"),
                }
            rows.append(row)
        analysis = analyze_rows(
            rows, draft["season"], existing, draft["source_sha256"], False,
            allow_pending_team=True, allow_pending_contact=True,
            ignore_team_name_suggestions=True,
        )
        if analysis["blocking_errors"] or analysis["unresolved_conflicts"]:
            raise HTTPException(status_code=409, detail="La validación final ha detectado bloqueos")
        job_id = new_id()
        base_operations = _build_import_operations(
            analysis, existing, job_id, {}, allow_pending_team=True,
            keep_family_candidates_separate=True, skip_payments=True,
        )
        enriched_existing = _operations_snapshot(existing, base_operations)
        enrichment_operations, enrichment = _historical_enrichment_operations(
            draft, enriched_existing, job_id, resolve_team_suggestions=False,
        )
        operations = [*base_operations, *enrichment_operations]
        if not operations:
            raise HTTPException(status_code=409, detail="No hay altas ni actualizaciones aplicables")
        enrichment.update({
            "created_players": sum(
                item["collection"] == "players" and item.get("before") is None
                for item in base_operations
            ),
            "created_families": sum(
                item["collection"] == "families" and item.get("before") is None
                for item in base_operations
            ),
            "created_inscriptions": sum(
                item["collection"] == "inscriptions" and item.get("before") is None
                for item in base_operations
            ),
            "base_operations": len(base_operations),
            "historical_operations": len(enrichment_operations),
            "team_assignment_policy": "pending_administrative_confirmation",
        })
        # Esta versión publica las fichas base y conserva los equipos históricos
        # como referencia; no crea ni asigna equipos por un texto del Excel.
        lock_id = f"historical-base-v4:{draft['season']}:{draft['source_sha256']}"
        try:
            await db.import_locks.insert_one({"_id": lock_id, "job_id": job_id, "created_at": now_iso()})
        except DuplicateKeyError as exc:
            raise HTTPException(status_code=409, detail="Este archivo ya fue importado") from exc
        job = {"id": job_id, "season": draft["season"], "file_sha256": draft["source_sha256"],
               "status": "applying", "summary": enrichment, "operations": operations,
               "created_by_user_id": _staging_actor(), "staging_id": draft_id,
               "created_at": now_iso(), "updated_at": now_iso()}
        try:
            await db.inscription_import_jobs.insert_one(job)
            await _apply_operations(operations)
            await db.inscription_import_jobs.update_one({"id": job_id}, {"$set": {"status": "applied", "updated_at": now_iso()}})
            await db.import_staging.update_one({"id": draft_id}, {"$set": {"status": "imported", "updated_at": datetime.now(timezone.utc)},
                                                                      "$unset": {"records": "", "incidents": "", "duplicates": ""}})
        except Exception as exc:
            await db.import_locks.delete_one({"_id": lock_id})
            await db.inscription_import_jobs.update_one({"id": job_id}, {"$set": {"status": "failed", "updated_at": now_iso()}})
            raise HTTPException(status_code=500, detail="La importación histórica fue revertida sin cambios parciales") from exc
        return {"ok": True, "job_id": job_id, "status": "applied", "operations": len(operations), "summary": enrichment}
    family_candidate_ids = {
        record_id for group in draft.get("family_candidates", []) if not group.get("decision")
        for record_id in (group.get("record_ids") or [])
    }
    rows = []
    for record in effective_records(draft):
        row = {key: value for key, value in record.items() if key in ALLOWED_RECORD_FIELDS}
        row["_row"] = record.get("source_row")
        if historical:
            row["_staging_id"] = record.get("id")
            row["equipo_id"] = record.get("equipo_id")
            row["_family_candidate_pending"] = record.get("id") in family_candidate_ids
        if record.get("identity_override"):
            row["external_id"] = record["identity_override"]
        bank = record.get("bank") or {}
        if bank.get("status") == "valid":
            row["_bank"] = {"iban_encrypted": bank.get("iban_encrypted"), "iban_last4": bank.get("iban_last4")}
        rows.append(row)
    existing = await _import_existing()
    analysis = analyze_rows(
        rows, draft["season"], existing, draft["source_sha256"], False,
        allow_pending_team=historical, allow_pending_contact=historical,
        ignore_team_name_suggestions=historical,
    )
    if analysis["blocking_errors"] or analysis["unresolved_conflicts"]:
        raise HTTPException(status_code=409, detail="La validación final ha detectado bloqueos")
    lock_id = f"{draft['season']}:{draft['source_sha256']}"
    job_id = new_id()
    operations = _build_import_operations(
        analysis, existing, job_id, {}, allow_pending_team=historical,
        keep_family_candidates_separate=historical, skip_payments=historical,
    )
    try:
        await db.import_locks.insert_one({"_id": lock_id, "job_id": job_id, "created_at": now_iso()})
    except DuplicateKeyError as exc:
        raise HTTPException(status_code=409, detail="Este archivo ya fue importado") from exc
    job = {"id": job_id, "season": draft["season"], "file_sha256": draft["source_sha256"],
           "status": "applying", "summary": analysis["summary"], "operations": operations,
           "created_by_user_id": _staging_actor(), "staging_id": draft_id,
           "created_at": now_iso(), "updated_at": now_iso()}
    try:
        await db.inscription_import_jobs.insert_one(job)
        await _apply_operations(operations)
        await db.inscription_import_jobs.update_one({"id": job_id}, {"$set": {"status": "applied", "updated_at": now_iso()}})
        await db.import_staging.update_one({"id": draft_id}, {"$set": {"status": "imported", "updated_at": datetime.now(timezone.utc)},
                                                                  "$unset": {"records": "", "incidents": "", "duplicates": ""}})
    except Exception as exc:
        await db.import_locks.delete_one({"_id": lock_id})
        await db.inscription_import_jobs.update_one({"id": job_id}, {"$set": {"status": "failed", "updated_at": now_iso()}})
        raise HTTPException(status_code=500, detail="La importación fue revertida sin cambios parciales") from exc
    return {"ok": True, "job_id": job_id, "status": "applied", "operations": len(operations)}


# ================= EXERCISE LIBRARY =================
EXERCISE_PUBLIC_FIELDS = {
    "_id": 0, "id": 1, "name": 1, "category": 1, "objective": 1, "description": 1,
    "instructions": 1, "recommended_duration": 1, "min_players": 1, "max_players": 1,
    "materials": 1, "intensity": 1, "recommended_space": 1, "safety_notes": 1,
    "image_url": 1, "author_id": 1, "author_name": 1, "team_ids": 1, "visibility": 1,
    "status": 1, "created_at": 1, "updated_at": 1, "usage_count": 1,
}
TEMPLATE_PUBLIC_FIELDS = {
    "_id": 0, "id": 1, "name": 1, "description": 1, "team_ids": 1, "visibility": 1,
    "status": 1, "planned_exercises": 1, "author_id": 1, "author_name": 1,
    "created_at": 1, "updated_at": 1,
}


class ExercisePayload(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    objective: Optional[str] = None
    description: Optional[str] = None
    instructions: List[str] = []
    recommended_duration: Optional[int] = None
    min_players: Optional[int] = None
    max_players: Optional[int] = None
    materials: List[str] = []
    intensity: Optional[str] = "medium"
    recommended_space: Optional[str] = None
    safety_notes: Optional[str] = None
    image_url: Optional[str] = None
    team_ids: List[str] = []
    visibility: str = "private"
    status: Optional[str] = None


class PlannedExercisePayload(BaseModel):
    exercise_id: str
    planned_duration: Optional[int] = None
    completed: Optional[bool] = None
    actual_duration: Optional[int] = None
    rating: Optional[str] = None
    observation: Optional[str] = None
    not_completed_reason: Optional[str] = None


class TrainingTemplatePayload(BaseModel):
    name: str
    description: Optional[str] = None
    team_ids: List[str] = []
    visibility: str = "private"
    status: str = "active"
    planned_exercises: List[PlannedExercisePayload] = []


def exercise_scope_query(actor: dict) -> dict:
    if actor.get("role") == "admin":
        return {}
    team_ids = ids(actor.get("assigned_team_ids") or [])
    return {"$or": [
        {"visibility": "club"},
        {"author_id": actor.get("id")},
        {"visibility": "teams", "team_ids": {"$in": team_ids}},
    ]}


async def validate_exercise_teams(team_ids: list[str], actor: dict) -> None:
    if not team_ids:
        return
    found = await db.teams.count_documents({"id": {"$in": team_ids}})
    if found != len(set(team_ids)):
        raise HTTPException(status_code=422, detail="Uno de los equipos no existe")
    if actor.get("role") != "admin":
        allowed = set(ids(actor.get("assigned_team_ids") or []))
        if not set(team_ids).issubset(allowed):
            raise HTTPException(status_code=403, detail="Uno de los equipos queda fuera de tu ámbito")


def can_manage_exercise(actor: dict, exercise: dict) -> bool:
    if actor.get("role") == "admin":
        return True
    if exercise.get("author_id") == actor.get("id"):
        return True
    return bool(set(ids(exercise.get("team_ids") or [])) & set(ids(actor.get("assigned_team_ids") or [])))


async def exercise_doc(identifier: str, *, manage: bool = False) -> dict:
    actor = current_user_context.get() or {}
    document = await db.exercises.find_one({"id": identifier}, EXERCISE_PUBLIC_FIELDS)
    if not document:
        raise HTTPException(status_code=404, detail="Ejercicio no encontrado")
    visible = actor.get("role") == "admin" or bool(await db.exercises.find_one(
        merge_query({"id": identifier}, exercise_scope_query(actor)), {"_id": 0, "id": 1},
    ))
    if not visible or (manage and not can_manage_exercise(actor, document)):
        raise HTTPException(status_code=403, detail="El ejercicio no pertenece a tu ámbito")
    return clean(document)


async def record_exercise_audit(action: str, identifier: str, detail: Optional[dict] = None) -> None:
    actor = current_user_context.get() or {}
    await db.internal_events.insert_one({
        "id": new_id(), "type": f"exercise.{action}", "exercise_id": identifier,
        "actor_id": actor.get("id"), "actor_role": actor.get("role"),
        "detail": detail or {}, "created_at": now_iso(),
    })


@api_router.get("/exercises/meta")
async def exercise_meta():
    return {
        "categories": EXERCISE_CATEGORIES, "intensities": INTENSITIES,
        "visibilities": VISIBILITIES, "states": EXERCISE_STATES, "ratings": RATINGS,
    }


@api_router.get("/exercises")
async def list_exercises(
    search: str = "", category: Optional[str] = None, objective: Optional[str] = None,
    team_id: Optional[str] = None, author_id: Optional[str] = None,
    status: str = "active", visibility: Optional[str] = None,
    sort: str = "name", page: int = Query(1, ge=1), page_size: int = Query(25, ge=1, le=100),
):
    actor = current_user_context.get() or {}
    clauses = [exercise_scope_query(actor)]
    if search.strip():
        clauses.append({"$or": [
            {"name": {"$regex": re.escape(search.strip()), "$options": "i"}},
            {"objective": {"$regex": re.escape(search.strip()), "$options": "i"}},
            {"description": {"$regex": re.escape(search.strip()), "$options": "i"}},
        ]})
    for field, value in (("category", category), ("objective", objective), ("author_id", author_id),
                         ("status", status), ("visibility", visibility)):
        if value and value != "all":
            clauses.append({field: value})
    if team_id:
        if actor.get("role") != "admin" and team_id not in ids(actor.get("assigned_team_ids") or []):
            raise HTTPException(status_code=403, detail="El equipo no pertenece a tu ámbito")
        clauses.append({"team_ids": team_id})
    query = {"$and": clauses}
    order = {
        "name": ("name", 1), "category": ("category", 1),
        "created_at": ("created_at", -1), "usage": ("usage_count", -1),
    }.get(sort, ("name", 1))
    total = await db.exercises.count_documents(query)
    rows = await db.exercises.find(query, EXERCISE_PUBLIC_FIELDS).sort(*order).skip(
        (page - 1) * page_size
    ).limit(page_size).to_list(page_size)
    return {"items": [clean(row) for row in rows], "page": page, "page_size": page_size, "total": total}


@api_router.post("/exercises")
async def create_exercise(payload: ExercisePayload):
    actor = current_user_context.get() or {}
    try:
        values = normalize_exercise(payload.model_dump())
    except ExerciseValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if values["visibility"] == "club" and actor.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administración puede publicar para todo el club")
    await validate_exercise_teams(values["team_ids"], actor)
    identifier = new_id()
    now = now_iso()
    document = {
        "id": identifier, **values, "status": "active", "author_id": actor.get("id"),
        "author_name": actor.get("username"), "usage_count": 0,
        "created_at": now, "updated_at": now,
    }
    await db.exercises.insert_one(document)
    await record_exercise_audit("created", identifier)
    return clean(document)


@api_router.get("/exercises/statistics")
async def get_exercise_statistics():
    actor = current_user_context.get() or {}
    exercises = await db.exercises.find(exercise_scope_query(actor), EXERCISE_PUBLIC_FIELDS).to_list(5000)
    trainings = await list_docs("trainings")
    return exercise_statistics(exercises, trainings)


@api_router.get("/exercises/{exercise_id}")
async def get_exercise(exercise_id: str):
    return await exercise_doc(exercise_id)


@api_router.put("/exercises/{exercise_id}")
async def update_exercise(exercise_id: str, payload: ExercisePayload):
    actor = current_user_context.get() or {}
    existing = await exercise_doc(exercise_id, manage=True)
    try:
        values = normalize_exercise(payload.model_dump(exclude_unset=True), partial=True)
    except ExerciseValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    try:
        validate_exercise_update(existing, values)
    except ExerciseValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    visibility = values.get("visibility", existing.get("visibility"))
    if visibility == "club" and actor.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administración puede publicar para todo el club")
    team_ids = values.get("team_ids", existing.get("team_ids") or [])
    await validate_exercise_teams(team_ids, actor)
    values["updated_at"] = now_iso()
    updated = await db.exercises.find_one_and_update(
        {"id": exercise_id}, {"$set": values}, return_document=ReturnDocument.AFTER,
        projection=EXERCISE_PUBLIC_FIELDS,
    )
    await record_exercise_audit("updated", exercise_id, {"fields": sorted(values)})
    return clean(updated)


@api_router.post("/exercises/{exercise_id}/duplicate")
async def duplicate_exercise(exercise_id: str):
    source = await exercise_doc(exercise_id)
    actor = current_user_context.get() or {}
    values = {key: source.get(key) for key in (
        "name", "category", "objective", "description", "instructions", "recommended_duration",
        "min_players", "max_players", "materials", "intensity", "recommended_space",
        "safety_notes", "image_url", "team_ids", "visibility",
    )}
    if values["visibility"] == "club" and actor.get("role") != "admin":
        values["visibility"] = "private"; values["team_ids"] = []
    values["name"] = f"{values['name']} (copia)"
    return await create_exercise(ExercisePayload(**values))


@api_router.post("/exercises/{exercise_id}/archive")
async def archive_exercise(exercise_id: str):
    await exercise_doc(exercise_id, manage=True)
    updated = await db.exercises.find_one_and_update(
        {"id": exercise_id}, {"$set": {"status": "archived", "updated_at": now_iso()}},
        return_document=ReturnDocument.AFTER, projection=EXERCISE_PUBLIC_FIELDS,
    )
    await record_exercise_audit("archived", exercise_id)
    return clean(updated)


@api_router.post("/exercises/{exercise_id}/restore")
async def restore_exercise(exercise_id: str):
    await exercise_doc(exercise_id, manage=True)
    updated = await db.exercises.find_one_and_update(
        {"id": exercise_id}, {"$set": {"status": "active", "updated_at": now_iso()}},
        return_document=ReturnDocument.AFTER, projection=EXERCISE_PUBLIC_FIELDS,
    )
    await record_exercise_audit("restored", exercise_id)
    return clean(updated)


async def template_doc(identifier: str, *, manage: bool = False) -> dict:
    actor = current_user_context.get() or {}
    query = exercise_scope_query(actor)
    document = await db.training_templates.find_one(
        merge_query({"id": identifier}, query), TEMPLATE_PUBLIC_FIELDS,
    )
    if not document:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    document = clean(document)
    if manage and not can_manage_exercise(actor, document):
        raise HTTPException(status_code=403, detail="La plantilla no pertenece a tu ámbito")
    return document


@api_router.get("/training-templates")
async def list_training_templates(status: str = "active"):
    actor = current_user_context.get() or {}
    query = merge_query({"status": status}, exercise_scope_query(actor))
    rows = await db.training_templates.find(query, TEMPLATE_PUBLIC_FIELDS).sort("name", 1).to_list(1000)
    return [clean(row) for row in rows]


async def accessible_exercise_map(identifiers: list[str]) -> dict[str, dict]:
    actor = current_user_context.get() or {}
    rows = await db.exercises.find(
        merge_query({"id": {"$in": identifiers}}, exercise_scope_query(actor)),
        EXERCISE_PUBLIC_FIELDS,
    ).to_list(100)
    return {row["id"]: clean(row) for row in rows}


def validate_exercises_for_training(exercises: dict[str, dict], identifiers: list[str], team_id: Optional[str]) -> None:
    for identifier in identifiers:
        exercise = exercises.get(identifier)
        if not exercise:
            raise HTTPException(status_code=422, detail="Uno de los ejercicios no está disponible")
        if exercise.get("visibility") == "teams" and team_id not in ids(exercise.get("team_ids") or []):
            raise HTTPException(status_code=403, detail="Uno de los ejercicios no pertenece al equipo del entrenamiento")


async def validate_training_template_reference(template_id: Optional[str]) -> None:
    if template_id:
        await template_doc(template_id)


@api_router.post("/training-templates")
async def create_training_template(payload: TrainingTemplatePayload):
    actor = current_user_context.get() or {}
    raw = payload.model_dump()
    exercises = await accessible_exercise_map([row["exercise_id"] for row in raw["planned_exercises"]])
    try:
        values = normalize_template(raw, exercises)
    except ExerciseValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if values["visibility"] == "club" and actor.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administración puede publicar para todo el club")
    await validate_exercise_teams(values["team_ids"], actor)
    now = now_iso()
    document = {
        "id": new_id(), **values, "author_id": actor.get("id"),
        "author_name": actor.get("username"), "created_at": now, "updated_at": now,
    }
    await db.training_templates.insert_one(document)
    return clean(document)


@api_router.put("/training-templates/{template_id}")
async def update_training_template(template_id: str, payload: TrainingTemplatePayload):
    existing = await template_doc(template_id, manage=True)
    actor = current_user_context.get() or {}
    raw = payload.model_dump()
    exercises = await accessible_exercise_map([row["exercise_id"] for row in raw["planned_exercises"]])
    try:
        values = normalize_template(raw, exercises)
    except ExerciseValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if values["visibility"] == "club" and actor.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administración puede publicar para todo el club")
    await validate_exercise_teams(values["team_ids"], actor)
    values["updated_at"] = now_iso()
    updated = await db.training_templates.find_one_and_update(
        {"id": existing["id"]}, {"$set": values}, return_document=ReturnDocument.AFTER,
        projection=TEMPLATE_PUBLIC_FIELDS,
    )
    return clean(updated)


@api_router.post("/training-templates/{template_id}/archive")
async def archive_training_template(template_id: str):
    await template_doc(template_id, manage=True)
    updated = await db.training_templates.find_one_and_update(
        {"id": template_id}, {"$set": {"status": "archived", "updated_at": now_iso()}},
        return_document=ReturnDocument.AFTER, projection=TEMPLATE_PUBLIC_FIELDS,
    )
    return clean(updated)


@api_router.post("/training-templates/{template_id}/restore")
async def restore_training_template(template_id: str):
    await template_doc(template_id, manage=True)
    updated = await db.training_templates.find_one_and_update(
        {"id": template_id}, {"$set": {"status": "active", "updated_at": now_iso()}},
        return_document=ReturnDocument.AFTER, projection=TEMPLATE_PUBLIC_FIELDS,
    )
    return clean(updated)


# ================= TRAININGS =================
class AsistenciaItem(BaseModel):
    player_id: str
    estado: str = "presente"  # presente, justificada, injustificada, lesion
    motivo: Optional[str] = None

    @field_validator("estado")
    @classmethod
    def validate_estado(cls, value: str):
        if value not in ATTENDANCE_STATES:
            raise ValueError("Estado de asistencia no válido")
        return value


class Training(BaseModel):
    fecha: Optional[str] = None
    hora: Optional[str] = None
    equipo_id: Optional[str] = None
    campo: Optional[str] = None
    asistencia: List[AsistenciaItem] = []
    ejercicios: Optional[str] = None
    planned_exercises: List[PlannedExercisePayload] = []
    session_template_id: Optional[str] = None
    observaciones: Optional[str] = None
    callup_id: Optional[str] = None


# Evaluaciones individuales: las escalas son ordinales y deliberadamente
# limitadas a 1--5. No representan mediciones clínicas ni predicciones.
EVALUATION_SCORE_FIELDS = (
    "participacion", "actitud", "esfuerzo", "comprension_tactica", "tecnica", "condicion_fisica",
)
EVALUATION_STATUSES = {"draft", "completed", "closed"}


class TrainingEvaluationPayload(BaseModel):
    training_id: str
    player_id: str
    asistencia: Optional[str] = None
    participacion: Optional[int] = Field(default=None, ge=1, le=5)
    actitud: Optional[int] = Field(default=None, ge=1, le=5)
    esfuerzo: Optional[int] = Field(default=None, ge=1, le=5)
    comprension_tactica: Optional[int] = Field(default=None, ge=1, le=5)
    tecnica: Optional[int] = Field(default=None, ge=1, le=5)
    condicion_fisica: Optional[int] = Field(default=None, ge=1, le=5)
    observaciones: Optional[str] = Field(default=None, max_length=2000)
    incidencias: Optional[str] = Field(default=None, max_length=2000)
    estado: str = "draft"
    fecha_evaluacion: Optional[str] = None

    @field_validator("asistencia")
    @classmethod
    def validate_attendance(cls, value: Optional[str]):
        if value is not None and value not in ATTENDANCE_STATES:
            raise ValueError("Estado de asistencia no válido")
        return value

    @field_validator("training_id", "player_id")
    @classmethod
    def validate_reference_ids(cls, value: str):
        value = value.strip()
        if not value:
            raise ValueError("La evaluación requiere referencias válidas")
        return value

    @field_validator("estado")
    @classmethod
    def validate_evaluation_status(cls, value: str):
        if value not in EVALUATION_STATUSES:
            raise ValueError("Estado de evaluación no válido")
        return value

    @field_validator("observaciones", "incidencias")
    @classmethod
    def normalize_evaluation_text(cls, value: Optional[str]):
        if value is None:
            return None
        value = value.strip()
        return value or None

    @field_validator("fecha_evaluacion")
    @classmethod
    def validate_evaluation_date(cls, value: Optional[str]):
        if value:
            try:
                datetime.fromisoformat(value.replace("Z", "+00:00"))
            except ValueError as exc:
                raise ValueError("Fecha de evaluación no válida") from exc
        return value


class TrainingEvaluationBulkPayload(BaseModel):
    training_id: str
    evaluations: List[TrainingEvaluationPayload] = Field(min_length=1, max_length=500)


_training_evaluation_write_lock = asyncio.Lock()


async def ensure_training_evaluation_indexes() -> None:
    """Inicialización idempotente del índice de integridad de evaluaciones.

    No transforma datos ni ejecuta migraciones: solo garantiza que la nueva
    colección no pueda aceptar dos evaluaciones activas para el mismo jugador
    y entrenamiento. La documentación de la colección describe su rollback.
    """
    await db.training_evaluations.create_index(
        [("training_id", 1), ("player_id", 1)],
        unique=True,
        name="training_player_unique",
    )
    await db.training_evaluations.create_index(
        [("player_id", 1), ("fecha_evaluacion", -1)],
        name="player_evaluation_history",
    )


async def _evaluation_training_context(training_id: str, player_id: Optional[str] = None) -> dict:
    """Resuelve entrenamiento, equipo, jugador y asistencia dentro del ámbito."""
    actor = current_user_context.get() or {}
    training = await get_doc("trainings", training_id)
    team_id = training.get("equipo_id")
    if not team_id:
        raise HTTPException(status_code=422, detail="El entrenamiento no tiene equipo asociado")
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=422, detail="El equipo del entrenamiento no existe")
    if actor.get("role") != "admin" and team_id not in set(ids(actor.get("assigned_team_ids") or [])):
        raise HTTPException(status_code=403, detail="El entrenamiento queda fuera de tu ámbito")
    player = None
    attendance = None
    if player_id:
        player = await db.players.find_one({"id": player_id}, {"_id": 0})
        if not player:
            raise HTTPException(status_code=404, detail="El jugador no existe")
        if player.get("equipo_id") != team_id:
            raise HTTPException(status_code=422, detail="El jugador no pertenece al equipo del entrenamiento")
        attendance = next((row for row in training.get("asistencia") or [] if row.get("player_id") == player_id), None)
        if attendance is None:
            raise HTTPException(status_code=422, detail="El jugador no tiene asistencia registrada en este entrenamiento")
    return {"actor": actor, "training": training, "team": team, "player": player, "attendance": attendance}


def _validate_evaluation_values(data: dict, attendance_state: str, *, require_complete: bool = False) -> None:
    """Valida la relación entre asistencia, escala y estado sin inventar datos."""
    if attendance_state not in ATTENDANCE_STATES:
        raise HTTPException(status_code=422, detail="La asistencia del jugador no es válida")
    scores = [data.get(field) for field in EVALUATION_SCORE_FIELDS]
    if attendance_state != "presente" and any(value is not None for value in scores):
        raise HTTPException(
            status_code=422,
            detail="No se pueden registrar puntuaciones para un jugador ausente; revisa primero la asistencia",
        )
    if require_complete and attendance_state == "presente" and any(value is None for value in scores):
        raise HTTPException(
            status_code=422,
            detail="La evaluación está incompleta: completa todos los criterios antes de cerrarla",
        )


def _evaluation_document(payload: TrainingEvaluationPayload, context: dict, actor: dict, *, status: Optional[str] = None) -> dict:
    data = payload.model_dump()
    attendance_state = context["attendance"].get("estado")
    requested_attendance = data.get("asistencia")
    if requested_attendance and requested_attendance != attendance_state:
        raise HTTPException(status_code=409, detail="La evaluación debe reutilizar la asistencia registrada")
    data["asistencia"] = attendance_state
    data["equipo_id"] = context["team"]["id"]
    data["temporada"] = context["team"].get("temporada") or context["training"].get("temporada")
    if not data["temporada"]:
        raise HTTPException(status_code=422, detail="No se ha podido determinar la temporada del entrenamiento")
    data["categoria"] = context["team"].get("categoria")
    data["evaluador_id"] = actor.get("id")
    data["evaluador_role"] = actor.get("role")
    data["fecha_evaluacion"] = data.get("fecha_evaluacion") or now_iso()
    data["estado"] = status or data.get("estado") or "draft"
    _validate_evaluation_values(data, attendance_state, require_complete=data["estado"] in {"completed", "closed"})
    if data["estado"] == "closed" and status != "closed":
        raise HTTPException(status_code=422, detail="El cierre debe confirmarse con el endpoint de cierre")
    return data


async def record_training_evaluation_audit(action: str, evaluation: dict, *, previous: Optional[dict] = None) -> None:
    actor = current_user_context.get() or {}
    changed = []
    if previous:
        changed = sorted(
            key for key in evaluation.keys()
            if key not in {"updated_at", "created_at"} and previous.get(key) != evaluation.get(key)
        )
    await db.internal_events.insert_one({
        "id": new_id(),
        "type": f"training_evaluation.{action}",
        "actor_user_id": actor.get("id"),
        "actor_role": actor.get("role"),
        "training_id": evaluation.get("training_id"),
        "player_id": evaluation.get("player_id"),
        "equipo_id": evaluation.get("equipo_id"),
        "evaluation_id": evaluation.get("id"),
        "status": evaluation.get("estado"),
        "detail": {"changed_fields": changed},
        "created_at": now_iso(),
        "result": "success",
    })


def _evaluation_public(document: dict, players_by_id: Optional[dict] = None) -> dict:
    row = dict(document)
    row.pop("_id", None)
    if players_by_id is not None:
        player = players_by_id.get(row.get("player_id"), {})
        row["player_name"] = f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip() or "—"
    return row


async def _training_evaluation_listing(training_id: str, status: Optional[str] = None) -> dict:
    context = await _evaluation_training_context(training_id)
    team_id = context["team"]["id"]
    players = await list_docs("players", {"equipo_id": team_id})
    players_by_id = {player.get("id"): player for player in players}
    attendance_by_player = {row.get("player_id"): row for row in context["training"].get("asistencia") or []}
    evaluations = await db.training_evaluations.find({"training_id": training_id}, {"_id": 0}).sort("updated_at", -1).to_list(5000)
    evaluations_by_player = {}
    for evaluation in evaluations:
        evaluations_by_player.setdefault(evaluation.get("player_id"), evaluation)
    rows = []
    for player in players:
        player_id = player.get("id")
        evaluation = evaluations_by_player.get(player_id)
        attendance = attendance_by_player.get(player_id)
        row = {
            "player_id": player_id,
            "player_name": f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip() or "—",
            "asistencia": attendance.get("estado") if attendance else None,
            "motivo_asistencia": attendance.get("motivo") if attendance else None,
            "evaluation": _evaluation_public(evaluation, players_by_id) if evaluation else None,
        }
        row["evaluation_status"] = (evaluation or {}).get("estado") if evaluation else "pending"
        if status and row["evaluation_status"] != status:
            continue
        rows.append(row)
    evaluated = [row for row in rows if row["evaluation"]]
    incomplete = [row for row in evaluated if row["evaluation_status"] == "draft"]
    absent = [row for row in rows if row.get("asistencia") in {"justificada", "injustificada", "lesion"}]
    return {
        "training": {"id": training_id, "fecha": context["training"].get("fecha"), "hora": context["training"].get("hora")},
        "team": {"id": team_id, "nombre": context["team"].get("nombre"), "categoria": context["team"].get("categoria"), "temporada": context["team"].get("temporada")},
        "players": rows,
        "summary": {
            "total": len(rows), "evaluated": len(evaluated), "pending": len(rows) - len(evaluated),
            "incomplete": len(incomplete), "absent": len(absent),
        },
    }


async def attendance_data(desde: Optional[str] = None, hasta: Optional[str] = None,
                          equipo_id: Optional[str] = None, player_id: Optional[str] = None):
    """Devuelve datos ya acotados por la identidad autenticada, nunca por el cliente."""
    actor = current_user_context.get() or {}
    trainings = await list_docs("trainings")
    players = await list_docs("players")
    teams = await list_docs("teams")
    allowed_players = {player.get("id") for player in players if player.get("id")}
    if player_id:
        if player_id not in allowed_players:
            raise HTTPException(status_code=403, detail="El jugador solicitado no pertenece a tu ámbito")
        allowed_players = {player_id}
    if equipo_id:
        allowed_teams = {team.get("id") for team in teams}
        if equipo_id not in allowed_teams:
            raise HTTPException(status_code=403, detail="El equipo solicitado no pertenece a tu ámbito")
        trainings = [training for training in trainings if training.get("equipo_id") == equipo_id]
        allowed_players = {player.get("id") for player in players if player.get("equipo_id") == equipo_id and player.get("id") in allowed_players}
    # Las familias y jugadores ven exclusivamente sus filas, incluso si el entrenamiento tiene más asistentes.
    if actor.get("role") in {"family", "player"}:
        trainings = [{**training, "asistencia": [row for row in training.get("asistencia", []) if row.get("player_id") in allowed_players]}
                     for training in trainings]
    return trainings, players, teams, allowed_players


@api_router.get("/attendance/summary")
async def get_attendance_summary(desde: Optional[str] = None, hasta: Optional[str] = None,
                                 equipo_id: Optional[str] = None, categoria: Optional[str] = None,
                                 player_id: Optional[str] = None, periodo: str = "weekly"):
    trainings, players, teams, allowed_players = await attendance_data(desde, hasta, equipo_id, player_id)
    if categoria:
        team_ids = {team.get("id") for team in teams if team.get("categoria") == categoria}
        trainings = [training for training in trainings if training.get("equipo_id") in team_ids]
        allowed_players = {player.get("id") for player in players if player.get("equipo_id") in team_ids and player.get("id") in allowed_players}
    threshold_doc = await db.settings.find_one({"id": SETTINGS_ID}, {"_id": 0, "attendance_alert_threshold": 1}) or {}
    threshold = threshold_doc.get("attendance_alert_threshold", 3)
    rows = attendance_rows(trainings, allowed_players, desde, hasta)
    names = {player.get("id"): f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip() for player in players}
    alerts = [{**alert, "player_name": names.get(alert["player_id"], "—")} for alert in repeated_absence_alerts(trainings, threshold, allowed_players, desde, hasta)]
    return {
        "summary": attendance_summary(trainings, allowed_players, desde, hasta),
        "trend": attendance_trend(trainings, allowed_players, periodo if periodo in {"weekly", "monthly"} else "weekly", desde, hasta),
        "players": [{"player_id": pid, "player_name": names.get(pid, "—"), **stats} for pid, stats in player_percentages(trainings, allowed_players, desde, hasta).items()],
        "alerts": alerts,
        "callup_comparison": callup_attendance_comparison(trainings, await list_docs("callups"), allowed_players),
        "recent": sorted(rows, key=lambda row: row.get("fecha") or "", reverse=True)[:12],
    }


def attendance_export_pdf(summary: dict, lang: str) -> io.BytesIO:
    labels = {
        "es": ("Informe de asistencia", "Periodo", "Presentes", "Justificadas", "Injustificadas", "Lesiones", "Porcentaje de presencia"),
        "eu": ("Asistentzia-txostena", "Aldia", "Bertaratutakoak", "Justifikatuak", "Justifikatu gabeak", "Lesioak", "Bertaratze-ehunekoa"),
    }
    title, period, present, justified, unjustified, injury, percentage = labels.get(lang, labels["es"])
    data = summary["summary"]
    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=18 * mm)
    styles = getSampleStyleSheet()
    heading = Table(
        [[pdf_logo(size_mm=16), Paragraph(title + " / " + labels["eu" if lang == "es" else "es"][0], styles["Title"])]],
        colWidths=[20 * mm, 140 * mm],
    )
    heading.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    story = [heading, Spacer(1, 8 * mm)]
    story.append(Table([[period, f"{data.get('desde') or '—'} — {data.get('hasta') or '—'}"],
                        [present, data["presente"]], [justified, data["justificada"]],
                        [unjustified, data["injustificada"]], [injury, data["lesion"]], [percentage, f"{data['porcentaje_presencia']}%"]],
                       colWidths=[75 * mm, 85 * mm], style=TableStyle([
                           ("GRID", (0, 0), (-1, -1), .35, colors.HexColor("#CBD5E1")),
                           ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
                           ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"), ("PADDING", (0, 0), (-1, -1), 7),
                       ])))
    document.build(story)
    buffer.seek(0)
    return buffer


@api_router.get("/attendance/export.pdf")
async def export_attendance_pdf(desde: Optional[str] = None, hasta: Optional[str] = None,
                                equipo_id: Optional[str] = None, player_id: Optional[str] = None, lang: str = "es"):
    summary = await get_attendance_summary(desde, hasta, equipo_id, None, player_id)
    return StreamingResponse(attendance_export_pdf(summary, "eu" if lang == "eu" else "es"), media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=asistencia.pdf"})


@api_router.get("/attendance/export.xlsx")
async def export_attendance_excel(desde: Optional[str] = None, hasta: Optional[str] = None,
                                  equipo_id: Optional[str] = None, player_id: Optional[str] = None):
    summary = await get_attendance_summary(desde, hasta, equipo_id, None, player_id)
    buffer = io.BytesIO()
    pd.DataFrame(summary["players"]).to_excel(buffer, index=False)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=asistencia.xlsx"})


@api_router.post("/trainings")
async def create_training(tr: Training):
    payload = tr.model_dump()
    exercise_ids = [row["exercise_id"] for row in payload.get("planned_exercises") or []]
    exercises = await accessible_exercise_map(exercise_ids)
    validate_exercises_for_training(exercises, exercise_ids, payload.get("equipo_id"))
    await validate_training_template_reference(payload.get("session_template_id"))
    try:
        payload["planned_exercises"] = normalize_planned_exercises(payload.get("planned_exercises"), exercises)
    except ExerciseValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    payload["attendance_history"] = []
    payload["attendance_updated_by"] = None
    created = await insert_doc("trainings", payload)
    if exercise_ids:
        await db.exercises.update_many({"id": {"$in": exercise_ids}}, {"$inc": {"usage_count": 1}})
        for identifier in exercise_ids:
            await record_exercise_audit("assigned", identifier, {"training_id": created["id"]})
    return created


@api_router.get("/trainings")
async def get_trainings(equipo_id: Optional[str] = None):
    query = {"equipo_id": equipo_id} if equipo_id else {}
    trs = await list_docs("trainings", query)
    teams = {t["id"]: t["nombre"] for t in await list_docs("teams")}
    for tr in trs:
        tr["equipo_nombre"] = teams.get(tr.get("equipo_id"), "—")
        a = tr.get("asistencia", [])
        tr["presentes"] = len([x for x in a if x.get("estado") == "presente"])
        tr["total_asistencia"] = len(a)
    return trs


@api_router.get("/trainings/{tr_id}")
async def get_training(tr_id: str):
    tr = await get_doc("trainings", tr_id)
    players = {p["id"]: p for p in await list_docs("players")}
    for item in tr.get("asistencia", []):
        p = players.get(item["player_id"], {})
        item["nombre"] = f"{p.get('nombre','')} {p.get('apellidos','')}".strip()
    return tr


@api_router.put("/trainings/{tr_id}")
async def edit_training(tr_id: str, tr: Training):
    existing = await get_doc("trainings", tr_id)
    payload = tr.model_dump()
    exercise_ids = [row["exercise_id"] for row in payload.get("planned_exercises") or []]
    exercises = await accessible_exercise_map(exercise_ids)
    validate_exercises_for_training(exercises, exercise_ids, payload.get("equipo_id"))
    if payload.get("session_template_id") != existing.get("session_template_id"):
        await validate_training_template_reference(payload.get("session_template_id"))
    existing_ids = {row.get("exercise_id") for row in existing.get("planned_exercises") or []}
    if any(exercises.get(identifier, {}).get("status") != "active" for identifier in set(exercise_ids) - existing_ids):
        raise HTTPException(status_code=422, detail="No se puede añadir un ejercicio archivado")
    try:
        payload["planned_exercises"] = normalize_planned_exercises(
            payload.get("planned_exercises"), exercises, allow_archived_existing=True,
        )
    except ExerciseValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    actor = current_user_context.get() or {}
    changes = attendance_history(existing.get("asistencia", []), payload.get("asistencia", []), actor)
    payload["attendance_history"] = [*(existing.get("attendance_history") or []), *changes]
    if changes:
        payload["attendance_updated_by"] = {"id": actor.get("id"), "role": actor.get("role"), "at": now_iso()}
    updated = await update_doc("trainings", tr_id, payload)
    added_ids = set(exercise_ids) - existing_ids
    if added_ids:
        await db.exercises.update_many({"id": {"$in": list(added_ids)}}, {"$inc": {"usage_count": 1}})
        for identifier in added_ids:
            await record_exercise_audit("assigned", identifier, {"training_id": tr_id})
    schedule_changed = any(existing.get(field) != updated.get(field) for field in ("fecha", "hora", "campo"))
    if schedule_changed:
        users = await users_for_team_players(updated.get("equipo_id"))
        await enqueue_notifications(users, "schedule.changed", "Cambio de entrenamiento / Entrenamendu aldaketa",
                                    str(updated.get("campo") or "Entrenamiento"), "/entrenamientos", "urgent",
                                    {"training_id": tr_id}, f"schedule.changed:training:{tr_id}:{updated.get('updated_at')}")
    return updated


@api_router.delete("/trainings/{tr_id}")
async def remove_training(tr_id: str):
    return await delete_doc("trainings", tr_id)


@api_router.post("/trainings/{tr_id}/duplicate")
async def duplicate_training(tr_id: str, data: Dict[str, Any]):
    source = await get_doc("trainings", tr_id)
    payload = {
        "fecha": data.get("fecha"), "hora": data.get("hora", source.get("hora")),
        "equipo_id": data.get("equipo_id", source.get("equipo_id")),
        "campo": data.get("campo", source.get("campo")), "asistencia": [],
        "ejercicios": source.get("ejercicios"),
        "planned_exercises": [
            {**row, "completed": None, "actual_duration": None, "rating": None,
             "observation": None, "not_completed_reason": None}
            for row in source.get("planned_exercises") or []
        ],
        "session_template_id": source.get("session_template_id"),
        "observaciones": data.get("observaciones"), "callup_id": None,
    }
    model = Training(**payload)
    duplicate_payload = model.model_dump()
    exercise_ids = [row["exercise_id"] for row in duplicate_payload.get("planned_exercises") or []]
    exercises = await accessible_exercise_map(exercise_ids)
    validate_exercises_for_training(exercises, exercise_ids, duplicate_payload.get("equipo_id"))
    await validate_training_template_reference(duplicate_payload.get("session_template_id"))
    try:
        duplicate_payload["planned_exercises"] = normalize_planned_exercises(
            duplicate_payload["planned_exercises"], exercises, allow_archived_existing=True,
        )
    except ExerciseValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    duplicate_payload["attendance_history"] = []
    duplicate_payload["attendance_updated_by"] = None
    created = await insert_doc("trainings", duplicate_payload)
    if exercise_ids:
        await db.exercises.update_many({"id": {"$in": exercise_ids}}, {"$inc": {"usage_count": 1}})
    return created


# ================= TRAINING EVALUATIONS =================
@api_router.get("/training-evaluations/training/{training_id}")
async def list_training_evaluations(training_id: str, status: Optional[str] = None):
    if status and status not in EVALUATION_STATUSES | {"pending"}:
        raise HTTPException(status_code=422, detail="Filtro de estado de evaluación no válido")
    return await _training_evaluation_listing(training_id, status)


@api_router.get("/training-evaluations/pending")
async def list_pending_training_evaluations(training_id: str):
    return await _training_evaluation_listing(training_id, "pending")


@api_router.get("/training-evaluations/player/{player_id}")
async def list_player_training_evaluations(player_id: str, training_id: Optional[str] = None):
    player = await get_doc("players", player_id)
    query = {"player_id": player_id}
    actor = current_user_context.get() or {}
    audit_query = {"type": {"$regex": "^training_evaluation\\."}, "player_id": player_id}
    if actor.get("role") != "admin":
        allowed_team_ids = ids(actor.get("assigned_team_ids") or [])
        query["equipo_id"] = {"$in": allowed_team_ids}
        audit_query["equipo_id"] = {"$in": allowed_team_ids}
    if training_id:
        await _evaluation_training_context(training_id, player_id)
        query["training_id"] = training_id
    evaluations = await db.training_evaluations.find(query, {"_id": 0}).sort("fecha_evaluacion", -1).to_list(5000)
    audits = await db.internal_events.find(
        audit_query,
        {"_id": 0, "type": 1, "actor_user_id": 1, "actor_role": 1, "training_id": 1,
         "evaluation_id": 1, "status": 1, "detail": 1, "created_at": 1},
    ).sort("created_at", -1).to_list(5000)
    return {
        "player": {"id": player.get("id"), "nombre": f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip()},
        "evaluations": [_evaluation_public(row) for row in evaluations],
        "audit": audits,
    }


@api_router.get("/training-evaluations/{evaluation_id}")
async def get_training_evaluation(evaluation_id: str):
    evaluation = await db.training_evaluations.find_one({"id": evaluation_id}, {"_id": 0})
    if not evaluation:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")
    await _evaluation_training_context(evaluation["training_id"], evaluation["player_id"])
    audits = await db.internal_events.find(
        {"type": {"$regex": "^training_evaluation\\."}, "evaluation_id": evaluation_id},
        {"_id": 0, "type": 1, "actor_user_id": 1, "actor_role": 1, "status": 1, "detail": 1, "created_at": 1},
    ).sort("created_at", -1).to_list(200)
    return {"evaluation": _evaluation_public(evaluation), "audit": audits}


@api_router.post("/training-evaluations")
async def create_training_evaluation(payload: TrainingEvaluationPayload):
    context = await _evaluation_training_context(payload.training_id, payload.player_id)
    actor = current_user_context.get() or {}
    data = _evaluation_document(payload, context, actor)
    if data.get("estado") == "closed":
        raise HTTPException(status_code=422, detail="El cierre debe confirmarse con el endpoint de cierre")
    async with _training_evaluation_write_lock:
        await ensure_training_evaluation_indexes()
        existing = await db.training_evaluations.find_one(
            {"training_id": payload.training_id, "player_id": payload.player_id}, {"_id": 0}
        )
        if existing:
            raise HTTPException(status_code=409, detail="Ya existe una evaluación para este jugador y entrenamiento")
        document = {"id": new_id(), **data, "created_at": now_iso(), "updated_at": now_iso()}
        try:
            await db.training_evaluations.insert_one(document)
        except DuplicateKeyError as exc:
            raise HTTPException(status_code=409, detail="Ya existe una evaluación para este jugador y entrenamiento") from exc
    await record_training_evaluation_audit("created", document)
    return _evaluation_public(document)


@api_router.post("/training-evaluations/bulk")
async def save_training_evaluations_bulk(payload: TrainingEvaluationBulkPayload):
    actor = current_user_context.get() or {}
    seen_players = set()
    contexts = []
    for item in payload.evaluations:
        if item.training_id != payload.training_id:
            raise HTTPException(status_code=422, detail="Todas las evaluaciones deben pertenecer al mismo entrenamiento")
        if item.player_id in seen_players:
            raise HTTPException(status_code=422, detail="La petición contiene jugadores duplicados")
        seen_players.add(item.player_id)
        context = await _evaluation_training_context(payload.training_id, item.player_id)
        data = _evaluation_document(item, context, actor)
        contexts.append((item, context, data))

    async with _training_evaluation_write_lock:
        await ensure_training_evaluation_indexes()
        existing_rows = await db.training_evaluations.find(
            {"training_id": payload.training_id, "player_id": {"$in": list(seen_players)}}, {"_id": 0}
        ).to_list(500)
        existing_by_player = {row.get("player_id"): row for row in existing_rows}
        for item, _context, data in contexts:
            existing = existing_by_player.get(item.player_id)
            if existing and existing.get("estado") == "closed":
                raise HTTPException(status_code=409, detail="No se puede editar una evaluación cerrada")
            if data.get("estado") == "closed":
                raise HTTPException(status_code=422, detail="El cierre debe confirmarse con el endpoint de cierre")
        saved = []
        now = now_iso()
        for item, _context, data in contexts:
            existing = existing_by_player.get(item.player_id)
            if existing:
                document = {**existing, **data, "id": existing["id"], "created_at": existing.get("created_at") or now, "updated_at": now}
                await db.training_evaluations.update_one({"id": existing["id"]}, {"$set": document})
                await record_training_evaluation_audit("updated", document, previous=existing)
            else:
                document = {"id": new_id(), **data, "created_at": now, "updated_at": now}
                await db.training_evaluations.insert_one(document)
                await record_training_evaluation_audit("created", document)
            saved.append(_evaluation_public(document))
    return {"saved": saved, "count": len(saved)}


@api_router.put("/training-evaluations/{evaluation_id}")
async def update_training_evaluation(evaluation_id: str, payload: TrainingEvaluationPayload):
    existing = await db.training_evaluations.find_one({"id": evaluation_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")
    if existing.get("estado") == "closed":
        raise HTTPException(status_code=409, detail="No se puede editar una evaluación cerrada")
    if payload.training_id != existing.get("training_id") or payload.player_id != existing.get("player_id"):
        raise HTTPException(status_code=409, detail="La evaluación no puede cambiar de jugador ni de entrenamiento")
    context = await _evaluation_training_context(payload.training_id, payload.player_id)
    data = _evaluation_document(payload, context, current_user_context.get() or {})
    if data.get("estado") == "closed":
        raise HTTPException(status_code=422, detail="El cierre debe confirmarse con el endpoint de cierre")
    document = {**existing, **data, "id": evaluation_id, "created_at": existing.get("created_at") or now_iso(), "updated_at": now_iso()}
    async with _training_evaluation_write_lock:
        await db.training_evaluations.update_one({"id": evaluation_id}, {"$set": document})
    await record_training_evaluation_audit("updated", document, previous=existing)
    return _evaluation_public(document)


@api_router.post("/training-evaluations/{evaluation_id}/close")
async def close_training_evaluation(evaluation_id: str):
    existing = await db.training_evaluations.find_one({"id": evaluation_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")
    if existing.get("estado") == "closed":
        return _evaluation_public(existing)
    context = await _evaluation_training_context(existing["training_id"], existing["player_id"])
    _validate_evaluation_values(existing, context["attendance"].get("estado"), require_complete=True)
    document = {**existing, "estado": "closed", "updated_at": now_iso()}
    async with _training_evaluation_write_lock:
        await db.training_evaluations.update_one({"id": evaluation_id}, {"$set": document})
    await record_training_evaluation_audit("closed", document, previous=existing)
    return _evaluation_public(document)


# ================= INTEGRAL STATISTICS =================
STATISTICS_STAFF_ROLES = {"admin", "coordinator", "coach"}


def _statistics_actor() -> dict:
    return current_user_context.get() or {}


def _statistics_filters(*, temporada: Optional[str] = None, categoria: Optional[str] = None,
                        equipo_id: Optional[str] = None, player_id: Optional[str] = None,
                        modalidad: Optional[str] = None, desde: Optional[str] = None,
                        hasta: Optional[str] = None, periodo: str = "weekly",
                        estado: Optional[str] = None, activo: Optional[bool] = None,
                        estado_partido: Optional[str] = None,
                        estado_entrenamiento: Optional[str] = None,
                        estado_convocatoria: Optional[str] = None) -> dict:
    for value, label in ((desde, "desde"), (hasta, "hasta")):
        if value:
            try:
                date.fromisoformat(str(value))
            except ValueError as exc:
                raise HTTPException(status_code=422, detail=f"Fecha {label} no válida") from exc
    if desde and hasta and desde > hasta:
        raise HTTPException(status_code=422, detail="El intervalo estadístico no es válido")
    if periodo not in {"weekly", "monthly"}:
        raise HTTPException(status_code=422, detail="El periodo debe ser weekly o monthly")
    if modalidad:
        modalidad = str(modalidad).upper()
        if modalidad not in {"F7", "F11"}:
            raise HTTPException(status_code=422, detail="Modalidad no válida")
    if activo is not None:
        estado = "activo" if activo else "inactivo"
    return {key: value for key, value in {
        "temporada": temporada, "categoria": categoria, "equipo_id": equipo_id,
        "player_id": player_id, "modalidad": modalidad, "desde": desde, "hasta": hasta,
        "period": periodo, "estado": estado, "estado_partido": estado_partido,
        "estado_entrenamiento": estado_entrenamiento, "estado_convocatoria": estado_convocatoria,
    }.items() if value not in (None, "", "all")}


async def _statistics_data(filters: dict) -> tuple[dict, dict]:
    actor = _statistics_actor()
    if actor.get("role") not in STATISTICS_STAFF_ROLES:
        raise HTTPException(status_code=403, detail="Las estadísticas internas están restringidas al personal autorizado")
    teams = await list_docs("teams")
    players = await list_docs("players")
    settings = await db.settings.find_one({"id": SETTINGS_ID}, {"_id": 0, "temporadas": 1}) or {}
    allowed_team_ids = {str(team.get("id")) for team in teams if team.get("id")}
    allowed_player_ids = {str(player.get("id")) for player in players if player.get("id")}
    if filters.get("equipo_id") and filters["equipo_id"] not in allowed_team_ids:
        raise HTTPException(status_code=403, detail="El equipo solicitado no pertenece a tu ámbito")
    if filters.get("player_id") and filters["player_id"] not in allowed_player_ids:
        raise HTTPException(status_code=403, detail="El jugador solicitado no pertenece a tu ámbito")
    available_categories = {str(team.get("categoria")) for team in teams if team.get("categoria")}
    available_seasons = {
        *[str(season) for season in settings.get("temporadas", []) if season],
        *[str(team.get("temporada")) for team in teams if team.get("temporada")],
    }
    if filters.get("categoria") and actor.get("role") != "admin" and filters["categoria"] not in available_categories:
        raise HTTPException(status_code=403, detail="La categoría solicitada no pertenece a tu ámbito")
    if filters.get("temporada") and actor.get("role") != "admin" and available_seasons and filters["temporada"] not in available_seasons:
        raise HTTPException(status_code=403, detail="La temporada solicitada no pertenece a tu ámbito")
    manual_stats = await list_docs("stats")
    player_names = {
        player.get("id"): f"{player.get('nombre') or ''} {player.get('apellidos') or ''}".strip()
        for player in players
    }
    for record in manual_stats:
        record["player_nombre"] = player_names.get(record.get("player_id"), "—")
    context = {
        "players": players, "teams": teams, "matches": await list_docs("matches"),
        "trainings": await list_docs("trainings"), "callups": await list_docs("callups"),
        "manual_stats": manual_stats,
    }
    options = {
        "seasons": sorted(available_seasons), "categories": sorted(available_categories),
        "teams": sorted([{"id": team.get("id"), "name": team.get("nombre"),
                           "category": team.get("categoria"), "season": team.get("temporada"),
                           "modality": team.get("modalidad")} for team in teams if team.get("id")],
                        key=lambda row: (row.get("name") or "").casefold()),
        "players": sorted([{"id": player.get("id"),
                             "name": f"{player.get('nombre') or ''} {player.get('apellidos') or ''}".strip(),
                             "team_id": player.get("equipo_id"), "category": player.get("categoria")}
                            for player in players if player.get("id")],
                           key=lambda row: row["name"].casefold()),
        "modalities": ["F7", "F11"],
    }
    return context, options


async def _build_statistics(filters: dict, *, paginated: bool = True, page: int = 1,
                            page_size: int = 25) -> tuple[dict, dict]:
    context, options = await _statistics_data(filters)
    result = calculate_statistics(**context, filters=filters)
    result["filter_options"] = options
    if paginated:
        result = paginate_player_rows(result, page, page_size)
    return result, options


def _statistics_export_rows(result: dict) -> list[dict]:
    return list(result.get("team_rows") or [])


def _statistics_export_totals(result: dict) -> dict:
    summary = result.get("summary") or {}
    attendance = result.get("attendance") or {}
    def value(key: str):
        item = summary.get(key, {})
        return item.get("value") if isinstance(item, dict) else item
    rate_item = attendance.get("porcentaje_presencia") or {}
    return {
        "players": value("active_players"), "teams": value("teams"),
        "sessions": value("training_sessions"), "present": attendance.get("presente"),
        "justified": attendance.get("justificada"), "unjustified": attendance.get("injustificada"),
        "injury": attendance.get("lesion"), "percentage": rate_item.get("value"),
        "matches_played": value("matches_played"), "wins": value("wins"),
        "draws": value("draws"), "losses": value("losses"),
    }


async def _export_statistics(format_name: str, filters: dict, lang: str):
    actor = _statistics_actor()
    result, options = await _build_statistics(filters, paginated=False)
    report = {
        "id": "statistics_integral",
        "name": {"es": "Estadísticas integrales", "eu": "Estatistika integralak"},
        "columns": ["team", "category", "modality", "players", "sessions", "present",
                     "justified", "unjustified", "injury", "percentage"],
        "generated_by": actor.get("username") or actor.get("id") or "—",
    }
    branding = await report_branding()
    renderer = generate_report_pdf if format_name == "pdf" else generate_report_xlsx
    content = renderer(report, _statistics_export_rows(result), _statistics_export_totals(result),
                       filters, options, branding, "eu" if lang == "eu" else "es")
    await db.internal_events.insert_one({
        "id": new_id(), "type": "statistics.exported", "actor_user_id": actor.get("id"),
        "actor_role": actor.get("role"), "format": format_name,
        "filter_keys": sorted(filters), "row_count": len(_statistics_export_rows(result)),
        "created_at": now_iso(), "result": "success",
    })
    media_type = "application/pdf" if format_name == "pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return StreamingResponse(io.BytesIO(content), media_type=media_type, headers={
        "Content-Disposition": f'attachment; filename="estadisticas-integrales.{format_name if format_name == "pdf" else "xlsx"}"',
        "Cache-Control": "no-store", "X-Content-Type-Options": "nosniff",
    })


@api_router.get("/statistics/options")
async def statistics_options():
    _, options = await _statistics_data({})
    return options


@api_router.get("/statistics/summary")
async def statistics_summary(temporada: Optional[str] = None, categoria: Optional[str] = None,
                             equipo_id: Optional[str] = None, player_id: Optional[str] = None,
                             modalidad: Optional[str] = None, desde: Optional[str] = None,
                             hasta: Optional[str] = None, periodo: str = "weekly",
                             estado: Optional[str] = None, activo: Optional[bool] = None,
                             estado_partido: Optional[str] = None,
                             estado_entrenamiento: Optional[str] = None,
                             estado_convocatoria: Optional[str] = None, page: int = 1,
                             page_size: int = 25):
    filters = _statistics_filters(temporada=temporada, categoria=categoria, equipo_id=equipo_id,
                                  player_id=player_id, modalidad=modalidad, desde=desde, hasta=hasta,
                                  periodo=periodo, estado=estado, activo=activo,
                                  estado_partido=estado_partido, estado_entrenamiento=estado_entrenamiento,
                                  estado_convocatoria=estado_convocatoria)
    result, _ = await _build_statistics(filters, page=page, page_size=page_size)
    return result


@api_router.get("/statistics/export.pdf")
async def statistics_export_pdf(temporada: Optional[str] = None, categoria: Optional[str] = None,
                                equipo_id: Optional[str] = None, player_id: Optional[str] = None,
                                modalidad: Optional[str] = None, desde: Optional[str] = None,
                                hasta: Optional[str] = None, periodo: str = "weekly",
                                estado: Optional[str] = None, activo: Optional[bool] = None,
                                estado_partido: Optional[str] = None,
                                estado_entrenamiento: Optional[str] = None,
                                estado_convocatoria: Optional[str] = None, lang: str = "es"):
    filters = _statistics_filters(temporada=temporada, categoria=categoria, equipo_id=equipo_id,
                                  player_id=player_id, modalidad=modalidad, desde=desde, hasta=hasta,
                                  periodo=periodo, estado=estado, activo=activo,
                                  estado_partido=estado_partido, estado_entrenamiento=estado_entrenamiento,
                                  estado_convocatoria=estado_convocatoria)
    return await _export_statistics("pdf", filters, lang)


@api_router.get("/statistics/export.xlsx")
async def statistics_export_xlsx(temporada: Optional[str] = None, categoria: Optional[str] = None,
                                 equipo_id: Optional[str] = None, player_id: Optional[str] = None,
                                 modalidad: Optional[str] = None, desde: Optional[str] = None,
                                 hasta: Optional[str] = None, periodo: str = "weekly",
                                 estado: Optional[str] = None, activo: Optional[bool] = None,
                                 estado_partido: Optional[str] = None,
                                 estado_entrenamiento: Optional[str] = None,
                                 estado_convocatoria: Optional[str] = None, lang: str = "es"):
    filters = _statistics_filters(temporada=temporada, categoria=categoria, equipo_id=equipo_id,
                                  player_id=player_id, modalidad=modalidad, desde=desde, hasta=hasta,
                                  periodo=periodo, estado=estado, activo=activo,
                                  estado_partido=estado_partido, estado_entrenamiento=estado_entrenamiento,
                                  estado_convocatoria=estado_convocatoria)
    return await _export_statistics("xlsx", filters, lang)


# ================= STATS (manual compatibility) =================
class PlayerStats(BaseModel):
    player_id: str
    temporada: Optional[str] = None
    partidos_convocado: Optional[int] = 0
    partidos_jugados: Optional[int] = 0
    minutos: Optional[int] = 0
    goles: Optional[int] = 0
    asistencias: Optional[int] = 0
    amarillas: Optional[int] = 0
    rojas: Optional[int] = 0
    porterias_cero: Optional[int] = 0
    posicion: Optional[str] = None
    valoracion: Optional[int] = None
    observaciones: Optional[str] = None


@api_router.post("/stats")
async def create_stats(stats: PlayerStats):
    return await insert_doc("stats", stats.model_dump())


@api_router.get("/stats")
async def get_stats(player_id: Optional[str] = None):
    query = {"player_id": player_id} if player_id else {}
    rows = await list_docs("stats", query)
    players = {p["id"]: f"{p.get('nombre','')} {p.get('apellidos','')}".strip() for p in await list_docs("players")}
    for r in rows:
        r["player_nombre"] = players.get(r.get("player_id"), "—")
    return rows


@api_router.put("/stats/{stats_id}")
async def edit_stats(stats_id: str, stats: PlayerStats):
    return await update_doc("stats", stats_id, stats.model_dump())


@api_router.delete("/stats/{stats_id}")
async def remove_stats(stats_id: str):
    return await delete_doc("stats", stats_id)


# ================= NOTIFICATIONS =================
async def generate_user_reminders(user: dict) -> None:
    now = datetime.now(timezone.utc)
    until = now + timedelta(hours=48)
    for collection, notification_type, link, label in (
        ("matches", "match.upcoming", "/partidos", "Partido próximo / Hurrengo partida"),
        ("trainings", "training.upcoming", "/entrenamientos", "Entrenamiento próximo / Hurrengo entrenamendua"),
    ):
        for item in await list_docs(collection):
            try:
                moment = datetime.fromisoformat(f"{item.get('fecha')}T{item.get('hora') or '00:00'}").replace(tzinfo=timezone.utc)
            except (TypeError, ValueError):
                continue
            if now <= moment <= until:
                detail = item.get("rival") or item.get("ejercicios") or item.get("campo") or "Ikas-Txiki"
                await enqueue_notifications([user], notification_type, label, str(detail), link, "high",
                                            {"id": item.get("id")}, f"{notification_type}:{item.get('id')}:{user.get('id')}")

    player_ids = set(await user_player_ids(user))
    for callup in await list_docs("callups"):
        pending = [row for row in callup.get("convocados", [])
                   if row.get("estado") in {"pending", "pendiente"}
                   and (not player_ids or row.get("player_id") in player_ids)]
        if pending:
            await enqueue_notifications([user], "callup.pending", "Convocatoria pendiente / Deialdia zain",
                                        f"{len(pending)} respuesta(s) pendiente(s)", "/convocatorias", "high",
                                        {"callup_id": callup.get("id")}, f"callup.pending:{callup.get('id')}:{user.get('id')}")

    if user.get("role") in {"family", "player"} and player_ids:
        for payment in await list_docs("payments") if user.get("role") == "family" else []:
            if payment.get("estado") in {"pendiente", "parcial", "devuelto"}:
                await enqueue_notifications([user], "payment.pending", "Pago pendiente / Ordainketa zain",
                                            str(payment.get("concepto") or "Cuota"), "/portal", "high",
                                            {"payment_id": payment.get("id")}, f"payment.pending:{payment.get('id')}:{user.get('id')}")
        for player in await list_docs("players"):
            if player.get("estado_documental") != "completo":
                await enqueue_notifications([user], "document.pending", "Documentación pendiente / Dokumentazioa zain",
                                            f"{player.get('nombre', '')} {player.get('apellidos', '')}".strip(), "/portal", "normal",
                                            {"player_id": player.get("id")}, f"document.pending:{player.get('id')}:{user.get('id')}")
        for authorization in await list_docs("authorizations"):
            if authorization.get("estado") == "pendiente":
                await enqueue_notifications([user], "authorization.pending", "Autorización pendiente / Baimena zain",
                                            str(authorization.get("tipo") or "Autorización"), "/portal", "normal",
                                            {"authorization_id": authorization.get("id")},
                                            f"authorization.pending:{authorization.get('id')}:{user.get('id')}")


@api_router.get("/notifications")
async def get_notifications(unread_only: bool = False):
    user = current_user_context.get() or {}
    await generate_user_reminders(user)
    notifications = await list_docs("notifications", {"read_at": None} if unread_only else None)
    now = now_iso()
    notifications = [item for item in notifications if not item.get("expires_at") or item["expires_at"] > now]
    return {"items": notifications, "unread": sum(1 for item in notifications if not item.get("read_at"))}


@api_router.patch("/notifications/{notification_id}/read")
async def read_notification(notification_id: str):
    query = merge_query({"id": notification_id}, await scope_for_collection("notifications"))
    result = await db.notifications.update_one(query, {"$set": {"read_at": now_iso()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notificación no encontrada")
    return {"ok": True}


@api_router.patch("/notifications/read-all")
async def read_all_notifications():
    query = merge_query({"read_at": None}, await scope_for_collection("notifications"))
    result = await db.notifications.update_many(query, {"$set": {"read_at": now_iso()}})
    return {"ok": True, "updated": result.modified_count}


@api_router.patch("/notifications/preferences")
async def update_notification_preferences(preferences: NotificationPreferences):
    user = current_user_context.get() or {}
    if not user.get("id"):
        raise HTTPException(status_code=422, detail="El usuario administrativo debe existir en la colección users")
    result = await db.users.update_one(
        {"$or": [{"id": user["id"]}, {"username": user.get("username")}]},
        {"$set": {"notification_preferences": preferences.model_dump(), "updated_at": now_iso()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=422, detail="El usuario debe existir en la colección users")
    return {"ok": True, "notification_preferences": preferences.model_dump()}


@api_router.get("/notifications/providers")
async def notification_providers():
    return provider_configuration()


# ================= COMMUNICATIONS =================
class Communication(BaseModel):
    destinatario_tipo: str = "equipo"  # equipo, categoria, individual
    destinatario_id: Optional[str] = None
    destinatario_nombre: Optional[str] = None
    canal: str = "email"  # email, telegram; whatsapp only for legacy records
    asunto: Optional[str] = None
    mensaje: Optional[str] = None
    enviado: bool = False
    fecha_envio: Optional[str] = None
    prioridad: str = "normal"

    @field_validator("canal")
    @classmethod
    def validate_channel(cls, value: str):
        if value not in {"email", "telegram", "whatsapp", "sms"}:
            raise ValueError("Canal no válido")
        return value

    @field_validator("destinatario_tipo")
    @classmethod
    def validate_recipient_type(cls, value: str):
        if value not in {"equipo", "categoria", "individual"}:
            raise ValueError("Destinatario no válido")
        return value

    @field_validator("prioridad")
    @classmethod
    def validate_priority(cls, value: str):
        if value not in {"low", "normal", "high", "urgent"}:
            raise ValueError("Prioridad no válida")
        return value


async def validate_communication_scope(data: Mapping[str, Any], user: Optional[Mapping[str, Any]] = None) -> None:
    """Validate a target using only the authenticated user's effective scope."""
    user = user or current_user_context.get() or {}
    target_type = data.get("destinatario_tipo")
    target_id = str(data.get("destinatario_id") or "").strip()
    if target_type not in {"equipo", "categoria", "individual"} or not target_id:
        raise HTTPException(status_code=422, detail="No se ha podido resolver el destinatario")
    if user.get("role") == "admin":
        return
    if user.get("role") not in {"coordinator", "coach"}:
        await record_security_event(
            "security.communication_scope.denied", user, "communication_recipients", "permission_denied",
        )
        raise HTTPException(status_code=403, detail="No tienes permiso para consultar estos destinatarios")

    team_ids = set(ids(user.get("assigned_team_ids") or []))
    allowed = False
    if target_type == "equipo":
        allowed = target_id in team_ids
    elif target_type == "categoria":
        own_categories = {
            normalized_key(value) for value in await db.teams.distinct(
                "categoria", {"id": {"$in": list(team_ids)}},
            ) if value
        } if team_ids else set()
        if normalized_key(target_id) in own_categories:
            category_team_ids = set(ids(await db.teams.distinct("id", {"categoria": target_id})))
            allowed = bool(category_team_ids) and category_team_ids.issubset(team_ids)
    elif target_type == "individual":
        player_ids = set(await user_player_ids(dict(user)))
        if target_id in player_ids:
            allowed = True
        elif player_ids:
            family_ids = set(ids(await db.players.distinct(
                "familia_id", {"id": {"$in": list(player_ids)}},
            )))
            allowed = target_id in family_ids
    if not allowed:
        await record_security_event(
            "security.communication_scope.denied", user, "communication_recipients", "outside_scope",
        )
        raise HTTPException(status_code=403, detail="No tienes permiso para consultar estos destinatarios")


def _contact_candidates(document: Mapping[str, Any], channel: str) -> list[dict]:
    fields = {
        "email": ("email", "email_formulario", "progenitor1_email", "progenitor2_email"),
        "sms": ("phone", "progenitor1_telefono", "progenitor2_telefono"),
        "telegram": ("telegram_chat_id",),
    }.get(channel, ())
    return [{**dict(document), "value": document.get(field)} for field in fields if document.get(field)]


async def communication_targets(data: dict, user: Optional[Mapping[str, Any]] = None,
                                *, validate_scope: bool = True
                                ) -> tuple[list[dict], list[str], dict[str, int]]:
    if validate_scope:
        await validate_communication_scope(data, user)
    target_type = data.get("destinatario_tipo")
    target_id = data.get("destinatario_id")
    channel = data.get("canal") or "email"
    team_ids: list[str] = []
    player_ids: list[str] = []
    family_ids: list[str] = []
    if target_type == "equipo" and target_id:
        team_ids = [target_id]
        player_ids = ids(await db.players.distinct("id", {"equipo_id": target_id}))
    elif target_type == "categoria" and target_id:
        team_ids = ids(await db.teams.distinct("id", {"categoria": target_id}))
        player_ids = ids(await db.players.distinct("id", {"equipo_id": {"$in": team_ids}}))
    elif target_type == "individual" and target_id:
        if await db.players.find_one({"id": target_id}):
            player_ids = [target_id]
        elif await db.families.find_one({"id": target_id}):
            family_ids = [target_id]
    if player_ids:
        family_ids = sorted(set(family_ids + ids(await db.players.distinct("familia_id", {"id": {"$in": player_ids}}))))
    users_all = await notification_users(team_ids, player_ids, family_ids)
    users = [candidate for candidate in users_all if usable_account(candidate)[0]]
    contacts: list[dict] = []
    for user_row in users:
        user_contacts = _contact_candidates(user_row, channel)
        if (user_row.get("notification_preferences") or {}).get(channel, True) is False:
            for candidate in user_contacts:
                candidate["communication_consents"] = {
                    **(candidate.get("communication_consents") or {}), channel: "no",
                }
        contacts.extend(user_contacts)
    eligibility_exclusions: dict[str, int] = defaultdict(int)
    player_projection = {
        "_id": 0, "active": 1, "estado": 1, "account_status": 1,
        "email_formulario": 1, "progenitor1_email": 1, "progenitor2_email": 1, "telegram_chat_id": 1,
        "progenitor1_telefono": 1, "progenitor2_telefono": 1,
        "communication_consents": 1, "consents": 1, "historical.consents": 1,
    }
    async for player in db.players.find({"id": {"$in": player_ids}}, player_projection):
        player_contacts = _contact_candidates(player, channel)
        if communication_record_active(player):
            contacts.extend(player_contacts)
        elif player_contacts:
            eligibility_exclusions["recipient_inactive"] += 1
    family_projection = {
        "_id": 0, "active": 1, "estado": 1, "account_status": 1,
        "progenitor1_email": 1, "progenitor2_email": 1, "telegram_chat_id": 1,
        "progenitor1_telefono": 1, "progenitor2_telefono": 1,
        "communication_consents": 1, "consents": 1, "historical.consents": 1,
    }
    async for family in db.families.find({"id": {"$in": family_ids}}, family_projection):
        family_contacts = _contact_candidates(family, channel)
        if communication_record_active(family):
            contacts.extend(family_contacts)
        elif family_contacts:
            eligibility_exclusions["recipient_inactive"] += 1
    destinations, consent_exclusions = consented_contacts(contacts, channel)
    for reason, count in eligibility_exclusions.items():
        consent_exclusions[reason] = consent_exclusions.get(reason, 0) + count
    return users, destinations, consent_exclusions


async def communication_target_context(data: dict, user: Optional[Mapping[str, Any]] = None,
                                       *, audit_exclusions: bool = False,
                                       validate_scope: bool = True) -> dict:
    user = user or current_user_context.get() or {}
    if validate_scope:
        await validate_communication_scope(data, user)
    target_type = data.get("destinatario_tipo")
    target_id = data.get("destinatario_id")
    team_ids: list[str] = []
    player_ids: list[str] = []
    family_ids: list[str] = []
    resolved_name = data.get("destinatario_nombre") or ""
    if target_type == "equipo" and target_id:
        team = await db.teams.find_one({"id": target_id}, {"_id": 0, "id": 1, "nombre": 1})
        if team:
            team_ids, resolved_name = [target_id], team.get("nombre") or resolved_name
            player_ids = ids(await db.players.distinct("id", {"equipo_id": target_id}))
    elif target_type == "categoria" and target_id:
        team_ids = ids(await db.teams.distinct("id", {"categoria": target_id}))
        player_ids = ids(await db.players.distinct("id", {"equipo_id": {"$in": team_ids}}))
        resolved_name = resolved_name or str(target_id)
    elif target_type == "individual" and target_id:
        player = await db.players.find_one({"id": target_id}, {"_id": 0, "id": 1, "nombre": 1, "apellidos": 1})
        family = None if player else await db.families.find_one(
            {"id": target_id}, {"_id": 0, "id": 1, "contacto_principal": 1, "progenitor1_nombre": 1},
        )
        if player:
            player_ids = [target_id]
            resolved_name = " ".join(value for value in (player.get("nombre"), player.get("apellidos")) if value)
        elif family:
            family_ids = [target_id]
            resolved_name = family.get("progenitor1_nombre") or family.get("contacto_principal") or resolved_name
    if player_ids:
        family_ids = sorted(set(family_ids + ids(await db.players.distinct("familia_id", {"id": {"$in": player_ids}}))))
    users_all = await notification_users(team_ids, player_ids, family_ids)
    _, destinations, consent_exclusions = await communication_targets(
        data, user, validate_scope=validate_scope,
    )
    if audit_exclusions and consent_exclusions:
        await record_security_event(
            "security.communication_consent.filtered", user, "communication_recipients",
            "consent_not_granted", aggregate=consent_exclusions,
        )
    return {
        "resolved_name": resolved_name,
        "summary": recipient_summary(
            users_all, len(destinations), team_count=len(team_ids), player_count=len(player_ids),
            family_count=len(family_ids), extra_exclusions=consent_exclusions,
        ),
    }


@api_router.post("/communications/recipients/preview")
async def preview_communication_recipients(comm: Communication):
    data = comm.model_dump()
    user = current_user_context.get() or {}
    context = await communication_target_context(data, user, audit_exclusions=True)
    if not context["resolved_name"]:
        raise HTTPException(status_code=422, detail="No se ha podido resolver el destinatario")
    return context


@api_router.post("/communications")
async def create_communication(comm: Communication):
    data = comm.model_dump()
    data.update({"enviado": False, "fecha_envio": None, "estado_envio": "pending", "error_envio": None})
    user = current_user_context.get() or {}
    await communication_target_context(data, user)
    created = await insert_doc("communications", data)
    users, destinations, consent_exclusions = await communication_targets(data, user)
    await enqueue_notifications(
        users, "communication.created", "Nueva comunicación / Komunikazio berria",
        str(data.get("asunto") or "Ikas-Txiki"), "/comunicacion", data.get("prioridad") or "normal",
        {"communication_id": created["id"]}, f"communication.created:{created['id']}",
    )
    logs = []
    if data.get("canal") == "email":
        # Re-resolve immediately before the provider boundary. A preview is
        # never treated as authorization for a later delivery.
        _, destinations, consent_exclusions = await communication_targets(data, user)
        if consent_exclusions:
            await record_security_event(
                "security.communication_consent.filtered", user, "communication_delivery",
                "consent_not_granted", aggregate=consent_exclusions,
            )
        if destinations:
            logs = [dispatch_email(destination, data.get("asunto") or "Ikas-Txiki", data.get("mensaje") or "")
                    for destination in destinations]
        else:
            logs = [{"id": new_id(), "channel": "email", "recipient": None, "provider": "smtp",
                     "status": "pending", "error": "recipient_missing", "created_at": now_iso(), "sent_at": None}]
    elif data.get("canal") == "telegram":
        _, destinations, consent_exclusions = await communication_targets(data, user)
        if consent_exclusions:
            await record_security_event(
                "security.communication_consent.filtered", user, "communication_delivery",
                "consent_not_granted", aggregate=consent_exclusions,
            )
        if destinations:
            text = "\n\n".join(part for part in (data.get("asunto"), data.get("mensaje")) if part)
            logs = [dispatch_telegram(destination, text) for destination in destinations]
        else:
            logs = [{"id": new_id(), "channel": "telegram", "recipient": None, "provider": "telegram_bot",
                     "status": "pending", "error": "telegram_not_linked", "created_at": now_iso(), "sent_at": None}]
    else:
        provider = provider_configuration().get(data.get("canal"), {"configured": False, "provider": "optional"})
        logs = [{"id": new_id(), "channel": data.get("canal"), "recipient": None,
                 "provider": provider["provider"], "status": "pending",
                 "error": "provider_optional_not_activated" if provider["configured"] else "provider_not_configured",
                 "created_at": now_iso(), "sent_at": None}]
    for log in logs:
        log["communication_id"] = created["id"]
    if logs:
        await db.delivery_logs.insert_many(logs)
    state = "sent" if logs and all(log["status"] == "sent" for log in logs) else (
        "failed" if any(log["status"] == "failed" for log in logs) else "pending"
    )
    await db.communications.update_one({"id": created["id"]}, {"$set": {
        "enviado": state == "sent", "estado_envio": state,
        "fecha_envio": now_iso() if state == "sent" else None,
        "error_envio": next((log["error"] for log in logs if log.get("error")), None),
        "delivery_log_ids": [log["id"] for log in logs],
    }})
    return await get_doc("communications", created["id"])


@api_router.get("/communications")
async def get_communications():
    items = await list_docs("communications")
    for item in items:
        context = await communication_target_context(item, validate_scope=False)
        item["destinatario_nombre_resuelto"] = context["resolved_name"] or item.get("destinatario_nombre")
    return items


@api_router.put("/communications/{comm_id}")
async def edit_communication(comm_id: str, comm: Communication):
    data = comm.model_dump()
    data.pop("enviado", None)
    data.pop("fecha_envio", None)
    return await update_doc("communications", comm_id, data)


@api_router.delete("/communications/{comm_id}")
async def remove_communication(comm_id: str):
    return await delete_doc("communications", comm_id)


# ================= SETTINGS / CONFIG =================
class Settings(BaseModel):
    club_nombre: Optional[str] = "Ikas-Txiki"
    club_logo: Optional[str] = None
    club_direccion: Optional[str] = None
    club_email: Optional[str] = None
    club_telefono: Optional[str] = None
    temporada_actual: Optional[str] = None
    temporadas: List[str] = []
    campos: List[str] = []
    entrenadores: List[str] = []
    cuota_base: Optional[float] = 0
    descuento_hermano: Optional[float] = 0
    attendance_alert_threshold: int = Field(default=3, ge=1, le=20)
    modalities: List[ModalityDefinition] = Field(default_factory=lambda: catalog_from_settings({}))

    @field_validator("modalities")
    @classmethod
    def validate_modalities(cls, values: List[ModalityDefinition]) -> List[ModalityDefinition]:
        return validate_compatibility_catalog(values)


SETTINGS_ID = "global"


@api_router.get("/settings")
async def get_settings():
    doc = await db.settings.find_one({"id": SETTINGS_ID}, {"_id": 0})
    if not doc:
        default = Settings().model_dump(mode="json")
        default_modalities = default.pop("modalities")
        default["id"] = SETTINGS_ID
        default["categories"] = CATEGORIES
        await db.settings.insert_one(dict(default))
        default["modalities"] = default_modalities
        return clean(default)
    doc["categories"] = CATEGORIES
    doc["modalities"] = [entry.model_dump(mode="json") for entry in catalog_from_settings(doc)]
    return doc


@api_router.put("/settings")
async def update_settings(settings: Settings):
    data = settings.model_dump(mode="json")
    data["id"] = SETTINGS_ID
    await db.settings.update_one({"id": SETTINGS_ID}, {"$set": data}, upsert=True)
    return await get_settings()


@api_router.get("/catalog-options")
async def get_catalog_options():
    """Catálogos operativos reutilizables sin exponer toda la configuración admin."""
    doc = await db.settings.find_one(
        {"id": SETTINGS_ID},
        {"_id": 0, "temporadas": 1, "campos": 1, "entrenadores": 1},
    ) or {}
    return {
        "temporadas": sorted({str(value) for value in doc.get("temporadas", []) if value}),
        "campos": sorted({str(value) for value in doc.get("campos", []) if value}),
        "entrenadores": sorted({str(value) for value in doc.get("entrenadores", []) if value}),
    }


async def _load_modality_catalog() -> list[ModalityDefinition]:
    settings = await db.settings.find_one({"id": SETTINGS_ID}, {"_id": 0, "modalities": 1}) or {}
    return catalog_from_settings(settings)


async def _save_modality_catalog(catalog: List[ModalityDefinition]) -> None:
    payload = [entry.model_dump(mode="json") for entry in validate_compatibility_catalog(catalog)]
    await db.settings.update_one(
        {"id": SETTINGS_ID},
        {"$set": {"modalities": payload, "updated_at": now_iso()}},
        upsert=True,
    )


async def _modality_usage(code: str) -> Dict[str, int]:
    query = {"modalidad": code}
    usage = {
        "teams": await db.teams.count_documents(query),
        "inscriptions": await db.inscriptions.count_documents(query),
        "players": await db.players.count_documents(query),
        "staging_records": await db.import_staging.count_documents({"records.modalidad": code, "status": "draft"}),
    }
    return {key: value for key, value in usage.items() if value}


def _modality_actor() -> Dict[str, Any]:
    return current_user_context.get() or {}


def _stamp_modality(entry: ModalityDefinition, actor: Dict[str, Any]) -> ModalityDefinition:
    return entry.model_copy(update={"updated_at": datetime.now(timezone.utc), "updated_by": actor.get("id")})


@api_router.get("/modalities")
async def get_modalities(include_inactive: bool = False):
    actor = _modality_actor()
    catalog = await _load_modality_catalog()
    if include_inactive and actor.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo el administrador puede consultar modalidades inactivas")
    visible = catalog if include_inactive else [entry for entry in catalog if entry.active]
    return [entry.model_dump(mode="json") for entry in visible]


@api_router.post("/modalities", status_code=201)
async def create_modality(request: ModalityCreateRequest):
    actor = _modality_actor()
    catalog = await _load_modality_catalog()
    if any(entry.code == request.code for entry in catalog):
        raise HTTPException(status_code=409, detail="El código de modalidad ya existe")
    entry = _stamp_modality(ModalityDefinition(**request.model_dump()), actor)
    try:
        updated = validate_compatibility_catalog([*catalog, entry])
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    await _save_modality_catalog(updated)
    return entry.model_dump(mode="json")


@api_router.put("/modalities/{code}")
async def update_modality(code: str, request: ModalityUpdateRequest):
    actor = _modality_actor()
    normalized_code = code.strip().upper()
    catalog = await _load_modality_catalog()
    current = next((entry for entry in catalog if entry.code == normalized_code), None)
    if not current:
        raise HTTPException(status_code=404, detail="Modalidad no encontrada")
    changes = request.model_dump(exclude_none=True)
    if changes.get("active") is False and current.active:
        usage = await _modality_usage(normalized_code)
        if usage:
            raise HTTPException(status_code=409, detail={"message": "La modalidad está en uso y no puede desactivarse", "usage": usage})
    try:
        candidate_data = current.model_dump()
        candidate_data.update(changes)
        candidate = _stamp_modality(ModalityDefinition.model_validate(candidate_data), actor)
        updated = validate_compatibility_catalog(
            [candidate if entry.code == normalized_code else entry for entry in catalog]
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    await _save_modality_catalog(updated)
    return candidate.model_dump(mode="json")


@api_router.patch("/modalities/{code}/status")
async def set_modality_status(code: str, request: ModalityStatusRequest):
    return await update_modality(code, ModalityUpdateRequest(active=request.active))


@api_router.post("/modalities/reorder")
async def reorder_modalities(request: ModalityReorderRequest):
    actor = _modality_actor()
    catalog = await _load_modality_catalog()
    if set(request.codes) != {entry.code for entry in catalog}:
        raise HTTPException(status_code=422, detail="El orden debe incluir exactamente todo el catálogo")
    positions = {code: (index + 1) * 10 for index, code in enumerate(request.codes)}
    updated = [_stamp_modality(entry.model_copy(update={"sort_order": positions[entry.code]}), actor) for entry in catalog]
    await _save_modality_catalog(updated)
    return [entry.model_dump(mode="json") for entry in validate_compatibility_catalog(updated)]


@api_router.get("/categories")
async def get_categories():
    return CATEGORIES


# ================= PROFESSIONAL REPORTS =================
class ReportPreviewRequest(BaseModel):
    report_id: str
    filters: Dict[str, Any] = Field(default_factory=dict)
    page: int = Field(default=1, ge=1)
    page_size: int = Field(default=25, ge=1, le=100)


class ReportExportRequest(BaseModel):
    report_id: str
    filters: Dict[str, Any] = Field(default_factory=dict)
    lang: str = "es"

    @field_validator("lang")
    @classmethod
    def validate_lang(cls, value: str) -> str:
        if value not in {"es", "eu"}:
            raise ValueError("Idioma no válido")
        return value


async def report_context() -> dict:
    """Carga exclusivamente documentos ya acotados por el usuario autenticado."""
    settings = await db.settings.find_one({"id": SETTINGS_ID}, {"_id": 0}) or {}
    async def projected(coll: str, fields: set[str], maximum: int = 5000) -> list[dict]:
        projection = {field: 1 for field in fields}
        projection["_id"] = 0
        scope = await scope_for_collection(coll)
        return await db[coll].find(scope or {}, projection).to_list(maximum)

    return {
        "players": await projected("players", {
            "id", "nombre", "apellidos", "fecha_nacimiento", "fecha_inscripcion", "fecha_alta",
            "fecha_baja", "equipo_id", "familia_id", "categoria", "modalidad", "dorsal",
            "posicion", "numero_licencia", "estado", "estado_documental",
            "doc_dni_jugador", "doc_dni_tutor", "doc_foto", "doc_autorizacion",
            "doc_justificante_pago", "doc_ficha_federativa", "talla_camiseta",
            "talla_pantalon", "talla_chandal", "equipacion_entregada",
            "fecha_entrega_equipacion", "equipamiento_items", "equipaciones",
            "progenitor1_nombre", "progenitor1_telefono", "progenitor1_email",
            "progenitor2_nombre", "progenitor2_telefono", "progenitor2_email",
        }, MAX_EXPORT_ROWS + 1),
        "families": await projected("families", {
            "id", "progenitor1_nombre", "progenitor1_telefono", "progenitor1_email",
            "progenitor2_nombre", "progenitor2_telefono", "progenitor2_email",
        }),
        "teams": await projected("teams", {
            "id", "nombre", "categoria", "modalidad", "temporada", "estado", "active",
            "limite_jugadores",
        }),
        "trainings": await projected("trainings", {
            "id", "equipo_id", "fecha", "hora", "campo", "callup_id",
            "asistencia.player_id", "asistencia.estado",
        }, MAX_EXPORT_ROWS + 1),
        "matches": await projected("matches", {
            "id", "temporada", "fecha", "hora", "equipo_id", "rival", "condicion",
            "tipo", "estado", "resultado_propio", "resultado_rival",
        }),
        "callups": await projected("callups", {
            "id", "match_id", "equipo_id", "convocados.player_id", "convocados.estado",
            "convocados.responded_at", "convocados.late",
        }),
        "inscriptions": await projected("inscriptions", {
            "id", "player_id", "created_at", "fecha_inscripcion", "nombre", "apellidos",
            "tipo", "temporada", "equipo_id", "categoria", "modalidad", "estado",
        }),
        "authorizations": await projected("authorizations", {
            "id", "player_id", "tipo", "estado", "fecha_firma", "fecha_caducidad",
        }),
        "payments": await projected("payments", {
            "id", "player_id", "created_at", "concepto", "importe_final", "forma_pago",
            "estado", "fecha_pago",
        }),
        "stats": await projected("stats", {
            "id", "player_id", "temporada", "partidos_convocado", "partidos_jugados",
            "minutos", "goles", "asistencias", "amarillas", "rojas", "valoracion",
        }),
        "settings": settings,
        "modalities": catalog_from_settings(settings),
    }


def report_filter_options(context: dict) -> dict:
    players, teams = context["players"], context["teams"]
    states = {
        value for collection in ("players", "matches", "inscriptions", "authorizations", "payments")
        for item in context.get(collection, []) for value in [item.get("estado")] if value
    }
    states.update({
        "activo", "baja", "lesionado", "pendiente_documentacion", "en_prueba",
        "programado", "jugado", "aplazado", "suspendido", "cancelado",
        "recibida", "revisada", "aceptada", "pendiente", "rechazada",
        "firmada", "caducada", "pagado", "parcial", "devuelto",
        "pending", "confirmed", "declined",
    })
    states.update(normalize_status(item.get("estado")) for callup in context.get("callups", [])
                  for item in callup.get("convocados", []) if item.get("estado"))
    return {
        "seasons": sorted({
            *[season for season in (context.get("settings") or {}).get("temporadas", []) if season],
            *[team.get("temporada") for team in teams if team.get("temporada")],
        }),
        "categories": sorted({value for value in [
            *(team.get("categoria") for team in teams), *(player.get("categoria") for player in players),
        ] if value}),
        "teams": sorted([{"id": team.get("id"), "name": team.get("nombre"), "category": team.get("categoria"),
                          "season": team.get("temporada"), "modality": normalize_modality(team.get("modalidad"), context["modalities"]).code}
                         for team in teams if team.get("id")], key=lambda item: (item.get("name") or "").casefold()),
        "modalities": [{"code": item.code, "name_es": item.name_es, "name_eu": item.name_eu}
                       for item in context["modalities"] if item.active],
        "players": sorted([{"id": player.get("id"), "name": f"{player.get('nombre') or ''} {player.get('apellidos') or ''}".strip(),
                            "team_id": player.get("equipo_id"), "category": player.get("categoria")}
                           for player in players if player.get("id")], key=lambda item: item["name"].casefold()),
        "states": sorted(states),
        "types": ["alta", "renovacion", "general", "imagen", "medica", "desplazamientos",
                  "recogida", "proteccion_datos"],
        "movements": ["alta", "baja"],
        "deliveries": ["delivered", "pending"],
        "contact_types": ["all", "phone", "email"],
        "payment_methods": ["domiciliacion", "transferencia", "efectivo", "bizum"],
    }


def enforce_report_filters(filters: dict, options: dict) -> None:
    protected = {
        "team_id": {item["id"] for item in options["teams"]},
        "player_id": {item["id"] for item in options["players"]},
        "category": set(options["categories"]), "season": set(options["seasons"]),
        "modality": {item["code"] for item in options["modalities"]}, "status": set(options["states"]),
    }
    for key, allowed in protected.items():
        if filters.get(key) and filters[key] not in allowed:
            raise HTTPException(status_code=403, detail="El filtro solicitado no pertenece a tu ámbito")
    for key in ("date_from", "date_to"):
        if filters.get(key):
            try:
                date.fromisoformat(str(filters[key]))
            except ValueError as exc:
                raise HTTPException(status_code=422, detail="Fecha de informe no válida") from exc


@api_router.get("/reports/catalog")
async def get_reports_catalog():
    actor = current_user_context.get() or {}
    context = await report_context()
    return {"reports": catalog_for_role(str(actor.get("role"))), "filter_options": report_filter_options(context)}


async def prepare_report(report_id: str, requested_filters: dict) -> dict:
    actor = current_user_context.get() or {}
    allowed_reports = {item["id"] for item in catalog_for_role(str(actor.get("role")))}
    if report_id not in allowed_reports:
        raise HTTPException(status_code=403, detail="Informe no autorizado")
    try:
        filters = validate_report_filters(report_id, requested_filters)
    except ReportValidationError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    context = await report_context()
    options = report_filter_options(context)
    enforce_report_filters(filters, options)
    try:
        definition, rows, totals = build_report(report_id, context, filters, str(actor.get("role")))
    except ReportValidationError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    return {"report": definition, "filters": filters, "rows": rows, "totals": totals,
            "filter_options": options}


@api_router.post("/reports/preview")
async def preview_report(request: ReportPreviewRequest):
    result = await prepare_report(request.report_id, request.filters)
    rows = result["rows"]
    page_rows, pagination = paginate(rows, request.page, request.page_size)
    return {**result, "rows": page_rows, "pagination": pagination}


async def report_branding() -> dict:
    return await db.settings.find_one({"id": SETTINGS_ID}, {"_id": 0, "club_nombre": 1, "club_logo": 1}) or {}


async def export_professional_report(request: ReportExportRequest, export_format: str):
    result = await prepare_report(request.report_id, request.filters)
    try:
        rows = enforce_export_limit(result["rows"])
    except ReportValidationError as exc:
        raise HTTPException(status_code=413, detail=str(exc)) from exc
    branding = await report_branding()
    renderer = generate_report_pdf if export_format == "pdf" else generate_report_xlsx
    content = renderer(result["report"], rows, result["totals"], result["filters"],
                       result["filter_options"], branding, request.lang)
    filename = safe_filename(request.report_id, request.lang, export_format)
    media_type = "application/pdf" if export_format == "pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return StreamingResponse(io.BytesIO(content), media_type=media_type, headers={
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store", "X-Content-Type-Options": "nosniff",
    })


@api_router.post("/reports/export.pdf")
async def export_professional_pdf(request: ReportExportRequest):
    return await export_professional_report(request, "pdf")


@api_router.post("/reports/export.xlsx")
async def export_professional_xlsx(request: ReportExportRequest):
    return await export_professional_report(request, "xlsx")


@api_router.get("/compute-category")
async def api_compute_category(fecha_nacimiento: str):
    return {"categoria": compute_category(fecha_nacimiento)}


# ================= DASHBOARD =================
@api_router.get("/dashboard")
async def dashboard(
    temporada: Optional[str] = None,
    categoria: Optional[str] = None,
    equipo_id: Optional[str] = None,
):
    user = current_user_context.get() or {}
    role = user.get("role", "player")

    async def permitted_docs(collection: str, resource: str):
        return await list_docs(collection) if has_permission(user, resource, "read") else []

    all_teams = await permitted_docs("teams", "teams")
    allowed_team_ids = {team.get("id") for team in all_teams if team.get("id")}
    if equipo_id and equipo_id not in allowed_team_ids:
        raise HTTPException(status_code=403, detail="El equipo solicitado no pertenece a tu ámbito")
    if categoria and role != "admin" and not any(team.get("categoria") == categoria for team in all_teams):
        raise HTTPException(status_code=403, detail="La categoría solicitada no pertenece a tu ámbito")
    allowed_seasons = {team.get("temporada") for team in all_teams if team.get("temporada")}
    if temporada and role != "admin" and allowed_seasons and temporada not in allowed_seasons:
        raise HTTPException(status_code=403, detail="La temporada solicitada no pertenece a tu ámbito")

    teams = [
        team for team in all_teams
        if (not equipo_id or team.get("id") == equipo_id)
        and (not categoria or team.get("categoria") == categoria)
        and (not temporada or team.get("temporada") == temporada)
    ]
    selected_team_ids = {team.get("id") for team in teams if team.get("id")}
    team_names = {team["id"]: team.get("nombre", "—") for team in teams}

    players = [
        player for player in await permitted_docs("players", "players")
        if (not selected_team_ids or player.get("equipo_id") in selected_team_ids)
        and (not categoria or player.get("categoria") == categoria)
    ] if (teams or not any((equipo_id, categoria, temporada))) else []
    selected_player_ids = {player.get("id") for player in players if player.get("id")}

    matches = [
        match for match in await permitted_docs("matches", "matches")
        if (not selected_team_ids or match.get("equipo_id") in selected_team_ids)
        and (not temporada or match.get("temporada") == temporada)
    ] if (teams or not any((equipo_id, categoria, temporada))) else []
    trainings = [
        training for training in await permitted_docs("trainings", "trainings")
        if not selected_team_ids or training.get("equipo_id") in selected_team_ids
    ] if (teams or not any((equipo_id, categoria, temporada))) else []
    club_events = [
        event for event in await permitted_docs("club_events", "calendar")
        if (not selected_team_ids or not event.get("equipo_id") or event.get("equipo_id") in selected_team_ids)
        and (not temporada or not event.get("temporada") or event.get("temporada") == temporada)
        and (not categoria or not event.get("categoria") or event.get("categoria") == categoria)
    ] if (teams or not any((equipo_id, categoria, temporada))) else []
    callups = [
        callup for callup in await permitted_docs("callups", "callups")
        if not selected_team_ids or callup.get("equipo_id") in selected_team_ids
    ] if (teams or not any((equipo_id, categoria, temporada))) else []

    payments = [
        payment for payment in await permitted_docs("payments", "payments")
        if not selected_player_ids or payment.get("player_id") in selected_player_ids
    ] if players else []
    auths = [
        authorization for authorization in await permitted_docs("authorizations", "authorizations")
        if not selected_player_ids or authorization.get("player_id") in selected_player_ids
    ] if players else []
    all_inscriptions = await permitted_docs("inscriptions", "inscriptions")
    inscriptions = all_inscriptions if role == "admin" and not any((equipo_id, categoria, temporada)) else [
        inscription for inscription in all_inscriptions
        if inscription.get("player_id") in selected_player_ids
    ]
    communications = await permitted_docs("communications", "communications")
    communication_targets = selected_player_ids | set(ids([user.get("family_id"), user.get("player_id")]))
    if role != "admin" or any((equipo_id, categoria, temporada)):
        communications = [
            communication for communication in communications
            if (
                communication.get("destinatario_tipo") == "equipo"
                and communication.get("destinatario_id") in selected_team_ids
            ) or (
                communication.get("destinatario_tipo") == "individual"
                and communication.get("destinatario_id") in communication_targets
            ) or (
                role == "admin" and communication.get("destinatario_tipo") == "categoria"
                and (not categoria or communication.get("destinatario_id") == categoria)
            )
        ]

    activos = [p for p in players if p.get("estado") == "activo"]
    pendientes_doc = [p for p in players if p.get("estado_documental") != "completo"]
    nuevas_inscripciones = [p for p in players if p.get("nueva_incorporacion")]
    inscripciones_pendientes = [p for p in players if p.get("estado") in ("pendiente_documentacion", "en_prueba")]
    pagos_pendientes = [p for p in payments if p.get("estado") in ("pendiente", "parcial")]
    auth_pendientes = [a for a in auths if a.get("estado") == "pendiente"]

    today = date.today().isoformat()
    proximos_partidos = sorted(
        [m for m in matches if m.get("fecha") and m.get("fecha") >= today and m.get("estado") == "programado"],
        key=lambda m: (m.get("fecha"), m.get("hora") or "")
    )[:5]
    for m in proximos_partidos:
        m["equipo_nombre"] = team_names.get(m.get("equipo_id"), "—")

    proximos_entrenamientos = sorted(
        [tr for tr in trainings if tr.get("fecha") and tr.get("fecha") >= today],
        key=lambda tr: (tr.get("fecha"), tr.get("hora") or "")
    )[:5]
    for tr in proximos_entrenamientos:
        tr["equipo_nombre"] = team_names.get(tr.get("equipo_id"), "—")

    insc_pendientes = [i for i in inscriptions if i.get("estado") in ("recibida", "pendiente", "revisada")]
    nuevas_insc = [i for i in inscriptions if i.get("tipo") == "alta" and not i.get("player_id")]

    pending_communications = [communication for communication in communications if not communication.get("enviado")]
    failed_communications = [
        communication for communication in communications
        if communication.get("estado") == "error" or communication.get("error_envio")
    ]
    latest_communications = sorted(
        communications,
        key=lambda communication: communication.get("fecha_envio") or communication.get("created_at") or "",
        reverse=True,
    )[:5]
    attendance = weekly_attendance(trainings, selected_player_ids or None)
    attendance_detail = attendance_summary(trainings, selected_player_ids or None)
    attendance_settings = await db.settings.find_one({"id": SETTINGS_ID}, {"_id": 0, "attendance_alert_threshold": 1, "temporadas": 1}) or {}
    attendance_alerts = repeated_absence_alerts(
        trainings, attendance_settings.get("attendance_alert_threshold", 3), selected_player_ids or None,
    )
    scoped_callup_players = selected_player_ids if role in {"family", "player"} else None
    callup_pending = pending_callups(callups, scoped_callup_players)
    callup_status = player_callup_status(callups, selected_player_ids) if role in {"family", "player"} else []
    calendar_events = aggregate_calendar_events(matches, trainings, club_events, teams)
    next_event = next_calendar_event(calendar_events)
    if next_event:
        next_event = {**next_event, "fecha_hora": f"{next_event['fecha']}T{next_event.get('hora') or '00:00'}:00+00:00"}
    incidents = [
        {"tipo": kind, "id": item.get("id"), "equipo_id": item.get("equipo_id"), "mensaje": item.get("observaciones")}
        for kind, rows in (("partido", matches), ("entrenamiento", trainings))
        for item in rows if item.get("observaciones")
    ][:5]

    alertas = []
    if failed_communications:
        alertas.append({"tipo": "error_comunicacion", "mensaje": f"{len(failed_communications)} comunicaciones fallidas"})
    if callup_pending["total"]:
        alertas.append({"tipo": "convocatoria", "mensaje": f"{callup_pending['total']} convocatorias pendientes de respuesta"})
    if pagos_pendientes:
        alertas.append({"tipo": "pago", "mensaje": f"{len(pagos_pendientes)} pagos pendientes"})
    if pendientes_doc:
        alertas.append({"tipo": "doc", "mensaje": f"{len(pendientes_doc)} jugadores con documentación incompleta"})
    if auth_pendientes:
        alertas.append({"tipo": "auth", "mensaje": f"{len(auth_pendientes)} autorizaciones pendientes de firma"})
    if insc_pendientes:
        alertas.append({"tipo": "inscripcion", "mensaje": f"{len(insc_pendientes)} inscripciones por revisar"})

    if pending_communications:
        alertas.append({"tipo": "comunicacion", "mensaje": f"{len(pending_communications)} comunicaciones pendientes"})
    if attendance_alerts:
        alertas.append({"tipo": "asistencia", "mensaje": f"{len(attendance_alerts)} alertas de asistencia"})

    available = [player for player in players if player.get("estado") == "activo"]
    absent = [player for player in players if player.get("estado") in {"lesionado", "baja"}]
    children = [
        {
            "id": player.get("id"), "nombre": player.get("nombre"),
            "apellidos": player.get("apellidos"), "equipo_id": player.get("equipo_id"),
            "categoria": player.get("categoria"), "estado": player.get("estado"),
        }
        for player in players
    ] if role == "family" else []

    return {
        "role": role,
        "scope": {"team_ids": sorted(selected_team_ids), "player_ids": sorted(selected_player_ids)},
        "filters": {"temporada": temporada, "categoria": categoria, "equipo_id": equipo_id},
        "filter_options": {
            "temporadas": sorted({
                *[season for season in attendance_settings.get("temporadas", []) if season],
                *[team.get("temporada") for team in all_teams if team.get("temporada")],
            }),
            "categorias": sorted({team.get("categoria") for team in all_teams if team.get("categoria")}),
            "equipos": [
                {"id": team.get("id"), "nombre": team.get("nombre"), "categoria": team.get("categoria"), "temporada": team.get("temporada")}
                for team in all_teams
            ],
        },
        "jugadores_activos": len(activos),
        "total_jugadores": len(players),
        "nuevas_inscripciones": len(nuevas_insc),
        "inscripciones_pendientes": len(insc_pendientes),
        "documentacion_pendiente": len(pendientes_doc),
        "pagos_pendientes": len(pagos_pendientes),
        "importe_pendiente": round(sum((p.get("importe_final") or 0) for p in pagos_pendientes), 2),
        "autorizaciones_pendientes": len(auth_pendientes),
        "proximos_partidos": proximos_partidos,
        "proximos_entrenamientos": proximos_entrenamientos,
        "siguiente_actividad": next_event,
        "ultimas_comunicaciones": latest_communications,
        "comunicaciones_pendientes": len(pending_communications),
        "comunicaciones_fallidas": len(failed_communications),
        "asistencia_semanal": attendance,
        "asistencia_resumen": attendance_detail,
        "alertas_asistencia": attendance_alerts,
        "convocatorias_pendientes": callup_pending,
        "estado_convocatorias": callup_status,
        "jugadores_disponibles": len(available),
        "jugadores_ausentes": len(absent),
        "hijos": children,
        "incidencias": incidents,
        "alertas": prioritized_alerts(alertas),
    }


@api_router.get("/search")
async def global_search(q: str):
    import unicodedata

    def _norm(s):
        s = str(s or "").lower()
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

    ql = _norm(q).strip()
    if not ql or len(ql) < 1:
        return []
    results = []

    teams = {t["id"]: t.get("nombre", "—") for t in await list_docs("teams")}

    players = await list_docs("players")
    for p in players:
        hay = _norm(f"{p.get('nombre','')} {p.get('apellidos','')} {p.get('dorsal','')} {p.get('posicion','')} {p.get('categoria','')} {p.get('numero_licencia','')} {p.get('progenitor1_telefono','')} {p.get('progenitor1_nombre','')} {p.get('email_formulario','')}")
        if ql in hay:
            results.append({"type": "player", "id": p["id"],
                            "title": f"{p.get('nombre','')} {p.get('apellidos','')}".strip(),
                            "subtitle": f"{p.get('categoria') or '—'} · {teams.get(p.get('equipo_id'), 'Sin equipo')}",
                            "route": "/jugadores"})

    for t in await list_docs("teams"):
        hay = _norm(f"{t.get('nombre','')} {t.get('categoria','')} {t.get('entrenador','')} {t.get('campo','')}")
        if ql in hay:
            results.append({"type": "team", "id": t["id"], "title": t.get("nombre", "—"),
                            "subtitle": f"{t.get('categoria') or '—'} · {t.get('entrenador') or ''}", "route": "/equipos"})

    for m in await list_docs("matches"):
        hay = _norm(f"{m.get('rival','')} {teams.get(m.get('equipo_id'),'')} {m.get('tipo','')} {m.get('jornada','')} {m.get('fecha','')}")
        if ql in hay:
            results.append({"type": "match", "id": m["id"],
                            "title": f"{teams.get(m.get('equipo_id'),'—')} vs {m.get('rival') or '—'}",
                            "subtitle": f"{m.get('fecha') or ''} · {m.get('hora') or ''}", "route": "/partidos"})

    for f in await list_docs("families"):
        hay = _norm(f"{f.get('progenitor1_nombre','')} {f.get('progenitor2_nombre','')} {f.get('progenitor1_telefono','')} {f.get('progenitor1_email','')} {f.get('contacto_principal','')} {f.get('domicilio','')}")
        if ql in hay:
            results.append({"type": "family", "id": f["id"],
                            "title": f.get("progenitor1_nombre") or f.get("contacto_principal") or "Familia",
                            "subtitle": f.get("progenitor1_telefono") or f.get("domicilio") or "", "route": "/familias"})

    for i in await list_docs("inscriptions"):
        hay = _norm(f"{i.get('nombre','')} {i.get('apellidos','')} {i.get('progenitor1_telefono','')} {i.get('email_formulario','')}")
        if ql in hay:
            results.append({"type": "inscription", "id": i["id"],
                            "title": f"{i.get('nombre','')} {i.get('apellidos','')}".strip(),
                            "subtitle": f"{i.get('estado','')} · {i.get('categoria') or ''}", "route": "/inscripciones"})

    player_names = {p["id"]: f"{p.get('nombre','')} {p.get('apellidos','')}".strip() for p in players}
    for pay in await list_docs("payments"):
        pname = player_names.get(pay.get("player_id"), "")
        hay = _norm(f"{pname} {pay.get('concepto','')} {pay.get('estado','')} {pay.get('forma_pago','')}")
        if ql in hay:
            results.append({"type": "payment", "id": pay["id"],
                            "title": pname or pay.get("concepto", "Pago"),
                            "subtitle": f"{pay.get('importe_final',0)} € · {pay.get('estado','')}", "route": "/pagos"})

    return results[:40]


@api_router.get("/")
async def root():
    return {"message": "Ikas-Txiki Manager API"}


# ================= DEMO SEED / CLEAR =================
ALL_COLLECTIONS = ["players", "families", "teams", "matches", "callups", "payments",
                   "authorizations", "inscriptions", "trainings", "training_evaluations", "stats", "communications"]


@api_router.post("/clear-all")
async def clear_all():
    for c in ALL_COLLECTIONS:
        await db[c].delete_many({})
    return {"ok": True}


@api_router.post("/seed-demo")
async def seed_demo():
    # wipe first to keep idempotent
    for c in ALL_COLLECTIONS:
        await db[c].delete_many({})

    from datetime import timedelta
    today = date.today()

    # Settings
    await db.settings.update_one({"id": SETTINGS_ID}, {"$set": {
        "id": SETTINGS_ID, "club_nombre": "Ikas-Txiki Futbol Eskola",
        "club_direccion": "Kiroldegia, Donostia", "club_email": "info@ikastxiki.eus",
        "club_telefono": "943 000 000", "temporada_actual": "2025-2026",
        "temporadas": ["2024-2025", "2025-2026"], "campos": ["Campo Municipal", "Anoeta B", "Pista cubierta"],
        "entrenadores": ["Mikel Agirre", "Jon Etxeberria", "Ane Garmendia"],
        "cuota_base": 180, "descuento_hermano": 30,
    }}, upsert=True)

    # Teams
    teams_def = [
        {"nombre": "Benjamín A", "categoria": "Benjamín", "entrenador": "Mikel Agirre", "horario": "L-X 17:30", "campo": "Campo Municipal"},
        {"nombre": "Alevín A", "categoria": "Alevín", "entrenador": "Jon Etxeberria", "horario": "M-J 18:00", "campo": "Anoeta B"},
        {"nombre": "Infantil A", "categoria": "Infantil", "entrenador": "Ane Garmendia", "horario": "M-J 19:00", "campo": "Campo Municipal"},
    ]
    teams = []
    for td in teams_def:
        t = await insert_doc("teams", Team(temporada="2025-2026", estado="activo", **td).model_dump())
        teams.append(t)

    # Families + Players
    players_def = [
        ("Unai", "Goikoetxea", "2016-03-12", 0, "Delantero", "9", "activo", "600111001", "unai.fam@mail.eus", "Calle Mayor 1"),
        ("Ane", "Lizarraga", "2016-07-05", 0, "Centrocampista", "8", "activo", "600111002", "ane.fam@mail.eus", "Av. Libertad 3"),
        ("Iker", "Mendizabal", "2015-11-20", 0, "Portero", "1", "lesionado", "600111003", "iker.fam@mail.eus", "Calle Río 5"),
        ("Maddi", "Sarriegi", "2014-02-18", 1, "Defensa", "4", "activo", "600111004", "maddi.fam@mail.eus", "Plaza Nueva 7"),
        ("Julen", "Aranburu", "2014-09-30", 1, "Delantero", "11", "activo", "600111005", "julen.fam@mail.eus", "Calle Sol 9"),
        ("Nora", "Etxeberria", "2013-05-14", 2, "Centrocampista", "10", "activo", "600111006", "nora.fam@mail.eus", "Av. Mar 2"),
        ("Aitor", "Goikoetxea", "2013-12-01", 2, "Defensa", "3", "pendiente_documentacion", "600111001", "unai.fam@mail.eus", "Calle Mayor 1"),
        ("Leire", "Otaegi", "2015-08-22", 0, "Delantera", "7", "en_prueba", "600111008", "leire.fam@mail.eus", "Calle Norte 4"),
    ]
    players = []
    for (nombre, ape, fnac, tidx, pos, dorsal, estado, tel, email, dom) in players_def:
        pdata = Player(
            nombre=nombre, apellidos=ape, fecha_nacimiento=fnac, equipo_id=teams[tidx]["id"],
            posicion=pos, dorsal=dorsal, estado=estado, progenitor1_telefono=tel,
            progenitor1_email=email, progenitor1_nombre=f"Familia {ape}", domicilio=dom,
            estado_documental="completo" if estado == "activo" else "pendiente",
            fecha_alta=str(today - timedelta(days=120)),
        ).model_dump()
        pdata["categoria"] = compute_category(fnac)
        p = await insert_doc("players", pdata)
        players.append(p)

    # Families
    for ape, tel, email, dom in [("Goikoetxea", "600111001", "unai.fam@mail.eus", "Calle Mayor 1"),
                                  ("Etxeberria", "600111006", "nora.fam@mail.eus", "Av. Mar 2")]:
        await insert_doc("families", Family(progenitor1_nombre=f"Familia {ape}", progenitor1_telefono=tel,
                                             progenitor1_email=email, domicilio=dom, contacto_principal=f"Familia {ape}",
                                             preferencia_comunicacion="whatsapp").model_dump())

    # Matches
    matches_def = [
        (teams[0]["id"], "C.D. Antiguoko", str(today + timedelta(days=3)), "10:00", "local", "liga", "programado", None, None, "1"),
        (teams[1]["id"], "Easo S.D.", str(today + timedelta(days=5)), "11:30", "visitante", "liga", "programado", None, None, "2"),
        (teams[2]["id"], "Real Sociedad B", str(today - timedelta(days=4)), "12:00", "local", "liga", "jugado", 3, 1, "1"),
    ]
    matches = []
    for (eid, rival, fecha, hora, cond, tipo, estado, rp, rr, jor) in matches_def:
        m = await insert_doc("matches", Match(equipo_id=eid, rival=rival, fecha=fecha, hora=hora, condicion=cond,
                                               tipo=tipo, estado=estado, resultado_propio=rp, resultado_rival=rr,
                                               jornada=jor, temporada="2025-2026", campo="Campo Municipal").model_dump())
        matches.append(m)

    # Trainings
    t0_players = [p for p in players if p["equipo_id"] == teams[0]["id"]]
    await insert_doc("trainings", Training(
        equipo_id=teams[0]["id"], campo="Campo Municipal", fecha=str(today + timedelta(days=1)), hora="17:30",
        asistencia=[{"player_id": p["id"], "estado": "presente"} for p in t0_players],
        ejercicios="Rondos + tiro a puerta").model_dump())
    t1_players = [p for p in players if p["equipo_id"] == teams[1]["id"]]
    await insert_doc("trainings", Training(
        equipo_id=teams[1]["id"], campo="Anoeta B", fecha=str(today - timedelta(days=2)), hora="18:00",
        asistencia=[{"player_id": p["id"], "estado": "presente" if i % 2 == 0 else "justificada"} for i, p in enumerate(t1_players)],
        ejercicios="Pase y control").model_dump())

    # Callup for first match
    await insert_doc("callups", Callup(
        match_id=matches[0]["id"], equipo_id=teams[0]["id"],
        convocados=[{"player_id": p["id"], "estado": "confirmado"} for p in t0_players],
        hora_quedada="09:15", lugar_quedada="Vestuarios Campo Municipal",
        material="Botas + agua", mensaje_familias="Convocatoria para el partido del fin de semana.").model_dump())

    # Payments
    for i, p in enumerate(players[:5]):
        estado = ["pagado", "pendiente", "pagado", "parcial", "pendiente"][i]
        await insert_doc("payments", Payment(player_id=p["id"], concepto="Cuota temporada", importe_base=180,
                                              descuento_hermano=30 if i in (3, 4) else 0, estado=estado,
                                              forma_pago="domiciliacion").model_dump())

    # Authorizations
    for i, p in enumerate(players[:4]):
        await insert_doc("authorizations", Authorization(player_id=p["id"], tipo=["general", "imagen", "medica", "desplazamientos"][i],
                                                          firmante=f"Familia {p['apellidos']}", estado="firmada" if i < 2 else "pendiente",
                                                          fecha_firma=str(today - timedelta(days=30)) if i < 2 else None).model_dump())

    # Inscriptions (pending)
    for nombre, ape, fnac, tel in [("Oihan", "Beristain", "2017-04-10", "600222001"),
                                    ("Irati", "Zubizarreta", "2016-10-02", "600222002")]:
        idata = Inscription(nombre=nombre, apellidos=ape, fecha_nacimiento=fnac, progenitor1_telefono=tel,
                            progenitor1_nombre=f"Familia {ape}", estado="recibida", tipo="alta").model_dump()
        idata["categoria"] = compute_category(fnac)
        await insert_doc("inscriptions", idata)

    # Stats
    for p in [players[5], players[0]]:
        await insert_doc("stats", PlayerStats(player_id=p["id"], temporada="2025-2026", partidos_convocado=8,
                                               partidos_jugados=7, minutos=420, goles=5, asistencias=3,
                                               amarillas=1, valoracion=8, posicion=p["posicion"]).model_dump())

    # Communications
    await insert_doc("communications", Communication(destinatario_tipo="equipo", destinatario_id=teams[0]["id"],
                                                      destinatario_nombre=teams[0]["nombre"], canal="whatsapp",
                                                      asunto="Entrenamiento del lunes", mensaje="Recordad traer botas de tacos.",
                                                      enviado=True, fecha_envio=now_iso()).model_dump())

    return {"ok": True, "teams": len(teams), "players": len(players), "matches": len(matches)}



# ================= EQUIPMENT =================

@api_router.get("/equipment")
async def get_equipment(equipo_id: Optional[str] = None, entregada: Optional[str] = None):
    """Devuelve todos los jugadores con sus datos de equipación."""
    query: Dict[str, Any] = {}
    if equipo_id:
        query["equipo_id"] = equipo_id
    if entregada is not None:
        query["equipacion_entregada"] = (entregada.lower() == "true")
    players = await list_docs("players", query)
    teams = {t["id"]: t["nombre"] for t in await list_docs("teams")}
    result = []
    for p in players:
        result.append({
            "id": p["id"],
            "nombre": p.get("nombre", ""),
            "apellidos": p.get("apellidos", ""),
            "categoria": p.get("categoria"),
            "equipo_id": p.get("equipo_id"),
            "equipo_nombre": teams.get(p.get("equipo_id"), "—"),
            "dorsal": p.get("dorsal"),
            "talla_camiseta": p.get("talla_camiseta"),
            "talla_pantalon": p.get("talla_pantalon"),
            "talla_chandal": p.get("talla_chandal"),
            "talla_medias": p.get("talla_medias"),
            "talla_calzado": p.get("talla_calzado"),
            "equipacion_entregada": p.get("equipacion_entregada", False),
            "fecha_entrega_equipacion": p.get("fecha_entrega_equipacion"),
            "observaciones_material": p.get("observaciones_material"),
            "segunda_equipacion": p.get("segunda_equipacion") or {},
            "historial_equipacion": p.get("historial_equipacion") or {},
            "estado": p.get("estado"),
        })
    return result


@api_router.put("/equipment/{player_id}")
async def update_equipment(player_id: str, data: Dict[str, Any]):
    """Actualiza solo los campos de equipación de un jugador."""
    allowed = {
        "dorsal", "talla_camiseta", "talla_pantalon", "talla_chandal",
        "talla_medias", "talla_calzado", "equipacion_entregada",
        "fecha_entrega_equipacion", "observaciones_material"
    }
    update = {k: v for k, v in data.items() if k in allowed}
    return await update_doc("players", player_id, update)


# ================= EXCEL IMPORT / EXPORT =================
EXPORT_COLLECTIONS = ALL_COLLECTIONS + ["settings"]


def _json_cell(value: Any, default: Any = None) -> Any:
    if default is None:
        default = {}
    if value is None:
        return default
    if isinstance(value, float) and math.isnan(value):
        return default
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return default
        try:
            return json.loads(text)
        except Exception:
            return default
    return default


def _display_value(value: Any, fallback: str = "") -> str:
    if value is None:
        return fallback
    if isinstance(value, float) and math.isnan(value):
        return fallback
    if isinstance(value, bool):
        return "Sí" if value else "No"
    text = str(value).strip()
    return text if text else fallback


def _decrypt_iban(value: Any) -> str:
    token = _display_value(value)
    if not token:
        return ""
    try:
        return _fernet(JWT_SECRET).decrypt(token.encode("ascii")).decode("ascii")
    except Exception:
        return ""


def _status_label(value: Any) -> str:
    text = _display_value(value).lower()
    labels = {
        "pendiente": "Pendiente",
        "pendiente_documentacion": "Pendiente documentación",
        "recibida": "Recibida",
        "activo": "Activo",
        "referencia": "Referencia",
        "firmada": "Firmada",
        "validada": "Validada",
    }
    return labels.get(text, _display_value(value))


def _join_list(value: Any) -> str:
    items = _json_cell(value, [])
    if isinstance(items, list):
        return ", ".join(_display_value(item) for item in items if _display_value(item))
    return _display_value(value)


def _season_pairs(value: Any) -> str:
    data = _json_cell(value, {})
    if not isinstance(data, dict):
        return _display_value(value)
    parts = []
    for season, season_value in sorted(data.items()):
        if isinstance(season_value, list):
            text = ", ".join(_display_value(v) for v in season_value if _display_value(v))
        else:
            text = _display_value(season_value)
        if text:
            parts.append(f"{season}: {text}")
    return " | ".join(parts)


def _dict_summary(value: Any) -> str:
    data = _json_cell(value, {})
    if not isinstance(data, dict):
        return _display_value(value)
    parts = []
    labels = {
        "due": "cuota",
        "paid": "pagado",
        "balance_reference": "referencia",
        "confirmed_debt": "deuda confirmada",
        "notifications": "avisos",
        "debit": "domiciliación",
        "images": "imagen",
        "signed": "firmado",
    }
    for key, raw in data.items():
        if key in {"source_present", "source_numeric"}:
            continue
        value_text = _display_value(raw)
        if value_text:
            parts.append(f"{labels.get(key, key)}: {value_text}")
    return " | ".join(parts)


def _authorization_state(auths: Mapping[str, dict], key: str) -> str:
    return _status_label(auths.get(key, {}).get("estado") or "Sin registro")


def _missing_notes(player: Mapping[str, Any], bank_rows: list[dict], auths: Mapping[str, dict], team: Mapping[str, Any]) -> str:
    notes = []
    if not team:
        notes.append("sin equipo")
    if not _display_value(player.get("progenitor1_email")) and not _display_value(player.get("progenitor2_email")):
        notes.append("sin email familiar")
    if not _display_value(player.get("progenitor1_telefono")) and not _display_value(player.get("progenitor2_telefono")):
        notes.append("sin teléfono familiar")
    if not bank_rows:
        notes.append("sin cuenta bancaria")
    if not _display_value(player.get("talla_camiseta")):
        notes.append("sin talla camiseta")
    pending_auths = [name for name, row in auths.items() if _display_value(row.get("estado")).lower() not in {"firmada", "validada"}]
    if pending_auths:
        notes.append(f"autorizaciones pendientes: {len(pending_auths)}")
    return "; ".join(notes) or "completo"


VISUAL_SHEET_META = {
    "Control operativo": "Vista de trabajo para saber qué falta por jugador y priorizar la gestión diaria.",
    "Equipos": "Resumen de equipos, límites y número de jugadores asignados.",
    "Jugadores por equipo": "Listado de plantillas para entregar o revisar con coordinación deportiva.",
    "Contactos familias": "Datos de contacto agrupados para administración y comunicación.",
    "Equipaciones": "Control de tallas, segunda equipación e historial de material.",
    "Cuentas bancarias": "Referencias bancarias para administración: IBAN completo y versión enmascarada.",
    "Historial deportivo": "Trayectoria por temporadas, federación, equipos y entrenamientos históricos.",
    "Autorizaciones resumen": "Estado de permisos por jugador para revisión administrativa.",
}


def _style_excel_sheet(writer, sheet_name: str, freeze: str = "A4", tab_color: str = "0F4C81", table_name: str = ""):
    ws = writer.book[sheet_name]
    ws.sheet_properties.tabColor = tab_color
    if ws.max_row >= 3:
        ws.freeze_panes = freeze
        header_fill = PatternFill("solid", fgColor=tab_color)
        thin = Side(style="thin", color="D8E2EE")
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        if ws.max_row > 3 and ws.max_column > 1:
            ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName=table_name or re.sub(r"[^A-Za-z0-9_]", "", sheet_name)[:20] or "Tabla", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
    for col_idx, column_cells in enumerate(ws.iter_cols(), start=1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in column_cells[:200]:
            max_len = max(max_len, len(_display_value(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, str):
                lowered = cell.value.lower()
                if lowered in {"pendiente", "pendiente documentación"} or "pendiente" in lowered:
                    cell.fill = PatternFill("solid", fgColor="FFF4CC")
                elif lowered in {"sí", "validada", "firmada", "activo", "completo"}:
                    cell.fill = PatternFill("solid", fgColor="E8F6EF")
                elif lowered in {"no", "sin registro"} or lowered.startswith("sin "):
                    cell.fill = PatternFill("solid", fgColor="FDEAEA")


def _write_visual_sheet(writer, sheet_name: str, rows: list[dict], columns: list[str], tab_color: str = "0F4C81"):
    df = pd.DataFrame(rows, columns=columns)
    safe_name = sheet_name[:31]
    df.to_excel(writer, sheet_name=safe_name, index=False, startrow=2)
    ws = writer.book[safe_name]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
    ws.cell(1, 1).value = sheet_name
    ws.cell(1, 1).font = Font(size=16, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=tab_color)
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(columns)))
    ws.cell(2, 1).value = VISUAL_SHEET_META.get(sheet_name, "")
    ws.cell(2, 1).font = Font(italic=True, color="5B677A")
    ws.cell(2, 1).fill = PatternFill("solid", fgColor="F3F6FA")
    ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24
    table_name = re.sub(r"[^A-Za-z0-9_]", "", f"Tabla_{safe_name}")[:30]
    _style_excel_sheet(writer, safe_name, tab_color=tab_color, table_name=table_name)


def _write_start_sheet(writer, metrics: list[dict], warnings: list[dict]):
    sheet_name = "Inicio"
    pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    ws = writer.book[sheet_name]
    ws.sheet_properties.tabColor = "0B3558"
    ws.merge_cells("A1:F1")
    ws["A1"] = "Ikas-Txiki Manager · Exportación operativa"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B3558")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws["A3"] = "Resumen rápido"
    ws["A3"].font = Font(size=13, bold=True, color="0B3558")
    metric_fill = PatternFill("solid", fgColor="EAF4FB")
    for idx, row in enumerate(metrics, start=4):
        ws.cell(idx, 1).value = row["Métrica"]
        ws.cell(idx, 2).value = row["Valor"]
        ws.cell(idx, 1).font = Font(bold=True, color="1F3652")
        ws.cell(idx, 1).fill = metric_fill
        ws.cell(idx, 2).fill = metric_fill
    ws["D3"] = "Avisos operativos"
    ws["D3"].font = Font(size=13, bold=True, color="0B3558")
    for idx, row in enumerate(warnings, start=4):
        ws.cell(idx, 4).value = row["Aviso"]
        ws.cell(idx, 5).value = row["Cantidad"]
        ws.cell(idx, 4).fill = PatternFill("solid", fgColor="FFF4CC" if row["Cantidad"] else "E8F6EF")
        ws.cell(idx, 5).fill = PatternFill("solid", fgColor="FFF4CC" if row["Cantidad"] else "E8F6EF")
    ws["A13"] = "Cómo usar este Excel"
    ws["A13"].font = Font(size=13, bold=True, color="0B3558")
    instructions = [
        "1. Usa 'Control operativo' para revisar qué falta por jugador.",
        "2. Usa 'Jugadores por equipo' para trabajar por plantilla.",
        "3. Usa 'Equipaciones', 'Cuentas bancarias' y 'Autorizaciones resumen' para gestión diaria.",
        "4. Las hojas técnicas están ocultas y sirven para restaurar/importar la base de datos.",
    ]
    for offset, text in enumerate(instructions, start=14):
        ws.cell(offset, 1).value = text
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 28
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


async def _excel_export_data() -> dict[str, list[dict]]:
    return {
        coll: await db[coll].find({}, {"_id": 0}).to_list(10000)
        for coll in EXPORT_COLLECTIONS
    }


def _visual_export_rows(data: dict[str, list[dict]]) -> dict[str, tuple[list[dict], list[str]]]:
    players = data.get("players", [])
    teams = {row.get("id"): row for row in data.get("teams", [])}
    payments_by_player = defaultdict(list)
    authorizations_by_player = defaultdict(dict)

    for row in data.get("payments", []):
        payments_by_player[row.get("player_id")].append(row)
    for row in data.get("authorizations", []):
        authorizations_by_player[row.get("player_id")][_display_value(row.get("tipo"))] = row

    roster_rows = []
    equipment_rows = []
    bank_rows = []
    sports_rows = []
    authorization_rows = []
    contact_rows = []
    control_rows = []

    for player in players:
        team = teams.get(player.get("equipo_id"), {})
        second_kit = _json_cell(player.get("segunda_equipacion"), {})
        sports = _json_cell(player.get("historial_deportivo"), {})
        fee_ref = _json_cell(player.get("referencias_cuotas"), {})
        permission_ref = _json_cell(player.get("referencias_permisos"), {})
        name = f"{_display_value(player.get('nombre'))} {_display_value(player.get('apellidos'))}".strip()
        player_bank_rows = payments_by_player.get(player.get("id"), [])
        auths = authorizations_by_player.get(player.get("id"), {})
        missing_notes = _missing_notes(player, player_bank_rows, auths, team)
        complete_auths = sum(1 for row in auths.values() if _display_value(row.get("estado")).lower() in {"firmada", "validada"})
        pending_auths = max(0, 6 - complete_auths)

        roster_rows.append({
            "Equipo": _display_value(team.get("nombre"), "Sin equipo"),
            "Temporada equipo": _display_value(team.get("temporada")),
            "Categoría equipo": _display_value(team.get("categoria")),
            "Jugador": name,
            "Fecha nacimiento": _display_value(player.get("fecha_nacimiento")),
            "Categoría jugador": _display_value(player.get("categoria")),
            "Modalidad": _display_value(player.get("modalidad")),
            "Posición": _display_value(player.get("posicion")),
            "Centro escolar": _display_value(player.get("centro_escolar")),
            "Estado": _status_label(player.get("estado")),
        })
        control_rows.append({
            "Jugador": name,
            "Equipo": _display_value(team.get("nombre"), "Sin equipo"),
            "Estado ficha": _status_label(player.get("estado")),
            "Autorizaciones pendientes": pending_auths,
            "Cuenta bancaria": "Sí" if player_bank_rows else "No",
            "IBAN visible": f"****{_display_value(player_bank_rows[0].get('iban_last4'))}" if player_bank_rows and _display_value(player_bank_rows[0].get("iban_last4")) else "",
            "Equipación 2ª": "Sí" if second_kit else "No",
            "Tallas básicas": "Sí" if _display_value(player.get("talla_camiseta")) and _display_value(player.get("talla_medias")) else "No",
            "Contacto familiar": "Sí" if _display_value(player.get("progenitor1_email")) or _display_value(player.get("progenitor1_telefono")) or _display_value(player.get("progenitor2_email")) or _display_value(player.get("progenitor2_telefono")) else "No",
            "Qué falta": missing_notes,
        })
        equipment_rows.append({
            "Jugador": name,
            "Categoría": _display_value(player.get("categoria")),
            "Talla camiseta": _display_value(player.get("talla_camiseta")),
            "Talla medias": _display_value(player.get("talla_medias")),
            "Nombre camiseta 2ª": _display_value(second_kit.get("shirt_name")),
            "Dorsal 2ª": _display_value(second_kit.get("number")),
            "Talla camiseta 2ª": _display_value(second_kit.get("shirt_size")),
            "Talla medias 2ª": _display_value(second_kit.get("socks_size")),
            "Historial equipación": _season_pairs(player.get("historial_equipacion")),
            "Equipación entregada": _display_value(player.get("equipacion_entregada")),
        })
        sports_rows.append({
            "Jugador": name,
            "Equipo actual": _display_value(team.get("nombre"), "Sin equipo"),
            "Equipo histórico referencia": _display_value(player.get("equipo_historico_referencia")),
            "Historial equipos": _season_pairs(player.get("historial_equipos")),
            "Historial federación": _season_pairs(player.get("historial_federacion")),
            "Programa 2025-2026": _display_value(sports.get("program_2025_2026")),
            "Categoría deportiva": _display_value(sports.get("playing_category")),
            "Posición": _display_value(sports.get("position") or player.get("posicion")),
            "Entrenamientos históricos": _join_list(player.get("historial_entrenamientos")),
        })
        authorization_rows.append({
            "Jugador": name,
            "General": _authorization_state(auths, "general"),
            "Médica": _authorization_state(auths, "medica"),
            "Imagen": _authorization_state(auths, "imagen"),
            "Protección datos": _authorization_state(auths, "proteccion_datos"),
            "Recogida": _authorization_state(auths, "recogida"),
            "Desplazamientos": _authorization_state(auths, "desplazamientos"),
            "Permisos históricos": _dict_summary(permission_ref),
        })
        contact_rows.append({
            "Jugador": name,
            "Familia": _display_value(player.get("familia_id")),
            "Domicilio": _display_value(player.get("domicilio")),
            "Progenitor 1": _display_value(player.get("progenitor1_nombre")),
            "Teléfono 1": _display_value(player.get("progenitor1_telefono")),
            "Email 1": _display_value(player.get("progenitor1_email")),
            "Progenitor 2": _display_value(player.get("progenitor2_nombre")),
            "Teléfono 2": _display_value(player.get("progenitor2_telefono")),
            "Email 2": _display_value(player.get("progenitor2_email")),
        })
        for payment in player_bank_rows:
            full_iban = _decrypt_iban(payment.get("iban_encrypted")) or _display_value(payment.get("iban"))
            bank_rows.append({
                "Jugador": name,
                "Titular cuenta": _display_value(payment.get("titular_cuenta")),
                "IBAN completo": full_iban,
                "IBAN enmascarado": f"****{_display_value(payment.get('iban_last4'))}" if _display_value(payment.get("iban_last4")) else "",
                "IBAN validado": _display_value(payment.get("iban_validado")),
                "Forma pago": _display_value(payment.get("forma_pago")),
                "Concepto": _display_value(payment.get("concepto")),
                "Estado": _display_value(payment.get("estado")),
                "Referencia histórica": _display_value(payment.get("historical_bank_reference")),
                "Deuda confirmada": _display_value(payment.get("confirmed_debt")),
                "Referencia cuotas": _dict_summary(fee_ref),
            })

    team_summary = []
    team_counts = defaultdict(int)
    for row in roster_rows:
        team_counts[row["Equipo"]] += 1
    for team in data.get("teams", []):
        team_summary.append({
            "Equipo": _display_value(team.get("nombre")),
            "Categoría": _display_value(team.get("categoria")),
            "Temporada": _display_value(team.get("temporada")),
            "Modalidad": _display_value(team.get("modalidad")),
            "Jugadores": team_counts[_display_value(team.get("nombre"))],
            "Límite jugadores": _display_value(team.get("limite_jugadores")),
            "Estado": _display_value(team.get("estado")),
            "Entrenador": _display_value(team.get("entrenador")),
            "Campo": _display_value(team.get("campo")),
        })

    metrics = [
            {"Métrica": "Jugadores", "Valor": len(players)},
            {"Métrica": "Familias", "Valor": len(data.get("families", []))},
            {"Métrica": "Equipos", "Valor": len(data.get("teams", []))},
            {"Métrica": "Referencias bancarias", "Valor": len(data.get("payments", []))},
            {"Métrica": "Autorizaciones", "Valor": len(data.get("authorizations", []))},
            {"Métrica": "Inscripciones", "Valor": len(data.get("inscriptions", []))},
    ]
    warnings = [
        {"Aviso": "Jugadores sin equipo", "Cantidad": sum(1 for row in control_rows if row["Equipo"] == "Sin equipo")},
        {"Aviso": "Jugadores sin cuenta bancaria", "Cantidad": sum(1 for row in control_rows if row["Cuenta bancaria"] == "No")},
        {"Aviso": "Jugadores con autorizaciones pendientes", "Cantidad": sum(1 for row in control_rows if row["Autorizaciones pendientes"])},
        {"Aviso": "Jugadores sin tallas básicas", "Cantidad": sum(1 for row in control_rows if row["Tallas básicas"] == "No")},
        {"Aviso": "Jugadores sin contacto familiar", "Cantidad": sum(1 for row in control_rows if row["Contacto familiar"] == "No")},
    ]

    return {
        "Inicio": (metrics, warnings),
        "Control operativo": (control_rows, ["Jugador", "Equipo", "Estado ficha", "Autorizaciones pendientes", "Cuenta bancaria", "IBAN visible", "Equipación 2ª", "Tallas básicas", "Contacto familiar", "Qué falta"], "C65911"),
        "Equipos": (team_summary, ["Equipo", "Categoría", "Temporada", "Modalidad", "Jugadores", "Límite jugadores", "Estado", "Entrenador", "Campo"], "0F4C81"),
        "Jugadores por equipo": (roster_rows, ["Equipo", "Temporada equipo", "Categoría equipo", "Jugador", "Fecha nacimiento", "Categoría jugador", "Modalidad", "Posición", "Centro escolar", "Estado"], "2F75B5"),
        "Contactos familias": (contact_rows, ["Jugador", "Familia", "Domicilio", "Progenitor 1", "Teléfono 1", "Email 1", "Progenitor 2", "Teléfono 2", "Email 2"], "548235"),
        "Equipaciones": (equipment_rows, ["Jugador", "Categoría", "Talla camiseta", "Talla medias", "Nombre camiseta 2ª", "Dorsal 2ª", "Talla camiseta 2ª", "Talla medias 2ª", "Historial equipación", "Equipación entregada"], "8064A2"),
        "Cuentas bancarias": (bank_rows, ["Jugador", "Titular cuenta", "IBAN completo", "IBAN enmascarado", "IBAN validado", "Forma pago", "Concepto", "Estado", "Referencia histórica", "Deuda confirmada", "Referencia cuotas"], "A64D79"),
        "Historial deportivo": (sports_rows, ["Jugador", "Equipo actual", "Equipo histórico referencia", "Historial equipos", "Historial federación", "Programa 2025-2026", "Categoría deportiva", "Posición", "Entrenamientos históricos"], "5B9BD5"),
        "Autorizaciones resumen": (authorization_rows, ["Jugador", "General", "Médica", "Imagen", "Protección datos", "Recogida", "Desplazamientos", "Permisos históricos"], "ED7D31"),
    }


def _flatten_for_excel(doc: dict) -> dict:
    out = {}
    for k, v in doc.items():
        if k == "_id":
            continue
        if isinstance(v, (list, dict)):
            out[k] = json.dumps(v, ensure_ascii=False)
        else:
            out[k] = v
    return out


def _unflatten_from_excel(row: dict) -> dict:
    out = {}
    for k, v in row.items():
        if v is None:
            continue
        if isinstance(v, float) and math.isnan(v):
            continue
        if isinstance(v, str):
            s = v.strip()
            if s == "":
                continue
            if (s.startswith("[") and s.endswith("]")) or (s.startswith("{") and s.endswith("}")):
                try:
                    out[k] = json.loads(s)
                    continue
                except Exception:
                    pass
            out[k] = v
        else:
            out[k] = v
    return out


@api_router.get("/export-excel")
async def export_excel():
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_data = await _excel_export_data()
        for sheet_name, payload in _visual_export_rows(export_data).items():
            if sheet_name == "Inicio":
                metrics, warnings = payload
                _write_start_sheet(writer, metrics, warnings)
                continue
            rows, columns, tab_color = payload
            _write_visual_sheet(writer, sheet_name, rows, columns, tab_color=tab_color)
        for coll in EXPORT_COLLECTIONS:
            docs = export_data.get(coll, [])
            rows = [_flatten_for_excel(d) for d in docs]
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name=coll[:31], index=False)
            writer.book[coll[:31]].sheet_state = "veryHidden"
        if not writer.book.sheetnames:
            pd.DataFrame().to_excel(writer, sheet_name="empty", index=False)
    buffer.seek(0)
    fname = f"ikastxiki_backup_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _safe_csv_name(value: str) -> str:
    name = re.sub(r"[^A-Za-z0-9ÁÉÍÓÚÜÑáéíóúüñ_-]+", "_", value).strip("_")
    return name or "datos"


@api_router.get("/export-csv")
async def export_csv():
    export_data = await _excel_export_data()
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for sheet_name, payload in _visual_export_rows(export_data).items():
            if sheet_name == "Inicio":
                metrics, warnings = payload
                rows = [{"Sección": "Resumen", **row} for row in metrics] + [
                    {"Sección": "Avisos", "Métrica": row["Aviso"], "Valor": row["Cantidad"]} for row in warnings
                ]
                df = pd.DataFrame(rows)
            else:
                rows, columns, _tab_color = payload
                df = pd.DataFrame(rows, columns=columns)
            archive.writestr(f"01_operativo/{_safe_csv_name(sheet_name)}.csv", df.to_csv(index=False))
        for coll in EXPORT_COLLECTIONS:
            rows = [_flatten_for_excel(d) for d in export_data.get(coll, [])]
            archive.writestr(f"02_backup_tecnico/{_safe_csv_name(coll)}.csv", pd.DataFrame(rows).to_csv(index=False))
    buffer.seek(0)
    fname = f"ikastxiki_csv_{date.today().isoformat()}.zip"
    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api_router.post("/import-excel")
async def import_excel(file: UploadFile = File(...)):
    content = await file.read()
    try:
        sheets = pd.read_excel(io.BytesIO(content), sheet_name=None)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {e}")

    summary = {}
    for coll, df in sheets.items():
        if coll not in EXPORT_COLLECTIONS:
            continue
        records = df.to_dict(orient="records")
        cleaned = [_unflatten_from_excel(r) for r in records]
        cleaned = [c for c in cleaned if c]
        if coll == "settings":
            for c in cleaned:
                c["id"] = SETTINGS_ID
                await db.settings.update_one({"id": SETTINGS_ID}, {"$set": c}, upsert=True)
            summary[coll] = len(cleaned)
            continue
        await db[coll].delete_many({})
        for c in cleaned:
            if not c.get("id"):
                c["id"] = new_id()
            c.setdefault("created_at", now_iso())
            c["updated_at"] = now_iso()
        if cleaned:
            await db[coll].insert_many(cleaned)
        summary[coll] = len(cleaned)
    return {"ok": True, "imported": summary}


# ================= HYBRID ASSISTANT =================
class AssistantHelpRequest(BaseModel):
    message: str = Field(min_length=1, max_length=1200)
    route: str = Field(default="/", max_length=80)
    language: str = "es"

    @field_validator("language")
    @classmethod
    def validate_assistant_language(cls, value: str):
        if value not in {"es", "eu"}:
            raise ValueError("Idioma no válido")
        return value


class AssistantProposalRequest(BaseModel):
    intent: str
    target_id: Optional[str] = None
    data: Dict[str, Any] = Field(default_factory=dict)


class AssistantInternalQueryRequest(BaseModel):
    intent: str
    target_id: Optional[str] = None


class AssistantConfirmRequest(BaseModel):
    confirmation_nonce: str = Field(min_length=20, max_length=120)


ASSISTANT_ALLOWED_FIELDS = {
    "player.create": {
        "nombre", "apellidos", "fecha_nacimiento", "centro_escolar", "categoria",
        "equipo_id", "dorsal", "posicion", "estado", "familia_id",
    },
    "player.update": {
        "nombre", "apellidos", "centro_escolar", "categoria", "equipo_id",
        "dorsal", "posicion", "estado", "familia_id",
    },
    "family.create": {
        "progenitor1_nombre", "progenitor1_telefono", "progenitor1_email",
        "progenitor2_nombre", "progenitor2_telefono", "progenitor2_email",
        "domicilio", "contacto_principal", "preferencia_comunicacion", "observaciones",
    },
    "family.update": {
        "progenitor1_nombre", "progenitor1_telefono", "progenitor1_email",
        "progenitor2_nombre", "progenitor2_telefono", "progenitor2_email",
        "domicilio", "contacto_principal", "preferencia_comunicacion", "observaciones",
    },
    "player.link_family": {"familia_id"},
    "player.assign_team": {"equipo_id"},
    "inscription.create": {
        "tipo", "nombre", "apellidos", "fecha_nacimiento", "centro_escolar",
        "nueva_incorporacion", "estado", "categoria", "temporada", "equipo_id",
        "modalidad", "equipamiento_items", "observaciones",
    },
    "attendance.update": {"training_id", "player_id", "estado", "motivo"},
    "callup.prepare": {
        "match_id", "equipo_id", "convocados", "hora_quedada", "lugar_quedada",
        "material", "mensaje_familias", "response_deadline", "player_self_response_allowed",
    },
    "equipment.update": {
        "dorsal", "talla_camiseta", "talla_pantalon", "talla_chandal",
        "talla_medias", "talla_calzado", "equipacion_entregada",
        "fecha_entrega_equipacion", "observaciones_material", "equipamiento_items",
    },
}

ASSISTANT_INTERNAL_QUERIES = {
    "player.summary": {"resource": "players", "target_required": True},
    "team.summary": {"resource": "teams", "target_required": True},
    "attendance.recent": {"resource": "attendance", "target_required": False},
}


def _assistant_user_id(user: Mapping[str, Any]) -> str:
    return str(user.get("id") or user.get("username") or "")


def _assistant_session(request: Request) -> str:
    return session_fingerprint(request.cookies.get("ikastxiki_session"))


def _assistant_rate_limit(user: Mapping[str, Any]) -> None:
    key = _assistant_user_id(user)
    now = time.monotonic()
    attempts = assistant_attempts[key]
    while attempts and now - attempts[0] > ASSISTANT_RATE_WINDOW_SECONDS:
        attempts.popleft()
    if len(attempts) >= ASSISTANT_RATE_MAX_REQUESTS:
        raise HTTPException(status_code=429, detail="Límite temporal del asistente alcanzado")
    attempts.append(now)


async def _assistant_existing(intent: str, target_id: Optional[str], data: Mapping[str, Any]) -> Optional[dict]:
    if intent in {"player.update", "player.link_family", "player.assign_team", "equipment.update"}:
        return await get_doc("players", str(target_id or ""))
    if intent == "family.update":
        return await get_doc("families", str(target_id or ""))
    if intent == "attendance.update":
        return await get_doc("trainings", str(data.get("training_id") or ""))
    return None


async def _validate_assistant_relations(intent: str, data: Mapping[str, Any], existing: Optional[dict]) -> None:
    team_id = data.get("equipo_id")
    if team_id:
        team = await get_doc("teams", str(team_id))
        if team.get("estado") not in (None, "", "activo", "active"):
            raise HTTPException(status_code=422, detail="El equipo no está activo")
        category = data.get("categoria") or (existing or {}).get("categoria")
        if category and team.get("categoria") and normalize_key(team.get("categoria")) != normalize_key(category):
            raise HTTPException(status_code=422, detail="El equipo no es compatible con la categoría")
    if data.get("familia_id"):
        await get_doc("families", str(data["familia_id"]))
    if intent == "attendance.update":
        if data.get("estado") not in ATTENDANCE_STATES:
            raise HTTPException(status_code=422, detail="Estado de asistencia no válido")
        player = await get_doc("players", str(data.get("player_id") or ""))
        if existing and existing.get("equipo_id") and player.get("equipo_id") != existing.get("equipo_id"):
            raise HTTPException(status_code=422, detail="El jugador no pertenece al equipo del entrenamiento")
    if intent == "callup.prepare":
        match = await get_doc("matches", str(data.get("match_id") or ""))
        if data.get("equipo_id") and match.get("equipo_id") != data.get("equipo_id"):
            raise HTTPException(status_code=422, detail="El partido no pertenece al equipo indicado")


async def _assistant_duplicate_hints(intent: str, data: Mapping[str, Any]) -> list[dict]:
    if intent not in {"player.create", "inscription.create"}:
        return []
    name = str(data.get("nombre") or "").strip()
    surname = str(data.get("apellidos") or "").strip()
    if not name:
        return []
    query = {"nombre": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}
    if surname:
        query["apellidos"] = {"$regex": f"^{re.escape(surname)}$", "$options": "i"}
    matches = await list_docs("players", query)
    return [{"id": row.get("id"), "label": "Posible coincidencia en tu ámbito"} for row in matches[:5]]


def _assistant_clean_data(intent: str, data: Mapping[str, Any]) -> dict:
    allowed = ASSISTANT_ALLOWED_FIELDS.get(intent)
    if allowed is None:
        raise HTTPException(status_code=422, detail="Acción guiada no permitida")
    unknown = set(data) - allowed
    if unknown:
        raise HTTPException(status_code=422, detail="La propuesta contiene campos no permitidos")
    return {key: value for key, value in data.items() if key in allowed}


@api_router.get("/assistant/capabilities")
async def assistant_capabilities():
    actor = current_user_context.get() or {}
    capabilities = []
    for intent, definition in ACTION_DEFINITIONS.items():
        if has_permission(actor, definition["resource"], definition["action"]):
            capabilities.append({
                "intent": intent, "required": definition["required"],
                "fields": sorted(ASSISTANT_ALLOWED_FIELDS[intent]),
            })
    return {
        "knowledge_version": KNOWLEDGE_VERSION,
        "provider_configured": assistant_provider.configured,
        "modules": available_modules(str(actor.get("role"))),
        "actions": capabilities,
        "queries": [
            {"intent": intent, "target_required": definition["target_required"]}
            for intent, definition in ASSISTANT_INTERNAL_QUERIES.items()
            if has_permission(actor, definition["resource"], "read")
        ],
        "conversation_persisted": False,
    }


@api_router.post("/assistant/help")
async def assistant_help(payload: AssistantHelpRequest):
    actor = current_user_context.get() or {}
    _assistant_rate_limit(actor)
    return answer_help(payload.message, actor, payload.route, payload.language, assistant_provider)


@api_router.post("/assistant/internal-query")
async def assistant_internal_query(payload: AssistantInternalQueryRequest):
    """Consultas cerradas y proyectadas; nunca se delegan al proveedor."""
    actor = current_user_context.get() or {}
    _assistant_rate_limit(actor)
    definition = ASSISTANT_INTERNAL_QUERIES.get(payload.intent)
    if not definition:
        raise HTTPException(status_code=422, detail="Consulta interna no permitida")
    enforce_permission(actor, definition["resource"], "read")
    if definition["target_required"] and not payload.target_id:
        raise HTTPException(status_code=422, detail="La consulta requiere un identificador")
    if payload.intent == "player.summary":
        player = await get_doc("players", str(payload.target_id))
        return {"channel": "internal", "result": {
            "id": player.get("id"), "nombre": player.get("nombre"),
            "apellidos": player.get("apellidos"), "categoria": player.get("categoria"),
            "equipo_id": player.get("equipo_id"), "estado": player.get("estado"),
        }}
    if payload.intent == "team.summary":
        team = await get_doc("teams", str(payload.target_id))
        players = await list_docs("players", {"equipo_id": team.get("id")})
        return {"channel": "internal", "result": {
            "id": team.get("id"), "nombre": team.get("nombre"),
            "categoria": team.get("categoria"), "modalidad": team.get("modalidad"),
            "temporada": team.get("temporada"), "jugadores": len(players),
        }}
    trainings, _, _, allowed_players = await attendance_data()
    rows = attendance_rows(trainings, allowed_players)
    return {"channel": "internal", "result": {
        "summary": attendance_summary(trainings, allowed_players),
        "recent": sorted(rows, key=lambda row: row.get("fecha") or "", reverse=True)[:10],
    }}


@api_router.post("/assistant/proposals")
async def assistant_create_proposal(payload: AssistantProposalRequest, request: Request):
    actor = current_user_context.get() or {}
    _assistant_rate_limit(actor)
    definition = ACTION_DEFINITIONS.get(payload.intent)
    if not definition:
        raise HTTPException(status_code=422, detail="Acción guiada no permitida")
    enforce_permission(actor, definition["resource"], definition["action"])
    data = _assistant_clean_data(payload.intent, payload.data)
    required_values = {**data, "target_id": payload.target_id}
    missing = [key for key in definition["required"] if required_values.get(key) in (None, "", [])]
    if missing:
        raise HTTPException(status_code=422, detail="Faltan campos obligatorios para preparar el cambio")
    existing = await _assistant_existing(payload.intent, payload.target_id, data)
    await _validate_assistant_relations(payload.intent, data, existing)
    duplicates = await _assistant_duplicate_hints(payload.intent, data)
    preview = {
        "operation": payload.intent, "target_id": payload.target_id,
        "changes": data, "possible_duplicates": duplicates,
        "requires_explicit_confirmation": True,
    }
    proposal = assistant_proposals.create(
        user_id=_assistant_user_id(actor), session_hash=_assistant_session(request),
        intent=payload.intent, data=data, target_id=payload.target_id,
        expected_version=(existing or {}).get("updated_at"), preview=preview,
    )
    return public_proposal(proposal)


async def _execute_assistant_proposal(proposal) -> dict:
    data = dict(proposal.data)
    intent = proposal.intent
    if intent == "player.create":
        model = Player(**data).model_dump()
        if model.get("fecha_nacimiento"):
            model["categoria"] = compute_category(model["fecha_nacimiento"])
        return await insert_doc("players", model)
    if intent == "player.update":
        existing = await get_doc("players", proposal.target_id)
        return await update_doc("players", proposal.target_id, {**existing, **data})
    if intent == "family.create":
        return await insert_doc("families", Family(**data).model_dump())
    if intent == "family.update":
        existing = await get_doc("families", proposal.target_id)
        return await update_doc("families", proposal.target_id, {**existing, **data})
    if intent in {"player.link_family", "player.assign_team"}:
        return await update_doc("players", proposal.target_id, data)
    if intent == "inscription.create":
        model = Inscription(**data).model_dump()
        if model.get("fecha_nacimiento"):
            model["categoria"] = compute_category(model["fecha_nacimiento"])
        return await insert_doc("inscriptions", model)
    if intent == "attendance.update":
        training = await get_doc("trainings", str(data["training_id"]))
        attendance = list(training.get("asistencia") or [])
        item = {"player_id": data["player_id"], "estado": data["estado"], "motivo": data.get("motivo")}
        index = next((index for index, value in enumerate(attendance) if value.get("player_id") == data["player_id"]), None)
        if index is None:
            attendance.append(item)
        else:
            attendance[index] = item
        actor = current_user_context.get() or {}
        changes = attendance_history(training.get("asistencia", []), attendance, actor)
        return await update_doc("trainings", training["id"], {
            **training, "asistencia": attendance,
            "attendance_history": [*(training.get("attendance_history") or []), *changes],
            "attendance_updated_by": {"id": actor.get("id"), "role": actor.get("role"), "at": now_iso()},
        })
    if intent == "callup.prepare":
        model = Callup(**data).model_dump()
        return await insert_doc("callups", model)
    if intent == "equipment.update":
        return await update_doc("players", proposal.target_id, data)
    raise HTTPException(status_code=422, detail="Acción guiada no permitida")


@api_router.post("/assistant/proposals/{proposal_id}/confirm")
async def assistant_confirm_proposal(proposal_id: str, payload: AssistantConfirmRequest,
                                     request: Request):
    if request.headers.get("X-Assistant-Confirm") != "true":
        raise HTTPException(status_code=403, detail="Falta la confirmación explícita de la operación")
    actor = current_user_context.get() or {}
    try:
        proposal = assistant_proposals.get(proposal_id, _assistant_user_id(actor), _assistant_session(request))
    except PermissionError:
        raise HTTPException(status_code=403, detail="La propuesta no pertenece a esta sesión")
    except TimeoutError:
        raise HTTPException(status_code=410, detail="La propuesta ha caducado")
    definition = ACTION_DEFINITIONS[proposal.intent]
    enforce_permission(actor, definition["resource"], definition["action"])
    existing = await _assistant_existing(proposal.intent, proposal.target_id, proposal.data)
    if existing and existing.get("updated_at") != proposal.expected_version:
        raise HTTPException(status_code=409, detail="El registro cambió después de la vista previa")
    await _validate_assistant_relations(proposal.intent, proposal.data, existing)
    try:
        assistant_proposals.consume(
            proposal_id, _assistant_user_id(actor), _assistant_session(request),
            payload.confirmation_nonce,
        )
    except PermissionError:
        raise HTTPException(status_code=403, detail="Confirmación inválida")
    except ValueError:
        raise HTTPException(status_code=409, detail="La propuesta ya fue utilizada o cancelada")
    result = await _execute_assistant_proposal(proposal)
    proposal.result = {"id": result.get("id"), "ok": True}
    await db.internal_events.insert_one({
        "id": new_id(), "type": "assistant.operation_confirmed",
        "actor_user_id": actor.get("id"), "actor_role": actor.get("role"),
        "operation": proposal.intent, "target_id": result.get("id") or proposal.target_id,
        "created_at": now_iso(), "result": "success",
    })
    return {"ok": True, "operation": proposal.intent, "result_id": result.get("id")}


@api_router.post("/assistant/proposals/{proposal_id}/cancel")
async def assistant_cancel_proposal(proposal_id: str, request: Request):
    actor = current_user_context.get() or {}
    try:
        proposal = assistant_proposals.cancel(
            proposal_id, _assistant_user_id(actor), _assistant_session(request),
        )
    except PermissionError:
        raise HTTPException(status_code=403, detail="La propuesta no pertenece a esta sesión")
    except TimeoutError:
        raise HTTPException(status_code=410, detail="La propuesta ha caducado")
    except ValueError:
        raise HTTPException(status_code=409, detail="La propuesta ya fue utilizada")
    return {"ok": True, "cancelled": proposal.cancelled}


app.include_router(api_router, dependencies=[Depends(authorize_request)])

cors_origins = os.environ.get('CORS_ORIGINS', 'https://ikasfutbase.cibermedida.es').split(',')
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def initialize_training_evaluation_indexes():
    await ensure_training_evaluation_indexes()


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
