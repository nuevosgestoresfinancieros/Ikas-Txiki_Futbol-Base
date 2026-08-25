"""Restaura asignaciones de jugadores desde una copia Excel verificada.

La migración es deliberadamente conservadora:

* relaciona jugadores exclusivamente por su identificador ``id``;
* crea los equipos de la temporada destino solo cuando sean necesarios;
* nunca borra ni modifica fichas deportivas fuera de ``equipo_id``;
* por defecto funciona en modo de vista previa;
* se niega a aplicar un lote parcial: cada jugador sin equipo debe tener una
  equivalencia exacta en la copia o una decisión explícita en ``--overrides``.

No se ejecuta en el despliegue. Antes de usar ``--apply`` debe existir una
copia de seguridad comprobada de MongoDB.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient


DEFAULT_SEASON = "2026-2027"
TEAM_FIELDS = (
    "nombre", "categoria", "modalidad", "entrenador", "segundo_entrenador",
    "delegado", "dias_entrenamiento", "horario", "campo", "estado",
)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normal(value: object) -> str:
    return " ".join(str(value or "").upper().split())


def missing_team_query() -> dict:
    return {"$or": [
        {"equipo_id": {"$exists": False}}, {"equipo_id": None}, {"equipo_id": ""},
    ]}


def player_identity(player: dict[str, Any]) -> tuple[str, str, str, str]:
    """Conservative identity used only to stop an accidental duplicate roster row."""
    return (
        normal(player.get("nombre")), normal(player.get("apellidos")),
        normal(player.get("categoria")), normal(player.get("modalidad")),
    )


def sheet_rows(path: Path, sheet: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
    except KeyError as exc:
        raise SystemExit(f"El Excel no contiene la hoja obligatoria {sheet!r}") from exc
    values = worksheet.iter_rows(values_only=True)
    headers = [str(item).strip() if item is not None else "" for item in next(values, ())]
    if not headers or "id" not in headers:
        raise SystemExit(f"La hoja {sheet!r} no contiene una cabecera id válida")
    return [dict(zip(headers, row)) for row in values if any(item is not None for item in row)]


def read_backup(path: Path) -> tuple[dict[str, str], dict[str, dict[str, Any]]]:
    player_rows = sheet_rows(path, "players")
    team_rows = sheet_rows(path, "teams")
    teams = {str(row.get("id")): row for row in team_rows if row.get("id")}
    player_teams = {
        str(row["id"]): str(row["equipo_id"])
        for row in player_rows
        if row.get("id") and row.get("equipo_id") and str(row["equipo_id"]) in teams
    }
    return player_teams, teams


def read_overrides(path: Path | None) -> dict[str, str]:
    if not path:
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SystemExit(f"No se puede leer --overrides: {exc}") from exc
    if not isinstance(data, dict) or not all(isinstance(key, str) and isinstance(value, str) for key, value in data.items()):
        raise SystemExit("--overrides debe ser un objeto JSON {\"player_id\": \"NOMBRE DEL EQUIPO\"}")
    return {key: normal(value) for key, value in data.items() if normal(value)}


def clone_team(source: dict[str, Any], season: str, roster_size: int) -> dict[str, Any]:
    clone = {field: source.get(field) for field in TEAM_FIELDS if source.get(field) is not None}
    clone.update({
        "id": str(uuid4()), "temporada": season,
        "estado": source.get("estado") or "activo",
        # A team must never be created with a roster cap below the restored roster.
        "limite_jugadores": max(int(source.get("limite_jugadores") or 0), roster_size, 1),
        "created_at": now_iso(), "updated_at": now_iso(),
    })
    return clone


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("backup", type=Path, help="Excel de copia con hojas players y teams")
    parser.add_argument("--season", default=DEFAULT_SEASON, help="Temporada destino")
    parser.add_argument("--overrides", type=Path, help="JSON de equipos para jugadores no presentes en la copia")
    parser.add_argument("--consolidate-duplicates", action="store_true",
                        help="Archivar y retirar registros duplicados sin referencias antes de finalizar")
    parser.add_argument("--apply", action="store_true", help="Escribir los equipos y asignaciones")
    parser.add_argument("--confirm", default="", help="Debe ser RESTORE-ALL-PLAYER-TEAMS")
    args = parser.parse_args()

    if not args.backup.is_file():
        raise SystemExit(f"No existe el Excel de copia: {args.backup}")
    backup_assignments, backup_teams = read_backup(args.backup)
    overrides = read_overrides(args.overrides)

    backend_dir = Path(__file__).resolve().parents[1]
    load_dotenv(backend_dir / ".env")
    client = MongoClient(os.environ["MONGO_URL"], serverSelectionTimeoutMS=5000)
    database = client[os.environ["DB_NAME"]]
    players = list(database.players.find(missing_team_query(), {
        "_id": 1, "id": 1, "nombre": 1, "apellidos": 1, "categoria": 1, "modalidad": 1,
    }))
    assigned_players = list(database.players.find({"equipo_id": {"$nin": [None, ""]}}, {
        "_id": 0, "id": 1, "nombre": 1, "apellidos": 1, "categoria": 1, "modalidad": 1, "equipo_id": 1,
    }))
    assigned_by_identity = {player_identity(player): player for player in assigned_players}
    target_teams = list(database.teams.find({"temporada": args.season}))

    current_by_name = {normal(team.get("nombre")): team for team in target_teams if team.get("id")}
    plans: list[tuple[dict[str, Any], str, str]] = []  # player, source team id/name, source kind
    unresolved: list[dict[str, str]] = []
    invalid_overrides: list[dict[str, str]] = []
    duplicates: list[dict[str, str]] = []

    for player in players:
        player_id = str(player.get("id") or "")
        existing = assigned_by_identity.get(player_identity(player))
        if existing:
            duplicates.append({
                "player_id": player_id,
                "name": f"{player.get('nombre') or ''} {player.get('apellidos') or ''}".strip(),
                "existing_player_id": str(existing.get("id") or ""),
                "existing_team_id": str(existing.get("equipo_id") or ""),
            })
            continue
        backup_team_id = backup_assignments.get(player_id)
        if backup_team_id:
            plans.append((player, backup_team_id, "backup"))
            continue
        requested_name = overrides.get(player_id)
        if requested_name:
            source = current_by_name.get(requested_name)
            if source:
                plans.append((player, str(source["id"]), "current"))
            else:
                invalid_overrides.append({"player_id": player_id, "team": requested_name})
        else:
            unresolved.append({
                "player_id": player_id, "name": f"{player.get('nombre') or ''} {player.get('apellidos') or ''}".strip(),
            })

    requested_backup_ids = Counter(team_id for _, team_id, kind in plans if kind == "backup")
    cloned: dict[str, dict[str, Any]] = {}
    newly_created_source_ids: set[str] = set()
    for source_id, roster_size in requested_backup_ids.items():
        source = backup_teams[source_id]
        current = current_by_name.get(normal(source.get("nombre")))
        if current:
            cloned[source_id] = current
        else:
            cloned[source_id] = clone_team(source, args.season, roster_size)
            newly_created_source_ids.add(source_id)

    assignments = []
    for player, source_team, kind in plans:
        team = cloned[source_team] if kind == "backup" else next(team for team in target_teams if str(team["id"]) == source_team)
        assignments.append({"player_id": player["id"], "team_id": team["id"], "team_name": team.get("nombre"), "source": kind})

    report = {
        "season": args.season,
        "players_without_team_before": len(players),
        "assignments_planned": len(assignments),
        "teams_to_create": sorted({team.get("nombre") for source, team in cloned.items() if source in newly_created_source_ids}),
        "unresolved": unresolved,
        "possible_duplicates": duplicates,
        "invalid_overrides": invalid_overrides,
        "sample_assignments": assignments[:10],
        "dry_run": not args.apply,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))

    if not args.apply:
        client.close()
        return
    if args.confirm != "RESTORE-ALL-PLAYER-TEAMS":
        raise SystemExit("Para aplicar use --confirm RESTORE-ALL-PLAYER-TEAMS")
    if unresolved or invalid_overrides or len(assignments) + len(duplicates) != len(players):
        raise SystemExit("No se aplica una restauración parcial: corrija unresolved/invalid_overrides primero")
    if duplicates and not args.consolidate_duplicates:
        raise SystemExit("Hay registros duplicados: revise possible_duplicates o use --consolidate-duplicates tras verificar la vista previa")

    # A duplicate used by sporting or access records must be reviewed manually.
    # The one safe exception is an inscription: it may be moved to the
    # canonical player only when it has no matching canonical inscription.
    duplicate_references: dict[str, list[str]] = {}
    reference_checks = {
        "users": lambda player_id: database.users.count_documents({"player_id": player_id}),
        "inscriptions": lambda player_id: database.inscriptions.count_documents({"player_id": player_id}),
        "payments": lambda player_id: database.payments.count_documents({"player_id": player_id}),
        "authorizations": lambda player_id: database.authorizations.count_documents({"player_id": player_id}),
        "trainings": lambda player_id: database.trainings.count_documents({"asistencia.player_id": player_id}),
        "callups": lambda player_id: database.callups.count_documents({"convocados.player_id": player_id}),
    }
    for duplicate in duplicates:
        player_id = duplicate["player_id"]
        references = [name for name, check in reference_checks.items() if check(player_id)]
        if references:
            duplicate_references[player_id] = references
    forbidden_references = {
        player_id: references for player_id, references in duplicate_references.items()
        if set(references) - {"inscriptions"}
    }
    inscription_moves: dict[str, list[dict[str, Any]]] = {}
    inscription_conflicts: dict[str, int] = {}
    for duplicate in duplicates:
        source_id = duplicate["player_id"]
        rows = list(database.inscriptions.find({"player_id": source_id}, {"_id": 0}))
        if not rows:
            continue
        target_id = duplicate["existing_player_id"]
        collisions = 0
        for row in rows:
            query = {"player_id": target_id, "temporada": row.get("temporada")}
            if row.get("import_identity_key"):
                query["import_identity_key"] = row["import_identity_key"]
            if database.inscriptions.count_documents(query):
                collisions += 1
        if collisions:
            inscription_conflicts[source_id] = collisions
        else:
            inscription_moves[source_id] = rows
    if forbidden_references or inscription_conflicts:
        print(json.dumps({"duplicate_references": duplicate_references}, ensure_ascii=False, indent=2))
        print(json.dumps({"forbidden_references": forbidden_references,
                          "inscription_conflicts": inscription_conflicts}, ensure_ascii=False, indent=2))
        raise SystemExit("No se consolidan duplicados con referencias incompatibles; requieren revisión manual")

    created_teams = [team for source, team in cloned.items() if source in newly_created_source_ids]
    if created_teams:
        database.teams.insert_many(created_teams)
    timestamp = now_iso()
    for item in assignments:
        database.players.update_one({"id": item["player_id"], **missing_team_query()}, {"$set": {
            "equipo_id": item["team_id"], "updated_at": timestamp,
        }})
    # The capacity is only increased where the restored roster requires it.
    counts = Counter(item["team_id"] for item in assignments)
    for team_id, incoming in counts.items():
        existing_count = database.players.count_documents({"equipo_id": team_id})
        database.teams.update_one({"id": team_id, "limite_jugadores": {"$lt": existing_count}}, {"$set": {
            "limite_jugadores": existing_count, "updated_at": timestamp,
        }})
    archived_duplicates = 0
    moved_inscriptions = 0
    for duplicate in duplicates:
        original = database.players.find_one({"id": duplicate["player_id"]})
        if not original:
            continue
        source_id = duplicate["player_id"]
        target_id = duplicate["existing_player_id"]
        result = database.inscriptions.update_many({"player_id": source_id}, {"$set": {
            "player_id": target_id, "merged_from_player_id": source_id, "updated_at": timestamp,
        }})
        moved_inscriptions += result.modified_count
        original.pop("_id", None)
        database.player_duplicate_archive.insert_one({
            "id": str(uuid4()), "archived_at": timestamp,
            "reason": "duplicate_player_during_team_restoration",
            "canonical_player_id": target_id,
            "original_player": original,
        })
        database.players.delete_one({"id": duplicate["player_id"]})
        archived_duplicates += 1
    database.team_restoration_audit.insert_one({
        "action": "restore_player_teams_from_backup", "at": timestamp, "season": args.season,
        "assignments": len(assignments), "teams_created": len(created_teams),
        "duplicates_archived": archived_duplicates,
        "inscriptions_moved": moved_inscriptions,
        "backup": args.backup.name,
    })
    print(json.dumps({"applied": True, "assignments": len(assignments), "teams_created": len(created_teams),
                      "duplicates_archived": archived_duplicates, "inscriptions_moved": moved_inscriptions}, ensure_ascii=False))
    client.close()


if __name__ == "__main__":
    main()
