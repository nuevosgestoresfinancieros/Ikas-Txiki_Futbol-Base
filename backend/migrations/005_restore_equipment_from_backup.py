"""Restaura equipaciones principales desde una copia Excel por ID exacto.

Solo actualiza valores presentes en la copia: talla de camiseta, talla de
medias, dorsal, nombre de camiseta e historial. La estructura histórica
``segunda_equipacion`` de la copia representa la equipación actual principal;
no se restaura como segunda equipación. Por defecto solo informa; ``--apply``
exige una confirmación explícita.
"""
from __future__ import annotations

import argparse
import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def nonempty(value: object) -> bool:
    return value not in (None, "", [], {}, "[]", "{}")


def structured(value: object) -> object:
    if not isinstance(value, str):
        return value
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return value


def backup_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["players"]
    except KeyError as exc:
        raise SystemExit("El Excel no contiene la hoja obligatoria 'players'") from exc
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(item).strip() if item is not None else "" for item in next(iterator, ())]
    required = {"id", "talla_camiseta", "talla_medias", "segunda_equipacion", "historial_equipacion"}
    if not required.issubset(headers):
        raise SystemExit("La hoja players no contiene las columnas de equipación esperadas")
    return [dict(zip(headers, row)) for row in iterator if any(item is not None for item in row)]


def fields_to_restore(row: dict[str, Any]) -> dict[str, Any]:
    changes: dict[str, Any] = {}
    for field in ("talla_camiseta", "talla_medias"):
        if nonempty(row.get(field)):
            changes[field] = row[field]
    current_kit = structured(row.get("segunda_equipacion"))
    if isinstance(current_kit, dict):
        for source, target in {
            "shirt_name": "nombre_camiseta", "number": "dorsal",
            "shirt_size": "talla_camiseta", "socks_size": "talla_medias",
        }.items():
            if nonempty(current_kit.get(source)):
                changes[target] = current_kit[source]
    history = structured(row.get("historial_equipacion"))
    if nonempty(history):
        changes["historial_equipacion"] = history
    return changes


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("backup", type=Path, help="Excel de copia con hoja players")
    parser.add_argument("--apply", action="store_true", help="Aplicar actualización")
    parser.add_argument("--confirm", default="", help="Debe ser RESTORE-EQUIPMENT-FROM-BACKUP")
    args = parser.parse_args()
    if not args.backup.is_file():
        raise SystemExit(f"No existe el Excel de copia: {args.backup}")

    rows = backup_rows(args.backup)
    updates = {str(row["id"]): fields_to_restore(row) for row in rows if row.get("id")}
    updates = {player_id: changes for player_id, changes in updates.items() if changes}

    backend_dir = Path(__file__).resolve().parents[1]
    load_dotenv(backend_dir / ".env")
    client = MongoClient(os.environ["MONGO_URL"], serverSelectionTimeoutMS=5000)
    database = client[os.environ["DB_NAME"]]
    existing_ids = set(database.players.distinct("id", {"id": {"$in": list(updates)}}))
    matched = {player_id: changes for player_id, changes in updates.items() if player_id in existing_ids}
    report = {
        "backup_players": len(rows),
        "players_with_equipment_data": len(updates),
        "players_matched": len(matched),
        "players_not_in_current_database": len(updates) - len(matched),
        "first_shirt_sizes": sum("talla_camiseta" in value for value in matched.values()),
        "first_socks_sizes": sum("talla_medias" in value for value in matched.values()),
        "primary_kit_names": sum("nombre_camiseta" in value for value in matched.values()),
        "primary_bibs": sum("dorsal" in value for value in matched.values()),
        "equipment_history": sum("historial_equipacion" in value for value in matched.values()),
        "main_dorsal_modified": bool(any("dorsal" in value for value in matched.values())),
        "delivery_status_modified": False,
        "dry_run": not args.apply,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if not args.apply:
        client.close()
        return
    if args.confirm != "RESTORE-EQUIPMENT-FROM-BACKUP":
        raise SystemExit("Para aplicar use --confirm RESTORE-EQUIPMENT-FROM-BACKUP")

    timestamp = now_iso()
    changed = 0
    for player_id, changes in matched.items():
        result = database.players.update_one({"id": player_id}, {"$set": {**changes, "updated_at": timestamp}})
        changed += result.modified_count
    database.equipment_restore_audit.insert_one({
        "action": "restore_equipment_from_backup", "at": timestamp,
        "backup": args.backup.name, "players_matched": len(matched), "players_changed": changed,
        "main_dorsal_modified": bool(any("dorsal" in value for value in matched.values())),
        "delivery_status_modified": False,
    })
    print(json.dumps({"applied": True, "players_changed": changed}, ensure_ascii=False))
    client.close()


if __name__ == "__main__":
    main()
