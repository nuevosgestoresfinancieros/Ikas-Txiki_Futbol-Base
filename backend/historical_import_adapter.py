"""Adaptador privado BBDD -> import_staging para la base histórica del club.

El módulo no escribe datos oficiales ni conserva el Excel original. Devuelve
registros normalizados, incidencias y un resumen agregado apto para auditoría.
"""
from __future__ import annotations

import io
import itertools
import re
import unicodedata
import uuid
from collections import defaultdict
from datetime import date, datetime
from difflib import SequenceMatcher
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from inscription_import_service import (
    ImportValidationError, _validate_archive, normalize_birthdate, normalize_key,
    normalize_text, valid_iban,
)
from import_staging_service import modality_suggestion, prepare_records


HISTORICAL_SHEET = "BBDD"
HISTORICAL_FORMAT = "historical_bbdd_v1"
OPERATIONAL_COLUMN_COUNT = 79
EXPECTED_TOTAL_COLUMNS = 129
ALLOWED_FORMULA_COLUMNS = {70, 71, 72}
COMPACT_TOTAL_COLUMNS = 64
COMPACT_ALLOWED_FORMULA_COLUMNS = {57, 58}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PHONE_RE = re.compile(r"^\d{9,15}$")

# Posición, cabecera original y destino semántico. Las columnas 80-129 quedan
# clasificadas como auxiliares y nunca se transforman en entidades oficiales.
HISTORICAL_COLUMN_MAPPING = (
    (1, "NOMBRE", "player.first_name"),
    (2, "APELLIDOS", "player.last_name"),
    (3, "FECHA DE NACIMIENTO", "player.birth_date"),
    (4, "AÑO DE NACIMIENTO", "player.birth_year_reference"),
    (5, "CATEGORÍA", "registration.category"),
    (6, "CENTRO DE PROCEDENCIA", "player.school_origin"),
    (7, "NOMBRE DEL PADRE", "contacts.father.name"),
    (8, "TELÉFONO DEL PADRE", "contacts.father.phone"),
    (9, "NOMBRE DE LA MADRE", "contacts.mother.name"),
    (10, "TELÉFONO DE LA MADRE", "contacts.mother.phone"),
    (11, "DIRECCIÓN", "family.address_candidate"),
    (12, "CORREO ELECTRÓNICO AITA", "contacts.father.email"),
    (13, "CORREO ELECTRÓNICO AMA", "contacts.mother.email"),
    (14, "RECIBIR NOTIFICACIONES", "consents.notifications"),
    (15, "TALLA 16&17", "equipment_history.2016-2017.shirt_size"),
    (16, "TALLA 17&18", "equipment_history.2017-2018.shirt_size"),
    (17, "NOMBRE CAMISETA 18&19", "equipment_history.2018-2019.shirt_name"),
    (18, "Dorsal 18&19", "equipment_history.2018-2019.number"),
    (19, "NOMBRE CAMISETA 19&20", "equipment_history.2019-2020.shirt_name"),
    (20, "Dorsal 19&20", "equipment_history.2019-2020.number"),
    (21, "TALLA 19&20", "equipment_history.2019-2020.shirt_size"),
    (22, "TALLA MEDIAS 19&20", "equipment_history.2019-2020.socks_size"),
    (23, "NOMBRE CAMISETA 20&21", "equipment_history.2020-2021.shirt_name"),
    (24, "DORSAL 20&21", "equipment_history.2020-2021.number"),
    (25, "TALLA 20&21", "equipment_history.2020-2021.shirt_size"),
    (26, "NOMBRE CAMISETA 21&22", "equipment_history.2021-2022.shirt_name"),
    (27, "DORSAL 21&22", "equipment_history.2021-2022.number"),
    (28, "TALLA 21&22", "equipment_history.2021-2022.shirt_size"),
    (29, "TALLA MEDIAS 21&22", "equipment_history.2021-2022.socks_size"),
    (30, "NOMBRE CAMISETA 22&23", "equipment_history.2022-2023.shirt_name"),
    (31, "DORSAL 22&23", "equipment_history.2022-2023.number"),
    (32, "TALLA 22&23", "equipment_history.2022-2023.shirt_size"),
    (33, "TALLA MEDIAS 22&23", "equipment_history.2022-2023.socks_size"),
    (34, "NOMBRE CAMISETA 23&24", "equipment_history.2023-2024.shirt_name"),
    (35, "DORSAL 23&24", "equipment_history.2023-2024.number"),
    (36, "TALLA 23&24", "equipment_history.2023-2024.shirt_size"),
    (37, "TALLA MEDIAS 23&24", "equipment_history.2023-2024.socks_size"),
    (38, "NOMBRE CAMISETA 24&25", "equipment_history.2024-2025.shirt_name"),
    (39, "DORSAL 24&25", "equipment_history.2024-2025.number"),
    (40, "TALLA 24&25", "equipment_history.2024-2025.shirt_size"),
    (41, "TALLA MEDIAS 24&25", "equipment_history.2024-2025.socks_size"),
    (42, "NOMBRE CAMISETA 25&26", "equipment_history.2025-2026.shirt_name"),
    (43, "DORSAL 25&26", "equipment_history.2025-2026.number"),
    (44, "TALLA 25&26", "equipment_history.2025-2026.shirt_size"),
    (45, "TALLA MEDIAS 25&26", "equipment_history.2025-2026.socks_size"),
    (46, "NOMBRE CAMISETA 26&27", "equipment.2026-2027.shirt_name"),
    (47, "DORSAL 26&27", "equipment.2026-2027.number"),
    (48, "TALLA 26&27", "equipment.2026-2027.shirt_size"),
    (49, "TALLA MEDIAS 26&27", "equipment.2026-2027.socks_size"),
    (50, "Eskola/futbolfederado 25/26", "sport_history.2025-2026.program"),
    (51, "EQUIPO 21&22", "team_history.2021-2022"),
    (52, "EQUIPO 23&24", "team_history.2023-2024"),
    (53, "EQUIPO 24&25", "team_history.2024-2025"),
    (54, "EQUIPO 25&26", "team_history.2025-2026"),
    (55, "EQUIPO 26&27", "registration.team_2026_2027"),
    (56, "CATEGORÍA DE JUEGO", "sport_history.playing_category"),
    (57, "FEDERADO 18&19", "federation_history.2018-2019"),
    (58, "FEDERADO 19&20", "federation_history.2019-2020"),
    (59, "FEDERADO 20&21", "federation_history.2020-2021"),
    (60, "FEDERADO 21&22", "federation_history.2021-2022"),
    (61, "FEDERADO 22&23", "federation_history.2022-2023"),
    (62, "FEDERADO 24&25", "federation_history.2024-2025"),
    (63, "FEDERADO 25&26", "federation_history.2025-2026"),
    (64, "FEDERADO 26&27", "federation_history.2026-2027"),
    (65, "DEMARCACIÓN", "sport_history.position"),
    (66, "TITULAR NÚMERO DE CUENTA", "bank.account_holder"),
    (67, "NÚMERO DE CUENTA", "bank.iban_candidate"),
    (68, "PERMISO PARA COBRAR DEL NÚMERO DE CUENTA", "consents.debit_permission"),
    (69, "Tipo cuota integra (I), fraccionada (f)", "fees.payment_schedule"),
    (70, "CUOTA A PAGAR", "fees.amount_due_unconfirmed"),
    (71, "CUOTA PAGADA", "fees.amount_paid_reference"),
    (72, "CUOTA PENDIENTE DE PAGO", "fees.balance_reference_unconfirmed"),
    (73, "Entrenamiento lunes 25/26", "schedule_history.2025-2026.monday"),
    (74, "entrenamiento martes 25/26", "schedule_history.2025-2026.tuesday"),
    (75, "entrenamiento miércoles 25/26", "schedule_history.2025-2026.wednesday"),
    (76, "entrenamiento jueves 25/26", "schedule_history.2025-2026.thursday"),
    (77, "entrenamiento viernes 25/26", "schedule_history.2025-2026.friday"),
    (78, "BESTE DATU INTERESGARRIAK/ OTROS DATOS DE INTERÉS", "sensitive_quarantine.other_notes"),
    (79, "PERMISO PARA PUBLICAR IMÁGENES", "consents.image_permission"),
)

# Exportación compacta recibida a partir de diciembre de 2025. Conserva el
# histórico deportivo/equipación y los datos económicos, pero ya no incluye
# nacimiento, contactos ni domicilio. No se inventan esos datos ausentes.
COMPACT_HEADERS = (
    "NOMBRE", "APELLIDOS", "TALLA 16&17", "TALLA 17&18",
    "NOMBRE CAMISETA 18&19", "Dorsal 18&19", "NOMBRE CAMISETA 19&20", "Dorsal 19&20",
    "TALLA 19&20", "TALLA MEDIAS 19&20", "NOMBRE CAMISETA 20&21", "DORSAL 20&21",
    "TALLA 20&21", "NOMBRE CAMISETA 21&22", "DORSAL 21&22", "TALLA 21&22",
    "TALLA MEDIAS 21&22", "NOMBRE CAMISETA 22&23", "DORSAL 22&23", "TALLA 22&23",
    "TALLA MEDIAS 22&23", "NOMBRE CAMISETA 23&24", "DORSAL 23&24", "TALLA 23&24",
    "TALLA MEDIAS 23&24", "NOMBRE CAMISETA 24&25", "DORSAL 24&25", "TALLA 24&25",
    "TALLA MEDIAS 24&25", "NOMBRE CAMISETA 25&26", "DORSAL 25&26", "TALLA 25&26",
    "TALLA MEDIAS 25&26", "SEGUNDA EQUIPACION NOMBRE CAMISETA 26&27",
    "SEGUNDA EQUIPACION DORSAL 26&27", "SEGUNDA EQUIPACION TALLA 26&27",
    "SEGUNDA EQUIPACION TALLA MEDIAS 26&27", "Eskola/futbolfederado 25/26",
    "EQUIPO 21&22", "EQUIPO 23&24", "EQUIPO 24&25", "EQUIPO 25&26", "EQUIPO 26&27",
    "CATEGORÍA DE JUEGO", "FEDERADO 18&19", "FEDERADO 19&20", "FEDERADO 20&21",
    "FEDERADO 21&22", "FEDERADO 22&23", "FEDERADO 24&25", "FEDERADO 25&26",
    "FEDERADO 26&27", "DEMARCACIÓN", "TITULAR NÚMERO DE CUENTA", "NÚMERO DE CUENTA",
    "PERMISO PARA COBRAR DEL NÚMERO DE CUENTA", "CUOTA A PAGAR", "CUOTA PENDIENTE DE PAGO",
    "Entrenamiento lunes 25/26", "entrenamiento martes 25/26", "entrenamiento miércoles 25/26",
    "entrenamiento jueves 25/26", "entrenamiento viernes 25/26", "PERMISO PARA PUBLICAR IMÁGENES",
)


def _norm(value: Any) -> str:
    return " ".join(unicodedata.normalize("NFKC", normalize_text(value)).split())


def _signal(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _norm(value)).encode("ascii", "ignore").decode().casefold()
    return re.sub(r"[^a-z0-9]+", "", text)


def _phone(value: Any) -> str:
    text = _norm(value)
    if not text:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    return re.sub(r"\D", "", text)


def _email(value: Any) -> str:
    return _norm(value).casefold()


def _tri_state(value: Any) -> str:
    key = normalize_key(value)
    if not key:
        return "unanswered"
    if key in {"si", "s", "yes", "bai", "1", "x"}:
        return "yes"
    if key in {"no", "ez", "0"}:
        return "no"
    return "unanswered"


def _number(value: Any) -> int | float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    text = _norm(value).replace("€", "").replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return None


def _seasonal(row: tuple, columns: Mapping[str, tuple[int, ...]]) -> dict:
    result = {}
    for season, indexes in columns.items():
        values = [_norm(row[index - 1]) for index in indexes]
        if any(values):
            result[season] = values[0] if len(values) == 1 else values
    return result


def _identity(record: Mapping[str, Any]) -> str:
    return "|".join((_signal(record.get("nombre")), _signal(record.get("apellidos")), str(record.get("fecha_nacimiento") or "")))


def _fuzzy_pairs(records: list[dict]) -> list[dict]:
    result = []
    for left, right in itertools.combinations(records, 2):
        if left.get("fecha_nacimiento") != right.get("fecha_nacimiento"):
            continue
        first = _signal(left.get("nombre")) + _signal(left.get("apellidos"))
        second = _signal(right.get("nombre")) + _signal(right.get("apellidos"))
        score = SequenceMatcher(None, first, second).ratio()
        if first != second and score >= 0.82:
            result.append({"record_ids": [left["id"], right["id"]], "score": round(score, 2), "decision": None})
    return result


def _family_candidates(records: list[dict]) -> list[dict]:
    keys = ("father_name", "father_phone", "mother_name", "mother_phone", "address", "father_email", "mother_email", "bank_holder", "bank")
    signals = []
    for record in records:
        contacts = record["historical"]["contacts"]
        bank = record["historical"]["bank_source"]
        signals.append({
            "father_name": _signal(contacts["father"]["name"]), "father_phone": _signal(contacts["father"]["phone"]),
            "mother_name": _signal(contacts["mother"]["name"]), "mother_phone": _signal(contacts["mother"]["phone"]),
            "address": _signal(record.get("domicilio")), "father_email": _signal(contacts["father"]["email"]),
            "mother_email": _signal(contacts["mother"]["email"]), "bank_holder": _signal(bank["holder"]),
            "bank": _signal(bank["candidate"]),
        })
    parent = list(range(len(records)))
    def find(value):
        while parent[value] != value:
            parent[value] = parent[parent[value]]; value = parent[value]
        return value
    def union(left, right):
        left, right = find(left), find(right)
        if left != right:
            parent[right] = left
    pair_signals = defaultdict(set)
    for left, right in itertools.combinations(range(len(records)), 2):
        matched = {key for key in keys if signals[left][key] and signals[left][key] == signals[right][key]}
        if len(matched) >= 2:
            union(left, right); pair_signals[(left, right)] = matched
    groups = defaultdict(list)
    for index in range(len(records)):
        groups[find(index)].append(index)
    result = []
    for indexes in groups.values():
        if len(indexes) < 2:
            continue
        evidence = sorted(set().union(*(pair_signals.get(tuple(sorted(pair)), set()) for pair in itertools.combinations(indexes, 2))))
        result.append({"record_ids": [records[index]["id"] for index in indexes], "signals": evidence, "decision": None})
    return result


def _compact_headers_match(sheet) -> bool:
    headers = [_norm(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=COMPACT_TOTAL_COLUMNS))]
    return all(normalize_key(actual) == normalize_key(expected) for actual, expected in zip(headers, COMPACT_HEADERS))


def _parse_compact_sheets(formula_sheet, value_sheet, id_factory) -> dict:
    """Normaliza la exportación compacta sin fabricar identidad o contactos."""
    formula_rows = formula_sheet.iter_rows(min_row=2, max_col=COMPACT_TOTAL_COLUMNS, values_only=True)
    value_rows = value_sheet.iter_rows(min_row=2, max_col=COMPACT_TOTAL_COLUMNS, values_only=True)
    records, auxiliary, blocked_formulas, materialized = [], [], [], []
    for source_row, (formula_row, value_row) in enumerate(zip(formula_rows, value_rows), 2):
        if not any(_norm(value) for value in value_row):
            continue
        if not (_norm(value_row[0]) and _norm(value_row[1])):
            auxiliary.append({"source_row": source_row, "nonempty_operational_cells": sum(bool(_norm(value)) for value in value_row)})
            continue
        formula_cells = []
        for column, formula in enumerate(formula_row, 1):
            if not (isinstance(formula, str) and formula.startswith("=")):
                continue
            cached = value_row[column - 1]
            if column in COMPACT_ALLOWED_FORMULA_COLUMNS and _number(cached) is not None:
                entry = {"source_row": source_row, "column": column, "value": _number(cached)}
                materialized.append(entry)
                formula_cells.append({"column": column, "status": "materialized", "value": entry["value"]})
            else:
                blocked_formulas.append({"source_row": source_row, "column": column})
                formula_cells.append({"column": column, "status": "blocked"})

        equipment_history = _seasonal(value_row, {
            "2016-2017": (3,), "2017-2018": (4,), "2018-2019": (5, 6),
            "2019-2020": (7, 8, 9, 10), "2020-2021": (11, 12, 13),
            "2021-2022": (14, 15, 16, 17), "2022-2023": (18, 19, 20, 21),
            "2023-2024": (22, 23, 24, 25), "2024-2025": (26, 27, 28, 29),
            "2025-2026": (30, 31, 32, 33), "2026-2027-second-kit": (34, 35, 36, 37),
        })
        current_second_kit = {
            "shirt_name": _norm(value_row[33]), "number": _norm(value_row[34]),
            "shirt_size": _norm(value_row[35]), "socks_size": _norm(value_row[36]),
        }
        record = {
            "id": id_factory(), "source_row": source_row,
            "nombre": _norm(value_row[0]), "apellidos": _norm(value_row[1]),
            "fecha_nacimiento": None, "tipo": "renovacion", "centro_escolar": "",
            "progenitor1_nombre": "", "progenitor1_telefono": "", "progenitor1_email": "",
            "progenitor2_nombre": "", "progenitor2_telefono": "", "progenitor2_email": "",
            "domicilio": "", "equipo_anterior": _norm(value_row[41]),
            "equipo": _norm(value_row[42]), "categoria": _norm(value_row[43]),
            "categoria_juego": _norm(value_row[43]), "modalidad": "",
            "equipamiento_items": [], "talla_camiseta": current_second_kit["shirt_size"],
            "talla_medias": current_second_kit["socks_size"], "iban": _norm(value_row[54]),
            "observaciones": "", "_phone_issue_count": 0,
            "_bank_issue": (not valid_iban(_norm(value_row[54]))) or any(
                unicodedata.category(char).startswith("C") for char in str(value_row[54] or "")
            ),
            "historical": {
                "birth_year_reference": "",
                "contacts": {"father": {"name": "", "phone": "", "email": ""},
                             "mother": {"name": "", "phone": "", "email": ""}},
                "equipment_current": current_second_kit,
                "equipment_history": equipment_history,
                "team_history": _seasonal(value_row, {
                    "2021-2022": (39,), "2023-2024": (40,), "2024-2025": (41,), "2025-2026": (42,),
                }),
                "federation_history": _seasonal(value_row, {
                    "2018-2019": (45,), "2019-2020": (46,), "2020-2021": (47,), "2021-2022": (48,),
                    "2022-2023": (49,), "2024-2025": (50,), "2025-2026": (51,), "2026-2027": (52,),
                }),
                "sport": {"program_2025_2026": _norm(value_row[37]),
                          "playing_category": _norm(value_row[43]), "position": _norm(value_row[52])},
                "bank_source": {"holder": _norm(value_row[53]), "candidate": _norm(value_row[54])},
                "fees": {"schedule": "", "due": _number(value_row[56]), "paid": None,
                         "balance_reference": _number(value_row[57]),
                         "source_present": [value_row[56] not in (None, ""), False, value_row[57] not in (None, "")],
                         "source_numeric": [isinstance(value_row[56], (int, float)) and not isinstance(value_row[56], bool),
                                            True, isinstance(value_row[57], (int, float)) and not isinstance(value_row[57], bool)],
                         "confirmed_debt": False},
                "consents": {"notifications": "unanswered", "debit": _tri_state(value_row[55]),
                             "images": _tri_state(value_row[63]), "signed": False},
                "schedule_history": [_norm(value_row[index]) for index in range(58, 63)],
                "sensitive_quarantine": {"other_notes_present": False},
                "formula_values": formula_cells,
            },
        }
        record["import_identity_key"] = _identity(record)
        records.append(record)

    exact_groups = [values for values in _group_by(records, _identity).values() if len(values) > 1]
    return {
        "source_format": HISTORICAL_FORMAT, "source_layout": "compact_64",
        "records": records, "auxiliary_rows": auxiliary, "exact_duplicate_groups": exact_groups,
        "fuzzy_matches": [], "family_candidates": [], "materialized_formulas": materialized,
        "blocked_formulas": blocked_formulas, "auxiliary_column_count": 0,
    }


def parse_historical_excel(content: bytes, id_factory) -> dict:
    _validate_archive(content)
    formula_book = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    value_book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if HISTORICAL_SHEET not in formula_book.sheetnames:
        raise ImportValidationError(f"Falta la hoja histórica '{HISTORICAL_SHEET}'")
    formula_sheet, value_sheet = formula_book[HISTORICAL_SHEET], value_book[HISTORICAL_SHEET]
    # Algunos exports de LibreOffice/Excel declaran una dimensión vacía aunque
    # contengan celdas. Forzamos el cálculo antes de decidir el formato.
    if formula_sheet.max_column is None:
        formula_sheet.reset_dimensions(); value_sheet.reset_dimensions()
        formula_sheet.calculate_dimension(force=True); value_sheet.calculate_dimension(force=True)
    if (formula_sheet.max_column or 0) >= COMPACT_TOTAL_COLUMNS and _compact_headers_match(formula_sheet):
        result = _parse_compact_sheets(formula_sheet, value_sheet, id_factory)
        formula_book.close(); value_book.close()
        return result
    if formula_sheet.max_column < EXPECTED_TOTAL_COLUMNS:
        raise ImportValidationError("La base histórica no coincide con los formatos verificados de 64 o 129 columnas")
    headers = [_norm(cell.value) for cell in next(formula_sheet.iter_rows(min_row=1, max_row=1, max_col=EXPECTED_TOTAL_COLUMNS))]
    for position, expected, _ in HISTORICAL_COLUMN_MAPPING:
        if normalize_key(headers[position - 1]) != normalize_key(expected):
            raise ImportValidationError(f"La columna histórica {position} no coincide con el formato verificado")
    formula_rows = formula_sheet.iter_rows(min_row=2, max_col=EXPECTED_TOTAL_COLUMNS, values_only=True)
    value_rows = value_sheet.iter_rows(min_row=2, max_col=EXPECTED_TOTAL_COLUMNS, values_only=True)
    records, auxiliary, blocked_formulas, materialized = [], [], [], []
    for source_row, (formula_row, value_row) in enumerate(zip(formula_rows, value_rows), 2):
        if not any(_norm(value) for value in value_row[:OPERATIONAL_COLUMN_COUNT]):
            continue
        formula_cells = []
        for column, formula in enumerate(formula_row, 1):
            if not (isinstance(formula, str) and formula.startswith("=")):
                continue
            if column not in ALLOWED_FORMULA_COLUMNS:
                blocked_formulas.append({"source_row": source_row, "column": column})
                formula_cells.append({"column": column, "status": "blocked"})
            else:
                cached = value_row[column - 1]
                if _number(cached) is None:
                    blocked_formulas.append({"source_row": source_row, "column": column})
                    formula_cells.append({"column": column, "status": "blocked_no_numeric_cache"})
                else:
                    entry = {"source_row": source_row, "column": column, "value": _number(cached)}
                    materialized.append(entry); formula_cells.append({"column": column, "status": "materialized", "value": entry["value"]})
        complete_identity = _norm(value_row[0]) and _norm(value_row[1]) and value_row[2] not in (None, "")
        if not complete_identity:
            auxiliary.append({"source_row": source_row, "nonempty_operational_cells": sum(bool(_norm(value)) for value in value_row[:79])})
            continue
        birth_date = normalize_birthdate(value_row[2])
        equipment_current = {
            "shirt_name": _norm(value_row[45]), "number": _norm(value_row[46]),
            "shirt_size": _norm(value_row[47]), "socks_size": _norm(value_row[48]),
        }
        record = {
            "id": id_factory(), "source_row": source_row, "nombre": _norm(value_row[0]),
            "apellidos": _norm(value_row[1]), "fecha_nacimiento": birth_date,
            "tipo": "renovacion", "centro_escolar": _norm(value_row[5]),
            "progenitor1_nombre": _norm(value_row[6]), "progenitor1_telefono": _phone(value_row[7]),
            "progenitor1_email": _email(value_row[11]), "progenitor2_nombre": _norm(value_row[8]),
            "progenitor2_telefono": _phone(value_row[9]), "progenitor2_email": _email(value_row[12]),
            "domicilio": _norm(value_row[10]), "equipo_anterior": _norm(value_row[53]),
            "equipo": _norm(value_row[54]), "categoria": _norm(value_row[4]), "categoria_juego": _norm(value_row[55]),
            "modalidad": "", "equipamiento_items": [], "talla_camiseta": equipment_current["shirt_size"],
            "talla_medias": equipment_current["socks_size"], "iban": _norm(value_row[66]),
            "observaciones": "", "_phone_issue_count": sum(
                raw not in (None, "") and not PHONE_RE.fullmatch(_phone(raw)) for raw in (value_row[7], value_row[9])
            ),
            "_bank_issue": (not valid_iban(_norm(value_row[66]))) or any(
                unicodedata.category(char).startswith("C") for char in str(value_row[66] or "")
            ),
            "historical": {
                "birth_year_reference": _norm(value_row[3]),
                "contacts": {"father": {"name": _norm(value_row[6]), "phone": _phone(value_row[7]), "email": _email(value_row[11])},
                             "mother": {"name": _norm(value_row[8]), "phone": _phone(value_row[9]), "email": _email(value_row[12])}},
                "equipment_current": equipment_current,
                "equipment_history": _seasonal(value_row, {
                    "2016-2017": (15,), "2017-2018": (16,), "2018-2019": (17,18), "2019-2020": (19,20,21,22),
                    "2020-2021": (23,24,25), "2021-2022": (26,27,28,29), "2022-2023": (30,31,32,33),
                    "2023-2024": (34,35,36,37), "2024-2025": (38,39,40,41), "2025-2026": (42,43,44,45),
                }),
                "team_history": _seasonal(value_row, {"2021-2022": (51,), "2023-2024": (52,), "2024-2025": (53,), "2025-2026": (54,)}),
                "federation_history": _seasonal(value_row, {"2018-2019": (57,), "2019-2020": (58,), "2020-2021": (59,), "2021-2022": (60,), "2022-2023": (61,), "2024-2025": (62,), "2025-2026": (63,), "2026-2027": (64,)}),
                "sport": {"program_2025_2026": _norm(value_row[49]), "playing_category": _norm(value_row[55]), "position": _norm(value_row[64])},
                "bank_source": {"holder": _norm(value_row[65]), "candidate": _norm(value_row[66])},
                "fees": {"schedule": _norm(value_row[68]), "due": _number(value_row[69]), "paid": _number(value_row[70]), "balance_reference": _number(value_row[71]),
                         "source_present": [value_row[index] not in (None, "") for index in (69, 70, 71)],
                         "source_numeric": [isinstance(value_row[index], (int, float)) and not isinstance(value_row[index], bool) for index in (69, 70, 71)],
                         "confirmed_debt": False},
                "consents": {"notifications": _tri_state(value_row[13]), "debit": _tri_state(value_row[67]), "images": _tri_state(value_row[78]), "signed": False},
                "schedule_history": [_norm(value_row[index]) for index in range(72, 77)],
                "sensitive_quarantine": {"other_notes_present": bool(_norm(value_row[77]))},
                "formula_values": formula_cells,
            },
        }
        record["import_identity_key"] = _identity(record)
        records.append(record)
    formula_book.close(); value_book.close()
    exact_groups = [values for values in _group_by(records, _identity).values() if len(values) > 1]
    fuzzy_matches = _fuzzy_pairs(records)
    family_candidates = _family_candidates(records)
    for item in fuzzy_matches + family_candidates:
        item["id"] = id_factory()
    result = {
        "source_format": HISTORICAL_FORMAT, "source_layout": "legacy_129", "records": records, "auxiliary_rows": auxiliary,
        "exact_duplicate_groups": exact_groups, "fuzzy_matches": fuzzy_matches,
        "family_candidates": family_candidates, "materialized_formulas": materialized,
        "blocked_formulas": blocked_formulas, "auxiliary_column_count": EXPECTED_TOTAL_COLUMNS - OPERATIONAL_COLUMN_COUNT,
    }
    return result


def _group_by(records: Iterable[dict], key_function) -> dict:
    groups = defaultdict(list)
    for record in records:
        groups[key_function(record)].append(record["id"])
    return groups


def historical_quality_summary(parsed: Mapping[str, Any]) -> dict:
    records = parsed["records"]
    invalid_emails = sum(
        bool(value) and not EMAIL_RE.fullmatch(value)
        for record in records for value in (record.get("progenitor1_email", ""), record.get("progenitor2_email", ""))
    )
    invalid_phones = sum(record.get("_phone_issue_count", 0) for record in records)
    invalid_ibans = sum(bool(record.get("_bank_issue")) for record in records)
    nonnumeric_fees = 0
    for record in records:
        # A non-empty economic source with no materialized numeric value is quarantined.
        fees = record["historical"]["fees"]
        values = (fees["due"], fees["paid"], fees["balance_reference"])
        nonnumeric_fees += sum(present and not numeric for present, numeric in zip(fees["source_present"], fees["source_numeric"]))
    consents = {}
    for kind in ("notifications", "debit", "images"):
        consents[kind] = {state: sum(record["historical"]["consents"][kind] == state for record in records) for state in ("yes", "no", "unanswered")}
    suggested_counts = defaultdict(int)
    for record in records:
        suggestion = modality_suggestion(record.get("categoria"))
        if record.get("equipo") and suggestion:
            suggested_counts[(record["equipo"], suggestion)] += 1
    suggested_capacities = [
        {"team": team, "modality_suggestion": modality, "count": count,
         "limit": 18 if modality == "F7" else 25, "over_capacity": count > (18 if modality == "F7" else 25)}
        for (team, modality), count in sorted(suggested_counts.items())
    ]
    return {
        "identified_players": len(records), "with_current_team": sum(bool(record.get("equipo")) for record in records),
        "without_current_team": sum(not record.get("equipo") for record in records),
        "auxiliary_rows_excluded": len(parsed["auxiliary_rows"]), "exact_duplicate_groups": len(parsed["exact_duplicate_groups"]),
        "fuzzy_match_pairs": len(parsed["fuzzy_matches"]), "family_candidate_groups": len(parsed["family_candidates"]),
        "players_in_family_candidates": sum(len(group["record_ids"]) for group in parsed["family_candidates"]),
        "invalid_email_cells": invalid_emails, "invalid_phone_cells": invalid_phones,
        "invalid_iban_cells": invalid_ibans, "materialized_formula_cells": len(parsed["materialized_formulas"]),
        "formula_player_rows": len({item["source_row"] for item in parsed["materialized_formulas"]} & {record["source_row"] for record in records}),
        "blocked_formula_cells": len(parsed["blocked_formulas"]), "nonnumeric_fee_values": nonnumeric_fees,
        "modality_pending": len(records), "suggested_capacity_teams": len(suggested_capacities),
        "suggested_teams_over_capacity": sum(item["over_capacity"] for item in suggested_capacities),
        "suggested_capacities": suggested_capacities, "consents": consents,
    }


def historical_simulation(parsed: Mapping[str, Any], existing_players: Iterable[Mapping[str, Any]] = ()) -> dict:
    compact = parsed.get("source_layout") == "compact_64"
    if compact:
        existing_by_name = defaultdict(list)
        for player in existing_players:
            existing_by_name[(_signal(player.get("nombre")), _signal(player.get("apellidos")))].append(player)
        exact = sum(
            len(existing_by_name[(_signal(record.get("nombre")), _signal(record.get("apellidos")))]) == 1
            for record in parsed["records"]
        )
    else:
        existing = {_identity(player): player for player in existing_players}
        exact = sum(record["import_identity_key"] in existing for record in parsed["records"])
    quality = historical_quality_summary(parsed)
    return {
        "mode": "simulation_only", "official_writes": 0,
        # El compacto no tiene fecha de nacimiento: solo puede enriquecer una
        # coincidencia nominal única y nunca debe proponer altas automáticas.
        "proposed_creates": 0 if compact else len(parsed["records"]) - exact,
        "profile_enrichment_candidates": exact if compact else 0,
        "unmatched_profile_rows": len(parsed["records"]) - exact if compact else 0,
        "exact_matches": exact,
        "fuzzy_matches": quality["fuzzy_match_pairs"], "family_candidates": quality["family_candidate_groups"],
        "conflicts": quality["fuzzy_match_pairs"] + quality["family_candidate_groups"],
        "blocked_records": sum(
            not record.get("equipo") or not record.get("modalidad")
            for record in parsed["records"]
        ),
        "omitted_auxiliary_rows": quality["auxiliary_rows_excluded"],
        "with_current_team": quality["with_current_team"], "without_current_team": quality["without_current_team"],
        "modality_pending": quality["modality_pending"],
        "suggested_capacity_teams": quality["suggested_capacity_teams"],
        "suggested_teams_over_capacity": quality["suggested_teams_over_capacity"],
        "data_quality": {key: quality[key] for key in ("invalid_email_cells", "invalid_phone_cells", "invalid_iban_cells", "nonnumeric_fee_values")},
        "formulas": {"materialized": quality["materialized_formula_cells"], "blocked": quality["blocked_formula_cells"]},
        "consents": quality["consents"],
        "planned_operations": {
            "players": quality["identified_players"], "registration_candidates": quality["with_current_team"],
            "team_assignments": quality["with_current_team"], "family_links_before_review": 0,
            "sport_histories": quality["identified_players"], "equipment_histories": quality["identified_players"],
            "consent_references": quality["identified_players"], "fee_references_unconfirmed": quality["identified_players"],
            "protected_bank_candidates": quality["identified_players"] - quality["invalid_iban_cells"],
            "official_payments": 0, "official_debts": 0, "official_receipts": 0, "official_remittances": 0,
        },
    }


def prepare_historical_staging(parsed: Mapping[str, Any], secret: str) -> tuple[list[dict], list[dict], list[dict]]:
    """Convierte el resultado histórico al formato de staging sin IBAN en claro."""
    rows = []
    by_source_row = {}
    for source in parsed["records"]:
        row = {key: value for key, value in source.items() if key != "historical"}
        row.update({"_row": source["source_row"], "_staging_id": source["id"]})
        rows.append(row); by_source_row[source["source_row"]] = source
    records, duplicates, incidents = prepare_records(rows, secret)
    for record in records:
        source = by_source_row[record["source_row"]]
        historical = dict(source["historical"])
        # El IBAN en claro nunca entra en staging. Conservamos únicamente el
        # titular para que la referencia bancaria cifrada sea identificable.
        bank_source = historical.pop("bank_source", {}) or {}
        historical["bank_reference"] = {"holder": bank_source.get("holder") or ""}
        record["historical"] = historical
        record["categoria_juego"] = source.get("categoria_juego")
        record["modality_suggestion"] = None if not source.get("categoria") else record.get("modality_suggestion")
        fees = record["historical"]["fees"]
        if any(present and not numeric for present, numeric in zip(fees["source_present"], fees["source_numeric"])):
            incidents.append({
                "id": str(uuid.uuid4()), "record_id": record["id"], "source_row": record["source_row"],
                "field": "fees", "code": "nonnumeric_fee_reference", "blocking": False, "resolution": "pending",
            })
    record_ids = {record["source_row"]: record["id"] for record in records}
    for blocked in parsed.get("blocked_formulas", []):
        if blocked["source_row"] in record_ids:
            incidents.append({
                "id": str(uuid.uuid4()), "record_id": record_ids[blocked["source_row"]],
                "source_row": blocked["source_row"], "field": "formula",
                "code": "formula_not_allowed", "blocking": True, "resolution": "pending",
            })
    return records, duplicates, incidents
