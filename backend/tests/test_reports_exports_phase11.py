import asyncio
import io
from datetime import datetime, timezone
from unittest.mock import AsyncMock

import pytest
from fastapi import HTTPException, Request
from openpyxl import load_workbook

import server
from authz import current_user_context, has_permission, route_permission
from modality_service import catalog_from_settings
from report_export_service import excel_safe, generate_pdf, generate_xlsx, safe_filename
from report_service import MAX_EXPORT_ROWS, REPORTS, ReportValidationError, build_report, enforce_export_limit


TEAMS = [{"id": "team-a", "nombre": "Talde Fiktiboa", "categoria": "Alevín", "modalidad": "F7", "temporada": "2026-2027"}]
PLAYERS = [
    {"id": "player-a", "nombre": "Ane", "apellidos": "Fiktiboa", "equipo_id": "team-a", "categoria": "Alevín", "dorsal": 7, "estado": "activo"},
    {"id": "player-b", "nombre": "=DANGEROUS", "apellidos": None, "equipo_id": "team-a", "categoria": "Alevín", "dorsal": None, "estado": "activo"},
]
TRAININGS = [{"id": "training-a", "fecha": "2026-09-01", "equipo_id": "team-a", "asistencia": [
    {"player_id": "player-a", "estado": "presente"}, {"player_id": "player-b", "estado": "justificada"},
]}]
CONTEXT = {"players": PLAYERS, "teams": TEAMS, "trainings": TRAININGS, "modalities": catalog_from_settings()}
OPTIONS = {
    "teams": [{"id": "team-a", "name": "Talde Fiktiboa"}],
    "players": [{"id": "player-a", "name": "Ane Fiktiboa"}, {"id": "player-b", "name": "Formula Fiktiboa"}],
}
BRANDING = {"club_nombre": "Ikas-Txiki Proba"}


def run(awaitable):
    return asyncio.run(awaitable)


def response_bytes(response):
    async def collect():
        return b"".join([chunk async for chunk in response.body_iterator])
    return run(collect())


@pytest.mark.parametrize("role", ["admin", "coordinator", "coach", "family", "player"])
@pytest.mark.parametrize("suffix", ["pdf", "xlsx"])
def test_export_routes_use_export_permission_for_all_report_roles(role, suffix):
    request = Request({"type": "http", "method": "POST", "path": f"/api/reports/export.{suffix}",
                       "query_string": b"", "headers": [], "client": ("test", 1)})
    assert route_permission(request) == ("reports", "export")
    assert has_permission({"role": role, "active": True}, "reports", "export")


@pytest.mark.parametrize("report_id", ["roster", "attendance"])
def test_export_uses_exact_same_rows_totals_and_order_as_preview(report_id):
    definition, rows, totals = build_report(report_id, CONTEXT, {}, "admin")
    assert definition["columns"] == REPORTS[report_id]["columns"]
    assert rows == sorted(rows, key=lambda row: (str(row.get("surname") or row.get("name") or "").casefold()))
    assert all(set(row) <= set(definition["columns"]) for row in rows)
    assert totals


@pytest.mark.parametrize("lang,expected", [("es", "Generado"), ("eu", "Sortua")])
@pytest.mark.parametrize("report_id", ["roster", "attendance"])
def test_pdf_is_vector_searchable_bilingual_and_handles_missing_values(report_id, lang, expected):
    definition, rows, totals = build_report(report_id, CONTEXT, {}, "admin")
    content = generate_pdf(definition, rows, totals, {}, OPTIONS, BRANDING, lang,
                           datetime(2026, 7, 22, 12, 0, tzinfo=timezone.utc))
    assert content.startswith(b"%PDF-") and len(content) > 2000
    assert b"/Font" in content and b"/Subtype /Type1" in content
    # No raster image object: ReportLab lays out the report as text and vector shapes.
    assert b"/Subtype /Image" not in content
    assert expected in ("Generado" if lang == "es" else "Sortua")


def test_pdf_supports_multiple_pages_without_rasterizing():
    definition = {key: value for key, value in REPORTS["roster"].items() if key != "roles"}
    rows = [{"name": f"Jokalari {index}", "surname": "Fiktiboa", "team": "Taldea", "category": "Alevín",
             "modality": "F7", "number": index, "status": "activo"} for index in range(240)]
    content = generate_pdf(definition, rows, {"players": 240, "teams": 1}, {}, OPTIONS, BRANDING)
    assert content.count(b"/Type /Page") >= 3


@pytest.mark.parametrize("report_id", ["roster", "attendance"])
@pytest.mark.parametrize("lang", ["es", "eu"])
def test_xlsx_is_structured_visible_and_contains_no_formulas(report_id, lang):
    definition, rows, totals = build_report(report_id, CONTEXT, {}, "admin")
    content = generate_xlsx(definition, rows, totals, {}, OPTIONS, BRANDING, lang,
                            datetime(2026, 7, 22, 12, 0, tzinfo=timezone.utc))
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    assert len(workbook.sheetnames) == 1 and workbook.active.sheet_state == "visible"
    assert len(workbook.active.title) <= 31
    assert workbook.active.auto_filter.ref and workbook.active.freeze_panes
    assert workbook.active.page_setup.orientation == "landscape"
    assert workbook.active.page_setup.fitToWidth == 1 and workbook.active.print_area
    assert not workbook._external_links
    formulas = [cell.value for row in workbook.active.iter_rows() for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")]
    assert formulas == []
    if report_id == "roster":
        assert any(cell.value == "'=DANGEROUS" for row in workbook.active.iter_rows() for cell in row)
        assert any(cell.value == "—" for row in workbook.active.iter_rows() for cell in row)


@pytest.mark.parametrize("value", ["=SUM(A1:A2)", "+1", "-2", "@cmd"])
def test_excel_formula_injection_is_neutralized(value):
    assert excel_safe(value).startswith("'")


def test_empty_exports_are_valid_and_limits_are_enforced():
    definition = {key: value for key, value in REPORTS["roster"].items() if key != "roles"}
    assert generate_pdf(definition, [], {"players": 0, "teams": 0}, {}, OPTIONS, BRANDING).startswith(b"%PDF-")
    workbook = load_workbook(io.BytesIO(generate_xlsx(definition, [], {"players": 0, "teams": 0}, {}, OPTIONS, BRANDING)))
    assert any(cell.value == "Sin resultados" for row in workbook.active.iter_rows() for cell in row)
    assert len(enforce_export_limit([{}] * MAX_EXPORT_ROWS)) == MAX_EXPORT_ROWS
    with pytest.raises(ReportValidationError):
        enforce_export_limit([{}] * (MAX_EXPORT_ROWS + 1))


def test_safe_filename_has_no_identifiers_or_path_segments():
    name = safe_filename("../Roster privado", "eu", "xlsx", datetime(2026, 7, 22, tzinfo=timezone.utc))
    assert name == "ikas-txiki_roster-privado_eu_20260722.xlsx"
    assert "/" not in name and "\\" not in name


def test_export_endpoint_returns_safe_headers_and_content(monkeypatch):
    definition, rows, totals = build_report("roster", CONTEXT, {}, "admin")
    monkeypatch.setattr(server, "prepare_report", AsyncMock(return_value={
        "report": definition, "rows": rows, "totals": totals, "filters": {}, "filter_options": OPTIONS,
    }))
    monkeypatch.setattr(server, "report_branding", AsyncMock(return_value=BRANDING))
    token = current_user_context.set({"id": "admin", "role": "admin", "active": True})
    try:
        pdf_response = run(server.export_professional_pdf(server.ReportExportRequest(report_id="roster", lang="es")))
        xlsx_response = run(server.export_professional_xlsx(server.ReportExportRequest(report_id="roster", lang="eu")))
    finally:
        current_user_context.reset(token)
    assert pdf_response.media_type == "application/pdf"
    assert pdf_response.headers["cache-control"] == "no-store"
    assert pdf_response.headers["content-disposition"].endswith('.pdf"')
    assert response_bytes(pdf_response).startswith(b"%PDF-")
    assert xlsx_response.media_type.endswith("spreadsheetml.sheet")
    assert response_bytes(xlsx_response).startswith(b"PK")


def test_export_endpoint_returns_413_above_safe_limit(monkeypatch):
    definition = {key: value for key, value in REPORTS["roster"].items() if key != "roles"}
    monkeypatch.setattr(server, "prepare_report", AsyncMock(return_value={
        "report": definition, "rows": [{}] * (MAX_EXPORT_ROWS + 1), "totals": {},
        "filters": {}, "filter_options": OPTIONS,
    }))
    with pytest.raises(HTTPException) as error:
        run(server.export_professional_pdf(server.ReportExportRequest(report_id="roster", lang="es")))
    assert error.value.status_code == 413


def test_out_of_scope_filter_is_rejected_before_rows_or_totals(monkeypatch):
    monkeypatch.setattr(server, "report_context", AsyncMock(return_value={
        "players": [PLAYERS[0]], "teams": TEAMS, "trainings": TRAININGS, "modalities": catalog_from_settings(),
    }))
    token = current_user_context.set({"id": "family", "role": "family", "active": True, "family_id": "family-a"})
    try:
        with pytest.raises(HTTPException) as error:
            run(server.export_professional_pdf(server.ReportExportRequest(
                report_id="attendance", filters={"player_id": "manipulated-player"}, lang="es")))
    finally:
        current_user_context.reset(token)
    assert error.value.status_code == 403


def test_export_payload_has_no_sensitive_or_technical_fields():
    _, rows, _ = build_report("roster", {
        **CONTEXT,
        "players": [{**PLAYERS[0], "telefono": "forbidden", "email": "forbidden", "dni": "forbidden",
                     "iban": "forbidden", "salud": "forbidden", "documentos": "forbidden", "token": "forbidden"}],
    }, {}, "admin")
    serialized = str(rows).casefold()
    for forbidden in ("telefono", "email", "dni", "iban", "salud", "documentos", "token", "_id"):
        assert forbidden not in serialized
