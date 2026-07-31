"""Análisis seguro y normalización de importaciones de inscripciones Excel."""
from __future__ import annotations

import base64
import hashlib
import io
import json
import re
import unicodedata
import zipfile
import zlib
from datetime import date, datetime, timezone
from typing import Any, Iterable, Mapping, Optional

from cryptography.fernet import Fernet, InvalidToken
from openpyxl import load_workbook


SEASON = "2026-2027"
MAX_FILE_SIZE = 10 * 1024 * 1024
MAX_UNCOMPRESSED_SIZE = 50 * 1024 * 1024
TEMPLATE_SHEET = "Inscripciones"
VALID_MODALITIES = {"F7", "F11"}
VALID_INSCRIPTION_TYPES = {"alta", "renovacion"}
SERIOUS_CODES = {
    "missing_required", "invalid_birthdate", "invalid_modality", "invalid_iban",
    "forbidden_team_column", "formula_not_allowed", "invalid_file",
}
MAX_ROWS = 5000

HEADER_ALIASES = {
    "id_externo": "external_id",
    "nombre": "nombre",
    "apellidos": "apellidos",
    "fecha_nacimiento": "fecha_nacimiento",
    "tipo_inscripcion": "tipo",
    "email_formulario": "email_formulario",
    "centro_escolar": "centro_escolar",
    "contacto_1_nombre": "progenitor1_nombre",
    "contacto_1_telefono": "progenitor1_telefono",
    "contacto_1_email": "progenitor1_email",
    "contacto_2_nombre": "progenitor2_nombre",
    "contacto_2_telefono": "progenitor2_telefono",
    "contacto_2_email": "progenitor2_email",
    "domicilio": "domicilio",
    "equipo_anterior": "equipo_anterior",
    "equipo_26_27": "equipo",
    "categoria": "categoria",
    "modalidad": "modalidad",
    "equipamiento": "equipamiento",
    "talla_camiseta": "talla_camiseta",
    "talla_pantalon": "talla_pantalon",
    "talla_chandal": "talla_chandal",
    "talla_medias": "talla_medias",
    "talla_calzado": "talla_calzado",
    "iban": "iban",
    "observaciones": "observaciones",
    "pendiente_octubre": "pendiente_octubre",
}
REQUIRED_FIELDS = {"nombre", "apellidos", "fecha_nacimiento", "equipo", "categoria", "modalidad"}
SAFE_PLAYER_FIELDS = {
    "nombre", "apellidos", "fecha_nacimiento", "email_formulario", "centro_escolar",
    "progenitor1_nombre", "progenitor1_telefono", "progenitor1_email",
    "progenitor2_nombre", "progenitor2_telefono", "progenitor2_email", "domicilio",
    "categoria", "modalidad", "talla_camiseta", "talla_pantalon", "talla_chandal",
    "talla_medias", "talla_calzado", "observaciones",
}


class ImportValidationError(ValueError):
    pass


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return str(value).strip()


def normalize_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", normalize_text(value)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def normalize_email(value: Any) -> str:
    return normalize_text(value).lower()


def normalize_phone(value: Any) -> str:
    return re.sub(r"\D", "", normalize_text(value))


def normalize_birthdate(value: Any) -> Optional[str]:
    text = normalize_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def normalize_iban(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value)).upper()


def valid_iban(value: str) -> bool:
    iban = normalize_iban(value)
    if not iban:
        return True
    if not re.fullmatch(r"ES\d{22}", iban):
        return False
    rearranged = iban[4:] + iban[:4]
    numeric = "".join(str(ord(char) - 55) if char.isalpha() else char for char in rearranged)
    return int(numeric) % 97 == 1


def parse_equipment(value: Any) -> list[str]:
    parts = re.split(r"[;,|]", normalize_text(value))
    result = []
    for part in parts:
        item = part.strip()
        if item and item.lower() not in {existing.lower() for existing in result}:
            result.append(item)
    return result


def identity_key(record: Mapping[str, Any]) -> str:
    external = normalize_key(record.get("external_id"))
    if external:
        return f"external:{external}"
    return "person:" + ":".join((
        normalize_key(record.get("nombre")), normalize_key(record.get("apellidos")),
        normalize_text(record.get("fecha_nacimiento")),
    ))


def family_key(record: Mapping[str, Any]) -> str:
    candidates = (
        normalize_email(record.get("progenitor1_email")), normalize_phone(record.get("progenitor1_telefono")),
        normalize_email(record.get("progenitor2_email")), normalize_phone(record.get("progenitor2_telefono")),
        normalize_key(record.get("domicilio")),
    )
    return next((value for value in candidates if value), identity_key(record))


def merge_nonempty(existing: Mapping[str, Any], incoming: Mapping[str, Any], allowed: Iterable[str]) -> dict:
    merged = dict(existing)
    for key in allowed:
        value = incoming.get(key)
        if value not in (None, "", []):
            merged[key] = value
    return merged


def sanitize_issue(issue: Mapping[str, Any]) -> dict:
    return {
        "row": issue.get("row"), "status": issue.get("status"), "severity": issue.get("severity"),
        "code": issue.get("code"), "message": issue.get("message"),
        "conflict_id": issue.get("conflict_id"), "allowed_decisions": issue.get("allowed_decisions") or [],
    }


def _validate_archive(content: bytes) -> None:
    if len(content) > MAX_FILE_SIZE:
        raise ImportValidationError("El archivo supera el límite de 10 MB")
    if not zipfile.is_zipfile(io.BytesIO(content)):
        raise ImportValidationError("El archivo debe ser un Excel .xlsx válido")
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        names = {item.filename.lower() for item in archive.infolist()}
        if any("vbaproject" in name or "externallinks" in name for name in names):
            raise ImportValidationError("No se permiten macros ni vínculos externos")
        if sum(item.file_size for item in archive.infolist()) > MAX_UNCOMPRESSED_SIZE:
            raise ImportValidationError("El contenido descomprimido supera el límite permitido")


def parse_excel(content: bytes) -> list[dict]:
    _validate_archive(content)
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    if TEMPLATE_SHEET not in workbook.sheetnames:
        raise ImportValidationError(f"Falta la hoja obligatoria '{TEMPLATE_SHEET}'")
    sheet = workbook[TEMPLATE_SHEET]
    raw_headers = [normalize_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    normalized_headers = [normalize_key(value) for value in raw_headers]
    if "equipo_25_26" in normalized_headers:
        raise ImportValidationError("No se admite EQUIPO 25&26; utiliza EQUIPO 26&27")
    mapped_headers = [HEADER_ALIASES.get(value) for value in normalized_headers]
    missing = sorted(REQUIRED_FIELDS - {value for value in mapped_headers if value})
    if missing:
        raise ImportValidationError("Faltan columnas obligatorias: " + ", ".join(missing))
    rows = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row_number > MAX_ROWS + 1:
            raise ImportValidationError(f"El archivo supera el límite de {MAX_ROWS} registros")
        if any(cell.data_type == "f" for cell in cells):
            rows.append({"_row": row_number, "_formula_error": True})
            continue
        values = [cell.value for cell in cells]
        if not any(normalize_text(value) for value in values):
            continue
        raw = {key: value for key, value in zip(mapped_headers, values) if key}
        record = {key: normalize_text(value) for key, value in raw.items()}
        record["_row"] = row_number
        record["fecha_nacimiento"] = normalize_birthdate(raw.get("fecha_nacimiento"))
        record["tipo"] = normalize_key(raw.get("tipo") or "alta")
        record["modalidad"] = normalize_text(raw.get("modalidad")).upper()
        record["email_formulario"] = normalize_email(raw.get("email_formulario"))
        record["progenitor1_email"] = normalize_email(raw.get("progenitor1_email"))
        record["progenitor2_email"] = normalize_email(raw.get("progenitor2_email"))
        record["progenitor1_telefono"] = normalize_phone(raw.get("progenitor1_telefono"))
        record["progenitor2_telefono"] = normalize_phone(raw.get("progenitor2_telefono"))
        record["iban"] = normalize_iban(raw.get("iban"))
        record["equipamiento_items"] = parse_equipment(raw.get("equipamiento"))
        rows.append(record)
    workbook.close()
    return rows


def _player_key(player: Mapping[str, Any]) -> str:
    return str(player.get("import_identity_key") or identity_key(player))


def _contact_values(record: Mapping[str, Any]) -> set[str]:
    return {value for value in (
        normalize_email(record.get("progenitor1_email")), normalize_email(record.get("progenitor2_email")),
        normalize_phone(record.get("progenitor1_telefono")), normalize_phone(record.get("progenitor2_telefono")),
    ) if value}


def analyze_rows(rows: list[dict], season: str, existing: Mapping[str, list[dict]],
                 file_sha256: str, duplicate_file: bool = False,
                 *, allow_pending_team: bool = False,
                 allow_pending_contact: bool = False) -> dict:
    if season != SEASON:
        raise ImportValidationError(f"La temporada permitida es {SEASON}")
    players = list(existing.get("players") or [])
    families = list(existing.get("families") or [])
    teams = list(existing.get("teams") or [])
    inscriptions = list(existing.get("inscriptions") or [])
    players_by_key: dict[str, list[dict]] = {}
    for player in players:
        players_by_key.setdefault(_player_key(player), []).append(player)
    teams_by_name: dict[str, list[dict]] = {}
    for team in teams:
        teams_by_name.setdefault(normalize_key(team.get("nombre")), []).append(team)
    families_by_key: dict[str, list[dict]] = {}
    for family in families:
        families_by_key.setdefault(family_key(family), []).append(family)
    inscription_keys = {
        (item.get("import_identity_key"), item.get("temporada")): item for item in inscriptions
        if item.get("import_identity_key")
    }
    seen: set[str] = set()
    issues: list[dict] = []
    row_results: list[dict] = []
    planned_team_counts: dict[str, int] = {}
    existing_team_counts = {
        str(team.get("id")): sum(1 for player in players if player.get("equipo_id") == team.get("id"))
        for team in teams
    }

    for record in rows:
        row = record.get("_row")
        if record.get("_formula_error"):
            issue = {"row": row, "status": "error", "severity": "serious", "code": "formula_not_allowed",
                     "message": "La fila contiene fórmulas; solo se admiten valores."}
            issues.append(issue); row_results.append({**issue, "record": record}); continue
        required_fields = REQUIRED_FIELDS - {"fecha_nacimiento"}
        if allow_pending_team:
            required_fields = required_fields - {"equipo"}
        missing = [field for field in required_fields if not record.get(field)]
        if missing:
            issue = {"row": row, "status": "error", "severity": "serious", "code": "missing_required",
                     "message": "Faltan campos obligatorios: " + ", ".join(sorted(missing))}
            issues.append(issue); row_results.append({**issue, "record": record}); continue
        if not record.get("fecha_nacimiento"):
            issue = {"row": row, "status": "error", "severity": "serious", "code": "invalid_birthdate",
                     "message": "La fecha de nacimiento no es válida."}
            issues.append(issue); row_results.append({**issue, "record": record}); continue
        if record.get("modalidad") not in VALID_MODALITIES:
            issue = {"row": row, "status": "error", "severity": "serious", "code": "invalid_modality",
                     "message": "La modalidad debe ser F7 o F11."}
            issues.append(issue); row_results.append({**issue, "record": record}); continue
        if record.get("tipo") not in VALID_INSCRIPTION_TYPES:
            issue = {"row": row, "status": "error", "severity": "serious", "code": "invalid_type",
                     "message": "El tipo debe ser alta o renovación."}
            issues.append(issue); row_results.append({**issue, "record": record}); continue
        invalid_emails = [record.get(field) for field in ("email_formulario", "progenitor1_email", "progenitor2_email")
                          if record.get(field) and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", record.get(field))]
        if invalid_emails and not allow_pending_contact:
            issue = {"row": row, "status": "error", "severity": "serious", "code": "invalid_email",
                     "message": "Hay una dirección de correo con formato no válido."}
            issues.append(issue); row_results.append({**issue, "record": record}); continue
        if invalid_emails and allow_pending_contact:
            for field in ("email_formulario", "progenitor1_email", "progenitor2_email"):
                if record.get(field) and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", record.get(field)):
                    record[field] = ""
            issues.append({"row": row, "status": "warning", "severity": "warning", "code": "invalid_email_pending",
                           "message": "Un correo queda pendiente de revisión y no se usará para comunicaciones."})
        if record.get("iban") and not valid_iban(record["iban"]):
            issue = {"row": row, "status": "error", "severity": "serious", "code": "invalid_iban",
                     "message": "El IBAN no supera la validación."}
            issues.append(issue); row_results.append({**issue, "record": record}); continue
        key = identity_key(record)
        record["import_identity_key"] = key
        if key in seen:
            result = {"row": row, "status": "duplicate", "severity": "warning", "code": "duplicate_in_file",
                      "message": "Registro duplicado dentro del archivo.", "record": record}
            issues.append(result); row_results.append(result); continue
        seen.add(key)
        matches = players_by_key.get(key, [])
        team_matches = teams_by_name.get(normalize_key(record.get("equipo")), []) if record.get("equipo") else []
        conflict_message = None
        if len(matches) > 1:
            conflict_message = "Hay más de un jugador existente con la misma identidad."
        elif len(team_matches) > 1:
            conflict_message = "Hay más de un equipo existente con el mismo nombre."
        elif len(families_by_key.get(family_key(record), [])) > 1:
            conflict_message = "Hay más de una familia existente para la misma referencia y requiere revisión."
        elif team_matches:
            team = team_matches[0]
            if team.get("categoria") and normalize_key(team.get("categoria")) != normalize_key(record.get("categoria")):
                conflict_message = "La categoría no coincide con la del equipo existente."
            if team.get("modalidad") and str(team.get("modalidad")).upper() != record.get("modalidad"):
                conflict_message = "La modalidad no coincide con la del equipo existente."
        if conflict_message:
            conflict_id = f"row:{row}"
            result = {"row": row, "status": "conflict", "severity": "conflict", "code": "manual_decision",
                      "message": conflict_message, "conflict_id": conflict_id, "allowed_decisions": ["skip"],
                      "record": record}
            issues.append(result); row_results.append(result); continue
        existing_player = matches[0] if matches else None
        if existing_player:
            merged = merge_nonempty(existing_player, record, SAFE_PLAYER_FIELDS)
            changed = any(merged.get(field) != existing_player.get(field) for field in SAFE_PLAYER_FIELDS)
            has_inscription = (key, season) in inscription_keys
            status = "update" if changed or not has_inscription else "unchanged"
            code = "existing_updated" if status == "update" else "no_changes"
            message = "Se actualizarán solo celdas con valor." if status == "update" else "No hay cambios respecto a la base actual."
        else:
            status, code, message = "create", "new_inscription", "Nueva inscripción."
        result = {"row": row, "status": status, "severity": "info", "code": code, "message": message,
                  "record": record}
        row_results.append(result)
        team_key = normalize_key(record.get("equipo"))
        if team_key:
            planned_team_counts[team_key] = planned_team_counts.get(team_key, 0) + (0 if existing_player else 1)
        if normalize_key(record.get("pendiente_octubre")) in {"si", "yes", "true", "pendiente", "x"}:
            issues.append({"row": row, "status": "warning", "severity": "warning", "code": "october_manual",
                           "message": "Pendiente de octubre: no se ha seleccionado ni deducido automáticamente."})

    for team_key, added in planned_team_counts.items():
        team = (teams_by_name.get(team_key) or [None])[0]
        modality = next((result["record"].get("modalidad") for result in row_results
                         if normalize_key(result.get("record", {}).get("equipo")) == team_key), None)
        limit = 18 if modality == "F7" else 25
        total = added + (existing_team_counts.get(str(team.get("id")), 0) if team else 0)
        if total > limit:
            issues.append({"row": None, "status": "warning", "severity": "warning", "code": "team_capacity",
                           "message": f"Un equipo {modality} alcanzaría {total} jugadores; el aviso comienza por encima de {limit}."})

    summary = {key: sum(1 for item in row_results if item.get("status") == key)
               for key in ("create", "update", "duplicate", "conflict", "error", "unchanged")}
    unique_count = len({identity_key(result["record"]) for result in row_results
                        if result.get("status") not in {"duplicate", "error"} and result.get("record")})
    if unique_count != 323:
        issues.append({"row": None, "status": "warning", "severity": "warning", "code": "expected_count",
                       "message": f"Se esperan 323 inscripciones únicas tras la revisión; el análisis contiene {unique_count}."})
    if duplicate_file:
        issues.append({"row": None, "status": "error", "severity": "serious", "code": "duplicate_file",
                       "message": "Este mismo archivo ya fue importado para la temporada seleccionada."})
    public_rows = [sanitize_issue(item) for item in row_results]
    public_issues = [sanitize_issue(item) for item in issues]
    return {
        "season": season, "file_sha256": file_sha256, "rows": row_results,
        "public_rows": public_rows, "issues": public_issues, "summary": summary,
        "unique_count": unique_count,
        "blocking_errors": sum(1 for item in issues if item.get("severity") == "serious"),
        "unresolved_conflicts": summary["conflict"], "duplicate_file": duplicate_file,
    }


def file_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _fernet(secret: str) -> Fernet:
    key = base64.urlsafe_b64encode(hashlib.sha256(secret.encode("utf-8")).digest())
    return Fernet(key)


def encode_plan(payload: Mapping[str, Any], secret: str) -> str:
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return _fernet(secret).encrypt(zlib.compress(raw)).decode("ascii")


def decode_plan(token: str, secret: str, max_age_seconds: int = 3600) -> dict:
    try:
        raw = _fernet(secret).decrypt(token.encode("ascii"), ttl=max_age_seconds)
        return json.loads(zlib.decompress(raw))
    except (InvalidToken, ValueError, json.JSONDecodeError, zlib.error) as exc:
        raise ImportValidationError("El análisis ha caducado o no es válido") from exc


def encrypt_iban(iban: str, secret: str) -> dict:
    normalized = normalize_iban(iban)
    if not normalized:
        return {}
    return {"iban_encrypted": _fernet(secret).encrypt(normalized.encode()).decode("ascii"),
            "iban_last4": normalized[-4:]}


def masked_iban(last4: Any) -> Optional[str]:
    value = normalize_text(last4)
    return f"ES•• •••• •••• •••• •••• {value}" if value else None
